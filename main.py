from PyQt6 import uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QListWidgetItem, QAbstractItemView
from PyQt6.QtCore import QSettings, QTimer, Qt
from PyQt6.QtGui import QIcon

import re
from PyQt6.QtWidgets import (
    QDialog, QFormLayout, QLineEdit, QComboBox, QLabel,
    QDialogButtonBox, QPushButton, QHBoxLayout, QCheckBox
)

import csv, json



from pathlib import Path
import sys
import os
import numpy as np
import pandas as pd

from core.state import AppState
from core import abf_io, filters
from core.plotting import PlotHost
from core import stim as stimdet   # stim detection
from core import peaks as peakdet   # peak detection
from core import metrics  # calculation of breath metrics

# Import version
from version_info import VERSION_STRING


ORG = "PlethApp"
APP = "PlethApp"

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        ui_file = Path(__file__).parent / "ui" / "pleth_app_layout_02_horizontal.ui"
        uic.loadUi(ui_file, self)

        # icon_path = Path(__file__).parent / "assets" / "plethapp_thumbnail_light_02.ico"
        icon_path = Path(__file__).parent / "assets" / "plethapp_thumbnail_dark_round.ico"
        self.setWindowIcon(QIcon(str(icon_path)))
        # after uic.loadUi(ui_file, self)
        from PyQt6.QtWidgets import QWidget
        for w in self.findChildren(QWidget):
            if w.property("startHidden") is True:
                w.hide()

        self.setWindowTitle(f"PlethAnalysis v{VERSION_STRING}")

        self.settings = QSettings(ORG, APP)
        self.state = AppState()
        self.single_panel_mode = False  # flips True after stim channel selection
        # store peaks per sweep
        self.state.peaks_by_sweep = {}
        self.state.breath_by_sweep = {}

        # Y2 plotting
        self.state.y2_metric_key = None
        self.state.y2_values_by_sweep = {}

        # Notch filter parameters
        self.notch_filter_lower = None
        self.notch_filter_upper = None

        # Filter order
        self.filter_order = 4  # Default Butterworth filter order

        # Outlier detection metrics (default set)
        self.outlier_metrics = ["if", "amp_insp", "amp_exp", "ti", "te", "area_insp", "area_exp"]

        # Eupnea detection parameters
        self.eupnea_min_duration = 2.0  # seconds - minimum sustained duration for eupnea region

        # --- Embed Matplotlib into MainPlot (QFrame in Designer) ---
        self.plot_host = PlotHost(self.MainPlot)
        layout = self.MainPlot.layout()
        if layout is None:
            from PyQt6.QtWidgets import QVBoxLayout
            layout = QVBoxLayout(self.MainPlot)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.plot_host)

        saved_geom = self.settings.value("geometry")
        if saved_geom:
            self.restoreGeometry(saved_geom)

        # --- Wire browse ---
        self.BrowseButton.clicked.connect(self.on_browse_clicked)

        # Add Ctrl+O shortcut - triggers different buttons based on active tab
        from PyQt6.QtGui import QShortcut, QKeySequence
        ctrl_o_shortcut = QShortcut(QKeySequence("Ctrl+O"), self)
        ctrl_o_shortcut.activated.connect(self.on_ctrl_o_pressed)

        # --- Wire channel selection (immediate application) ---
        self.AnalyzeChanSelect.currentIndexChanged.connect(self.on_analyze_channel_changed)
        self.StimChanSelect.currentIndexChanged.connect(self.on_stim_channel_changed)


        # --- Wire filter controls ---
        self._redraw_timer = QTimer(self)
        self._redraw_timer.setSingleShot(True)
        self._redraw_timer.setInterval(150)       # ms
        self._redraw_timer.timeout.connect(self.redraw_main_plot)

        # filters: commit-on-finish, not per key
        self.LowPassVal.editingFinished.connect(self.update_and_redraw)
        self.HighPassVal.editingFinished.connect(self.update_and_redraw)
        self.FilterOrderSpin.valueChanged.connect(self.update_and_redraw)

        # checkboxes toggled immediately, but we debounce the draw
        self.LowPass_checkBox.toggled.connect(self.update_and_redraw)
        self.HighPass_checkBox.toggled.connect(self.update_and_redraw)
        self.InvertSignal_checkBox.toggled.connect(self.update_and_redraw)

        # Spectral Analysis button
        self.SpectralAnalysisButton.clicked.connect(self.on_spectral_analysis_clicked)

        # Outlier Threshold button
        self.OutlierThreshButton.clicked.connect(self.on_outlier_thresh_clicked)

        # Eupnea Threshold button
        self.EupneaThreshButton.clicked.connect(self.on_eupnea_thresh_clicked)

        # --- Wire sweep navigation ---
        self.PrevSweepButton.clicked.connect(self.on_prev_sweep)
        self.NextSweepButton.clicked.connect(self.on_next_sweep)
        self.SnaptoSweepButton.clicked.connect(self.on_snap_to_sweep)
        

        # --- Wire window navigation ---
        self.PrevWindowButton.clicked.connect(self.on_prev_window)
        self.NextWindowButton.clicked.connect(self.on_next_window)
        self.SnaptoWindowButton.clicked.connect(self.on_snap_to_window)
        self.WindowRangeValue.setText("20")# Default window length
        # overlap settings for window stepping
        self._win_overlap_frac = 0.10   # 10% of the window length
        self._win_min_overlap_s = 0.50  # but at least 0.5 s overlap

        self._win_left = None# Track current window left edge (in "display time" coordinates)


        # --- Peak-detect UI wiring ---
        self.ApplyPeakFindPushButton.setEnabled(False)  # stays disabled until threshold typed
        self.ApplyPeakFindPushButton.clicked.connect(self.on_apply_peak_find_clicked)
        # --- live threshold line on the main plot ---
        self._threshold_value = None
        self._thresh_line_artists = []
        self.ThreshVal.textChanged.connect(self._on_thresh_text_changed)


        self.ThreshVal.textChanged.connect(self._maybe_enable_peak_apply)
        self.PeakPromValue.textChanged.connect(self._maybe_enable_peak_apply)
        self.MinPeakDistValue.textChanged.connect(self._maybe_enable_peak_apply)

        # Default for refractory period / min peak distance (seconds)
        self.MinPeakDistValue.setText("0.05")

        # Default values for eupnea and apnea thresholds
        self.EupneaThresh.setText("5.0")  # Hz - breathing below this is eupnea
        self.ApneaThresh.setText("0.5")   # seconds - gaps longer than this are apnea
        self.OutlierSD.setText("3.0")     # SD - standard deviations for outlier detection

        # Connect signals for eupnea/apnea/outlier threshold changes to trigger redraw
        self.EupneaThresh.textChanged.connect(self._on_region_threshold_changed)
        self.ApneaThresh.textChanged.connect(self._on_region_threshold_changed)
        self.OutlierSD.textChanged.connect(self._on_region_threshold_changed)

        # --- y2 metric dropdown (choices only; plotting later) ---
        self.y2plot_dropdown.clear()
        self.state.y2_values_by_sweep.clear()
        self.plot_host.clear_y2()

        self.y2plot_dropdown.addItem("None", userData=None)
        for label, key in metrics.METRIC_SPECS:
            self.y2plot_dropdown.addItem(label, userData=key)

        # ADD/DELETE Peak Mode: track selection in state
        self.state.y2_metric_key = None
        self.y2plot_dropdown.currentIndexChanged.connect(self.on_y2_metric_changed)

        self._add_peaks_mode = False
        self._delete_peaks_mode = False
        self._peak_edit_half_win_s = 0.08

        self.addPeaksButton.setCheckable(True)
        self.deletePeaksButton.setCheckable(True)
        self.addPeaksButton.toggled.connect(self.on_add_peaks_toggled)
        self.deletePeaksButton.toggled.connect(self.on_delete_peaks_toggled)

        # --- Sighs (manual markers) ---
        self.state.sigh_by_sweep = {}     # map: sweep_idx -> np.ndarray of PEAK indices marked as sighs
        self._add_sigh_mode = False
        self._sigh_artists = []         # matplotlib artists for sigh overlay

        # --- Omit-sweep state ---
        self.state.omitted_sweeps = set()   # set of int sweep indices

        # --- Wire omit button ---
        self.OmitSweepButton.clicked.connect(self.on_omit_sweep_clicked)


        # Button in your UI: objectName 'addSighButton'
        self.addSighButton.setCheckable(True)
        self.addSighButton.toggled.connect(self.on_add_sigh_toggled)

        # --- Move Point mode ---
        self._move_point_mode = False
        self._selected_point = None  # {'type': 'peak'|'onset'|'offset'|'exp', 'index': int, 'artist': Line2D}
        self._move_point_artist = None  # Visual marker for selected point
        self._key_press_cid = None  # Connection ID for matplotlib key events
        self._motion_cid = None  # Connection ID for mouse motion events

        # --- Sniffing regions (manual markers) ---
        self.state.sniff_regions_by_sweep = {}  # map: sweep_idx -> list of (start_time, end_time) tuples
        self._mark_sniff_mode = False
        self._sniff_start_x = None  # X-coordinate where drag started
        self._sniff_drag_artist = None  # Visual indicator while dragging
        self._sniff_artists = []  # Matplotlib artists for sniff overlays
        self._sniff_edge_mode = None  # 'start' or 'end' if dragging an edge, None if creating new region
        self._sniff_region_index = None  # Index of region being edited
        self._release_cid = None  # Connection ID for mouse release events
        self._is_dragging = False  # Track if currently dragging a point

        self.movePointButton.setCheckable(True)
        self.movePointButton.toggled.connect(self.on_move_point_toggled)

        # Mark Sniff button
        self.markSniffButton.toggled.connect(self.on_mark_sniff_toggled)

        #wire save analyzed data button
        self.SaveAnalyzedDataButton.clicked.connect(self.on_save_analyzed_clicked)

        # Wire view summary button to show PDF preview
        self.ViewSummary_pushButton.clicked.connect(self.on_view_summary_clicked)



        # Defaults: 0.5–20 Hz band, all off initially
        self.HighPassVal.setText("0.5")
        self.LowPassVal.setText("20")
        self.HighPass_checkBox.setChecked(False)
        self.LowPass_checkBox.setChecked(True)
        self.InvertSignal_checkBox.setChecked(False)

        # Push defaults into state (no-op if no data yet)
        self.update_and_redraw()
        self._refresh_omit_button_label()

        # Connect matplotlib toolbar to turn off edit modes
        self.plot_host.set_toolbar_callback(self._turn_off_all_edit_modes)


        # --- Curation tab wiring ---
        self.FilePathButton.clicked.connect(self.on_curation_choose_dir_clicked)
        # Enable multiple selection for both list widgets
        self.FileList.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.FilestoConsolidateList.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        # --- Curation: move buttons ---
        self.moveAllRight.clicked.connect(self.on_move_all_right)
        self.moveSingleRight.clicked.connect(self.on_move_selected_right)
        self.moveSingleLeft.clicked.connect(self.on_move_selected_left)
        self.moveAllLeft.clicked.connect(self.on_move_all_left)
        # Wire up search/filter for the detected files list
        self.FileListSearchBox.textChanged.connect(self._filter_file_list)
        self.FileListSearchBox.setPlaceholderText("Filter by keywords (e.g., 'gfp 2.5mW' or 'gfp, chr2')...")
        # Wire consolidate button
        self.ConsolidateSaveDataButton.clicked.connect(self.on_consolidate_save_data_clicked)
        


        # optional: keep a handle to the chosen dir
        self._curation_dir = None

    # ---------- File browse ----------
    def closeEvent(self, event):
        """Save window geometry on close."""
        self.settings.setValue("geometry", self.saveGeometry())
        super().closeEvent(event)

    def keyPressEvent(self, event):
        """Handle keyboard events for move point mode."""
        from PyQt6.QtCore import Qt

        if self._move_point_mode and self._selected_point:
            if event.key() in (Qt.Key.Key_Left, Qt.Key.Key_Right):
                # Move point left or right
                direction = -1 if event.key() == Qt.Key.Key_Left else 1
                self._move_selected_point(direction)
                event.accept()
                return
            elif event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
                # Save the moved point
                self._save_moved_point()
                event.accept()
                return
            elif event.key() == Qt.Key.Key_Escape:
                # Cancel move
                self._cancel_move_point()
                event.accept()
                return

        # Pass to parent for other key handling
        super().keyPressEvent(event)

    def _on_canvas_key_press(self, event):
        """Handle matplotlib canvas key events for move point mode."""
        if not self._move_point_mode or not self._selected_point:
            return

        if event.key in ('left', 'right'):
            # Move point left or right
            direction = -1 if event.key == 'left' else 1
            self._move_selected_point(direction)
        elif event.key in ('enter', 'return'):
            # Save the moved point
            self._save_moved_point()
        elif event.key == 'escape':
            # Cancel move
            self._cancel_move_point()

    def _on_canvas_motion(self, event):
        """Handle mouse motion for click-and-drag point movement."""
        if not self._move_point_mode:
            return

        # Toolbar already disabled when entering move mode

        if not self._selected_point:
            return

        # Only move if mouse button is pressed (dragging)
        if event.button != 1 or event.xdata is None or event.inaxes is None:
            return

        st = self.state
        s = self._selected_point['sweep']
        t, y = self._current_trace()
        if t is None or y is None:
            return

        # Get time basis
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
            t_plot = t - t0
        else:
            t_plot = t

        # Find closest sample to mouse position
        import numpy as np
        new_idx = int(np.clip(np.searchsorted(t_plot, float(event.xdata)), 0, len(t_plot) - 1))

        # Constrain movement between adjacent peaks
        new_idx = self._constrain_to_peak_boundaries(new_idx, s)

        # Update the point to new position
        self._update_point_position(new_idx, t_plot, y, s)

    def _on_canvas_release(self, event):
        """Handle mouse release - auto-save the moved point."""
        if not self._move_point_mode or not self._selected_point:
            return

        if event.button == 1:  # Left click release
            # Auto-save and recompute metrics
            self._save_moved_point(recompute_metrics=True)

    def on_ctrl_o_pressed(self):
        """Handle Ctrl+O shortcut - triggers different buttons based on active tab."""
        current_tab = self.Tabs.currentIndex()
        if current_tab == 0:  # Analysis tab
            self.on_browse_clicked()
        elif current_tab == 1:  # Curation tab
            self.on_curation_choose_dir_clicked()

    def on_browse_clicked(self):
        last_dir = self.settings.value("last_dir", str(Path.home()))
        if not Path(str(last_dir)).exists():
            last_dir = str(Path.home())

        path, _ = QFileDialog.getOpenFileName(
            self, "Select File", last_dir, "Data Files (*.abf *.smrx);;ABF Files (*.abf);;SMRX Files (*.smrx);;All Files (*.*)"
        )
        if not path:
            return
        self.settings.setValue("last_dir", str(Path(path).parent))
        self.BrowseFilePath.setText(path)
        self.load_file(Path(path))

    def load_file(self, path: Path):
        from PyQt6.QtWidgets import QProgressDialog, QApplication
        from PyQt6.QtCore import Qt

        # Determine file type for progress dialog title
        file_type = path.suffix.upper()[1:]  # .abf -> ABF, .smrx -> SMRX

        # Create progress dialog
        progress = QProgressDialog(f"Loading file...\n{path.name}", None, 0, 100, self)
        progress.setWindowTitle(f"Opening {file_type} File")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)  # Show immediately
        progress.setCancelButton(None)  # No cancel button
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()

        def update_progress(current, total, message):
            """Callback to update progress dialog."""
            progress.setValue(current)
            progress.setLabelText(f"{message}\n{path.name}")
            QApplication.processEvents()

        try:
            # Load data file (supports .abf and .smrx)
            sr, sweeps_by_ch, ch_names, t = abf_io.load_data_file(path, progress_callback=update_progress)
        except Exception as e:
            progress.close()
            QMessageBox.critical(self, "Load error", str(e))
            return
        finally:
            progress.close()


        st = self.state
        st.in_path = path
        st.sr_hz = sr
        st.sweeps = sweeps_by_ch
        st.channel_names = ch_names
        st.t = t
        st.sweep_idx = 0
        self._win_left = None

        # Reset peak results and trace cache
        if not hasattr(st, "peaks_by_sweep"):
            st.peaks_by_sweep = {}
            st.breath_by_sweep = {}
        else:
            st.peaks_by_sweep.clear()
            self.state.sigh_by_sweep.clear()
            st.breath_by_sweep.clear()
            self.state.omitted_sweeps.clear()
            self._refresh_omit_button_label()

        # Clear global trace cache when loading new file
        self._global_trace_cache = {}




        # Reset Apply button and its enable logic
        self.ApplyPeakFindPushButton.setEnabled(False)
        self._maybe_enable_peak_apply()


        


        # Fill combos safely (no signal during population)
        self.AnalyzeChanSelect.blockSignals(True)
        self.AnalyzeChanSelect.clear()
        self.AnalyzeChanSelect.addItem("All Channels")  # First option for grid view
        self.AnalyzeChanSelect.addItems(ch_names)
        self.AnalyzeChanSelect.setCurrentIndex(0)  # default = "All Channels" (grid mode)
        self.AnalyzeChanSelect.blockSignals(False)

        self.StimChanSelect.blockSignals(True)
        self.StimChanSelect.clear()
        self.StimChanSelect.addItem("None")        # default
        self.StimChanSelect.addItems(ch_names)
        self.StimChanSelect.setCurrentIndex(0)     # select "None"
        self.StimChanSelect.blockSignals(False)

        #Clear peaks
        self.state.peaks_by_sweep.clear()
        self.state.sigh_by_sweep.clear()
        self.state.breath_by_sweep.clear()

        #Clear omitted sweeps
        self.state.omitted_sweeps.clear()
        self._refresh_omit_button_label()





        # Start in grid mode (All Channels view)
        st.analyze_chan = None  # None = grid mode showing all channels
        self.single_panel_mode = False  # Start in grid mode

        # No stim selected by default
        st.stim_chan = None
        st.stim_onsets_by_sweep.clear()
        st.stim_offsets_by_sweep.clear()
        st.stim_spans_by_sweep.clear()
        st.stim_metrics_by_sweep.clear()

        # Start in multi-panel (all channels) view
        self.single_panel_mode = False
        self.plot_host.clear_saved_view("grid")  # fresh autoscale for grid
        self.plot_all_channels()

    def _proc_key(self, chan: str, sweep: int):
        st = self.state
        return (
            chan, sweep,
            st.use_low,  st.low_hz,
            st.use_high, st.high_hz,
            st.use_mean_sub, st.mean_val,
            st.use_invert,
            self.filter_order,
            self.notch_filter_lower, self.notch_filter_upper
        )

    def plot_all_channels(self):
        """Plot the current sweep for every channel, one panel per channel."""
        st = self.state
        if not st.channel_names or st.t is None:
            return
        s = max(0, min(st.sweep_idx, next(iter(st.sweeps.values())).shape[1] - 1))

        traces = []
        for ch_name in st.channel_names:
            Y = st.sweeps[ch_name]           # (n_samples, n_sweeps)
            y = Y[:, s]                      # current sweep (1D)
            # No filtering in preview - show raw data to avoid distorting stimulus channels
            traces.append((st.t, y, ch_name))

        # Adaptive downsampling: only for very long traces
        # No downsampling for recordings < 100k samples (~100 seconds at 1kHz)
        # Use 50k points for longer recordings to maintain good visual quality
        max_pts = None if len(st.t) < 100000 else 50000

        self.plot_host.show_multi_grid(
            traces,
            title=f"All channels | sweep {s+1}",
            max_points_per_trace=max_pts
        )

    def on_analyze_channel_changed(self, idx: int):
        """Apply analyze channel selection immediately."""
        st = self.state
        if not st.channel_names:
            return

        # Check if "All Channels" was selected (idx 0)
        if idx == 0:
            # Switch to grid mode (multi-channel view)
            if self.single_panel_mode:
                self.single_panel_mode = False
                st.analyze_chan = None

                # Clear stimulus data but keep the channel selected in dropdown
                # so it will be recomputed when switching back to single channel
                st.stim_onsets_by_sweep.clear()
                st.stim_offsets_by_sweep.clear()
                st.stim_spans_by_sweep.clear()
                st.stim_metrics_by_sweep.clear()

                st.proc_cache.clear()

                # Clear saved view to force fresh autoscale for grid mode
                self.plot_host.clear_saved_view("grid")
                self.plot_host.clear_saved_view("single")

                # Switch to grid plot
                self.plot_all_channels()
        elif 0 < idx <= len(st.channel_names):
            # Switch to single channel view
            new_chan = st.channel_names[idx - 1]  # -1 because idx 0 is "All Channels"
            if new_chan != st.analyze_chan or not self.single_panel_mode:
                st.analyze_chan = new_chan
                st.proc_cache.clear()
                st.peaks_by_sweep.clear()
                st.sigh_by_sweep.clear()
                if hasattr(st, 'breath_by_sweep'):
                    st.breath_by_sweep.clear()

                # Clear sniffing regions
                if hasattr(st, 'sniff_regions_by_sweep'):
                    st.sniff_regions_by_sweep.clear()

                # Reset navigation to first sweep
                st.sweep_idx = 0
                st.window_start_s = 0.0

                # Switch to single panel mode
                if not self.single_panel_mode:
                    self.single_panel_mode = True

                # If a stimulus channel is selected, recompute it for the current sweep
                if st.stim_chan is not None:
                    self._compute_stim_for_current_sweep()

                # Clear saved view to force fresh autoscale for single mode
                self.plot_host.clear_saved_view("single")
                self.plot_host.clear_saved_view("grid")

                self.ApplyPeakFindPushButton.setEnabled(False)
                self._maybe_enable_peak_apply()
                self.redraw_main_plot()

    def on_stim_channel_changed(self, idx: int):
        """Apply stimulus channel selection immediately."""
        st = self.state
        if not st.channel_names:
            return

        # idx 0 = "None", idx 1+ = channel names
        new_stim = None if idx == 0 else st.channel_names[idx - 1]
        if new_stim != st.stim_chan:
            st.stim_chan = new_stim

            # Clear stimulus detection results
            st.stim_onsets_by_sweep.clear()
            st.stim_offsets_by_sweep.clear()
            st.stim_spans_by_sweep.clear()
            st.stim_metrics_by_sweep.clear()

            # Compute stimulus for current sweep if a channel is selected
            if new_stim is not None:
                self._compute_stim_for_current_sweep()

            # Clear saved view to force fresh autoscale when stimulus changes
            self.plot_host.clear_saved_view("single")

            st.proc_cache.clear()
            self.redraw_main_plot()


    def _compute_stim_for_current_sweep(self, thresh: float = 1.0):
        st = self.state
        if not st.stim_chan or st.stim_chan not in st.sweeps:
            return
        Y = st.sweeps[st.stim_chan]
        s = max(0, min(st.sweep_idx, Y.shape[1] - 1))
        y = Y[:, s]
        t = st.t

        on_idx, off_idx, spans_s, metrics = stimdet.detect_threshold_crossings(y, t, thresh=thresh)
        st.stim_onsets_by_sweep[s] = on_idx
        st.stim_offsets_by_sweep[s] = off_idx
        st.stim_spans_by_sweep[s] = spans_s
        st.stim_metrics_by_sweep[s] = metrics

        # Debug print
        if metrics:
            pw = metrics.get("pulse_width_s")
            dur = metrics.get("duration_s")
            hz = metrics.get("freq_hz")
            msg = f"[stim] sweep {s}: width={pw:.6f}s, duration={dur:.6f}s"
            if hz:
                msg += f", freq={hz:.3f}Hz"
            print(msg)

    # ---------- Filters & redraw ----------
    def update_and_redraw(self, *args):
        st = self.state

        # checkboxes
        st.use_low       = self.LowPass_checkBox.isChecked()
        st.use_high      = self.HighPass_checkBox.isChecked()
        # Mean subtraction is now controlled from Spectral Analysis dialog
        # st.use_mean_sub is set directly in the dialog handlers
        st.use_invert    = self.InvertSignal_checkBox.isChecked()

        # Filter order
        self.filter_order = self.FilterOrderSpin.value()


        # Peaks/breaths no longer valid if filters change
        if hasattr(self.state, "peaks_by_sweep"):
            self.state.peaks_by_sweep.clear()
        if hasattr(self.state, "breath_by_sweep"):
            self.state.breath_by_sweep.clear()
        self._maybe_enable_peak_apply()  # re-enable Apply Peak if threshold is valid

        # Peaks/breaths/y2 no longer valid if filters change
        if hasattr(self.state, "peaks_by_sweep"):
            self.state.peaks_by_sweep.clear()
        if hasattr(self.state, "breath_by_sweep"):
            self.state.breath_by_sweep.clear()
            self.state.y2_values_by_sweep.clear()
            self.plot_host.clear_y2()



        def _val_if_enabled(line, checked: bool, cast=float, default=None):
            """Return a numeric value only if the box is checked and a value exists."""
            if not checked:
                return None
            txt = line.text().strip()
            if not txt:
                return None
            try:
                return cast(txt)
            except ValueError:
                return None

        # only take values if box is checked AND entry is valid
        st.low_hz  = _val_if_enabled(self.LowPassVal, st.use_low, float, None)
        st.high_hz = _val_if_enabled(self.HighPassVal, st.use_high, float, None)
        # Mean subtraction value is now controlled from Spectral Analysis dialog
        # st.mean_win_s is set directly in the dialog handlers

        # If the checkbox is checked but the box is empty/invalid, disable that filter automatically
        if st.use_low and st.low_hz is None:
            st.use_low = False
        if st.use_high and st.high_hz is None:
            st.use_high = False
        # Mean subtraction validation is handled in Spectral Analysis dialog

        # Invalidate processed cache
        st.proc_cache.clear()

        # Debounce redraw
        self._redraw_timer.start()

    def _current_trace(self):
        """Return (t, y_proc) for analyze channel & current sweep, using cached processing."""
        st = self.state
        if not st.analyze_chan or st.analyze_chan not in st.sweeps:
            return None, None

        Y = st.sweeps[st.analyze_chan]
        s = max(0, min(st.sweep_idx, Y.shape[1] - 1))
        key = self._proc_key(st.analyze_chan, s)

        # Fast path: reuse processed data if settings didn't change
        if key in st.proc_cache:
            return st.t, st.proc_cache[key]

        # Compute once, cache, and return
        y = Y[:, s]
        y2 = filters.apply_all_1d(
            y, st.sr_hz,
            st.use_low,  st.low_hz,
            st.use_high, st.high_hz,
            st.use_mean_sub, st.mean_val,
            st.use_invert,
            order=self.filter_order
        )

        # Apply notch filter if configured
        if self.notch_filter_lower is not None and self.notch_filter_upper is not None:
            y2 = self._apply_notch_filter(y2, st.sr_hz, self.notch_filter_lower, self.notch_filter_upper)

        st.proc_cache[key] = y2
        return st.t, y2

    def redraw_main_plot(self):
        st = self.state
        if st.t is None:
            return

        if self.single_panel_mode:
            t, y = self._current_trace()
            if t is None:
                return
            s = max(0, min(st.sweep_idx, next(iter(st.sweeps.values())).shape[1] - 1))
            spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []

            # Normalize time to first stim onset if available
            if st.stim_chan and spans:
                t0 = spans[0][0]
                t_plot = t - t0
                spans_plot = [(a - t0, b - t0) for (a, b) in spans]
            else:
                t0 = 0.0
                t_plot = t
                spans_plot = spans

            # base trace
            self.plot_host.show_trace_with_spans(
                t_plot, y, spans_plot,
                title=f"{st.analyze_chan or ''} | sweep {s+1}",
                max_points=None
            )

            # Clear any existing region overlays (will be recomputed if needed)
            self.plot_host.clear_region_overlays()

            # Overlay peaks for the current sweep (if computed)
            pks = getattr(st, "peaks_by_sweep", {}).get(s, None)
            if pks is not None and len(pks):
                t_peaks = t_plot[pks]
                y_peaks = y[pks]
                self.plot_host.update_peaks(t_peaks, y_peaks, size=24)
            else:
                self.plot_host.clear_peaks()

            # ---- Sigh markers (stars) tied to PEAK indices ----
            sigh_idx = getattr(st, "sigh_by_sweep", {}).get(s, None)
            if sigh_idx is not None and len(sigh_idx):
                import numpy as np
                t_sigh = t_plot[sigh_idx]

                # add a small vertical offset (default 3% of current y-span)
                offset_frac = float(getattr(self, "_sigh_offset_frac", 0.07))
                try:
                    y_span = float(np.nanmax(y) - np.nanmin(y))
                    y_off  = offset_frac * (y_span if np.isfinite(y_span) and y_span > 0 else 1.0)
                except Exception:
                    y_off = offset_frac
                y_sigh = y[sigh_idx] + y_off

                # filled orange star, a hair bigger so it pops
                self.plot_host.update_sighs(
                    t_sigh, y_sigh,
                    size=110,
                    color="#ff9f1a",   # warm orange fill
                    edge="#a35400",    # slightly darker edge
                    filled=True
                )
            else:
                self.plot_host.clear_sighs()



            # Overlay breath markers (if computed)
            br = getattr(st, "breath_by_sweep", {}).get(s, None)
            if br:
                on_idx  = br.get("onsets",  [])
                off_idx = br.get("offsets", [])
                ex_idx  = br.get("expmins", [])
                exoff_idx= br.get("expoffs", [])

                t_on  = t_plot[on_idx]  if len(on_idx)  else None
                y_on  = y[on_idx]       if len(on_idx)  else None
                t_off = t_plot[off_idx] if len(off_idx) else None
                y_off = y[off_idx]      if len(off_idx) else None
                t_exp = t_plot[ex_idx]  if len(ex_idx)  else None
                y_exp = y[ex_idx]       if len(ex_idx)  else None
                t_exof = t_plot[exoff_idx] if len(exoff_idx) else None
                y_exof = y[exoff_idx]      if len(exoff_idx) else None

                self.plot_host.update_breath_markers(
                    t_on=t_on, y_on=y_on,
                    t_off=t_off, y_off=y_off,
                    t_exp=t_exp, y_exp=y_exp,
                    t_exoff=t_exof, y_exoff=y_exof,
                    size=36
                )
            else:
                self.plot_host.clear_breath_markers()

            # update sigh markers (if any)
            # self._update_sigh_artists(t_plot, y, s)

            # Update sniff region overlays
            self._update_sniff_artists(t_plot, s)

            # ---- Y2 metric (if selected and available for this sweep) ----
            key = getattr(st, "y2_metric_key", None)
            if key:
                arr = st.y2_values_by_sweep.get(s, None)
                if arr is not None and len(arr) == len(t):
                    # Use the same time axis (t_plot) you've just plotted
                    # label = "IF (breaths/min)" if key == "if" else key
                    label = "IF (Hz)" if key == "if" else key
                    # self.plot_host.add_or_update_y2(t_plot, arr, label=label, max_points=None)
                    self.plot_host.add_or_update_y2(t_plot, arr, label=label, color="#39FF14", max_points=None)
                    self.plot_host.fig.tight_layout()
                else:
                    self.plot_host.clear_y2()
                    self.plot_host.fig.tight_layout()
            else:
                self.plot_host.clear_y2()
                self.plot_host.fig.tight_layout()

            # ---- Automatic Region Overlays (Eupnea & Apnea) ----
            try:
                # Compute eupnea and apnea masks for this sweep
                br = getattr(st, "breath_by_sweep", {}).get(s, None)
                if br and len(t) > 100:  # Only if we have breath data and sufficient points
                    # Extract breath events
                    pks = getattr(st, "peaks_by_sweep", {}).get(s, None)
                    on_idx = br.get("onsets", [])
                    off_idx = br.get("offsets", [])
                    ex_idx = br.get("expmins", [])
                    exoff_idx = br.get("expoffs", [])

                    # Get eupnea and apnea thresholds from UI (with defaults)
                    eupnea_thresh = self._parse_float(self.EupneaThresh) or 5.0  # Hz
                    apnea_thresh = self._parse_float(self.ApneaThresh) or 0.5    # seconds

                    # Get sniff regions for this sweep
                    sniff_regions = self.state.sniff_regions_by_sweep.get(s, [])

                    # Compute eupnea regions using UI threshold (excluding sniff regions)
                    eupnea_mask = metrics.detect_eupnic_regions(
                        t, y, st.sr_hz, pks, on_idx, off_idx, ex_idx, exoff_idx,
                        freq_threshold_hz=eupnea_thresh,
                        min_duration_sec=self.eupnea_min_duration,
                        sniff_regions=sniff_regions
                    )

                    # Compute apnea regions using UI threshold
                    apnea_mask = metrics.detect_apneas(
                        t, y, st.sr_hz, pks, on_idx, off_idx, ex_idx, exoff_idx,
                        min_apnea_duration_sec=apnea_thresh
                    )

                    # Identify problematic breaths using outlier detection
                    outlier_mask = None
                    failure_mask = None
                    try:
                        from core.breath_outliers import identify_problematic_breaths
                        import numpy as np

                        # Convert indices to arrays (handle both lists and arrays)
                        peaks_arr = np.array(pks) if pks is not None and len(pks) > 0 else np.array([])
                        onsets_arr = np.array(on_idx) if on_idx is not None and len(on_idx) > 0 else np.array([])
                        offsets_arr = np.array(off_idx) if off_idx is not None and len(off_idx) > 0 else np.array([])
                        expmins_arr = np.array(ex_idx) if ex_idx is not None and len(ex_idx) > 0 else np.array([])
                        expoffs_arr = np.array(exoff_idx) if exoff_idx is not None and len(exoff_idx) > 0 else np.array([])

                        # Get all computed metrics for this sweep
                        metrics_dict = {}
                        for metric_key in ["if", "amp_insp", "amp_exp", "ti", "te", "area_insp", "area_exp"]:
                            if metric_key in metrics.METRICS:
                                metric_arr = metrics.METRICS[metric_key](
                                    t, y, st.sr_hz, peaks_arr, onsets_arr, offsets_arr, expmins_arr, expoffs_arr
                                )
                                metrics_dict[metric_key] = metric_arr

                        # Get outlier threshold from UI (with default)
                        outlier_sd = self._parse_float(self.OutlierSD) or 3.0  # SD

                        # Identify problematic breaths (returns separate masks for outliers and failures)
                        outlier_mask, failure_mask = identify_problematic_breaths(
                            t, y, st.sr_hz, peaks_arr, onsets_arr, offsets_arr,
                            expmins_arr, expoffs_arr, metrics_dict, outlier_threshold=outlier_sd,
                            outlier_metrics=self.outlier_metrics
                        )

                    except Exception as outlier_error:
                        print(f"Warning: Could not detect breath outliers: {outlier_error}")
                        import traceback
                        traceback.print_exc()

                    # Apply the overlays using plot time (with stim normalization if applicable)
                    # outlier_mask = orange highlighting for statistical outliers
                    # failure_mask = red highlighting for calculation failures (NaN values)
                    self.plot_host.update_region_overlays(t_plot, eupnea_mask, apnea_mask, outlier_mask, failure_mask)
                else:
                    self.plot_host.clear_region_overlays()
            except Exception as e:
                # Graceful fallback if overlay computation fails
                print(f"Warning: Could not compute region overlays: {e}")
                self.plot_host.clear_region_overlays()

            self._refresh_threshold_lines()

            # If this sweep is omitted, dim the plot and hide markers
            if s in st.omitted_sweeps:
                fig = self.plot_host.fig
                if fig and fig.axes:
                    ax = fig.axes[0]
                    # hide detail markers so the dimming reads clearly
                    self.plot_host.clear_peaks()
                    self.plot_host.clear_breath_markers()
                    self._dim_axes_for_omitted(ax, label=True)
                    self.plot_host.fig.tight_layout()
                    self.plot_host.canvas.draw_idle()


    def _maybe_enable_peak_apply(self):
        """
        Enable ApplyPeakFindPushButton if there's a numeric threshold AND we have data.
        We keep it simple: whenever user edits any peak param, allow pressing Apply again.
        """
        st = self.state
        has_data = bool(st.channel_names)
        txt = self.ThreshVal.text().strip()
        try:
            float(txt)
            ok_thresh = True
        except Exception:
            ok_thresh = False

        self.ApplyPeakFindPushButton.setEnabled(has_data and ok_thresh)

    ##################################################
    ##threshold plotting                            ##
    ##################################################
    def _on_thresh_text_changed(self, *_):
        """
        Called whenever the ThreshVal text changes.
        Parses the float (or clears if invalid) and refreshes the dashed line(s).
        """
        self._threshold_value = self._parse_float(self.ThreshVal)  # None if invalid/empty
        self._refresh_threshold_lines()

    def _on_region_threshold_changed(self, *_):
        """
        Called whenever eupnea or apnea threshold values change.
        Redraws the current sweep to update region overlays.
        """
        # Simply redraw current sweep, which will use the new threshold values
        self.redraw_main_plot()

    def _refresh_threshold_lines(self):
        """
        (Re)draw the dashed threshold line on the current visible axes.
        Called after typing in ThreshVal and after redraws that clear the axes.
        """
        fig = getattr(self.plot_host, "fig", None)
        canvas = getattr(self.plot_host, "canvas", None)
        if fig is None or canvas is None or not fig.axes:
            return

        # Remove previous lines if they exist
        for ln in getattr(self, "_thresh_line_artists", []):
            try:
                ln.remove()
            except Exception:
                pass
        self._thresh_line_artists = []

        y = getattr(self, "_threshold_value", None)
        if y is None:
            # Nothing to draw
            canvas.draw_idle()
            return

        # Decide which axes to draw on:
        # - In single-panel mode, draw only on the first axes (main plot).
        # - In grid mode, you can choose all axes or just the first; here we do first only.
        axes = fig.axes[:1] if getattr(self, "single_panel_mode", False) else fig.axes[:1]

        for ax in axes:
            # line = ax.axhline(
            #     y,
            #     linestyle=(0, (5, 5)),  # dashed
            #     linewidth=1.2,
            #     alpha=0.9,
            #     zorder=5,
            # )
            line = ax.axhline(
                y,
                color="red",              # make it red
                linewidth=1.0,            # a touch thinner (optional)
                linestyle=(0, (2, 2)),    # smaller dash pattern: 2 on, 2 off
                alpha=0.95,
                zorder=5,
            )

            self._thresh_line_artists.append(line)

        canvas.draw_idle()


    ##################################################
    ##Sweep navigation                              ##
    ##################################################
    def _num_sweeps(self) -> int:
        """Return total sweep count from the first channel (0 if no data)."""
        st = self.state
        if not st.sweeps:
            return 0
        first = next(iter(st.sweeps.values()))
        return int(first.shape[1]) if first is not None else 0

    def on_prev_sweep(self):
        st = self.state
        n = self._num_sweeps()
        if n == 0:
            return
        if st.sweep_idx > 0:
            st.sweep_idx -= 1
            # recompute stim spans for this sweep if a stim channel is selected
            if st.stim_chan:
                self._compute_stim_for_current_sweep()
            self._refresh_omit_button_label()
            self.redraw_main_plot()

    def on_next_sweep(self):
        st = self.state
        n = self._num_sweeps()
        if n == 0:
            return
        if st.sweep_idx < n - 1:
            st.sweep_idx += 1
            if st.stim_chan:
                self._compute_stim_for_current_sweep()
            self._refresh_omit_button_label()
            self.redraw_main_plot()

    def on_snap_to_sweep(self):
        # clear saved zoom so the next draw autoscales to full sweep range
        self.plot_host.clear_saved_view("single" if self.single_panel_mode else "grid")
        self._refresh_omit_button_label()
        self.redraw_main_plot()

    ##################################################
    ##Window navigation (relative to current window)##
    ##################################################
    def _parse_window_seconds(self) -> float:
        """Read WindowRangeValue (seconds). Returns a positive float, default 20."""
        try:
            val = float(self.WindowRangeValue.text().strip())
            if val > 0:
                return val
        except Exception:
            pass
        return 20.0

    def _window_step(self, W: float) -> float:
        """
        Step size when paging windows: W - overlap,
        where overlap is max(min_overlap, frac * W).
        """
        overlap = max(self._win_min_overlap_s, self._win_overlap_frac * W)
        step = max(0.0, W - overlap)
        # avoid zero step for tiny windows
        if step <= 0:
            step = 0.9 * W
        return step

    def on_snap_to_window(self):
        """Jump to start of current sweep (normalized domain if applicable)."""
        t = self._current_t_plot()
        if t is None or t.size == 0:
            return
        W = self._parse_window_seconds()
        left = float(t[0])
        self._set_window(left=left, width=W)

    def on_next_window(self):
        """Step forward; if stepping past end, first show the last full window once,
        then on the next press hop to the first window of the next sweep."""
        t = self._current_t_plot()
        if t is None or t.size == 0:
            return

        W = self._parse_window_seconds()
        step = self._window_step(W)

        # Initialize left edge if needed
        if self._win_left is None:
            ax = self.plot_host.fig.axes[0] if self.plot_host.fig.axes else None
            self._win_left = float(ax.get_xlim()[0]) if ax else float(t[0])

        # Use an effective width that never exceeds this sweep's duration
        dur = float(t[-1] - t[0])
        W_eff = min(W, max(1e-6, dur))
        max_left = float(t[-1]) - W_eff
        eps = 1e-9

        # Normal step within this sweep?
        if self._win_left + step <= max_left + eps:
            self._set_window(left=self._win_left + step, width=W_eff)
            return

        # Not enough room for a full step:
        # 1) If we're not yet at the last full window, show it once.
        if self._win_left < max_left - eps:
            self._set_window(left=max_left, width=W_eff)
            return

        # 2) Already at last full window -> hop to next sweep if possible
        s_count = self._sweep_count()
        if self.state.sweep_idx < s_count - 1:
            self.state.sweep_idx += 1
            if self.state.stim_chan:
                self._compute_stim_for_current_sweep()
            self.redraw_main_plot()

            t2 = self._current_t_plot()
            if t2 is None or t2.size == 0:
                return
            dur2 = float(t2[-1] - t2[0])
            W_eff2 = min(W, max(1e-6, dur2))
            self._set_window(left=float(t2[0]), width=W_eff2)
        else:
            # No next sweep: stay clamped at the last full window
            self._set_window(left=max_left, width=W_eff)

    def on_prev_window(self):
        """Step backward; if stepping before start, first show the first full window once,
        then on the next press hop to the last window of the previous sweep."""
        t = self._current_t_plot()
        if t is None or t.size == 0:
            return

        W = self._parse_window_seconds()
        step = self._window_step(W)

        if self._win_left is None:
            ax = self.plot_host.fig.axes[0] if self.plot_host.fig.axes else None
            self._win_left = float(ax.get_xlim()[0]) if ax else float(t[0])

        dur = float(t[-1] - t[0])
        W_eff = min(W, max(1e-6, dur))
        min_left = float(t[0])
        eps = 1e-9

        # Normal step within this sweep?
        if self._win_left - step >= min_left - eps:
            self._set_window(left=self._win_left - step, width=W_eff)
            return

        # Not enough room for a full step:
        # 1) If we're not yet at the first full window, show it once.
        if self._win_left > min_left + eps:
            self._set_window(left=min_left, width=W_eff)
            return

        # 2) Already at first window -> hop to previous sweep if possible
        if self.state.sweep_idx > 0:
            self.state.sweep_idx -= 1
            if self.state.stim_chan:
                self._compute_stim_for_current_sweep()
            self.redraw_main_plot()

            t2 = self._current_t_plot()
            if t2 is None or t2.size == 0:
                return
            dur2 = float(t2[-1] - t2[0])
            W_eff2 = min(W, max(1e-6, dur2))
            last_left = max(float(t2[0]), float(t2[-1]) - W_eff2)
            self._set_window(left=last_left, width=W_eff2)
        else:
            # No previous sweep: stay clamped at the first window
            self._set_window(left=min_left, width=W_eff)

    def _sweep_count(self) -> int:
        st = self.state
        if not st.sweeps:
            return 0
        any_ch = next(iter(st.sweeps.values()))
        return any_ch.shape[1]

    def _current_t_plot(self):
        """Time axis exactly like the one used in redraw (normalized if stim spans exist)."""
        st = self.state
        if st.t is None:
            return None
        s = max(0, min(st.sweep_idx, self._sweep_count()-1))
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
            return st.t - t0
        return st.t

    def _set_window(self, left: float, width: float):
        """Apply x-limits and remember left edge for subsequent steps."""
        ax = self.plot_host.fig.axes[0] if self.plot_host.fig.axes else None
        if ax is None:
            return
        right = left + max(0.01, float(width))
        ax.set_xlim(left, right)
        self._win_left = float(left)
        self.plot_host.fig.tight_layout()
        self.plot_host.canvas.draw_idle()

    ##################################################
    ##Peak detection parameters                     ##
    ##################################################

    def _parse_float(self, line_edit):
        txt = line_edit.text().strip()
        if not txt:
            return None
        try:
            return float(txt)
        except ValueError:
            return None

    def _on_peak_param_changed(self, *args):
        """Enable Apply if threshold is a valid number and we have data."""
        th = self._parse_float(self.ThreshVal)
        has_data = (self.state.t is not None) and (self.state.analyze_chan in (self.state.sweeps or {}))
        self.ApplyPeakFindPushButton.setEnabled(th is not None and has_data)



    def _get_processed_for(self, chan: str, sweep_idx: int):
        """Return processed y for (channel, sweep_idx) using the same cache key logic."""
        st = self.state
        Y = st.sweeps[chan]
        s = max(0, min(sweep_idx, Y.shape[1]-1))
        key = (chan, s, st.use_low, st.low_hz, st.use_high, st.high_hz, st.use_mean_sub, st.mean_val, st.use_invert,
               self.notch_filter_lower, self.notch_filter_upper)
        if key in st.proc_cache:
            return st.proc_cache[key]
        y = Y[:, s]
        y2 = filters.apply_all_1d(
            y, st.sr_hz,
            st.use_low,  st.low_hz,
            st.use_high, st.high_hz,
            st.use_mean_sub, st.mean_val,
            st.use_invert
        )

        # Apply notch filter if configured
        if self.notch_filter_lower is not None and self.notch_filter_upper is not None:
            y2 = self._apply_notch_filter(y2, st.sr_hz, self.notch_filter_lower, self.notch_filter_upper)

        st.proc_cache[key] = y2
        return y2

    def _apply_notch_filter(self, y, sr_hz, lower_freq, upper_freq):
        """Apply a notch (band-stop) filter to remove frequencies between lower_freq and upper_freq."""
        from scipy import signal
        import numpy as np

        print(f"[notch-filter] Applying notch filter: {lower_freq:.2f} - {upper_freq:.2f} Hz (sr={sr_hz} Hz)")

        # Design a butterworth band-stop filter
        nyquist = sr_hz / 2.0
        low = lower_freq / nyquist
        high = upper_freq / nyquist

        # Ensure frequencies are in valid range (0, 1)
        low = np.clip(low, 0.001, 0.999)
        high = np.clip(high, 0.001, 0.999)

        if low >= high:
            print(f"[notch-filter] Invalid frequency range: {lower_freq}-{upper_freq} Hz")
            return y

        try:
            # Design 4th order Butterworth band-stop filter
            sos = signal.butter(4, [low, high], btype='bandstop', output='sos')
            # Apply filter (sos format is more numerically stable)
            y_filtered = signal.sosfiltfilt(sos, y)
            print(f"[notch-filter] Filter applied successfully. Signal range before: [{y.min():.3f}, {y.max():.3f}], after: [{y_filtered.min():.3f}, {y_filtered.max():.3f}]")
            return y_filtered
        except Exception as e:
            print(f"[notch-filter] Error applying filter: {e}")
            return y


    def on_apply_peak_find_clicked(self):
        """
        Run peak detection on the ANALYZE channel for ALL sweeps,
        store indices per sweep, and redraw current sweep with peaks + breath markers.
        """
        st = self.state
        if not st.channel_names or not st.analyze_chan or st.analyze_chan not in st.sweeps:
            self.ApplyPeakFindPushButton.setEnabled(False)
            return

        # Parse UI parameters
        def _num(line):
            txt = line.text().strip()
            return float(txt) if txt else None

        thresh = _num(self.ThreshVal)                 # required (button enabled only if valid)
        prom   = _num(self.PeakPromValue)             # optional
        min_d  = _num(self.MinPeakDistValue)          # seconds
        direction = "up"  # Always detect peaks above threshold for breathing signals

        min_dist_samples = None
        if min_d is not None and min_d > 0:
            min_dist_samples = max(1, int(round(min_d * st.sr_hz)))

        # Detect on ALL sweeps for the analyze channel
        any_chan = next(iter(st.sweeps.values()))
        n_sweeps = any_chan.shape[1]
        st.peaks_by_sweep.clear()
        st.breath_by_sweep.clear()
        # st.sigh_by_sweep.clear()


        for s in range(n_sweeps):
            y_proc = self._get_processed_for(st.analyze_chan, s)
            pks, breaths = peakdet.detect_peaks_and_breaths(
                y=y_proc, sr_hz=st.sr_hz,
                thresh=thresh,
                prominence=prom,
                min_dist_samples=min_dist_samples,
                direction=direction,
            )
            st.peaks_by_sweep[s] = pks
            st.breath_by_sweep[s] = breaths  # dict with 'onsets', 'offsets', 'expmins'

        # Disable until parameters change again and redraw current sweep
        self.ApplyPeakFindPushButton.setEnabled(False)
        # If a Y2 metric is selected, recompute it now that peaks/breaths changed
        if getattr(self.state, "y2_metric_key", None):
            self._compute_y2_all_sweeps()
            self.plot_host.clear_y2()

        self.redraw_main_plot()

    ##################################################
    ##y2 plotting                                   ##
    ##################################################
    def _compute_y2_all_sweeps(self):
        """Compute active y2 metric for ALL sweeps on the analyze channel."""
        st = self.state
        key = getattr(st, "y2_metric_key", None)
        if not key:
            st.y2_values_by_sweep.clear()
            return
        if key not in metrics.METRICS:
            st.y2_values_by_sweep.clear()
            return
        if st.t is None or st.analyze_chan not in st.sweeps:
            st.y2_values_by_sweep.clear()
            return

        fn = metrics.METRICS[key]
        any_ch = next(iter(st.sweeps.values()))
        n_sweeps = any_ch.shape[1]
        st.y2_values_by_sweep = {}

        for s in range(n_sweeps):
            y_proc = self._get_processed_for(st.analyze_chan, s)
            # pull peaks/breaths if available
            pks = getattr(st, "peaks_by_sweep", {}).get(s, None)
            # breaths = getattr(st, "breath_by_sweep", {}).get(s, {}) if hasattr(st, "breath_by_sweep") else {}
            breaths = getattr(st, "breath_by_sweep", {}).get(s, {})
            on = breaths.get("onsets", None)
            off = breaths.get("offsets", None)
            exm = breaths.get("expmins", None)
            exo = breaths.get("expoffs", None)

            y2 = fn(st.t, y_proc, st.sr_hz, pks, on, off, exm, exo)
            st.y2_values_by_sweep[s] = y2

    def on_y2_metric_changed(self, idx: int):
        key = self.y2plot_dropdown.itemData(idx)
        self.state.y2_metric_key = key  # None or e.g. "if"

        # Recompute Y2 (needs peaks/breaths for most metrics; IF falls back to peaks)
        self._compute_y2_all_sweeps()

        # Force a redraw of current sweep
        # Also reset Y2 axis so it rescales to new data
        self.plot_host.clear_y2()
        self.redraw_main_plot()

    ##################################################
    ## Turn Off All Edit Modes ##
    ##################################################
    def _turn_off_all_edit_modes(self):
        """Turn off all edit modes (add/delete peaks, add sigh, move point, mark sniff)."""
        # Turn off Add Peaks mode
        if getattr(self, "_add_peaks_mode", False):
            self._add_peaks_mode = False
            self.addPeaksButton.blockSignals(True)
            self.addPeaksButton.setChecked(False)
            self.addPeaksButton.blockSignals(False)
            self.addPeaksButton.setText("Add Peaks")

        # Turn off Delete Peaks mode
        if getattr(self, "_delete_peaks_mode", False):
            self._delete_peaks_mode = False
            self.deletePeaksButton.blockSignals(True)
            self.deletePeaksButton.setChecked(False)
            self.deletePeaksButton.blockSignals(False)
            self.deletePeaksButton.setText("Delete Peaks")

        # Turn off Add Sigh mode
        if getattr(self, "_add_sigh_mode", False):
            self._add_sigh_mode = False
            self.addSighButton.blockSignals(True)
            self.addSighButton.setChecked(False)
            self.addSighButton.blockSignals(False)
            self.addSighButton.setText("ADD/DEL Sigh")

        # Turn off Move Point mode
        if getattr(self, "_move_point_mode", False):
            self._move_point_mode = False
            self.movePointButton.blockSignals(True)
            self.movePointButton.setChecked(False)
            self.movePointButton.blockSignals(False)
            self.movePointButton.setText("Move Point")
            # Clean up any selected point visualization
            if self._move_point_artist:
                self._move_point_artist.remove()
                self._move_point_artist = None
            self._selected_point = None

        # Turn off Mark Sniff mode
        if getattr(self, "_mark_sniff_mode", False):
            self._mark_sniff_mode = False
            self.markSniffButton.blockSignals(True)
            self.markSniffButton.setChecked(False)
            self.markSniffButton.blockSignals(False)
            self.markSniffButton.setText("Mark Sniff")
            # Disconnect matplotlib events if connected
            if self._motion_cid is not None:
                self.plot_host.canvas.mpl_disconnect(self._motion_cid)
                self._motion_cid = None
            if self._release_cid is not None:
                self.plot_host.canvas.mpl_disconnect(self._release_cid)
                self._release_cid = None
            # Clear drag state
            if self._sniff_drag_artist:
                self._sniff_drag_artist.remove()
                self._sniff_drag_artist = None
            self._sniff_start_x = None
            self._sniff_edge_mode = None
            self._sniff_region_index = None

        # Clear click callback and reset cursor
        self.plot_host.clear_click_callback()
        self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)

    ##################################################
    ##ADD Peaks Button##
    ##################################################
    # def on_add_peaks_toggled(self, checked: bool):
    #     """Enter/exit Add Peaks mode, mutually exclusive with Delete mode."""
    #     self._add_peaks_mode = checked

    #     if checked:
    #         # Turn OFF delete mode visually and internally, without triggering its slot.
    #         if getattr(self, "_delete_peaks_mode", False):
    #             self._delete_peaks_mode = False
    #             self.deletePeaksButton.blockSignals(True)
    #             self.deletePeaksButton.setChecked(False)
    #             self.deletePeaksButton.blockSignals(False)
    #             self.deletePeaksButton.setText("Delete Peaks")

    #         self.addPeaksButton.setText("Add Peaks (ON)")
    #         self.plot_host.set_click_callback(self._on_plot_click_add_peak)
    #         self.plot_host.setCursor(Qt.CursorShape.CrossCursor)
    #     else:
    #         self.addPeaksButton.setText("Add Peaks")
    #         # Only clear callbacks/cursor if no other edit mode is active
    #         if not getattr(self, "_delete_peaks_mode", False):
    #             self.plot_host.clear_click_callback()
    #             self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)

    def on_add_peaks_toggled(self, checked: bool):
        self._add_peaks_mode = checked

        if checked:
            # Turn off matplotlib toolbar modes (zoom, pan)
            self.plot_host.turn_off_toolbar_modes()

            # turn OFF delete mode
            if getattr(self, "_delete_peaks_mode", False):
                self._delete_peaks_mode = False
                self.deletePeaksButton.blockSignals(True)
                self.deletePeaksButton.setChecked(False)
                self.deletePeaksButton.blockSignals(False)
                self.deletePeaksButton.setText("Delete Peaks")

            # turn OFF sigh mode + reset its label
            if getattr(self, "_add_sigh_mode", False):
                self._add_sigh_mode = False
                self.addSighButton.blockSignals(True)
                self.addSighButton.setChecked(False)
                self.addSighButton.blockSignals(False)
                self.addSighButton.setText("ADD/DEL Sigh")

            # turn OFF move point mode
            if getattr(self, "_move_point_mode", False):
                self._move_point_mode = False
                self.movePointButton.blockSignals(True)
                self.movePointButton.setChecked(False)
                self.movePointButton.blockSignals(False)
                self.movePointButton.setText("Move Point")

            # turn OFF mark sniff mode
            if getattr(self, "_mark_sniff_mode", False):
                self._mark_sniff_mode = False
                self.markSniffButton.blockSignals(True)
                self.markSniffButton.setChecked(False)
                self.markSniffButton.blockSignals(False)
                self.markSniffButton.setText("Mark Sniff")

            self.addPeaksButton.setText("Add Peaks (ON) [Shift=Del, Ctrl=Sigh]")
            self.plot_host.set_click_callback(self._on_plot_click_add_peak)
            self.plot_host.setCursor(Qt.CursorShape.CrossCursor)
        else:
            self.addPeaksButton.setText("Add Peaks")
            if not getattr(self, "_delete_peaks_mode", False) and not getattr(self, "_add_sigh_mode", False):
                self.plot_host.clear_click_callback()
                self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)



    def _on_plot_click_add_peak(self, xdata, ydata, event, _force_mode=None):
        # Only when "Add Peaks (ON)" or force mode is 'add'
        if _force_mode != 'add' and not getattr(self, "_add_peaks_mode", False):
            return
        if event.inaxes is None or xdata is None:
            return

        # Check for modifier keys (but not if we're already in force mode to prevent recursion)
        if _force_mode is None:
            from PyQt6.QtWidgets import QApplication
            from PyQt6.QtCore import Qt
            modifiers = QApplication.keyboardModifiers()
            shift_held = bool(modifiers & Qt.KeyboardModifier.ShiftModifier)
            ctrl_held = bool(modifiers & Qt.KeyboardModifier.ControlModifier)

            # Shift toggles to delete mode
            if shift_held:
                self._on_plot_click_delete_peak(xdata, ydata, event, _force_mode='delete')
                return

            # Ctrl switches to add sigh mode
            if ctrl_held:
                self._on_plot_click_add_sigh(xdata, ydata, event, _force_mode='sigh')
                return

        import numpy as np
        st = self.state
        if st.t is None or st.analyze_chan not in st.sweeps:
            return

        # Current sweep + processed trace (what user sees)
        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        t, y = self._current_trace()
        if t is None or y is None:
            return

        # Use the same time basis as the plot (normalized if stim)
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
            t_plot = t - t0
        else:
            t_plot = t

        # ±080 ms search window around click
        half_win_s = 0.08
        half_win_n = max(1, int(round(half_win_s * st.sr_hz)))
        i_center = int(np.clip(np.searchsorted(t_plot, float(xdata)), 0, len(t_plot) - 1))
        i0 = max(0, i_center - half_win_n)
        i1 = min(len(y) - 1, i_center + half_win_n)
        if i1 <= i0:
            return

        # Find local maximum (breathing signals always use upward peaks)
        seg = y[i0:i1 + 1]
        loc = int(np.argmax(seg))
        i_peak = i0 + loc

        # ---- NEW: enforce minimum separation from existing peaks ----
        # Use UI "MinPeakDistValue" if valid; fallback to 0.05 s
        try:
            sep_s = float(self.MinPeakDistValue.text().strip())
            if not (sep_s > 0):
                raise ValueError
        except Exception:
            sep_s = 0.05
        sep_n = max(1, int(round(sep_s * st.sr_hz)))

        pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
        if pks.size and np.any(np.abs(pks - i_peak) <= sep_n):
            print(f"[add-peak] Rejected: candidate within {sep_s:.3f}s of an existing peak.")
            return

        # Insert, sort, and store
        pks_new = np.sort(np.append(pks, i_peak))
        st.peaks_by_sweep[s] = pks_new

        # Recompute breath markers with the updated peaks
        breaths = peakdet.compute_breath_events(y, pks_new, st.sr_hz)  # uses default deadband
        st.breath_by_sweep[s] = breaths

        # Recompute Y2 metric if selected
        if getattr(st, "y2_metric_key", None):
            self._compute_y2_all_sweeps()

        # Refresh plot
        self.redraw_main_plot()

    
    # def on_delete_peaks_toggled(self, checked: bool):
    #     """Enter/exit Delete Peaks mode, mutually exclusive with Add mode."""
    #     self._delete_peaks_mode = checked

    #     if checked:
    #         # Turn OFF add mode visually and internally, without triggering its slot.
    #         if getattr(self, "_add_peaks_mode", False):
    #             self._add_peaks_mode = False
    #             self.addPeaksButton.blockSignals(True)
    #             self.addPeaksButton.setChecked(False)
    #             self.addPeaksButton.blockSignals(False)
    #             self.addPeaksButton.setText("Add Peaks")

    #         self.deletePeaksButton.setText("Delete Peaks (ON)")
    #         self.plot_host.set_click_callback(self._on_plot_click_delete_peak)
    #         self.plot_host.setCursor(Qt.CursorShape.CrossCursor)
    #     else:
    #         self.deletePeaksButton.setText("Delete Peaks")
    #         # Only clear callbacks/cursor if no other edit mode is active
    #         if not getattr(self, "_add_peaks_mode", False):
    #             self.plot_host.clear_click_callback()
    #             self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)

    def on_delete_peaks_toggled(self, checked: bool):
        self._delete_peaks_mode = checked

        if checked:
            # Turn off matplotlib toolbar modes (zoom, pan)
            self.plot_host.turn_off_toolbar_modes()

            # turn OFF add mode
            if getattr(self, "_add_peaks_mode", False):
                self._add_peaks_mode = False
                self.addPeaksButton.blockSignals(True)
                self.addPeaksButton.setChecked(False)
                self.addPeaksButton.blockSignals(False)
                self.addPeaksButton.setText("Add Peaks")

            # turn OFF sigh mode + reset its label
            if getattr(self, "_add_sigh_mode", False):
                self._add_sigh_mode = False
                self.addSighButton.blockSignals(True)
                self.addSighButton.setChecked(False)
                self.addSighButton.blockSignals(False)
                self.addSighButton.setText("ADD/DEL Sigh")

            # turn OFF move point mode
            if getattr(self, "_move_point_mode", False):
                self._move_point_mode = False
                self.movePointButton.blockSignals(True)
                self.movePointButton.setChecked(False)
                self.movePointButton.blockSignals(False)
                self.movePointButton.setText("Move Point")

            # turn OFF mark sniff mode
            if getattr(self, "_mark_sniff_mode", False):
                self._mark_sniff_mode = False
                self.markSniffButton.blockSignals(True)
                self.markSniffButton.setChecked(False)
                self.markSniffButton.blockSignals(False)
                self.markSniffButton.setText("Mark Sniff")

            self.deletePeaksButton.setText("Delete Peaks (ON) [Shift=Add, Ctrl=Sigh]")
            self.plot_host.set_click_callback(self._on_plot_click_delete_peak)
            self.plot_host.setCursor(Qt.CursorShape.CrossCursor)
        else:
            self.deletePeaksButton.setText("Delete Peaks")
            if not getattr(self, "_add_peaks_mode", False) and not getattr(self, "_add_sigh_mode", False):
                self.plot_host.clear_click_callback()
                self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)



    def _on_plot_click_delete_peak(self, xdata, ydata, event, _force_mode=None):
        # Only when "Delete Peaks (ON)" or force mode is 'delete'
        if _force_mode != 'delete' and (not getattr(self, "_delete_peaks_mode", False) or getattr(event, "button", 1) != 1):
            return
        if event.inaxes is None or xdata is None:
            return

        # Check for modifier keys (but not if we're already in force mode to prevent recursion)
        if _force_mode is None:
            from PyQt6.QtWidgets import QApplication
            from PyQt6.QtCore import Qt
            modifiers = QApplication.keyboardModifiers()
            shift_held = bool(modifiers & Qt.KeyboardModifier.ShiftModifier)
            ctrl_held = bool(modifiers & Qt.KeyboardModifier.ControlModifier)

            # Shift toggles to add mode
            if shift_held:
                self._on_plot_click_add_peak(xdata, ydata, event, _force_mode='add')
                return

            # Ctrl switches to add sigh mode
            if ctrl_held:
                self._on_plot_click_add_sigh(xdata, ydata, event, _force_mode='sigh')
                return

        st = self.state
        if st.t is None or st.analyze_chan not in st.sweeps:
            return

        # Current sweep & processed trace
        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        t, y = self._current_trace()
        if t is None or y is None:
            return

        # Match the plotted time basis (normalized if stim spans exist)
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
            t_plot = t - t0
        else:
            t_plot = t

        # Find the index corresponding to the click position
        i_click = int(np.clip(np.searchsorted(t_plot, float(xdata)), 0, len(t_plot) - 1))

        # Existing peaks for this sweep
        pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
        if pks.size == 0:
            print("[delete-peak] No peaks to delete in this sweep.")
            return

        # Find the closest peak to the click position
        distances = np.abs(pks - i_click)
        closest_idx = np.argmin(distances)
        closest_peak = pks[closest_idx]

        # Optional: Only delete if within reasonable distance (e.g., ±80ms window)
        half_win_s = float(getattr(self, "_peak_edit_half_win_s", 0.08))
        half_win_n = max(1, int(round(half_win_s * st.sr_hz)))

        if distances[closest_idx] > half_win_n:
            print(f"[delete-peak] Closest peak is too far ({distances[closest_idx]} samples > {half_win_n} samples).")
            return

        # Delete only the closest peak
        pks_new = np.delete(pks, closest_idx)
        print(f"[delete-peak] Deleted peak at index {closest_peak} (distance: {distances[closest_idx]} samples)")
        st.peaks_by_sweep[s] = pks_new

        # Recompute breaths with updated peaks (fallback to your available signature)
        try:
            breaths = peakdet.compute_breath_events(y, pks_new, st.sr_hz)
        except TypeError:
            breaths = peakdet.compute_breath_events(y, pks_new)  # older signature

        st.breath_by_sweep[s] = breaths

        # If a Y2 metric is selected, recompute
        if getattr(st, "y2_metric_key", None):
            self._compute_y2_all_sweeps()

        # Redraw
        self.redraw_main_plot()


    # def on_add_sigh_toggled(self, checked: bool):
    #     """Enter/exit Add Sigh mode; mutually exclusive with Add/Delete Peaks modes."""
    #     self._add_sigh_mode = checked

    #     if checked:
    #         # Turn OFF other edit modes visually and internally (no signal storms)
    #         if getattr(self, "_add_peaks_mode", False):
    #             self._add_peaks_mode = False
    #             self.addPeaksButton.blockSignals(True)
    #             self.addPeaksButton.setChecked(False)
    #             self.addPeaksButton.blockSignals(False)
    #             self.addPeaksButton.setText("Add Peaks")

    #         if getattr(self, "_delete_peaks_mode", False):
    #             self._delete_peaks_mode = False
    #             self.deletePeaksButton.blockSignals(True)
    #             self.deletePeaksButton.setChecked(False)
    #             self.deletePeaksButton.blockSignals(False)
    #             self.deletePeaksButton.setText("Delete Peaks")

    #         self.addSighButton.setText("Add Sigh (ON)")
    #         self.plot_host.set_click_callback(self._on_plot_click_add_sigh)
    #         self.plot_host.setCursor(Qt.CursorShape.CrossCursor)
    #     else:
    #         self.addSighButton.setText("Add Sigh")
    #         # Only clear callbacks/cursor if no other edit mode is active
    #         if not getattr(self, "_add_peaks_mode", False) and not getattr(self, "_delete_peaks_mode", False):
    #             self.plot_host.clear_click_callback()
    #             self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)


    # def _on_plot_click_add_sigh(self, xdata, ydata, event):
    #     """Add a sigh marker at the nearest breath-midpoint/peak/sample on the current sweep."""
    #     if not getattr(self, "_add_sigh_mode", False):
    #         return
    #     if event.inaxes is None or xdata is None:
    #         return

    #     import numpy as np
    #     st = self.state
    #     if st.t is None or st.analyze_chan not in st.sweeps:
    #         return

    #     # Current sweep + processed trace (what's displayed)
    #     s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
    #     t, y = self._current_trace()
    #     if t is None or y is None:
    #         return

    #     # Match the plotted time basis (stim-normalized if present)
    #     spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
    #     if st.stim_chan and spans:
    #         t0 = spans[0][0]
    #         t_plot = t - t0
    #     else:
    #         t_plot = t

    #     # Start at nearest time index to the click
    #     i_click = int(np.clip(np.searchsorted(t_plot, float(xdata)), 0, len(t_plot) - 1))

    #     # Prefer snapping to breath "midpoints" if we have breaths
    #     target_idx = i_click
    #     br = st.breath_by_sweep.get(s, None)
    #     if br is not None and "onsets" in br and len(br["onsets"]) >= 2:
    #         on = np.asarray(br["onsets"], dtype=int)
    #         mids = (on[:-1] + on[1:]) // 2
    #         j = int(np.argmin(np.abs(mids - i_click)))
    #         target_idx = int(mids[j])
    #     else:
    #         # Else snap to nearest detected peak if available
    #         pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #         if pks.size:
    #             j = int(np.argmin(np.abs(pks - i_click)))
    #             target_idx = int(pks[j])

    #     # Insert into sigh list (unique, sorted)
    #     sighs = np.asarray(st.sigh_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #     if sighs.size == 0 or np.min(np.abs(sighs - target_idx)) > 0:
    #         new_sighs = np.sort(np.append(sighs, target_idx))
    #         st.sigh_by_sweep[s] = new_sighs

    #         # Redraw overlay
    #         self.redraw_main_plot()


    # def _update_sigh_artists(self, t_plot, y, sweep_idx: int):
    #     """(Re)draw star markers for currently stored sighs (current sweep only)."""
    #     # Clear existing artists
    #     fig = getattr(self.plot_host, "fig", None)
    #     canvas = getattr(self.plot_host, "canvas", None)
    #     if fig is None or canvas is None or not fig.axes:
    #         return
    #     for art in getattr(self, "_sigh_artists", []):
    #         try:
    #             art.remove()
    #         except Exception:
    #             pass
    #     self._sigh_artists = []

    #     # Nothing to draw?
    #     s = int(sweep_idx)
    #     sighs = self.state.sigh_by_sweep.get(s, None)
    #     if sighs is None or len(sighs) == 0:
    #         canvas.draw_idle()
    #         return

    #     import numpy as np
    #     sighs = np.asarray(sighs, dtype=int)
    #     sighs = sighs[(sighs >= 0) & (sighs < len(y))]
    #     if sighs.size == 0:
    #         canvas.draw_idle()
    #         return

    #     ax = fig.axes[0]
    #     t_s = t_plot[sighs]
    #     y_s = y[sighs]

    #     # Star markers (low-profile but visible)
    #     sc = ax.scatter(
    #         t_s, y_s,
    #         marker="*",
    #         s=90,               # marker size (points^2)
    #         linewidths=0.9,
    #         edgecolors="#0d1117",
    #         facecolors="#ffcc66",
    #         zorder=7,
    #     )
    #     self._sigh_artists.append(sc)
    #     canvas.draw_idle()

    # def on_add_sigh_toggled(self, checked: bool):
    #     self._add_sigh_mode = checked

    #     # make it mutually exclusive with add/delete peaks modes
    #     if checked:
    #         if getattr(self, "_add_peaks_mode", False):
    #             self._add_peaks_mode = False
    #             self.addPeaksButton.blockSignals(True)
    #             self.addPeaksButton.setChecked(False)
    #             self.addPeaksButton.blockSignals(False)
    #             self.addPeaksButton.setText("Add Peaks")

    #         if getattr(self, "_delete_peaks_mode", False):
    #             self._delete_peaks_mode = False
    #             self.deletePeaksButton.blockSignals(True)
    #             self.deletePeaksButton.setChecked(False)
    #             self.deletePeaksButton.blockSignals(False)
    #             self.deletePeaksButton.setText("Delete Peaks")

    #         self.addSighButton.setText("Add Sigh (ON)")
    #         self.plot_host.set_click_callback(self._on_plot_click_add_sigh)
    #         self.plot_host.setCursor(Qt.CursorShape.CrossCursor)
    #     else:
    #         self.addSighButton.setText("Add Sigh")
    #         # release the click handler only if no other edit mode active
    #         if not getattr(self, "_add_peaks_mode", False) and not getattr(self, "_delete_peaks_mode", False):
    #             self.plot_host.clear_click_callback()
    #             self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)

    def _on_plot_click_add_sigh(self, xdata, ydata, event, _force_mode=None):
        # Only in sigh mode or force mode is 'sigh'
        if _force_mode != 'sigh' and (not getattr(self, "_add_sigh_mode", False) or event.inaxes is None or xdata is None):
            return
        if event.inaxes is None or xdata is None:
            return

        st = self.state
        if st.t is None or st.analyze_chan not in st.sweeps:
            return

        # Current sweep + processed trace (what you're seeing)
        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        t, y = self._current_trace()
        if t is None or y is None:
            return

        # Use same time basis as the plot (normalized if stim)
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
            t_plot = t - t0
        else:
            t_plot = t

        # We ONLY snap to existing detected peaks
        import numpy as np
        pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
        if pks.size == 0:
            print("[sigh] No peaks detected in this sweep—cannot place a sigh.")
            return

        # Find the nearest PEAK index to the click
        i_click = int(np.clip(np.searchsorted(t_plot, float(xdata)), 0, len(t_plot) - 1))
        i_nearest_peak = int(pks[np.argmin(np.abs(pks - i_click))])

        # Toggle: add if absent, remove if present
        current = set(map(int, st.sigh_by_sweep.get(s, [])))
        if i_nearest_peak in current:
            current.remove(i_nearest_peak)
            print(f"[sigh] Removed sigh at peak index {i_nearest_peak} (t={t_plot[i_nearest_peak]:.3f}s)")
        else:
            current.add(i_nearest_peak)
            print(f"[sigh] Added sigh at peak index {i_nearest_peak} (t={t_plot[i_nearest_peak]:.3f}s)")

        st.sigh_by_sweep[s] = np.array(sorted(current), dtype=int)

        # Redraw to see star(s)
        self.redraw_main_plot()

    def on_add_sigh_toggled(self, checked: bool):
        self._add_sigh_mode = checked

        if checked:
            # Turn off matplotlib toolbar modes (zoom, pan)
            self.plot_host.turn_off_toolbar_modes()

            # turn OFF other modes (without triggering their slots)
            if getattr(self, "_add_peaks_mode", False):
                self._add_peaks_mode = False
                self.addPeaksButton.blockSignals(True)
                self.addPeaksButton.setChecked(False)
                self.addPeaksButton.blockSignals(False)
                self.addPeaksButton.setText("Add Peaks")

            if getattr(self, "_delete_peaks_mode", False):
                self._delete_peaks_mode = False
                self.deletePeaksButton.blockSignals(True)
                self.deletePeaksButton.setChecked(False)
                self.deletePeaksButton.blockSignals(False)
                self.deletePeaksButton.setText("Delete Peaks")

            # turn OFF move point mode
            if getattr(self, "_move_point_mode", False):
                self._move_point_mode = False
                self.movePointButton.blockSignals(True)
                self.movePointButton.setChecked(False)
                self.movePointButton.blockSignals(False)
                self.movePointButton.setText("Move Point")

            # turn OFF mark sniff mode
            if getattr(self, "_mark_sniff_mode", False):
                self._mark_sniff_mode = False
                self.markSniffButton.blockSignals(True)
                self.markSniffButton.setChecked(False)
                self.markSniffButton.blockSignals(False)
                self.markSniffButton.setText("Mark Sniff")

            self.addSighButton.setText("Add Sigh (ON)")
            self.plot_host.set_click_callback(self._on_plot_click_add_sigh)
            self.plot_host.setCursor(Qt.CursorShape.CrossCursor)
        else:
            self.addSighButton.setText("Add Sigh")
            # only clear if no other edit mode is active
            if not getattr(self, "_add_peaks_mode", False) and not getattr(self, "_delete_peaks_mode", False):
                self.plot_host.clear_click_callback()
                self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)


    def on_move_point_toggled(self, checked: bool):
        """Enter/exit Move Point mode; mutually exclusive with other edit modes."""
        self._move_point_mode = checked

        if checked:
            # Turn off matplotlib toolbar modes (zoom, pan)
            self.plot_host.turn_off_toolbar_modes()

            # Turn OFF other edit modes
            if getattr(self, "_add_peaks_mode", False):
                self._add_peaks_mode = False
                self.addPeaksButton.blockSignals(True)
                self.addPeaksButton.setChecked(False)
                self.addPeaksButton.blockSignals(False)
                self.addPeaksButton.setText("Add Peaks")

            if getattr(self, "_delete_peaks_mode", False):
                self._delete_peaks_mode = False
                self.deletePeaksButton.blockSignals(True)
                self.deletePeaksButton.setChecked(False)
                self.deletePeaksButton.blockSignals(False)
                self.deletePeaksButton.setText("Delete Peaks")

            if getattr(self, "_add_sigh_mode", False):
                self._add_sigh_mode = False
                self.addSighButton.blockSignals(True)
                self.addSighButton.setChecked(False)
                self.addSighButton.blockSignals(False)
                self.addSighButton.setText("ADD/DEL Sigh")

            # turn OFF mark sniff mode
            if getattr(self, "_mark_sniff_mode", False):
                self._mark_sniff_mode = False
                self.markSniffButton.blockSignals(True)
                self.markSniffButton.setChecked(False)
                self.markSniffButton.blockSignals(False)
                self.markSniffButton.setText("Mark Sniff")

            self.movePointButton.setText("Move Point (ON)")
            self.plot_host.set_click_callback(self._on_plot_click_move_point)
            self.plot_host.setCursor(Qt.CursorShape.CrossCursor)

            # Connect matplotlib events
            self._key_press_cid = self.plot_host.canvas.mpl_connect('key_press_event', self._on_canvas_key_press)
            self._motion_cid = self.plot_host.canvas.mpl_connect('motion_notify_event', self._on_canvas_motion)
            self._release_cid = self.plot_host.canvas.mpl_connect('button_release_event', self._on_canvas_release)

            # Disable matplotlib's built-in toolbar - turn off any active modes
            # The toolbar has a mode attribute we can check
            if hasattr(self.plot_host.toolbar, 'mode') and self.plot_host.toolbar.mode != '':
                # There's an active mode - turn it off by calling the same method again (toggle)
                if self.plot_host.toolbar.mode == 'pan/zoom':
                    self.plot_host.toolbar.pan()
                elif self.plot_host.toolbar.mode == 'zoom rect':
                    self.plot_host.toolbar.zoom()

            self.plot_host.canvas.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
            self.plot_host.canvas.setFocus()
        else:
            self.movePointButton.setText("Move Point")

            # Disconnect matplotlib events
            if self._key_press_cid is not None:
                self.plot_host.canvas.mpl_disconnect(self._key_press_cid)
                self._key_press_cid = None
            if self._motion_cid is not None:
                self.plot_host.canvas.mpl_disconnect(self._motion_cid)
                self._motion_cid = None
            if self._release_cid is not None:
                self.plot_host.canvas.mpl_disconnect(self._release_cid)
                self._release_cid = None

            # Re-enable toolbar (user can re-select zoom/pan if they want)
            self.plot_host.canvas.toolbar.setEnabled(True)

            # Clear selected point
            self._selected_point = None
            if self._move_point_artist:
                try:
                    self._move_point_artist.remove()
                except:
                    pass
                self._move_point_artist = None
                self.plot_host.canvas.draw_idle()
            # Only clear if no other edit mode is active
            if not getattr(self, "_add_peaks_mode", False) and not getattr(self, "_delete_peaks_mode", False) and not getattr(self, "_add_sigh_mode", False):
                self.plot_host.clear_click_callback()
                self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)

    def _on_plot_click_move_point(self, xdata, ydata, event):
        """Select a point (peak/onset/offset/exp) to move."""
        if not getattr(self, "_move_point_mode", False):
            return
        if event.inaxes is None or xdata is None:
            return

        # No need to check toolbar here - it's disabled when entering move mode

        st = self.state
        if st.t is None or st.analyze_chan not in st.sweeps:
            return

        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        t, y = self._current_trace()
        if t is None or y is None:
            return

        # Get time basis (normalized if stim)
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
            t_plot = t - t0
        else:
            t_plot = t

        # Find closest point among all types (within a reasonable distance)
        import numpy as np
        pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
        breaths = st.breath_by_sweep.get(s, {})

        # Debug: print what's in the breaths dict
        print(f"[move-point-debug] breath_by_sweep keys for sweep {s}: {list(breaths.keys()) if breaths else 'None'}")

        onsets = np.asarray(breaths.get('onsets', np.array([], dtype=int)), dtype=int)
        offsets = np.asarray(breaths.get('offsets', np.array([], dtype=int)), dtype=int)
        expmins = np.asarray(breaths.get('expmins', np.array([], dtype=int)), dtype=int)
        expoffs = np.asarray(breaths.get('expoffs', np.array([], dtype=int)), dtype=int)

        # Debug: print what's available
        print(f"[move-point-debug] Available points - peaks: {pks.size}, onsets: {onsets.size}, offsets: {offsets.size}, expmins: {expmins.size}, expoffs: {expoffs.size}")

        # Find closest point (only consider points within 0.5 seconds)
        max_distance = 0.5  # seconds
        candidates = []

        if pks.size > 0:
            dists = np.abs(t_plot[pks] - xdata)
            min_idx = np.argmin(dists)
            if dists[min_idx] < max_distance:
                candidates.append(('peak', pks[min_idx], dists[min_idx]))

        if onsets.size > 0:
            dists = np.abs(t_plot[onsets] - xdata)
            min_idx = np.argmin(dists)
            if dists[min_idx] < max_distance:
                candidates.append(('onset', onsets[min_idx], dists[min_idx]))

        if offsets.size > 0:
            dists = np.abs(t_plot[offsets] - xdata)
            min_idx = np.argmin(dists)
            if dists[min_idx] < max_distance:
                candidates.append(('offset', offsets[min_idx], dists[min_idx]))

        if expmins.size > 0:
            dists = np.abs(t_plot[expmins] - xdata)
            min_idx = np.argmin(dists)
            if dists[min_idx] < max_distance:
                candidates.append(('expmin', expmins[min_idx], dists[min_idx]))

        if expoffs.size > 0:
            dists = np.abs(t_plot[expoffs] - xdata)
            min_idx = np.argmin(dists)
            if dists[min_idx] < max_distance:
                candidates.append(('expoff', expoffs[min_idx], dists[min_idx]))

        if not candidates:
            print("[move-point] No points within 0.5s of click location - click closer to a point")
            return

        # Debug: print all candidates
        print(f"[move-point-debug] Candidates within range: {[(c[0], f'{c[2]:.3f}s') for c in candidates]}")

        # Select closest
        point_type, idx, dist = min(candidates, key=lambda x: x[2])

        # Store selection (keep original_index for finding it later)
        self._selected_point = {'type': point_type, 'index': idx, 'sweep': s, 'original_index': idx}

        # Visual feedback - match point type color and make it larger with yellow outline
        if self._move_point_artist:
            try:
                self._move_point_artist.remove()
            except:
                pass

        # Define colors and markers for each point type
        color_map = {
            'peak': 'red',
            'onset': '#2ecc71',    # green
            'offset': '#f39c12',   # orange
            'expmin': '#1f78b4',   # blue
            'expoff': '#9b59b6'    # purple
        }

        ax = event.inaxes
        point_color = color_map.get(point_type, 'cyan')
        self._move_point_artist, = ax.plot([t_plot[idx]], [y[idx]], 'o',
                                           color=point_color, markersize=18,
                                           markeredgecolor='yellow', markeredgewidth=3,
                                           zorder=100)
        self.plot_host.canvas.draw_idle()
        print(f"[move-point] Selected {point_type} at index {idx}")

    def _constrain_to_peak_boundaries(self, new_idx, s):
        """Constrain point movement to stay between adjacent peaks."""
        if not self._selected_point:
            return new_idx

        import numpy as np
        st = self.state
        point_type = self._selected_point['type']
        original_idx = self._selected_point['original_index']

        # Only constrain non-peak points
        if point_type == 'peak':
            return new_idx

        # Get all peaks
        pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
        if pks.size < 2:
            return new_idx

        # Find the breath cycle this point belongs to (between which two peaks)
        peak_before = pks[pks <= original_idx]
        peak_after = pks[pks > original_idx]

        min_bound = peak_before[-1] if len(peak_before) > 0 else 0
        max_bound = peak_after[0] if len(peak_after) > 0 else len(st.sweeps[st.analyze_chan][s]) - 1

        # Constrain new_idx to be within these bounds
        return int(np.clip(new_idx, min_bound, max_bound))

    def _move_selected_point(self, direction):
        """Move the selected point left or right by 1 sample (for arrow keys)."""
        if not self._selected_point:
            return

        old_idx = self._selected_point['index']
        new_idx = old_idx + direction

        # Get trace for bounds checking
        t, y = self._current_trace()
        if t is None or y is None:
            return

        if new_idx < 0 or new_idx >= len(t):
            print("[move-point] Cannot move beyond trace bounds")
            return

        # Constrain to peak boundaries
        st = self.state
        s = self._selected_point['sweep']
        new_idx = self._constrain_to_peak_boundaries(new_idx, s)

        # Get time basis
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
            t_plot = t - t0
        else:
            t_plot = t

        self._update_point_position(new_idx, t_plot, y, s)

    def _update_point_position(self, new_idx, t_plot, y, s):
        """Update point to new index position (shared by arrow keys and drag)."""
        if not self._selected_point:
            return

        st = self.state
        point_type = self._selected_point['type']
        old_idx = self._selected_point['index']

        # Update index
        self._selected_point['index'] = new_idx

        # Update the actual data array
        import numpy as np
        if point_type == 'peak':
            pks = st.peaks_by_sweep.get(s, np.array([], dtype=int))
            if pks.size > 0:
                old_idx_estimate = self._selected_point.get('original_index', old_idx)
                distances = np.abs(pks - old_idx_estimate)
                replace_idx = np.argmin(distances)
                pks[replace_idx] = new_idx
        elif point_type in ('onset', 'offset', 'expmin', 'expoff'):
            breaths = st.breath_by_sweep.get(s, {})
            key_map = {'onset': 'onsets', 'offset': 'offsets', 'expmin': 'expmins', 'expoff': 'expoffs'}
            key = key_map[point_type]
            if key in breaths:
                arr = breaths[key]
                if arr.size > 0:
                    old_idx_estimate = self._selected_point.get('original_index', old_idx)
                    distances = np.abs(arr - old_idx_estimate)
                    replace_idx = np.argmin(distances)
                    arr[replace_idx] = new_idx

        # Update visual marker
        if self._move_point_artist:
            self._move_point_artist.set_data([t_plot[new_idx]], [y[new_idx]])

        # Update scatter plot markers
        breaths = st.breath_by_sweep.get(s, {})
        on_idx = breaths.get('onsets', np.array([], dtype=int))
        off_idx = breaths.get('offsets', np.array([], dtype=int))
        ex_idx = breaths.get('expmins', np.array([], dtype=int))
        exoff_idx = breaths.get('expoffs', np.array([], dtype=int))

        t_on = t_plot[on_idx] if len(on_idx) else None
        y_on = y[on_idx] if len(on_idx) else None
        t_off = t_plot[off_idx] if len(off_idx) else None
        y_off = y[off_idx] if len(off_idx) else None
        t_exp = t_plot[ex_idx] if len(ex_idx) else None
        y_exp = y[ex_idx] if len(ex_idx) else None
        t_exof = t_plot[exoff_idx] if len(exoff_idx) else None
        y_exof = y[exoff_idx] if len(exoff_idx) else None

        self.plot_host.update_breath_markers(
            t_on=t_on, y_on=y_on,
            t_off=t_off, y_off=y_off,
            t_exp=t_exp, y_exp=y_exp,
            t_exoff=t_exof, y_exoff=y_exof,
            size=36
        )

        # Update peaks scatter plot
        pks = st.peaks_by_sweep.get(s, np.array([], dtype=int))
        if len(pks) > 0:
            self.plot_host.update_peaks(t_plot[pks], y[pks], size=50)

        # Just update the canvas
        self.plot_host.canvas.draw_idle()

    def _save_moved_point(self, recompute_metrics=False):
        """Save the moved point position and clear selection."""
        if not self._selected_point:
            return

        # Point has already been updated during movement, just need to clear selection
        point_type = self._selected_point['type']
        new_idx = self._selected_point['index']

        print(f"[move-point] Saved {point_type} at index {new_idx}")

        # Clear selection and glowing marker
        self._selected_point = None
        if self._move_point_artist:
            try:
                self._move_point_artist.remove()
            except:
                pass
            self._move_point_artist = None

        # Recompute metrics if requested (after drag release)
        if recompute_metrics:
            # Trigger eupnea/outlier region recalculation by calling redraw
            self.redraw_main_plot()
            # Toolbar stays disabled during move mode
        else:
            # Just redraw to remove the glowing marker
            self.plot_host.canvas.draw_idle()

    def _cancel_move_point(self):
        """Cancel the point move operation and restore original position."""
        if not self._selected_point:
            return

        # Restore original position
        st = self.state
        s = self._selected_point['sweep']
        point_type = self._selected_point['type']
        original_idx = self._selected_point.get('original_index')
        current_idx = self._selected_point['index']

        if original_idx != current_idx:
            import numpy as np
            # Restore the point to its original position
            if point_type == 'peak':
                pks = st.peaks_by_sweep.get(s, np.array([], dtype=int))
                if pks.size > 0:
                    distances = np.abs(pks - current_idx)
                    replace_idx = np.argmin(distances)
                    pks[replace_idx] = original_idx
            elif point_type in ('onset', 'offset', 'expmin', 'expoff'):
                breaths = st.breath_by_sweep.get(s, {})
                key_map = {'onset': 'onsets', 'offset': 'offsets', 'expmin': 'expmins', 'expoff': 'expoffs'}
                key = key_map[point_type]
                if key in breaths:
                    arr = breaths[key]
                    if arr.size > 0:
                        distances = np.abs(arr - current_idx)
                        replace_idx = np.argmin(distances)
                        arr[replace_idx] = original_idx

        # Clear selection
        self._selected_point = None
        if self._move_point_artist:
            try:
                self._move_point_artist.remove()
            except:
                pass
            self._move_point_artist = None

        # Redraw to show restored position
        self.plot_host.canvas.draw_idle()
        print("[move-point] Move cancelled, position restored")

    # ========== Sniffing Region Marking ==========

    def on_mark_sniff_toggled(self, checked: bool):
        """Enter/exit Mark Sniff mode."""
        self._mark_sniff_mode = checked

        if checked:
            # Turn off matplotlib toolbar modes (zoom, pan)
            self.plot_host.turn_off_toolbar_modes()

            # Turn OFF other edit modes
            if getattr(self, "_add_peaks_mode", False):
                self._add_peaks_mode = False
                self.addPeaksButton.blockSignals(True)
                self.addPeaksButton.setChecked(False)
                self.addPeaksButton.blockSignals(False)
                self.addPeaksButton.setText("Add Peaks")

            if getattr(self, "_delete_peaks_mode", False):
                self._delete_peaks_mode = False
                self.deletePeaksButton.blockSignals(True)
                self.deletePeaksButton.setChecked(False)
                self.deletePeaksButton.blockSignals(False)
                self.deletePeaksButton.setText("Delete Peaks")

            if getattr(self, "_add_sigh_mode", False):
                self._add_sigh_mode = False
                self.addSighButton.blockSignals(True)
                self.addSighButton.setChecked(False)
                self.addSighButton.blockSignals(False)
                self.addSighButton.setText("ADD/DEL Sigh")

            if getattr(self, "_move_point_mode", False):
                self._move_point_mode = False
                self.movePointButton.blockSignals(True)
                self.movePointButton.setChecked(False)
                self.movePointButton.blockSignals(False)
                self.movePointButton.setText("Move Point")

            self.markSniffButton.setText("Mark Sniff (ON) [Shift=Delete]")
            self.plot_host.set_click_callback(self._on_plot_click_mark_sniff)
            self.plot_host.setCursor(Qt.CursorShape.CrossCursor)

            # Connect matplotlib events for drag functionality
            self._motion_cid = self.plot_host.canvas.mpl_connect('motion_notify_event', self._on_sniff_drag)
            self._release_cid = self.plot_host.canvas.mpl_connect('button_release_event', self._on_sniff_release)
        else:
            self.markSniffButton.setText("Mark Sniff")

            # Disconnect matplotlib events
            if self._motion_cid is not None:
                self.plot_host.canvas.mpl_disconnect(self._motion_cid)
                self._motion_cid = None
            if self._release_cid is not None:
                self.plot_host.canvas.mpl_disconnect(self._release_cid)
                self._release_cid = None

            # Clear drag artist
            if self._sniff_drag_artist:
                try:
                    self._sniff_drag_artist.remove()
                except:
                    pass
                self._sniff_drag_artist = None
                self.plot_host.canvas.draw_idle()

            # Only clear if no other edit mode is active
            if not getattr(self, "_add_peaks_mode", False) and not getattr(self, "_delete_peaks_mode", False) and not getattr(self, "_add_sigh_mode", False) and not getattr(self, "_move_point_mode", False):
                self.plot_host.clear_click_callback()
                self.plot_host.setCursor(Qt.CursorShape.ArrowCursor)

    def _on_plot_click_mark_sniff(self, xdata, ydata, event):
        """Start marking a sniffing region (click-and-drag) or grab an edge to adjust.
        Shift+click on a region to delete it."""
        if not getattr(self, "_mark_sniff_mode", False):
            return
        if event.inaxes is None or xdata is None:
            return

        # Check if Shift key is held
        from PyQt6.QtCore import Qt
        shift_held = (QApplication.keyboardModifiers() & Qt.KeyboardModifier.ShiftModifier)

        # Check if click is near an existing region edge
        st = self.state
        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        regions = self.state.sniff_regions_by_sweep.get(s, [])

        # Convert to plot time for comparison
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
        else:
            t0 = 0.0

        # SHIFT+CLICK: Delete region
        if shift_held and regions:
            for i, (start_time, end_time) in enumerate(regions):
                plot_start = start_time - t0
                plot_end = end_time - t0

                # Check if click is INSIDE this region
                if plot_start <= xdata <= plot_end:
                    # Delete this region
                    del self.state.sniff_regions_by_sweep[s][i]
                    print(f"[mark-sniff] Deleted region {i}: {start_time:.3f} - {end_time:.3f} s")
                    self.redraw_main_plot()
                    return

        # Edge detection threshold (in plot time units)
        edge_threshold = 0.3  # seconds

        # Check each region for edge proximity (for adjusting edges)
        for i, (start_time, end_time) in enumerate(regions):
            plot_start = start_time - t0
            plot_end = end_time - t0

            # Check if near start edge
            if abs(xdata - plot_start) < edge_threshold:
                self._sniff_edge_mode = 'start'
                self._sniff_region_index = i
                self._sniff_start_x = plot_end  # The other edge stays fixed
                print(f"[mark-sniff] Grabbed START edge of region {i}")
                return

            # Check if near end edge
            if abs(xdata - plot_end) < edge_threshold:
                self._sniff_edge_mode = 'end'
                self._sniff_region_index = i
                self._sniff_start_x = plot_start  # The other edge stays fixed
                print(f"[mark-sniff] Grabbed END edge of region {i}")
                return

        # Not near any edge - start creating new region
        self._sniff_edge_mode = None
        self._sniff_region_index = None
        self._sniff_start_x = xdata
        print(f"[mark-sniff] Started new region at x={xdata:.3f}")

    def _on_sniff_drag(self, event):
        """Update visual indicator while dragging to mark sniffing region."""
        if not getattr(self, "_mark_sniff_mode", False):
            return
        if self._sniff_start_x is None or event.inaxes is None or event.xdata is None:
            return

        # Get plot axes
        ax = self.plot_host.ax_main
        if ax is None:
            return

        # Remove previous drag indicator
        if self._sniff_drag_artist:
            try:
                self._sniff_drag_artist.remove()
            except:
                pass

        # Draw semi-transparent purple rectangle
        x_start = self._sniff_start_x
        x_end = event.xdata
        x_left = min(x_start, x_end)
        x_right = max(x_start, x_end)

        self._sniff_drag_artist = ax.axvspan(x_left, x_right, alpha=0.3, color='purple', zorder=10)
        self.plot_host.canvas.draw_idle()

    def _on_sniff_release(self, event):
        """Finalize the sniffing region when mouse is released."""
        if not getattr(self, "_mark_sniff_mode", False):
            return
        if self._sniff_start_x is None or event.inaxes is None or event.xdata is None:
            return

        x_start = self._sniff_start_x
        x_end = event.xdata
        x_left = min(x_start, x_end)
        x_right = max(x_start, x_end)

        # Minimum width check (avoid accidental clicks)
        if abs(x_right - x_left) < 0.1:  # Less than 0.1 seconds
            print(f"[mark-sniff] Region too small, ignoring")
            self._sniff_start_x = None
            self._sniff_edge_mode = None
            self._sniff_region_index = None
            if self._sniff_drag_artist:
                try:
                    self._sniff_drag_artist.remove()
                except:
                    pass
                self._sniff_drag_artist = None
                self.plot_host.canvas.draw_idle()
            return

        # Convert from normalized time back to actual time if needed
        st = self.state
        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []

        if st.stim_chan and spans:
            t0 = spans[0][0]
            # x_left and x_right are in normalized time, convert to actual time
            actual_start = x_left + t0
            actual_end = x_right + t0
        else:
            actual_start = x_left
            actual_end = x_right

        # Snap to nearest breath events
        actual_start, actual_end = self._snap_sniff_to_breath_events(s, actual_start, actual_end)

        # Initialize regions list if needed
        if s not in self.state.sniff_regions_by_sweep:
            self.state.sniff_regions_by_sweep[s] = []

        # Handle edge dragging vs new region creation
        if self._sniff_edge_mode is not None and self._sniff_region_index is not None:
            # Editing existing region - update it
            old_start, old_end = self.state.sniff_regions_by_sweep[s][self._sniff_region_index]
            if self._sniff_edge_mode == 'start':
                # Update start edge, keep end fixed
                self.state.sniff_regions_by_sweep[s][self._sniff_region_index] = (actual_start, old_end)
                print(f"[mark-sniff] Updated START edge of region {self._sniff_region_index}: {actual_start:.3f} - {old_end:.3f} s")
            else:  # 'end'
                # Update end edge, keep start fixed
                self.state.sniff_regions_by_sweep[s][self._sniff_region_index] = (old_start, actual_end)
                print(f"[mark-sniff] Updated END edge of region {self._sniff_region_index}: {old_start:.3f} - {actual_end:.3f} s")
        else:
            # Creating new region - add it
            self.state.sniff_regions_by_sweep[s].append((actual_start, actual_end))
            print(f"[mark-sniff] Added new sniff region: {actual_start:.3f} - {actual_end:.3f} s (sweep {s})")

        # Merge overlapping regions
        self._merge_sniff_regions(s)

        # Clear drag state
        self._sniff_start_x = None
        self._sniff_edge_mode = None
        self._sniff_region_index = None
        if self._sniff_drag_artist:
            try:
                self._sniff_drag_artist.remove()
            except:
                pass
            self._sniff_drag_artist = None

        # Redraw to show permanent overlay
        self.redraw_main_plot()

    def _snap_sniff_to_breath_events(self, sweep_idx: int, start_time: float, end_time: float):
        """Snap sniff region edges to nearest breath events.

        Left edge (start) snaps to nearest inspiratory onset.
        Right edge (end) snaps to nearest expiratory offset.
        """
        import numpy as np

        # Get current trace to convert indices to times
        t, y = self._current_trace()
        if t is None:
            print("[mark-sniff] No trace available for snapping")
            return start_time, end_time

        # Get breath events for this sweep
        breaths = self.state.breath_by_sweep.get(sweep_idx, {})
        onsets = np.asarray(breaths.get('onsets', []), dtype=int)
        expoffs = np.asarray(breaths.get('expoffs', []), dtype=int)

        snapped_start = start_time
        snapped_end = end_time

        # Snap start to nearest onset
        if onsets.size > 0:
            onset_times = t[onsets]
            distances = np.abs(onset_times - start_time)
            nearest_idx = np.argmin(distances)

            # Only snap if within reasonable distance (e.g., 1 second)
            if distances[nearest_idx] < 1.0:
                snapped_start = onset_times[nearest_idx]
                print(f"[mark-sniff] Snapped START to onset at {snapped_start:.3f}s (was {start_time:.3f}s)")

        # Snap end to nearest expiratory offset
        if expoffs.size > 0:
            expoff_times = t[expoffs]
            distances = np.abs(expoff_times - end_time)
            nearest_idx = np.argmin(distances)

            # Only snap if within reasonable distance (e.g., 1 second)
            if distances[nearest_idx] < 1.0:
                snapped_end = expoff_times[nearest_idx]
                print(f"[mark-sniff] Snapped END to expiratory offset at {snapped_end:.3f}s (was {end_time:.3f}s)")

        return snapped_start, snapped_end

    def _merge_sniff_regions(self, sweep_idx: int):
        """Merge overlapping or adjacent sniffing regions for a given sweep."""
        regions = self.state.sniff_regions_by_sweep.get(sweep_idx, [])
        if len(regions) <= 1:
            return

        # Sort regions by start time
        regions = sorted(regions, key=lambda x: x[0])

        # Merge overlapping regions
        merged = []
        current_start, current_end = regions[0]

        for start, end in regions[1:]:
            if start <= current_end:  # Overlapping or adjacent
                # Merge by extending current region
                current_end = max(current_end, end)
                print(f"[mark-sniff] Merged overlapping regions into: {current_start:.3f} - {current_end:.3f} s")
            else:
                # No overlap - save current and start new
                merged.append((current_start, current_end))
                current_start, current_end = start, end

        # Add the last region
        merged.append((current_start, current_end))

        # Update state
        self.state.sniff_regions_by_sweep[sweep_idx] = merged
        print(f"[mark-sniff] After merging: {len(merged)} region(s) on sweep {sweep_idx}")

    def _update_sniff_artists(self, t_plot, sweep_idx: int):
        """(Re)draw purple overlays for marked sniffing regions (current sweep only)."""
        # Clear existing artists
        for art in self._sniff_artists:
            try:
                art.remove()
            except:
                pass
        self._sniff_artists = []

        # Get sniff regions for this sweep
        s = int(sweep_idx)
        regions = self.state.sniff_regions_by_sweep.get(s, [])
        if not regions:
            return

        # Get time offset for normalization (if stim channel)
        st = self.state
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if st.stim_chan and spans:
            t0 = spans[0][0]
        else:
            t0 = 0.0

        # Draw each region as a semi-transparent purple rectangle
        ax = self.plot_host.ax_main
        if ax is None:
            return
        for (start_time, end_time) in regions:
            # Convert to plot time (normalized if stim)
            plot_start = start_time - t0
            plot_end = end_time - t0

            # Draw overlay
            artist = ax.axvspan(plot_start, plot_end, alpha=0.25, color='purple', zorder=5, label='Sniffing')
            self._sniff_artists.append(artist)

        self.plot_host.canvas.draw_idle()

    def on_spectral_analysis_clicked(self):
        """Open spectral analysis dialog and optionally apply notch filter."""
        st = self.state
        if st.t is None or not st.analyze_chan or st.analyze_chan not in st.sweeps:
            QMessageBox.warning(self, "Spectral Analysis", "Please load data and select an analyze channel first.")
            return

        # Get current sweep data
        t, y = self._current_trace()
        if t is None or y is None:
            QMessageBox.warning(self, "Spectral Analysis", "No data available for current sweep.")
            return

        # Get stimulation spans for current sweep if available
        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        stim_spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []

        # Open dialog
        dlg = self.SpectralAnalysisDialog(parent=self, t=t, y=y, sr_hz=st.sr_hz, stim_spans=stim_spans, parent_window=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            # Get filter parameters
            lower, upper = dlg.get_filter_params()
            print(f"[spectral-dialog] Dialog accepted. Filter params: lower={lower}, upper={upper}")

            if lower is not None and upper is not None:
                # Apply notch filter
                self.notch_filter_lower = lower
                self.notch_filter_upper = upper
                print(f"[notch-filter] Set notch filter: {lower:.2f} - {upper:.2f} Hz")

                # Clear processing cache to force recomputation with new filter
                st.proc_cache.clear()
                print(f"[notch-filter] Cleared processing cache")

                # Redraw to show filtered signal
                self.redraw_main_plot()
                print(f"[notch-filter] Main plot redrawn")

            else:
                print("[notch-filter] No filter applied (lower or upper is None)")
        else:
            print("[spectral-dialog] Dialog was not accepted (user cancelled or closed)")

    def on_outlier_thresh_clicked(self):
        """Open dialog to select which metrics to use for outlier detection."""
        # Get all available metrics from core.metrics
        from core.metrics import METRICS

        # Filter to only numeric metrics (exclude region detection functions)
        numeric_metrics = {k: v for k, v in METRICS.items()
                          if k not in ["eupnic", "apnea", "regularity"]}

        # Create and show dialog
        dlg = self.OutlierMetricsDialog(parent=self,
                                        available_metrics=list(numeric_metrics.keys()),
                                        selected_metrics=self.outlier_metrics)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            # Update selected metrics
            self.outlier_metrics = dlg.get_selected_metrics()
            print(f"[outlier-metrics] Updated outlier detection metrics: {self.outlier_metrics}")

            # Redraw to apply new outlier detection
            self.redraw_main_plot()

    def on_eupnea_thresh_clicked(self):
        """Open dialog to configure eupnea detection parameters."""
        # Create and show dialog
        dlg = self.EupneaParamsDialog(
            parent=self,
            freq_threshold=self._parse_float(self.EupneaThresh) or 5.0,
            min_duration=self.eupnea_min_duration
        )
        if dlg.exec() == QDialog.DialogCode.Accepted:
            # Get parameters from dialog
            freq_thresh, min_dur = dlg.get_params()

            # Update UI and internal state
            self.EupneaThresh.setText(str(freq_thresh))
            self.eupnea_min_duration = min_dur

            print(f"[eupnea-params] Updated: freq_threshold={freq_thresh} Hz, min_duration={min_dur} s")

            # Redraw to apply new eupnea detection
            self.redraw_main_plot()

    def _refresh_omit_button_label(self):
        """Update Omit button text based on whether current sweep is omitted."""
        s = max(0, min(self.state.sweep_idx, self._sweep_count() - 1))
        if s in self.state.omitted_sweeps:
            self.OmitSweepButton.setText("Un-omit Sweep")
            self.OmitSweepButton.setToolTip("This sweep will be excluded from saving and stats.")
        else:
            self.OmitSweepButton.setText("Omit Sweep")
            self.OmitSweepButton.setToolTip("Mark this sweep to be excluded from saving and stats.")

    def on_omit_sweep_clicked(self):
        """Toggle omission for the current sweep and refresh plot/label."""
        st = self.state
        if self._sweep_count() == 0:
            return
        s = max(0, min(st.sweep_idx, self._sweep_count() - 1))
        if s in st.omitted_sweeps:
            st.omitted_sweeps.remove(s)
            try: self.statusbar.showMessage(f"Sweep {s+1}: included", 3000)
            except Exception: pass
        else:
            st.omitted_sweeps.add(s)
            try: self.statusbar.showMessage(f"Sweep {s+1}: omitted", 3000)
            except Exception: pass

        self._refresh_omit_button_label()
        self.redraw_main_plot()

    def _dim_axes_for_omitted(self, ax, label=True):
        """Grey overlay + 'OMITTED' watermark on given axes."""
        x0, x1 = ax.get_xlim()
        ax.axvspan(x0, x1, ymin=0.0, ymax=1.0, color="#6c7382", alpha=0.22, zorder=50)
        if label:
            ax.text(0.5, 0.5, "OMITTED",
                    transform=ax.transAxes, ha="center", va="center",
                    fontsize=22, weight="bold", color="#d7dce7", alpha=0.65, zorder=60)


    ##################################################
    ##Save Data to File                             ##
    ##################################################
    def _load_save_dialog_history(self) -> dict:
        """Load autocomplete history for the Save Data dialog from QSettings."""
        history = {
            'strain': [],
            'virus': [],
            'location': [],
            'stim': [],
            'power': [],
            'animal': []
        }

        for key in history.keys():
            saved_list = self.settings.value(f"save_history/{key}", [])
            if isinstance(saved_list, str):
                # Single value, convert to list
                saved_list = [saved_list] if saved_list else []
            history[key] = saved_list if saved_list else []

        return history

    def _update_save_dialog_history(self, vals: dict):
        """Update autocomplete history with new values from the Save Data dialog."""
        max_history = 20  # Keep last 20 unique values for each field

        for key in ['strain', 'virus', 'location', 'stim', 'power', 'animal']:
            value = vals.get(key, '').strip()
            if not value:
                continue

            # Get current history
            current = self.settings.value(f"save_history/{key}", [])
            if isinstance(current, str):
                current = [current] if current else []
            elif not isinstance(current, list):
                current = []

            # Remove value if already exists (we'll re-add at front)
            if value in current:
                current.remove(value)

            # Add to front
            current.insert(0, value)

            # Trim to max_history
            current = current[:max_history]

            # Save back
            self.settings.setValue(f"save_history/{key}", current)

    def _sanitize_token(self, s: str) -> str:
        if not s:
            return ""
        s = s.strip()
        s = s.replace(" ", "_")
        # allow alnum, underscore, hyphen, dot
        s = re.sub(r"[^A-Za-z0-9._-]+", "", s)
        # squeeze repeats of underscores/hyphens
        s = re.sub(r"_+", "_", s)
        s = re.sub(r"-+", "-", s)
        return s


    ##################################################
    ##Eupnea Parameters Dialog                      ##
    ##################################################
    class EupneaParamsDialog(QDialog):
        def __init__(self, parent=None, freq_threshold=5.0, min_duration=2.0):
            from PyQt6.QtWidgets import (QVBoxLayout, QHBoxLayout, QLabel,
                                        QDoubleSpinBox, QPushButton, QGroupBox)
            from PyQt6.QtCore import Qt

            super().__init__(parent)
            self.setWindowTitle("Eupnea Detection Parameters")
            self.resize(450, 300)

            # Main layout
            main_layout = QVBoxLayout(self)

            # Title
            title = QLabel("Configure Eupnea Detection")
            title.setStyleSheet("font-size: 14pt; font-weight: bold; margin-bottom: 10px;")
            main_layout.addWidget(title)

            # Description
            desc = QLabel("Eupnea refers to normal, regular breathing patterns. "
                         "Configure the criteria used to identify eupneic regions:")
            desc.setWordWrap(True)
            desc.setStyleSheet("color: #B0B0B0; margin-bottom: 20px;")
            main_layout.addWidget(desc)

            # Parameters group
            params_group = QGroupBox("Detection Parameters")
            params_layout = QVBoxLayout()

            # Frequency threshold
            freq_layout = QHBoxLayout()
            freq_layout.addWidget(QLabel("Maximum Frequency (Hz):"))
            self.freq_spin = QDoubleSpinBox()
            self.freq_spin.setRange(0.1, 20.0)
            self.freq_spin.setValue(freq_threshold)
            self.freq_spin.setDecimals(1)
            self.freq_spin.setSingleStep(0.5)
            self.freq_spin.setToolTip("Breathing must be below this frequency to be considered eupneic")
            freq_layout.addWidget(self.freq_spin)
            freq_layout.addWidget(QLabel("(typical: 3-5 Hz)"))
            freq_layout.addStretch()
            params_layout.addLayout(freq_layout)

            # Min duration
            dur_layout = QHBoxLayout()
            dur_layout.addWidget(QLabel("Minimum Duration (s):"))
            self.dur_spin = QDoubleSpinBox()
            self.dur_spin.setRange(0.5, 10.0)
            self.dur_spin.setValue(min_duration)
            self.dur_spin.setDecimals(1)
            self.dur_spin.setSingleStep(0.5)
            self.dur_spin.setToolTip("Region must sustain these criteria for at least this long")
            dur_layout.addWidget(self.dur_spin)
            dur_layout.addWidget(QLabel("(typical: 2-3 s)"))
            dur_layout.addStretch()
            params_layout.addLayout(dur_layout)

            params_group.setLayout(params_layout)
            main_layout.addWidget(params_group)

            # Info about visual indicators
            visual_info = QLabel("Green overlay indicates detected eupneic regions on the main plot.")
            visual_info.setWordWrap(True)
            visual_info.setStyleSheet("color: #2ecc71; font-style: italic; margin-top: 15px;")
            main_layout.addWidget(visual_info)

            main_layout.addStretch()

            # Dialog buttons
            button_layout = QHBoxLayout()

            reset_btn = QPushButton("Reset to Defaults")
            reset_btn.clicked.connect(self.reset_to_defaults)
            button_layout.addWidget(reset_btn)

            button_layout.addStretch()

            cancel_btn = QPushButton("Cancel")
            cancel_btn.clicked.connect(self.reject)
            button_layout.addWidget(cancel_btn)

            ok_btn = QPushButton("OK")
            ok_btn.setDefault(True)
            ok_btn.clicked.connect(self.accept)
            button_layout.addWidget(ok_btn)

            main_layout.addLayout(button_layout)

        def reset_to_defaults(self):
            """Reset parameters to default values."""
            self.freq_spin.setValue(5.0)
            self.dur_spin.setValue(2.0)

        def get_params(self):
            """Return (freq_threshold, min_duration) tuple."""
            return (self.freq_spin.value(), self.dur_spin.value())


    ##################################################
    ##Outlier Metrics Selection Dialog              ##
    ##################################################
    class OutlierMetricsDialog(QDialog):
        def __init__(self, parent=None, available_metrics=None, selected_metrics=None):
            from PyQt6.QtWidgets import (QVBoxLayout, QHBoxLayout, QLabel, QCheckBox,
                                        QPushButton, QScrollArea, QWidget)
            from PyQt6.QtCore import Qt

            super().__init__(parent)
            self.setWindowTitle("Select Outlier Detection Metrics")
            self.resize(500, 600)

            # Store available metrics
            self.available_metrics = available_metrics or []
            self.selected_metrics = set(selected_metrics or [])

            # Metric descriptions
            self.metric_descriptions = {
                "if": "Instantaneous Frequency (Hz) - breath rate",
                "amp_insp": "Inspiratory Amplitude - peak height",
                "amp_exp": "Expiratory Amplitude - trough depth",
                "ti": "Inspiratory Time (s) - inhalation duration",
                "te": "Expiratory Time (s) - exhalation duration",
                "area_insp": "Inspiratory Area - integral during inhalation",
                "area_exp": "Expiratory Area - integral during exhalation",
                "vent_proxy": "Ventilation Proxy - breathing effort estimate",
                "d1": "D1 - duty cycle (Ti/Ttot)",
                "d2": "D2 - expiratory fraction (Te/Ttot)"
            }

            # Main layout
            main_layout = QVBoxLayout(self)

            # Title label
            title = QLabel("Select which metrics to use for outlier detection:")
            title.setStyleSheet("font-size: 12pt; font-weight: bold; margin-bottom: 10px;")
            main_layout.addWidget(title)

            # Info label
            info = QLabel("Breaths with values beyond ±N standard deviations (set in SD field) "
                         "for ANY selected metric will be flagged as outliers.")
            info.setWordWrap(True)
            info.setStyleSheet("color: #B0B0B0; margin-bottom: 15px;")
            main_layout.addWidget(info)

            # Scroll area for checkboxes
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll_content = QWidget()
            scroll_layout = QVBoxLayout(scroll_content)

            # Create checkboxes
            self.checkboxes = {}
            for metric in self.available_metrics:
                checkbox = QCheckBox(metric)
                checkbox.setChecked(metric in self.selected_metrics)

                # Add description if available
                if metric in self.metric_descriptions:
                    checkbox.setText(f"{metric} - {self.metric_descriptions[metric]}")

                checkbox.setStyleSheet("margin: 5px 0px;")
                self.checkboxes[metric] = checkbox
                scroll_layout.addWidget(checkbox)

            scroll_layout.addStretch()
            scroll.setWidget(scroll_content)
            main_layout.addWidget(scroll)

            # Quick selection buttons
            quick_buttons = QHBoxLayout()
            select_all_btn = QPushButton("Select All")
            select_all_btn.clicked.connect(self.select_all)
            quick_buttons.addWidget(select_all_btn)

            deselect_all_btn = QPushButton("Deselect All")
            deselect_all_btn.clicked.connect(self.deselect_all)
            quick_buttons.addWidget(deselect_all_btn)

            default_btn = QPushButton("Reset to Default")
            default_btn.clicked.connect(self.reset_to_default)
            quick_buttons.addWidget(default_btn)

            main_layout.addLayout(quick_buttons)

            # Dialog buttons
            button_layout = QHBoxLayout()
            button_layout.addStretch()

            cancel_btn = QPushButton("Cancel")
            cancel_btn.clicked.connect(self.reject)
            button_layout.addWidget(cancel_btn)

            ok_btn = QPushButton("OK")
            ok_btn.setDefault(True)
            ok_btn.clicked.connect(self.accept)
            button_layout.addWidget(ok_btn)

            main_layout.addLayout(button_layout)

        def select_all(self):
            """Check all metric checkboxes."""
            for checkbox in self.checkboxes.values():
                checkbox.setChecked(True)

        def deselect_all(self):
            """Uncheck all metric checkboxes."""
            for checkbox in self.checkboxes.values():
                checkbox.setChecked(False)

        def reset_to_default(self):
            """Reset to default metric selection."""
            default_metrics = ["if", "amp_insp", "amp_exp", "ti", "te", "area_insp", "area_exp"]
            for metric, checkbox in self.checkboxes.items():
                checkbox.setChecked(metric in default_metrics)

        def get_selected_metrics(self):
            """Return list of selected metric keys."""
            return [metric for metric, checkbox in self.checkboxes.items() if checkbox.isChecked()]


    ##################################################
    ##Spectral Analysis Dialog                      ##
    ##################################################
    class SpectralAnalysisDialog(QDialog):
        def __init__(self, parent=None, t=None, y=None, sr_hz=None, stim_spans=None, parent_window=None):
            from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QDoubleSpinBox, QPushButton, QMessageBox
            from PyQt6.QtCore import Qt
            import numpy as np

            super().__init__(parent)
            self.setWindowTitle("Spectral Analysis & Notch Filter")
            self.resize(1400, 900)

            self.parent_window = parent_window  # Reference to main window for sweep navigation
            self.t = t
            self.y = y
            self.sr_hz = sr_hz
            self.stim_spans = stim_spans  # List of (start, end) tuples for stimulation periods
            self.notch_lower = None
            self.notch_upper = None

            # Normalize time to stim onset if stim available
            self.t_offset = 0
            if stim_spans and len(stim_spans) > 0:
                self.t_offset = stim_spans[0][0]  # First stim onset

            # Main layout with tight spacing
            main_layout = QVBoxLayout(self)
            main_layout.setSpacing(5)
            main_layout.setContentsMargins(5, 5, 5, 5)

            # Control panel at top
            control_layout = QHBoxLayout()
            control_layout.setSpacing(5)

            # Sweep navigation controls
            if parent_window:
                self.prev_sweep_btn = QPushButton("◄ Prev Sweep")
                self.prev_sweep_btn.clicked.connect(self.on_prev_sweep)
                control_layout.addWidget(self.prev_sweep_btn)

                self.sweep_label = QLabel(f"Sweep: {getattr(parent_window.state, 'sweep_idx', 0) + 1}")
                self.sweep_label.setStyleSheet("color: black; font-size: 12pt; font-weight: bold; background-color: white; padding: 2px 8px; border-radius: 3px;")
                control_layout.addWidget(self.sweep_label)

                self.next_sweep_btn = QPushButton("Next Sweep ►")
                self.next_sweep_btn.clicked.connect(self.on_next_sweep)
                control_layout.addWidget(self.next_sweep_btn)

                control_layout.addWidget(QLabel("  |  "))  # Separator

            # Notch filter controls
            control_layout.addWidget(QLabel("Notch Filter (Hz):"))

            self.lower_freq_spin = QDoubleSpinBox()
            self.lower_freq_spin.setRange(0.0, sr_hz/2 if sr_hz else 100)
            self.lower_freq_spin.setValue(0.0)
            self.lower_freq_spin.setDecimals(2)
            self.lower_freq_spin.setSuffix(" Hz")
            control_layout.addWidget(QLabel("Lower:"))
            control_layout.addWidget(self.lower_freq_spin)

            self.upper_freq_spin = QDoubleSpinBox()
            self.upper_freq_spin.setRange(0.0, sr_hz/2 if sr_hz else 100)
            self.upper_freq_spin.setValue(0.0)
            self.upper_freq_spin.setDecimals(2)
            self.upper_freq_spin.setSuffix(" Hz")
            control_layout.addWidget(QLabel("Upper:"))
            control_layout.addWidget(self.upper_freq_spin)

            self.apply_filter_btn = QPushButton("Apply Filter")
            self.apply_filter_btn.clicked.connect(self.on_apply_filter)
            control_layout.addWidget(self.apply_filter_btn)

            self.reset_filter_btn = QPushButton("Reset Filter")
            self.reset_filter_btn.clicked.connect(self.on_reset_filter)
            control_layout.addWidget(self.reset_filter_btn)

            # Add separator
            control_layout.addWidget(QLabel("  |  "))

            # Mean Subtraction controls
            from PyQt6.QtWidgets import QCheckBox
            self.mean_subtract_cb = QCheckBox("Mean Subtraction")
            self.mean_subtract_cb.setChecked(parent_window.state.use_mean_sub if parent_window else False)
            self.mean_subtract_cb.toggled.connect(self.on_mean_subtract_toggled)
            control_layout.addWidget(self.mean_subtract_cb)

            self.mean_window_spin = QDoubleSpinBox()
            self.mean_window_spin.setRange(0.1, 100.0)
            # Use mean_val attribute which is the window size in seconds
            self.mean_window_spin.setValue(parent_window.state.mean_val if (parent_window and parent_window.state.mean_val) else 10.0)
            self.mean_window_spin.setDecimals(1)
            self.mean_window_spin.setSuffix(" s")
            self.mean_window_spin.setEnabled(self.mean_subtract_cb.isChecked())
            self.mean_window_spin.valueChanged.connect(self.on_mean_window_changed)
            control_layout.addWidget(self.mean_window_spin)

            control_layout.addStretch()
            main_layout.addLayout(control_layout)

            # Plot area with matplotlib
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
            from matplotlib.figure import Figure

            # Create figure with dark background
            self.figure = Figure(figsize=(14, 10), facecolor='#2b2b2b')
            self.canvas = FigureCanvasQTAgg(self.figure)
            self.canvas.setStyleSheet("background-color: #2b2b2b;")
            main_layout.addWidget(self.canvas)

            # Buttons at bottom
            btn_layout = QHBoxLayout()
            btn_layout.addStretch()

            self.close_btn = QPushButton("Close")
            self.close_btn.clicked.connect(self.accept)
            btn_layout.addWidget(self.close_btn)

            main_layout.addLayout(btn_layout)

            # Initial plot
            self.update_plots()

        def update_plots(self):
            """Generate power spectrum and wavelet plots."""
            import numpy as np
            from scipy import signal
            import matplotlib.pyplot as plt

            if self.y is None or len(self.y) == 0:
                return

            self.figure.clear()

            # Create subplots with aligned axes: power spectrum on top, wavelet on bottom
            # Let matplotlib handle spacing automatically with tight_layout
            from matplotlib.gridspec import GridSpec
            gs = GridSpec(2, 1, figure=self.figure)
            ax1 = self.figure.add_subplot(gs[0])
            ax2 = self.figure.add_subplot(gs[1])

            # Power Spectrum (Welch method)
            if self.sr_hz:
                # Increase resolution 5x: longer nperseg for smoother curves
                nperseg = min(163840, len(self.y)//2)  # 5x increase from 32768 to 163840
                noverlap = int(nperseg * 0.90)  # 90% overlap for smooth estimate

                # All sweeps concatenated (if parent window available)
                if self.parent_window and hasattr(self.parent_window.state, 'sweeps') and self.parent_window.state.analyze_chan:
                    try:
                        sweeps_dict = self.parent_window.state.sweeps
                        analyze_chan = self.parent_window.state.analyze_chan

                        if analyze_chan in sweeps_dict:
                            sweeps_data = sweeps_dict[analyze_chan]  # Shape: (n_samples, n_sweeps)

                            # Concatenate all sweeps
                            all_sweeps_concat = []
                            for sweep_idx in range(sweeps_data.shape[1]):
                                all_sweeps_concat.append(sweeps_data[:, sweep_idx])

                            if all_sweeps_concat:
                                concatenated = np.concatenate(all_sweeps_concat)

                                # Compute Welch PSD on concatenated data
                                freqs_all, psd_all = signal.welch(concatenated, fs=self.sr_hz, nperseg=nperseg, noverlap=noverlap)
                                mask_all = freqs_all <= 30
                                ax1.plot(freqs_all[mask_all], psd_all[mask_all], 'magenta', linewidth=2, label='All Sweeps', alpha=0.8)
                    except Exception as e:
                        print(f"[spectral] Failed to compute all-sweeps spectrum: {e}")

                # Full trace spectrum
                freqs, psd = signal.welch(self.y, fs=self.sr_hz, nperseg=nperseg, noverlap=noverlap)

                # Only plot frequencies up to 30 Hz (respiratory range)
                mask = freqs <= 30
                freqs_plot = freqs[mask]
                psd_plot = psd[mask]

                ax1.plot(freqs_plot, psd_plot, 'cyan', linewidth=2, label='Current Trace')

                # If stim spans provided, compute spectrum during and after stim
                if self.stim_spans and len(self.stim_spans) > 0:
                    # During stim: from first laser onset to last laser offset
                    first_stim_start = self.stim_spans[0][0]
                    last_stim_end = self.stim_spans[-1][1]

                    stim_start_idx = np.searchsorted(self.t, first_stim_start)
                    stim_end_idx = np.searchsorted(self.t, last_stim_end)

                    stim_data = None
                    post_stim_data = None

                    if stim_end_idx > stim_start_idx:
                        stim_data = self.y[stim_start_idx:stim_end_idx]

                    # Post-stim: everything after last laser offset
                    if stim_end_idx < len(self.y):
                        post_stim_data = self.y[stim_end_idx:]

                    # During stim spectrum (use adaptive nperseg if data is short)
                    if stim_data is not None and len(stim_data) > 256:
                        nperseg_stim = min(nperseg, len(stim_data)//2)
                        noverlap_stim = int(nperseg_stim * 0.90)
                        freqs_stim, psd_stim = signal.welch(stim_data, fs=self.sr_hz, nperseg=nperseg_stim, noverlap=noverlap_stim)
                        mask_stim = freqs_stim <= 30
                        ax1.plot(freqs_stim[mask_stim], psd_stim[mask_stim], 'orange', linewidth=2, label='During Stim', alpha=0.8)

                    # Post-stim spectrum (use adaptive nperseg if data is short)
                    if post_stim_data is not None and len(post_stim_data) > 256:
                        nperseg_post = min(nperseg, len(post_stim_data)//2)
                        noverlap_post = int(nperseg_post * 0.90)
                        freqs_post, psd_post = signal.welch(post_stim_data, fs=self.sr_hz, nperseg=nperseg_post, noverlap=noverlap_post)
                        mask_post = freqs_post <= 30
                        ax1.plot(freqs_post[mask_post], psd_post[mask_post], 'lime', linewidth=2, label='Post-Stim', alpha=0.8)

                # Add labels with padding to prevent cutoff
                ax1.set_xlabel('Frequency (Hz)', color='white', fontsize=16, fontweight='bold', labelpad=10)
                ax1.set_ylabel('Power Spectral Density', color='white', fontsize=16, fontweight='bold', labelpad=10)
                ax1.set_title('Power Spectrum (Welch Method)', color='white', fontsize=18, fontweight='bold', pad=15)
                ax1.set_xlim([0, 30])
                ax1.grid(True, alpha=0.3, color='gray', linestyle='--')
                ax1.set_facecolor('#1a1a1a')
                ax1.tick_params(colors='white', labelsize=13, width=2, length=6, pad=8)

                # Set white spines with thicker lines
                for spine in ax1.spines.values():
                    spine.set_edgecolor('white')
                    spine.set_linewidth(2)

                # Highlight notch filter region if set
                if self.notch_lower is not None and self.notch_upper is not None:
                    ax1.axvspan(self.notch_lower, self.notch_upper, alpha=0.3, color='red', label='Notch Filter')

                ax1.legend(facecolor='#2b2b2b', edgecolor='white', labelcolor='white', fontsize=11)

            # Wavelet Analysis (Continuous Wavelet Transform)
            try:
                if self.sr_hz and self.t is not None and len(self.y) > 0:
                    print(f"[wavelet] Computing CWT for {len(self.y)} samples at {self.sr_hz} Hz")

                    # Downsample for faster computation if signal is very long
                    downsample_factor = 1
                    if len(self.y) > 100000:  # If more than 100k samples
                        downsample_factor = max(1, len(self.y) // 50000)
                        y_ds = self.y[::downsample_factor]
                        t_ds = self.t[::downsample_factor]
                        print(f"[wavelet] Downsampling by factor {downsample_factor} to {len(y_ds)} samples")
                    else:
                        y_ds = self.y
                        t_ds = self.t

                    # Create frequency array from 0.5 Hz to 30 Hz (fewer frequencies for speed)
                    frequencies = np.linspace(0.5, 30, 50)  # Restricted to respiratory range

                    # Compute CWT using FFT-based convolution for speed
                    cwtmatr = np.zeros((len(frequencies), len(y_ds)))

                    for i, freq in enumerate(frequencies):
                        # Create Complex Morlet wavelet for this frequency
                        w = 6.0  # Standard Morlet parameter
                        sigma = w / (2 * np.pi * freq)  # Time domain width

                        # Limit wavelet length for speed
                        max_wavelet_samples = min(int(10 * sigma * self.sr_hz / downsample_factor), len(y_ds) // 2)
                        wavelet_time = np.arange(-max_wavelet_samples, max_wavelet_samples) / (self.sr_hz / downsample_factor)

                        # Complex Morlet wavelet (optimal for oscillatory respiratory signals)
                        wavelet = np.exp(2j * np.pi * freq * wavelet_time) * np.exp(-wavelet_time**2 / (2 * sigma**2))
                        wavelet = wavelet / np.sqrt(sigma * np.sqrt(np.pi))  # Normalize

                        # FFT-based convolution (much faster)
                        convolved = signal.fftconvolve(y_ds, wavelet, mode='same')
                        cwtmatr[i, :] = np.abs(convolved)

                    print(f"[wavelet] CWT matrix shape: {cwtmatr.shape}, min={cwtmatr.min():.2e}, max={cwtmatr.max():.2e}")

                    # Use percentile-based color scaling to handle bright transients (like sniffing bouts)
                    # This prevents one bright spot from washing out the rest of the signal
                    vmin = 0
                    vmax = np.percentile(cwtmatr, 95)  # Use 95th percentile instead of max
                    print(f"[wavelet] Color scale: vmin={vmin:.2e}, vmax (95th percentile)={vmax:.2e}, actual max={cwtmatr.max():.2e}")

                    # Plot scalogram (use downsampled time array, normalized to stim onset)
                    t_plot_start = t_ds[0] - self.t_offset
                    t_plot_end = t_ds[-1] - self.t_offset
                    im = ax2.imshow(cwtmatr, extent=[t_plot_start, t_plot_end, frequencies[0], frequencies[-1]],
                               cmap='hot', aspect='auto', interpolation='bilinear', origin='lower',
                               vmin=vmin, vmax=vmax)

                    # Add vertical lines for stim onset and offset
                    if self.stim_spans and len(self.stim_spans) > 0:
                        stim_start_rel = self.stim_spans[0][0] - self.t_offset
                        stim_end_rel = self.stim_spans[-1][1] - self.t_offset  # Use last span's end time
                        ax2.axvline(x=stim_start_rel, color='lime', linewidth=2.5, linestyle='--', alpha=0.9)
                        ax2.axvline(x=stim_end_rel, color='lime', linewidth=2.5, linestyle='--', alpha=0.9)
                        ax2.legend(['Stim On/Offset'], facecolor='#2b2b2b', edgecolor='white', labelcolor='white', fontsize=10, loc='upper right')

                    # Add labels with padding to prevent cutoff
                    ax2.set_xlabel('Time (s, rel. to stim onset)', color='white', fontsize=16, fontweight='bold', labelpad=10)
                    ax2.set_ylabel('Frequency (Hz)', color='white', fontsize=16, fontweight='bold', labelpad=10)
                    ax2.set_title('Wavelet Analysis (Scalogram)', color='white', fontsize=18, fontweight='bold', pad=15)
                    ax2.set_ylim([0, 30])
                    ax2.set_facecolor('#1a1a1a')
                    ax2.tick_params(colors='white', labelsize=13, width=2, length=6, pad=8)

                    # Set white spines with thicker lines
                    for spine in ax2.spines.values():
                        spine.set_edgecolor('white')
                        spine.set_linewidth(2)

                    # Add colorbar
                    cbar = self.figure.colorbar(im, ax=ax2, pad=0.02)
                    cbar.set_label('Magnitude', color='white', fontsize=13, fontweight='bold', labelpad=10)
                    cbar.ax.tick_params(colors='white', labelsize=11)
                    # Make colorbar outline white
                    cbar.outline.set_edgecolor('white')
                    cbar.outline.set_linewidth(2)

                    print("[wavelet] Scalogram plotted successfully")

            except Exception as e:
                import traceback
                error_msg = f'Wavelet analysis error: {str(e)}'
                print(f"[wavelet] ERROR: {error_msg}")
                traceback.print_exc()

                ax2.text(0.5, 0.5, error_msg,
                        ha='center', va='center', transform=ax2.transAxes, color='white', fontsize=12,
                        bbox=dict(boxstyle='round', facecolor='red', alpha=0.5))
                ax2.set_facecolor('#1a1a1a')
                ax2.set_xlabel('Time (s)', color='white', fontsize=14, fontweight='bold')
                ax2.set_ylabel('Frequency (Hz)', color='white', fontsize=14, fontweight='bold')
                ax2.set_title('Wavelet Analysis (Error)', color='white', fontsize=16, fontweight='bold')
                ax2.tick_params(colors='white', labelsize=12, width=2, length=6)
                for spine in ax2.spines.values():
                    spine.set_edgecolor('white')
                    spine.set_linewidth(2)

            # Use tight_layout with padding to prevent text cutoff
            self.figure.tight_layout(pad=2.0)
            self.canvas.draw()

        def on_apply_filter(self):
            """Store the notch filter settings."""
            from PyQt6.QtWidgets import QMessageBox

            self.notch_lower = self.lower_freq_spin.value()
            self.notch_upper = self.upper_freq_spin.value()

            if self.notch_lower >= self.notch_upper:
                QMessageBox.warning(self, "Invalid Range", "Lower frequency must be less than upper frequency.")
                return

            self.update_plots()
            QMessageBox.information(self, "Filter Applied",
                                  f"Notch filter set to {self.notch_lower:.2f} - {self.notch_upper:.2f} Hz.\n"
                                  "Close this dialog to apply the filter to your signal.")

        def on_reset_filter(self):
            """Reset the notch filter."""
            self.notch_lower = None
            self.notch_upper = None
            self.lower_freq_spin.setValue(0.0)
            self.upper_freq_spin.setValue(0.0)
            self.update_plots()

        def on_mean_subtract_toggled(self, checked):
            """Handle mean subtraction checkbox toggle."""
            if self.parent_window:
                self.parent_window.state.use_mean_sub = checked
                self.mean_window_spin.setEnabled(checked)
                # Update main window and this dialog
                self.parent_window.update_and_redraw()
                self._load_sweep_data()
                self.update_plots()

        def on_mean_window_changed(self, value):
            """Handle mean subtraction window size change."""
            if self.parent_window:
                self.parent_window.state.mean_val = value
                # Update main window and this dialog
                self.parent_window.update_and_redraw()
                self._load_sweep_data()
                self.update_plots()

        def get_filter_params(self):
            """Return the notch filter parameters."""
            return self.notch_lower, self.notch_upper

        def on_prev_sweep(self):
            """Navigate to previous sweep."""
            if not self.parent_window:
                return

            # Move to previous sweep
            if self.parent_window.state.sweep_idx > 0:
                self.parent_window.state.sweep_idx -= 1
                # Re-extract data for new sweep
                self._load_sweep_data()
                self.update_plots()

        def on_next_sweep(self):
            """Navigate to next sweep."""
            if not self.parent_window:
                return

            # Move to next sweep
            sweep_count = self.parent_window._sweep_count()
            if self.parent_window.state.sweep_idx < sweep_count - 1:
                self.parent_window.state.sweep_idx += 1
                # Re-extract data for new sweep
                self._load_sweep_data()
                self.update_plots()

        def _load_sweep_data(self):
            """Reload data for current sweep from parent window."""
            if not self.parent_window:
                return

            # Get current sweep data (already filtered by _current_trace)
            t_all, y_all = self.parent_window._current_trace()
            if t_all is None or y_all is None:
                return

            # Update instance variables
            self.t = t_all
            self.y = y_all

            # Get stim spans for this sweep
            if hasattr(self.parent_window.state, 'stim_markers') and self.parent_window.state.stim_markers:
                sweep_idx = self.parent_window.state.sweep_idx
                if sweep_idx in self.parent_window.state.stim_markers:
                    self.stim_spans = self.parent_window.state.stim_markers[sweep_idx]
                else:
                    self.stim_spans = None

            # Update sweep label if it exists
            if hasattr(self, 'sweep_label'):
                self.sweep_label.setText(f"Sweep: {self.parent_window.state.sweep_idx + 1}")

            # Calculate stim onset offset for time normalization
            if self.stim_spans and len(self.stim_spans) > 0:
                self.t_offset = self.stim_spans[0][0]  # Use first stim onset
            else:
                self.t_offset = 0.0


    class SaveMetaDialog(QDialog):
        def __init__(self, abf_name: str, channel: str, parent=None, auto_stim: str = "", history: dict = None):
            super().__init__(parent)
            self.setWindowTitle("Save analyzed data — name builder")

            self._abf_name = abf_name
            self._channel = channel
            self._history = history or {}

            lay = QFormLayout(self)

            # Mouse Strain with autocomplete
            from PyQt6.QtWidgets import QCompleter
            from PyQt6.QtCore import Qt as QtCore_Qt

            self.le_strain = QLineEdit(self)
            self.le_strain.setPlaceholderText("e.g., VgatCre")
            if self._history.get('strain'):
                completer = QCompleter(self._history['strain'], self)
                completer.setCaseSensitivity(QtCore_Qt.CaseSensitivity.CaseInsensitive)
                completer.setCompletionMode(QCompleter.CompletionMode.InlineCompletion)
                self.le_strain.setCompleter(completer)
            lay.addRow("Mouse Strain:", self.le_strain)

            # Virus with autocomplete
            self.le_virus = QLineEdit(self)
            self.le_virus.setPlaceholderText("e.g., ConFoff-ChR2")
            if self._history.get('virus'):
                completer = QCompleter(self._history['virus'], self)
                completer.setCaseSensitivity(QtCore_Qt.CaseSensitivity.CaseInsensitive)
                completer.setCompletionMode(QCompleter.CompletionMode.InlineCompletion)
                self.le_virus.setCompleter(completer)
            lay.addRow("Virus:", self.le_virus)

            # Location with autocomplete
            self.le_location = QLineEdit(self)
            self.le_location.setPlaceholderText("e.g., preBotC or RTN")
            if self._history.get('location'):
                completer = QCompleter(self._history['location'], self)
                completer.setCaseSensitivity(QtCore_Qt.CaseSensitivity.CaseInsensitive)
                completer.setCompletionMode(QCompleter.CompletionMode.InlineCompletion)
                self.le_location.setCompleter(completer)
            lay.addRow("Location:", self.le_location)

            # Stimulation type (can be auto-populated) with autocomplete
            self.le_stim = QLineEdit(self)
            self.le_stim.setPlaceholderText("e.g., 20Hz10s15ms or 15msPulse")
            if auto_stim:
                self.le_stim.setText(auto_stim)
            if self._history.get('stim'):
                completer = QCompleter(self._history['stim'], self)
                completer.setCaseSensitivity(QtCore_Qt.CaseSensitivity.CaseInsensitive)
                completer.setCompletionMode(QCompleter.CompletionMode.InlineCompletion)
                self.le_stim.setCompleter(completer)
            lay.addRow("Stimulation type:", self.le_stim)

            # Laser power with autocomplete
            self.le_power = QLineEdit(self)
            self.le_power.setPlaceholderText("e.g., 8mW")
            if self._history.get('power'):
                completer = QCompleter(self._history['power'], self)
                completer.setCaseSensitivity(QtCore_Qt.CaseSensitivity.CaseInsensitive)
                completer.setCompletionMode(QCompleter.CompletionMode.InlineCompletion)
                self.le_power.setCompleter(completer)
            lay.addRow("Laser power:", self.le_power)

            self.cb_sex = QComboBox(self)
            self.cb_sex.addItems(["", "M", "F", "Unknown"])
            lay.addRow("Sex:", self.cb_sex)

            # Animal ID with autocomplete
            self.le_animal = QLineEdit(self)
            self.le_animal.setPlaceholderText("e.g., 25121004")
            if self._history.get('animal'):
                completer = QCompleter(self._history['animal'], self)
                completer.setCaseSensitivity(QtCore_Qt.CaseSensitivity.CaseInsensitive)
                completer.setCompletionMode(QCompleter.CompletionMode.InlineCompletion)
                self.le_animal.setCompleter(completer)
            lay.addRow("Animal ID:", self.le_animal)

            # Read-only info
            self.lbl_abf = QLabel(abf_name, self)
            self.lbl_chn = QLabel(channel or "", self)
            lay.addRow("ABF file:", self.lbl_abf)
            lay.addRow("Channel:", self.lbl_chn)

            # NEW: choose location checkbox
            self.cb_choose_dir = QCheckBox("Let me choose where to save", self)
            self.cb_choose_dir.setToolTip("If unchecked, files go to a 'Pleth_App_Analysis' folder automatically.")
            lay.addRow("", self.cb_choose_dir)

            # Live preview
            self.lbl_preview = QLabel("", self)
            self.lbl_preview.setStyleSheet("color:#b6bfda;")
            lay.addRow("Preview:", self.lbl_preview)

            # Buttons
            btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)
            lay.addRow(btns)
            btns.accepted.connect(self.accept)
            btns.rejected.connect(self.reject)

            # Update preview on change
            self.le_strain.textChanged.connect(self._update_preview)
            self.le_virus.textChanged.connect(self._update_preview)
            self.le_location.textChanged.connect(self._update_preview)
            self.le_stim.textChanged.connect(self._update_preview)
            self.le_power.textChanged.connect(self._update_preview)
            self.cb_sex.currentTextChanged.connect(self._update_preview)
            self.le_animal.textChanged.connect(self._update_preview)

            self._update_preview()

        # --- Helpers: light canonicalization + sanitization ---
        def _norm_token(self, s: str) -> str:
            s0 = (s or "").strip()
            if not s0:
                return ""
            s1 = s0.replace(" ", "")
            s1 = re.sub(r"(?i)chr\s*2", "ChR2", s1)            # chr2 -> ChR2
            s1 = re.sub(r"(?i)gcamp\s*6f", "GCaMP6f", s1)      # gcamp6f -> GCaMP6f
            s1 = re.sub(r"(?i)([A-Za-z0-9_-]*?)cre$", lambda m: (m.group(1) or "") + "Cre", s1)  # ...cre -> ...Cre
            return s1

        def _san(self, s: str) -> str:
            s = (s or "").strip()
            s = s.replace(" ", "_")
            s = re.sub(r"[^A-Za-z0-9._-]+", "", s)
            s = re.sub(r"_+", "_", s)
            s = re.sub(r"-+", "-", s)
            return s

        def _update_preview(self):
            # Read & normalize
            strain = self._norm_token(self.le_strain.text())
            virus  = self._norm_token(self.le_virus.text())
            location = self.le_location.text().strip()

            stim   = self.le_stim.text().strip()
            power  = self.le_power.text().strip()
            sex    = self.cb_sex.currentText().strip()
            animal = self.le_animal.text().strip()
            abf    = self._abf_name
            ch     = self._channel

            # Sanitize for filename
            strain_s = self._san(strain)
            virus_s  = self._san(virus)
            location_s = self._san(location)
            stim_s   = self._san(stim)
            power_s  = self._san(power)
            sex_s    = self._san(sex)
            animal_s = self._san(animal)
            abf_s    = self._san(abf)
            ch_s     = self._san(ch)

            # STANDARD ORDER:
            # Strain_Virus_Location_Sex_Animal_Stim_Power_ABF_Channel
            parts = [p for p in (strain_s, virus_s, location_s, sex_s, animal_s, stim_s, power_s, abf_s, ch_s) if p]
            preview = "_".join(parts) if parts else "analysis"
            self.lbl_preview.setText(preview)

        def values(self) -> dict:
            return {
                "strain": self.le_strain.text().strip(),
                "virus":  self.le_virus.text().strip(),
                "location": self.le_location.text().strip(),
                "stim":   self.le_stim.text().strip(),
                "power":  self.le_power.text().strip(),
                "sex":    self.cb_sex.currentText().strip(),
                "animal": self.le_animal.text().strip(),
                "abf":    self._abf_name,
                "chan":   self._channel,
                "preview": self.lbl_preview.text().strip(),
                "choose_dir": bool(self.cb_choose_dir.isChecked()),
            }


    def _suggest_stim_string(self) -> str:
        """
        Build a stim name like '20Hz10s15ms' from detected stim metrics
        or '15msPulse' / '5sPulse' for single pulses.
        Rounding:
        - freq_hz -> nearest Hz
        - duration_s -> nearest second
        - pulse_width_s -> nearest millisecond (or nearest second if >1s)
        """
        st = self.state
        if not getattr(st, "stim_chan", None):
            return ""

        # Make sure current sweep has metrics if possible
        try:
            self._compute_stim_for_current_sweep()
        except Exception:
            pass

        # Prefer current sweep, else first one that has metrics
        m = st.stim_metrics_by_sweep.get(st.sweep_idx) if hasattr(st, "stim_metrics_by_sweep") else None
        if not m:
            for _s, _m in getattr(st, "stim_metrics_by_sweep", {}).items():
                if _m:
                    m = _m
                    break
        if not m:
            return ""

        freq_hz   = m.get("freq_hz", None)
        dur_s     = m.get("duration_s", None)
        pw_s      = m.get("pulse_width_s", None)

        def _round_int(x):  # safe int rounding
            if x is None:
                return None
            try:
                return int(round(float(x)))
            except Exception:
                return None

        # TRAIN: if freq present and > 0, use Freq + Duration + PW
        if freq_hz is not None and np.isfinite(freq_hz) and freq_hz > 0:
            F = _round_int(freq_hz)              # Hz
            D = _round_int(dur_s)                # s
            MS = _round_int((pw_s or 0) * 1000)  # ms

            parts = []
            if F is not None and F > 0:  parts.append(f"{F}Hz")
            if D is not None and D > 0:  parts.append(f"{D}s")
            if MS is not None and MS > 0: parts.append(f"{MS}ms")
            return "".join(parts) if parts else ""

        # SINGLE PULSE (no usable freq)
        # Prefer pulse width; if missing, fall back to duration
        width_s = pw_s if (pw_s is not None and np.isfinite(pw_s)) else dur_s
        if width_s is None:
            return ""

        if width_s < 1.0:
            ms = _round_int(width_s * 1000)
            if ms is None or ms <= 0:
                ms = 1
            return f"{ms}msPulse"
        else:
            secs = _round_int(width_s)
            if secs is None or secs <= 0:
                secs = 1
            return f"{secs}sPulse"



    def on_save_analyzed_clicked(self):
        """Save analyzed data to disk after prompting for location/name."""
        from PyQt6.QtWidgets import QProgressDialog, QApplication
        from PyQt6.QtCore import Qt

        # Create progress dialog
        progress = QProgressDialog("Preparing data export...", None, 0, 100, self)
        progress.setWindowTitle("PlethAnalysis")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)  # Show immediately
        progress.setValue(0)
        QApplication.processEvents()

        try:
            self._export_all_analyzed_data(preview_only=False, progress_dialog=progress)
        finally:
            progress.close()

    def on_view_summary_clicked(self):
        """Display interactive preview of the PDF summary without saving."""
        from PyQt6.QtWidgets import QProgressDialog, QApplication
        from PyQt6.QtCore import Qt

        # Create progress dialog
        progress = QProgressDialog("Generating summary preview...", None, 0, 100, self)
        progress.setWindowTitle("PlethAnalysis")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)  # Show immediately
        progress.setValue(0)
        QApplication.processEvents()

        try:
            self._export_all_analyzed_data(preview_only=True, progress_dialog=progress)
        finally:
            progress.close()


    # metrics we won't include in CSV exports and PDFs
    _EXCLUDE_FOR_CSV = {"d1", "d2", "eupnic", "apnea", "regularity"}

    def _metric_keys_in_order(self):
        """Return metric keys in the UI order (from metrics.METRIC_SPECS)."""
        return [key for (_, key) in metrics.METRIC_SPECS]

    def _compute_metric_trace(self, key, t, y, sr_hz, peaks, breaths):
        """
        Call the metric function, passing expoffs if it exists.
        Falls back to legacy signature when needed.
        """
        fn = metrics.METRICS[key]
        on  = breaths.get("onsets")   if breaths else None
        off = breaths.get("offsets")  if breaths else None
        exm = breaths.get("expmins")  if breaths else None
        exo = breaths.get("expoffs")  if breaths else None
        try:
            return fn(t, y, sr_hz, peaks, on, off, exm, exo)  # new signature
        except TypeError:
            return fn(t, y, sr_hz, peaks, on, off, exm)       # legacy signature

    def _get_stim_masks(self, s: int):
        """
        Build (baseline_mask, stim_mask, post_mask) boolean arrays over st.t for sweep s.
        Uses union of all stim spans for 'stim'.
        """
        st = self.state
        t = st.t
        spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
        if not spans:
            # no stim: whole trace is baseline; stim/post empty
            B = np.ones_like(t, dtype=bool)
            Z = np.zeros_like(t, dtype=bool)
            return B, Z, Z

        starts = np.array([a for (a, _) in spans], dtype=float)
        ends   = np.array([b for (_, b) in spans], dtype=float)
        t0 = np.min(starts)
        t1 = np.max(ends)

        stim_mask = np.zeros_like(t, dtype=bool)
        for (a, b) in spans:
            stim_mask |= (t >= a) & (t <= b)

        baseline_mask = t < t0
        post_mask     = t > t1
        return baseline_mask, stim_mask, post_mask

    # def _nanmean_sem(self, X: np.ndarray, axis: int = 0):
    #     """Return (nanmean, nansem) along axis; SEM uses ddof=1 where n>=2 else NaN."""
    #     with np.errstate(invalid="ignore"):
    #         mean = np.nanmean(X, axis=axis)
    #         n    = np.sum(np.isfinite(X), axis=axis)
    #         # std with ddof=1; guard n<2
    #         std  = np.nanstd(X, axis=axis, ddof=1)
    #         sem  = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #     return mean, sem

    def _nanmean_sem(self, X, axis=0):
        """
        Robust mean/SEM that avoids NumPy RuntimeWarnings when there are
        0 or 1 finite values along the chosen axis.
        """
        import numpy as np

        A = np.asarray(X, dtype=float)
        if A.size == 0:
            return np.nan, np.nan

        # Move the target axis to 0 so we can index rows easily
        A0 = np.moveaxis(A, axis, 0)      # shape: (N, ...rest...)
        # Collapse the rest to a single trailing dimension for simplicity
        tail = int(np.prod(A0.shape[1:])) or 1
        A2 = A0.reshape(A0.shape[0], tail)  # (N, T)

        finite = np.isfinite(A2)
        n = finite.sum(axis=0)              # (T,)

        mean = np.full((tail,), np.nan, dtype=float)
        sem  = np.full((tail,), np.nan, dtype=float)

        # rows/columns (along axis) with at least one finite value
        msk_mean = n > 0
        if np.any(msk_mean):
            mean[msk_mean] = np.nanmean(A2[:, msk_mean], axis=0)

        # rows/columns with at least two finite values (only here compute SEM)
        msk_sem = n >= 2
        if np.any(msk_sem):
            std = np.nanstd(A2[:, msk_sem], axis=0, ddof=1)
            sem[msk_sem] = std / np.sqrt(n[msk_sem])

        # reshape back to the original tail and move axis back
        mean = mean.reshape(A0.shape[1:])
        sem  = sem.reshape(A0.shape[1:])
        # move axis back to original position (only if result has enough dimensions)
        # After reduction along one axis, result has (A.ndim - 1) dimensions
        # We can only moveaxis if result is multi-dimensional
        if mean.ndim > 1:
            mean = np.moveaxis(mean, 0, axis)
            sem  = np.moveaxis(sem, 0, axis)

        return mean, sem







    # def _export_all_analyzed_data(self):
    #     """
    #     Save:
    #     1) <base>_bundle.npz  (downsampled Y_proc + y2 traces, peaks/breaths, stim spans, meta)
    #     2) <base>_means_by_time.csv  (downsampled time-series: per-sweep (optional), mean, sem)
    #         + appended block of normalized per-sweep traces/mean/sem (suffix *_norm)
    #     3) <base>_breaths.csv  (WIDE per-breath table: ALL | BASELINE | STIM | POST blocks)
    #         + appended duplicate blocks with normalized values (all headers suffixed *_norm)
    #     4) <base>_summary.pdf  (figure; unchanged here)
    #     """
    #     st = self.state
    #     if not getattr(self, "_save_dir", None) or not getattr(self, "_save_base", None):
    #         QMessageBox.warning(self, "Save analyzed data", "Choose a save location/name first.")
    #         return
    #     if st.t is None or not st.analyze_chan or st.analyze_chan not in st.sweeps:
    #         QMessageBox.warning(self, "Save analyzed data", "No analyzed data available.")
    #         return

    #     import numpy as np, csv, json
    #     from PyQt6.QtCore import Qt
    #     from PyQt6.QtWidgets import QApplication

    #     # ---------- knobs ----------
    #     DS_TARGET_HZ    = 50.0
    #     CSV_FLUSH_EVERY = 2000
    #     INCLUDE_TRACES  = bool(getattr(self, "_csv_include_traces", True))
    #     # normalization window (seconds before t=0); override via self._norm_window_s if you like
    #     NORM_BASELINE_WINDOW_S = float(getattr(self, "_norm_window_s", 10.0))
    #     EPS_BASE = 1e-12  # avoid divide-by-zero

    #     # ---------- basics ----------
    #     # any_ch   = next(iter(st.sweeps.values()))
    #     # n_sweeps = any_ch.shape[1]
    #     # N        = len(st.t)
    #     any_ch   = next(iter(st.sweeps.values()))
    #     n_sweeps = any_ch.shape[1]
    #     kept_sweeps = [s for s in range(n_sweeps) if s not in st.omitted_sweeps]
    #     S = len(kept_sweeps)
    #     if S == 0:
    #         QMessageBox.warning(self, "Save analyzed data", "All sweeps are omitted. Nothing to save.")
    #         return


    #     # Downsample index used for NPZ + time CSV
    #     ds_step = max(1, int(round(float(st.sr_hz) / DS_TARGET_HZ))) if st.sr_hz else 1
    #     ds_idx  = np.arange(0, N, ds_step, dtype=int)
    #     M       = len(ds_idx)

    #     # Global stim zero and duration (union across ALL sweeps)
    #     global_s0, global_s1 = None, None
    #     if st.stim_chan:
    #         for s in range(n_sweeps):
    #             spans = st.stim_spans_by_sweep.get(s, [])
    #             if spans:
    #                 starts = [a for (a, _) in spans]
    #                 ends   = [b for (_, b) in spans]
    #                 m0 = float(min(starts))
    #                 m1 = float(max(ends))
    #                 global_s0 = m0 if global_s0 is None else min(global_s0, m0)
    #                 global_s1 = m1 if global_s1 is None else max(global_s1, m1)
    #     have_global_stim = (global_s0 is not None and global_s1 is not None)
    #     global_dur = (global_s1 - global_s0) if have_global_stim else None

    #     # Time for NPZ (raw) and for CSV (normalized to global_s0 if present)
    #     t_ds_raw = st.t[ds_idx]
    #     csv_t0   = (global_s0 if have_global_stim else 0.0)
    #     t_ds_csv = (st.t - csv_t0)[ds_idx]

    #     # ---------- containers ----------
    #     Y_proc_ds = np.full((M, S), np.nan, dtype=float)
    #     peaks_by_sweep, on_by_sweep, off_by_sweep, exm_by_sweep, exo_by_sweep = [], [], [], [], []

    #     all_keys     = self._metric_keys_in_order()
    #     y2_ds_by_key = {k: np.full((M, S), np.nan, dtype=float) for k in all_keys}

    #     # Keep full-res processed pleth (for other uses)
    #     Y_full_by_sweep = []

    #     # ---------- fill per-sweep ----------
    #     # for s in range(n_sweeps):
    #     #     y_proc = self._get_processed_for(st.analyze_chan, s)
    #     #     Y_full_by_sweep.append(y_proc)
    #     #     Y_proc_ds[:, s] = y_proc[ds_idx]

    #     #     pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #     #     br  = st.breath_by_sweep.get(s, None)
    #     #     if br is None and pks.size:
    #     #         br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #     #         st.breath_by_sweep[s] = br
    #     #     if br is None:
    #     #         br = {
    #     #             "onsets":  np.array([], dtype=int),
    #     #             "offsets": np.array([], dtype=int),
    #     #             "expmins": np.array([], dtype=int),
    #     #             "expoffs": np.array([], dtype=int),
    #     #         }

    #     #     peaks_by_sweep.append(pks)
    #     #     on_by_sweep.append(np.asarray(br.get("onsets",  []), dtype=int))
    #     #     off_by_sweep.append(np.asarray(br.get("offsets", []), dtype=int))
    #     #     exm_by_sweep.append(np.asarray(br.get("expmins", []), dtype=int))
    #     #     exo_by_sweep.append(np.asarray(br.get("expoffs", []), dtype=int))

    #     #     for k in all_keys:
    #     #         y2 = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #     #         if y2 is not None and len(y2) == N:
    #     #             y2_ds_by_key[k][:, s] = y2[ds_idx]
    #     for col, s in enumerate(kept_sweeps):
    #         y_proc = self._get_processed_for(st.analyze_chan, s)
    #         Y_proc_ds[:, col] = y_proc[ds_idx]
    #         pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #         br  = st.breath_by_sweep.get(s, None)
    #         if br is None and pks.size:
    #             br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #             st.breath_by_sweep[s] = br
    #         if br is None:
    #             br = {"onsets": np.array([], dtype=int), "offsets": np.array([], dtype=int),
    #                 "expmins": np.array([], dtype=int), "expoffs": np.array([], dtype=int)}

    #         peaks_by_sweep.append(pks)
    #         on_by_sweep.append(np.asarray(br.get("onsets",  []), dtype=int))
    #         off_by_sweep.append(np.asarray(br.get("offsets", []), dtype=int))
    #         exm_by_sweep.append(np.asarray(br.get("expmins", []), dtype=int))
    #         exo_by_sweep.append(np.asarray(br.get("expoffs", []), dtype=int))

    #         for k in all_keys:
    #             y2 = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #             if y2 is not None and len(y2) == N:
    #                 y2_ds_by_key[k][:, col] = y2[ds_idx]



    #     # ---------- (1) NPZ bundle (downsampled) ----------
    #     base     = self._save_dir / self._save_base
    #     npz_path = base.with_name(base.name + "_bundle.npz")

    #     stim_obj = np.empty(n_sweeps, dtype=object)
    #     # for s in range(n_sweeps):
    #     for s in kept_sweeps:
    #         spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
    #         stim_obj[s] = np.array(spans, dtype=float) if spans else np.array([], dtype=float).reshape(0, 2)

    #     peaks_obj = np.array(peaks_by_sweep, dtype=object)
    #     on_obj    = np.array(on_by_sweep,  dtype=object)
    #     off_obj   = np.array(off_by_sweep, dtype=object)
    #     exm_obj   = np.array(exm_by_sweep, dtype=object)
    #     exo_obj   = np.array(exo_by_sweep, dtype=object)

    #     y2_kwargs_ds = {f"y2_{k}_ds": y2_ds_by_key[k] for k in all_keys}

    #     meta = {
    #         "analyze_channel": st.analyze_chan,
    #         "sr_hz": float(st.sr_hz),
    #         "n_sweeps": int(n_sweeps),
    #         "abf_path": str(getattr(st, "in_path", "")),
    #         "ui_meta": getattr(self, "_save_meta", {}),
    #         "excluded_for_csv": sorted(list(self._EXCLUDE_FOR_CSV)),
    #         "ds_target_hz": float(DS_TARGET_HZ),
    #         "ds_step": int(ds_step),
    #         "csv_time_zero": float(csv_t0),
    #         "csv_includes_traces": bool(INCLUDE_TRACES),
    #         "norm_window_s": float(NORM_BASELINE_WINDOW_S),
    #     }

    #     np.savez_compressed(
    #         npz_path,
    #         t_ds=t_ds_raw,
    #         Y_proc_ds=Y_proc_ds,
    #         peaks_by_sweep=peaks_obj,
    #         onsets_by_sweep=on_obj,
    #         offsets_by_sweep=off_obj,
    #         expmins_by_sweep=exm_obj,
    #         expoffs_by_sweep=exo_obj,
    #         stim_spans_by_sweep=stim_obj,
    #         meta_json=json.dumps(meta),
    #         **y2_kwargs_ds,
    #     )

    #     # ---------- helpers for normalization ----------
    #     def _per_sweep_baseline_for_time(A_ds: np.ndarray) -> np.ndarray:
    #         """
    #         A_ds: (M,S) downsampled metric matrix.
    #         Returns b[S]: mean over last NORM_BASELINE_WINDOW_S before 0; fallback to first W after 0.
    #         """
    #         b = np.full((A_ds.shape[1],), np.nan, dtype=float)
    #         mask_pre  = (t_ds_csv >= -NORM_BASELINE_WINDOW_S) & (t_ds_csv < 0.0)
    #         mask_post = (t_ds_csv >= 0.0) & (t_ds_csv <=  NORM_BASELINE_WINDOW_S)
    #         for s in range(A_ds.shape[1]):
    #             col = A_ds[:, s]
    #             vals = col[mask_pre]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = col[mask_post]
    #                 vals = vals[np.isfinite(vals)]
    #             if vals.size:
    #                 b[s] = float(np.mean(vals))
    #         return b

    #     def _normalize_matrix_by_baseline(A_ds: np.ndarray, b: np.ndarray) -> np.ndarray:
    #         out = np.full_like(A_ds, np.nan)
    #         for s in range(A_ds.shape[1]):
    #             bs = b[s]
    #             if np.isfinite(bs) and abs(bs) > EPS_BASE:
    #                 out[:, s] = A_ds[:, s] / bs
    #         return out

    #     # ---------- (2) Per-time CSV (raw + normalized appended) ----------
    #     csv_time_path = base.with_name(base.name + "_means_by_time.csv")
    #     keys_for_csv  = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]

    #     # Build normalized stacks per metric
    #     y2_ds_by_key_norm = {}
    #     baseline_by_key   = {}
    #     for k in keys_for_csv:
    #         b = _per_sweep_baseline_for_time(y2_ds_by_key[k])
    #         baseline_by_key[k] = b
    #         y2_ds_by_key_norm[k] = _normalize_matrix_by_baseline(y2_ds_by_key[k], b)

    #     # headers: raw first (unchanged), then the same pattern with *_norm suffix
    #     header = ["t"]
    #     for k in keys_for_csv:
    #         if INCLUDE_TRACES:
    #             header += [f"{k}_s{j+1}" for j in range(n_sweeps)]
    #         header += [f"{k}_mean", f"{k}_sem"]

    #     # normalized headers appended
    #     for k in keys_for_csv:
    #         if INCLUDE_TRACES:
    #             header += [f"{k}_norm_s{j+1}" for j in range(n_sweeps)]
    #         header += [f"{k}_norm_mean", f"{k}_norm_sem"]

    #     self.setCursor(Qt.CursorShape.WaitCursor)
    #     try:
    #         with open(csv_time_path, "w", newline="") as f:
    #             w = csv.writer(f)
    #             w.writerow(header)

    #             for i in range(M):
    #                 row = [f"{t_ds_csv[i]:.9f}"]

    #                 # RAW block
    #                 for k in keys_for_csv:
    #                     col = y2_ds_by_key[k][i, :]
    #                     if INCLUDE_TRACES:
    #                         row += [f"{v:.9g}" if np.isfinite(v) else "" for v in col]
    #                     m, sem = self._mean_sem_1d(col)
    #                     row += [f"{m:.9g}", f"{sem:.9g}"]

    #                 # NORMALIZED block
    #                 for k in keys_for_csv:
    #                     colN = y2_ds_by_key_norm[k][i, :]
    #                     if INCLUDE_TRACES:
    #                         row += [f"{v:.9g}" if np.isfinite(v) else "" for v in colN]
    #                     mN, semN = self._mean_sem_1d(colN)
    #                     row += [f"{mN:.9g}", f"{semN:.9g}"]

    #                 w.writerow(row)
    #                 if (i % CSV_FLUSH_EVERY) == 0:
    #                     QApplication.processEvents()
    #     finally:
    #         self.unsetCursor()

    #     # ---------- (3) Per-breath CSV (WIDE) ----------
    #     breaths_path = base.with_name(base.name + "_breaths.csv")

    #     BREATH_COLS = [
    #         "sweep", "breath", "t", "region",
    #         "if", "amp_insp", "amp_exp", "area_insp", "area_exp",
    #         "ti", "te", "vent_proxy",
    #     ]
    #     def _headers_for_block(suffix: str | None) -> list[str]:
    #         if not suffix: return BREATH_COLS[:]
    #         return [f"{c}_{suffix}" for c in BREATH_COLS]

    #     def _headers_for_block_norm(suffix: str | None) -> list[str]:
    #         # duplicate the same block but suffix *all* column names with _norm
    #         base = _headers_for_block(suffix)
    #         return [h + "_norm" for h in base]

    #     rows_all, rows_bl, rows_st, rows_po = [], [], [], []
    #     rows_all_N, rows_bl_N, rows_st_N, rows_po_N = [], [], [], []

    #     need_keys = ["if", "amp_insp", "amp_exp", "area_insp", "area_exp", "ti", "te", "vent_proxy"]

    #     for s in range(n_sweeps):
    #         y_proc = self._get_processed_for(st.analyze_chan, s)
    #         pks    = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #         br     = st.breath_by_sweep.get(s, None)
    #         if br is None and pks.size:
    #             br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #             st.breath_by_sweep[s] = br
    #         if br is None:
    #             br = {"onsets": np.array([], dtype=int)}

    #         on = np.asarray(br.get("onsets", []), dtype=int)
    #         if on.size < 2:
    #             continue

    #         mids = (on[:-1] + on[1:]) // 2

    #         traces = {}
    #         for k in need_keys:
    #             if k in metrics.METRICS:
    #                 traces[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #             else:
    #                 traces[k] = None

    #         # Per-sweep breath-based baselines for normalization (use breath midpoints)
    #         # Window: last W seconds before t=0 (fallback: first W seconds after 0)
    #         # Compute t_rel for each breath midpoint once:
    #         t_rel_all = (st.t[mids] - (global_s0 if have_global_stim else 0.0)).astype(float)
    #         mask_pre_b  = (t_rel_all >= -NORM_BASELINE_WINDOW_S) & (t_rel_all < 0.0)
    #         mask_post_b = (t_rel_all >=  0.0) & (t_rel_all <= NORM_BASELINE_WINDOW_S)

    #         b_by_k = {}
    #         for k in need_keys:
    #             arr = traces.get(k, None)
    #             if arr is None or len(arr) != N:
    #                 b_by_k[k] = np.nan
    #                 continue
    #             vals = arr[mids[mask_pre_b]]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = arr[mids[mask_post_b]]
    #                 vals = vals[np.isfinite(vals)]
    #             b_by_k[k] = float(np.mean(vals)) if vals.size else np.nan

    #         for i, idx in enumerate(mids, start=1):
    #             t_rel = float(st.t[int(idx)] - (global_s0 if have_global_stim else 0.0))

    #             # ----- RAW: ALL
    #             row_all = [str(s + 1), str(i), f"{t_rel:.9g}", "all"]
    #             for k in need_keys:
    #                 v = np.nan
    #                 arr = traces.get(k, None)
    #                 if arr is not None and len(arr) == N:
    #                     v = arr[int(idx)]
    #                 row_all.append(f"{v:.9g}" if np.isfinite(v) else "")
    #             rows_all.append(row_all)

    #             # ----- NORM: ALL (duplicate id columns + normalized metrics)
    #             row_allN = [str(s + 1), str(i), f"{t_rel:.9g}", "all"]
    #             for k in need_keys:
    #                 v = np.nan
    #                 arr = traces.get(k, None)
    #                 if arr is not None and len(arr) == N:
    #                     v = arr[int(idx)]
    #                 b = b_by_k.get(k, np.nan)
    #                 vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
    #                 row_allN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
    #             rows_all_N.append(row_allN)

    #             if have_global_stim:
    #                 if t_rel < 0:
    #                     tgt_list = rows_bl; tgt_listN = rows_bl_N; region = "Baseline"
    #                 elif 0.0 <= t_rel <= global_dur:
    #                     tgt_list = rows_st; tgt_listN = rows_st_N; region = "Stim"
    #                 else:
    #                     tgt_list = rows_po; tgt_listN = rows_po_N; region = "Post"

    #                 # RAW regional row
    #                 row_reg = [str(s + 1), str(i), f"{t_rel:.9g}", region]
    #                 for k in need_keys:
    #                     v = np.nan
    #                     arr = traces.get(k, None)
    #                     if arr is not None and len(arr) == N:
    #                         v = arr[int(idx)]
    #                     row_reg.append(f"{v:.9g}" if np.isfinite(v) else "")
    #                 tgt_list.append(row_reg)

    #                 # NORM regional row
    #                 row_regN = [str(s + 1), str(i), f"{t_rel:.9g}", region]
    #                 for k in need_keys:
    #                     v = np.nan
    #                     arr = traces.get(k, None)
    #                     if arr is not None and len(arr) == N:
    #                         v = arr[int(idx)]
    #                     b = b_by_k.get(k, np.nan)
    #                     vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
    #                     row_regN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
    #                 tgt_listN.append(row_regN)

    #     def _pad_row(row, want_len):
    #         if row is None: return [""] * want_len
    #         if len(row) < want_len: return row + [""] * (want_len - len(row))
    #         return row

    #     headers_all = _headers_for_block(None)
    #     headers_bl  = _headers_for_block("baseline")
    #     headers_st  = _headers_for_block("stim")
    #     headers_po  = _headers_for_block("post")

    #     headers_allN = _headers_for_block_norm(None)
    #     headers_blN  = _headers_for_block_norm("baseline")
    #     headers_stN  = _headers_for_block_norm("stim")
    #     headers_poN  = _headers_for_block_norm("post")

    #     have_stim_blocks = have_global_stim and (len(rows_bl) + len(rows_st) + len(rows_po) > 0)

    #     with open(breaths_path, "w", newline="") as f:
    #         w = csv.writer(f)
    #         if not have_stim_blocks:
    #             # RAW + NORM (ALL only)
    #             full_header = headers_all + [""] + headers_allN
    #             w.writerow(full_header)
    #             L = max(len(rows_all), len(rows_all_N))
    #             LA = len(headers_all); LAN = len(headers_allN)
    #             for i in range(L):
    #                 ra  = rows_all[i]   if i < len(rows_all)   else None
    #                 raN = rows_all_N[i] if i < len(rows_all_N) else None
    #                 row = _pad_row(ra, LA) + [""] + _pad_row(raN, LAN)
    #                 w.writerow(row)
    #         else:
    #             # RAW blocks, then NORM blocks
    #             full_header = (
    #                 headers_all + [""] + headers_bl + [""] + headers_st + [""] + headers_po + [""] +
    #                 headers_allN + [""] + headers_blN + [""] + headers_stN + [""] + headers_poN
    #             )
    #             w.writerow(full_header)

    #             L = max(
    #                 len(rows_all), len(rows_bl), len(rows_st), len(rows_po),
    #                 len(rows_all_N), len(rows_bl_N), len(rows_st_N), len(rows_po_N),
    #             )
    #             LA = len(headers_all); LB = len(headers_bl); LS = len(headers_st); LP = len(headers_po)
    #             LAN = len(headers_allN); LBN = len(headers_blN); LSN = len(headers_stN); LPN = len(headers_poN)

    #             for i in range(L):
    #                 ra  = rows_all[i]   if i < len(rows_all)   else None
    #                 rb  = rows_bl[i]    if i < len(rows_bl)    else None
    #                 rs  = rows_st[i]    if i < len(rows_st)    else None
    #                 rp  = rows_po[i]    if i < len(rows_po)    else None
    #                 raN = rows_all_N[i] if i < len(rows_all_N) else None
    #                 rbN = rows_bl_N[i]  if i < len(rows_bl_N)  else None
    #                 rsN = rows_st_N[i]  if i < len(rows_st_N)  else None
    #                 rpN = rows_po_N[i]  if i < len(rows_po_N)  else None

    #                 row = (
    #                     _pad_row(ra, LA) + [""] +
    #                     _pad_row(rb, LB) + [""] +
    #                     _pad_row(rs, LS) + [""] +
    #                     _pad_row(rp, LP) + [""] +
    #                     _pad_row(raN, LAN) + [""] +
    #                     _pad_row(rbN, LBN) + [""] +
    #                     _pad_row(rsN, LSN) + [""] +
    #                     _pad_row(rpN, LPN)
    #                 )
    #                 w.writerow(row)

    #     # ---------- (4) Summary PDF ----------
    #     keys_for_timeplots = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]
    #     label_by_key = {key: label for (label, key) in metrics.METRIC_SPECS if key in keys_for_timeplots}
    #     pdf_path = base.with_name(base.name + "_summary.pdf")
    #     try:
    #         self._save_metrics_summary_pdf(
    #             out_path=pdf_path,
    #             t_ds_csv=t_ds_csv,
    #             y2_ds_by_key=y2_ds_by_key,
    #             keys_for_csv=keys_for_timeplots,
    #             label_by_key=label_by_key,
    #             stim_zero=(global_s0 if have_global_stim else None),
    #             stim_dur=(global_dur if have_global_stim else None),
    #         )
    #     except Exception as e:
    #         print(f"[save][summary-pdf] skipped: {e}")

    #     # ---------- done ----------
    #     msg = f"Saved:\n- {npz_path.name}\n- {csv_time_path.name}\n- {breaths_path.name}\n- {pdf_path.name}"
    #     print("[save]", msg)
    #     try:
    #         self.statusbar.showMessage(msg, 6000)
    #     except Exception:
    #         pass

    # def _export_all_analyzed_data(self):
    #     """
    #     Save:
    #     1) <base>_bundle.npz  (downsampled Y_proc + y2 traces, peaks/breaths, stim spans, meta)
    #     2) <base>_means_by_time.csv  (downsampled time-series: per-sweep (optional), mean, sem)
    #         + appended block of normalized per-sweep traces/mean/sem (suffix *_norm)
    #     3) <base>_breaths.csv  (WIDE per-breath table: ALL | BASELINE | STIM | POST blocks)
    #         + appended duplicate blocks with normalized values (all headers suffixed *_norm)
    #     4) <base>_summary.pdf  (figure)
    #     """
    #     st = self.state
    #     if not getattr(self, "_save_dir", None) or not getattr(self, "_save_base", None):
    #         QMessageBox.warning(self, "Save analyzed data", "Choose a save location/name first.")
    #         return
    #     if st.t is None or not st.analyze_chan or st.analyze_chan not in st.sweeps:
    #         QMessageBox.warning(self, "Save analyzed data", "No analyzed data available.")
    #         return

    #     import numpy as np, csv, json
    #     from PyQt6.QtCore import Qt
    #     from PyQt6.QtWidgets import QApplication

    #     # ---------- knobs ----------
    #     DS_TARGET_HZ    = 50.0
    #     CSV_FLUSH_EVERY = 2000
    #     INCLUDE_TRACES  = bool(getattr(self, "_csv_include_traces", True))
    #     NORM_BASELINE_WINDOW_S = float(getattr(self, "_norm_window_s", 10.0))
    #     EPS_BASE = 1e-12

    #     # ---------- basics ----------
    #     any_ch    = next(iter(st.sweeps.values()))
    #     n_sweeps  = int(any_ch.shape[1])
    #     N         = int(len(st.t))
    #     kept_sweeps = [s for s in range(n_sweeps) if s not in getattr(st, "omitted_sweeps", set())]
    #     S = len(kept_sweeps)
    #     if S == 0:
    #         QMessageBox.warning(self, "Save analyzed data", "All sweeps are omitted. Nothing to save.")
    #         return

    #     # Downsample index used for NPZ + time CSV
    #     ds_step = max(1, int(round(float(st.sr_hz) / DS_TARGET_HZ))) if st.sr_hz else 1
    #     ds_idx  = np.arange(0, N, ds_step, dtype=int)
    #     M       = int(len(ds_idx))

    #     # Global stim zero and duration (union across KEPT sweeps)
    #     global_s0, global_s1 = None, None
    #     if st.stim_chan:
    #         for s in kept_sweeps:
    #             spans = st.stim_spans_by_sweep.get(s, [])
    #             if spans:
    #                 starts = [a for (a, _) in spans]
    #                 ends   = [b for (_, b) in spans]
    #                 m0 = float(min(starts)); m1 = float(max(ends))
    #                 global_s0 = m0 if global_s0 is None else min(global_s0, m0)
    #                 global_s1 = m1 if global_s1 is None else max(global_s1, m1)
    #     have_global_stim = (global_s0 is not None and global_s1 is not None)
    #     global_dur = (global_s1 - global_s0) if have_global_stim else None

    #     # Time for NPZ (raw) and for CSV (normalized to global_s0 if present)
    #     t_ds_raw = st.t[ds_idx]
    #     csv_t0   = (global_s0 if have_global_stim else 0.0)
    #     t_ds_csv = (st.t - csv_t0)[ds_idx]

    #     # ---------- containers ----------
    #     all_keys     = self._metric_keys_in_order()
    #     Y_proc_ds    = np.full((M, S), np.nan, dtype=float)
    #     y2_ds_by_key = {k: np.full((M, S), np.nan, dtype=float) for k in all_keys}

    #     peaks_by_sweep, on_by_sweep, off_by_sweep, exm_by_sweep, exo_by_sweep = [], [], [], [], []

    #     # ---------- fill per kept sweep ----------
    #     for col, s in enumerate(kept_sweeps):
    #         y_proc = self._get_processed_for(st.analyze_chan, s)
    #         Y_proc_ds[:, col] = y_proc[ds_idx]

    #         pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #         br  = st.breath_by_sweep.get(s, None)
    #         if br is None and pks.size:
    #             # backfill breaths for this sweep so exports are consistent
    #             br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #             st.breath_by_sweep[s] = br
    #         if br is None:
    #             br = {
    #                 "onsets":  np.array([], dtype=int),
    #                 "offsets": np.array([], dtype=int),
    #                 "expmins": np.array([], dtype=int),
    #                 "expoffs": np.array([], dtype=int),
    #             }

    #         peaks_by_sweep.append(pks)
    #         on_by_sweep.append(np.asarray(br.get("onsets",  []), dtype=int))
    #         off_by_sweep.append(np.asarray(br.get("offsets", []), dtype=int))
    #         exm_by_sweep.append(np.asarray(br.get("expmins", []), dtype=int))
    #         exo_by_sweep.append(np.asarray(br.get("expoffs", []), dtype=int))

    #         for k in all_keys:
    #             y2 = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #             if y2 is not None and len(y2) == N:
    #                 y2_ds_by_key[k][:, col] = y2[ds_idx]

    #     # ---------- (1) NPZ bundle (downsampled) ----------
    #     base     = self._save_dir / self._save_base
    #     npz_path = base.with_name(base.name + "_bundle.npz")

    #     # Pack stim spans in KEPT order (align with columns)
    #     stim_obj = np.empty(S, dtype=object)
    #     for col, s in enumerate(kept_sweeps):
    #         spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
    #         stim_obj[col] = np.array(spans, dtype=float).reshape(-1, 2) if spans else np.empty((0, 2), dtype=float)

    #     peaks_obj = np.array(peaks_by_sweep, dtype=object)
    #     on_obj    = np.array(on_by_sweep,  dtype=object)
    #     off_obj   = np.array(off_by_sweep, dtype=object)
    #     exm_obj   = np.array(exm_by_sweep, dtype=object)
    #     exo_obj   = np.array(exo_by_sweep, dtype=object)

    #     y2_kwargs_ds = {f"y2_{k}_ds": y2_ds_by_key[k] for k in all_keys}

    #     meta = {
    #         "analyze_channel": st.analyze_chan,
    #         "sr_hz": float(st.sr_hz),
    #         "n_sweeps_total": int(n_sweeps),
    #         "n_sweeps_kept": int(S),
    #         "kept_sweeps": [int(s) for s in kept_sweeps],                  # original indices
    #         "omitted_sweeps": sorted(int(x) for x in getattr(st, "omitted_sweeps", set())),
    #         "abf_path": str(getattr(st, "in_path", "")),
    #         "ui_meta": getattr(self, "_save_meta", {}),
    #         "excluded_for_csv": sorted(list(self._EXCLUDE_FOR_CSV)),
    #         "ds_target_hz": float(DS_TARGET_HZ),
    #         "ds_step": int(ds_step),
    #         "csv_time_zero": float(csv_t0),
    #         "csv_includes_traces": bool(INCLUDE_TRACES),
    #         "norm_window_s": float(NORM_BASELINE_WINDOW_S),
    #     }

    #     np.savez_compressed(
    #         npz_path,
    #         t_ds=t_ds_raw,
    #         Y_proc_ds=Y_proc_ds,
    #         peaks_by_sweep=peaks_obj,
    #         onsets_by_sweep=on_obj,
    #         offsets_by_sweep=off_obj,
    #         expmins_by_sweep=exm_obj,
    #         expoffs_by_sweep=exo_obj,
    #         stim_spans_by_sweep=stim_obj,
    #         meta_json=json.dumps(meta),
    #         **y2_kwargs_ds,
    #     )

    #     # ---------- helpers for normalization ----------
    #     def _per_sweep_baseline_for_time(A_ds: np.ndarray) -> np.ndarray:
    #         """
    #         A_ds: (M,S) downsampled metric matrix.
    #         Returns b[S]: mean over last NORM_BASELINE_WINDOW_S before 0; fallback to first W after 0.
    #         """
    #         b = np.full((A_ds.shape[1],), np.nan, dtype=float)
    #         mask_pre  = (t_ds_csv >= -NORM_BASELINE_WINDOW_S) & (t_ds_csv < 0.0)
    #         mask_post = (t_ds_csv >=  0.0) & (t_ds_csv <= NORM_BASELINE_WINDOW_S)
    #         for sidx in range(A_ds.shape[1]):
    #             col = A_ds[:, sidx]
    #             vals = col[mask_pre]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = col[mask_post]
    #                 vals = vals[np.isfinite(vals)]
    #             if vals.size:
    #                 b[sidx] = float(np.mean(vals))
    #         return b

    #     def _normalize_matrix_by_baseline(A_ds: np.ndarray, b: np.ndarray) -> np.ndarray:
    #         out = np.full_like(A_ds, np.nan)
    #         for sidx in range(A_ds.shape[1]):
    #             bs = b[sidx]
    #             if np.isfinite(bs) and abs(bs) > EPS_BASE:
    #                 out[:, sidx] = A_ds[:, sidx] / bs
    #         return out

    #     # ---------- (2) Per-time CSV (raw + normalized appended) ----------
    #     csv_time_path = base.with_name(base.name + "_means_by_time.csv")
    #     keys_for_csv  = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]

    #     # Build normalized stacks per metric
    #     y2_ds_by_key_norm = {}
    #     baseline_by_key   = {}
    #     for k in keys_for_csv:
    #         b = _per_sweep_baseline_for_time(y2_ds_by_key[k])
    #         baseline_by_key[k] = b
    #         y2_ds_by_key_norm[k] = _normalize_matrix_by_baseline(y2_ds_by_key[k], b)

    #     # headers: raw first, then the same pattern with *_norm suffix
    #     header = ["t"]
    #     for k in keys_for_csv:
    #         if INCLUDE_TRACES:
    #             header += [f"{k}_s{j+1}" for j in range(S)]
    #         header += [f"{k}_mean", f"{k}_sem"]

    #     for k in keys_for_csv:
    #         if INCLUDE_TRACES:
    #             header += [f"{k}_norm_s{j+1}" for j in range(S)]
    #         header += [f"{k}_norm_mean", f"{k}_norm_sem"]

    #     self.setCursor(Qt.CursorShape.WaitCursor)
    #     try:
    #         with open(csv_time_path, "w", newline="") as f:
    #             w = csv.writer(f)
    #             w.writerow(header)

    #             for i in range(M):
    #                 row = [f"{t_ds_csv[i]:.9f}"]

    #                 # RAW block
    #                 for k in keys_for_csv:
    #                     col = y2_ds_by_key[k][i, :]
    #                     if INCLUDE_TRACES:
    #                         row += [f"{v:.9g}" if np.isfinite(v) else "" for v in col]
    #                     m, sem = self._mean_sem_1d(col)
    #                     row += [f"{m:.9g}", f"{sem:.9g}"]

    #                 # NORMALIZED block
    #                 for k in keys_for_csv:
    #                     colN = y2_ds_by_key_norm[k][i, :]
    #                     if INCLUDE_TRACES:
    #                         row += [f"{v:.9g}" if np.isfinite(v) else "" for v in colN]
    #                     mN, semN = self._mean_sem_1d(colN)
    #                     row += [f"{mN:.9g}", f"{semN:.9g}"]

    #                 w.writerow(row)
    #                 if (i % CSV_FLUSH_EVERY) == 0:
    #                     QApplication.processEvents()
    #     finally:
    #         self.unsetCursor()

    #     # ---------- (3) Per-breath CSV (WIDE) ----------
    #     breaths_path = base.with_name(base.name + "_breaths.csv")

    #     BREATH_COLS = [
    #         "sweep", "breath", "t", "region",
    #         "if", "amp_insp", "amp_exp", "area_insp", "area_exp",
    #         "ti", "te", "vent_proxy",
    #     ]
    #     def _headers_for_block(suffix: str | None) -> list[str]:
    #         if not suffix: return BREATH_COLS[:]
    #         return [f"{c}_{suffix}" for c in BREATH_COLS]

    #     def _headers_for_block_norm(suffix: str | None) -> list[str]:
    #         base_cols = _headers_for_block(suffix)
    #         return [h + "_norm" for h in base_cols]

    #     rows_all, rows_bl, rows_st, rows_po = [], [], [], []
    #     rows_all_N, rows_bl_N, rows_st_N, rows_po_N = [], [], [], []

    #     need_keys = ["if", "amp_insp", "amp_exp", "area_insp", "area_exp", "ti", "te", "vent_proxy"]

    #     for s in kept_sweeps:
    #         y_proc = self._get_processed_for(st.analyze_chan, s)
    #         pks    = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #         br     = st.breath_by_sweep.get(s, None)
    #         if br is None and pks.size:
    #             br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #             st.breath_by_sweep[s] = br
    #         if br is None:
    #             br = {"onsets": np.array([], dtype=int)}

    #         on = np.asarray(br.get("onsets", []), dtype=int)
    #         if on.size < 2:
    #             continue

    #         mids = (on[:-1] + on[1:]) // 2

    #         traces = {}
    #         for k in need_keys:
    #             if k in metrics.METRICS:
    #                 traces[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #             else:
    #                 traces[k] = None

    #         # Per-sweep breath-based baselines (use breath midpoints)
    #         t_rel_all = (st.t[mids] - (global_s0 if have_global_stim else 0.0)).astype(float)
    #         mask_pre_b  = (t_rel_all >= -NORM_BASELINE_WINDOW_S) & (t_rel_all < 0.0)
    #         mask_post_b = (t_rel_all >=  0.0) & (t_rel_all <= NORM_BASELINE_WINDOW_S)

    #         b_by_k = {}
    #         for k in need_keys:
    #             arr = traces.get(k, None)
    #             if arr is None or len(arr) != N:
    #                 b_by_k[k] = np.nan
    #                 continue
    #             vals = arr[mids[mask_pre_b]]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = arr[mids[mask_post_b]]
    #                 vals = vals[np.isfinite(vals)]
    #             b_by_k[k] = float(np.mean(vals)) if vals.size else np.nan

    #         for i, idx in enumerate(mids, start=1):
    #             t_rel = float(st.t[int(idx)] - (global_s0 if have_global_stim else 0.0))

    #             # ----- RAW: ALL
    #             row_all = [str(s + 1), str(i), f"{t_rel:.9g}", "all"]
    #             for k in need_keys:
    #                 v = np.nan
    #                 arr = traces.get(k, None)
    #                 if arr is not None and len(arr) == N:
    #                     v = arr[int(idx)]
    #                 row_all.append(f"{v:.9g}" if np.isfinite(v) else "")
    #             rows_all.append(row_all)

    #             # ----- NORM: ALL
    #             row_allN = [str(s + 1), str(i), f"{t_rel:.9g}", "all"]
    #             for k in need_keys:
    #                 v = np.nan
    #                 arr = traces.get(k, None)
    #                 if arr is not None and len(arr) == N:
    #                     v = arr[int(idx)]
    #                 b = b_by_k.get(k, np.nan)
    #                 vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
    #                 row_allN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
    #             rows_all_N.append(row_allN)

    #             if have_global_stim:
    #                 if t_rel < 0:
    #                     tgt_list = rows_bl; tgt_listN = rows_bl_N; region = "Baseline"
    #                 elif 0.0 <= t_rel <= global_dur:
    #                     tgt_list = rows_st; tgt_listN = rows_st_N; region = "Stim"
    #                 else:
    #                     tgt_list = rows_po; tgt_listN = rows_po_N; region = "Post"

    #                 # RAW regional row
    #                 row_reg = [str(s + 1), str(i), f"{t_rel:.9g}", region]
    #                 for k in need_keys:
    #                     v = np.nan
    #                     arr = traces.get(k, None)
    #                     if arr is not None and len(arr) == N:
    #                         v = arr[int(idx)]
    #                     row_reg.append(f"{v:.9g}" if np.isfinite(v) else "")
    #                 tgt_list.append(row_reg)

    #                 # NORM regional row
    #                 row_regN = [str(s + 1), str(i), f"{t_rel:.9g}", region]
    #                 for k in need_keys:
    #                     v = np.nan
    #                     arr = traces.get(k, None)
    #                     if arr is not None and len(arr) == N:
    #                         v = arr[int(idx)]
    #                     b = b_by_k.get(k, np.nan)
    #                     vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
    #                     row_regN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
    #                 tgt_listN.append(row_regN)

    #     def _pad_row(row, want_len):
    #         if row is None: return [""] * want_len
    #         if len(row) < want_len: return row + [""] * (want_len - len(row))
    #         return row

    #     headers_all = _headers_for_block(None)
    #     headers_bl  = _headers_for_block("baseline")
    #     headers_st  = _headers_for_block("stim")
    #     headers_po  = _headers_for_block("post")

    #     headers_allN = _headers_for_block_norm(None)
    #     headers_blN  = _headers_for_block_norm("baseline")
    #     headers_stN  = _headers_for_block_norm("stim")
    #     headers_poN  = _headers_for_block_norm("post")

    #     have_stim_blocks = have_global_stim and (len(rows_bl) + len(rows_st) + len(rows_po) > 0)

    #     with open(breaths_path, "w", newline="") as f:
    #         w = csv.writer(f)
    #         if not have_stim_blocks:
    #             # RAW + NORM (ALL only)
    #             full_header = headers_all + [""] + headers_allN
    #             w.writerow(full_header)
    #             L = max(len(rows_all), len(rows_all_N))
    #             LA = len(headers_all); LAN = len(headers_allN)
    #             for i in range(L):
    #                 ra  = rows_all[i]   if i < len(rows_all)   else None
    #                 raN = rows_all_N[i] if i < len(rows_all_N) else None
    #                 row = _pad_row(ra, LA) + [""] + _pad_row(raN, LAN)
    #                 w.writerow(row)
    #         else:
    #             # RAW blocks, then NORM blocks
    #             full_header = (
    #                 headers_all + [""] + headers_bl + [""] + headers_st + [""] + headers_po + [""] +
    #                 headers_allN + [""] + headers_blN + [""] + headers_stN + [""] + headers_poN
    #             )
    #             w.writerow(full_header)

    #             L = max(
    #                 len(rows_all), len(rows_bl), len(rows_st), len(rows_po),
    #                 len(rows_all_N), len(rows_bl_N), len(rows_st_N), len(rows_po_N),
    #             )
    #             LA = len(headers_all); LB = len(headers_bl); LS = len(headers_st); LP = len(headers_po)
    #             LAN = len(headers_allN); LBN = len(headers_blN); LSN = len(headers_stN); LPN = len(headers_poN)

    #             for i in range(L):
    #                 ra  = rows_all[i]   if i < len(rows_all)   else None
    #                 rb  = rows_bl[i]    if i < len(rows_bl)    else None
    #                 rs  = rows_st[i]    if i < len(rows_st)    else None
    #                 rp  = rows_po[i]    if i < len(rows_po)    else None
    #                 raN = rows_all_N[i] if i < len(rows_all_N) else None
    #                 rbN = rows_bl_N[i]  if i < len(rows_bl_N)  else None
    #                 rsN = rows_st_N[i]  if i < len(rows_st_N)  else None
    #                 rpN = rows_po_N[i]  if i < len(rows_po_N)  else None

    #                 row = (
    #                     _pad_row(ra, LA) + [""] +
    #                     _pad_row(rb, LB) + [""] +
    #                     _pad_row(rs, LS) + [""] +
    #                     _pad_row(rp, LP) + [""] +
    #                     _pad_row(raN, LAN) + [""] +
    #                     _pad_row(rbN, LBN) + [""] +
    #                     _pad_row(rsN, LSN) + [""] +
    #                     _pad_row(rpN, LPN)
    #                 )
    #                 w.writerow(row)

    #     # ---------- (4) Summary PDF ----------
    #     keys_for_timeplots = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]
    #     label_by_key = {key: label for (label, key) in metrics.METRIC_SPECS if key in keys_for_timeplots}
    #     pdf_path = base.with_name(base.name + "_summary.pdf")
    #     try:
    #         self._save_metrics_summary_pdf(
    #             out_path=pdf_path,
    #             t_ds_csv=t_ds_csv,
    #             y2_ds_by_key=y2_ds_by_key,
    #             keys_for_csv=keys_for_timeplots,
    #             label_by_key=label_by_key,
    #             stim_zero=(global_s0 if have_global_stim else None),
    #             stim_dur=(global_dur if have_global_stim else None),
    #         )
    #     except Exception as e:
    #         print(f"[save][summary-pdf] skipped: {e}")

    #     # ---------- done ----------
    #     msg = f"Saved:\n- {npz_path.name}\n- {csv_time_path.name}\n- {breaths_path.name}\n- {pdf_path.name}"
    #     print("[save]", msg)
    #     try:
    #         self.statusbar.showMessage(msg, 6000)
    #     except Exception:
    #         pass

    # def _export_all_analyzed_data(self):
    #     """
    #     Save:
    #     1) <base>_bundle.npz  (downsampled Y_proc + y2 traces, peaks/breaths, stim spans, meta)
    #     2) <base>_means_by_time.csv  (downsampled time-series: per-sweep (optional), mean, sem)
    #         + appended block of normalized per-sweep traces/mean/sem (suffix *_norm)
    #     3) <base>_breaths.csv  (WIDE per-breath table: ALL | BASELINE | STIM | POST blocks)
    #         + appended duplicate blocks with normalized values (all headers suffixed *_norm)
    #         + NEW column 'is_sigh' indicating breath contains a sigh-marked peak (1/0)
    #     4) <base>_summary.pdf  (figure)
    #     """
    #     st = self.state
    #     if not getattr(self, "_save_dir", None) or not getattr(self, "_save_base", None):
    #         QMessageBox.warning(self, "Save analyzed data", "Choose a save location/name first.")
    #         return
    #     if st.t is None or not st.analyze_chan or st.analyze_chan not in st.sweeps:
    #         QMessageBox.warning(self, "Save analyzed data", "No analyzed data available.")
    #         return

    #     import numpy as np, csv, json
    #     from PyQt6.QtCore import Qt
    #     from PyQt6.QtWidgets import QApplication

    #     # ---------- knobs ----------
    #     DS_TARGET_HZ    = 50.0
    #     CSV_FLUSH_EVERY = 2000
    #     INCLUDE_TRACES  = bool(getattr(self, "_csv_include_traces", True))
    #     NORM_BASELINE_WINDOW_S = float(getattr(self, "_norm_window_s", 10.0))
    #     EPS_BASE = 1e-12

    #     # ---------- basics ----------
    #     any_ch    = next(iter(st.sweeps.values()))
    #     n_sweeps  = int(any_ch.shape[1])
    #     N         = int(len(st.t))
    #     kept_sweeps = [s for s in range(n_sweeps) if s not in getattr(st, "omitted_sweeps", set())]
    #     S = len(kept_sweeps)
    #     if S == 0:
    #         QMessageBox.warning(self, "Save analyzed data", "All sweeps are omitted. Nothing to save.")
    #         return

    #     # Downsample index used for NPZ + time CSV
    #     ds_step = max(1, int(round(float(st.sr_hz) / DS_TARGET_HZ))) if st.sr_hz else 1
    #     ds_idx  = np.arange(0, N, ds_step, dtype=int)
    #     M       = int(len(ds_idx))

    #     # Global stim zero and duration (union across KEPT sweeps)
    #     global_s0, global_s1 = None, None
    #     if st.stim_chan:
    #         for s in kept_sweeps:
    #             spans = st.stim_spans_by_sweep.get(s, [])
    #             if spans:
    #                 starts = [a for (a, _) in spans]
    #                 ends   = [b for (_, b) in spans]
    #                 m0 = float(min(starts)); m1 = float(max(ends))
    #                 global_s0 = m0 if global_s0 is None else min(global_s0, m0)
    #                 global_s1 = m1 if global_s1 is None else max(global_s1, m1)
    #     have_global_stim = (global_s0 is not None and global_s1 is not None)
    #     global_dur = (global_s1 - global_s0) if have_global_stim else None

    #     # Time for NPZ (raw) and for CSV (normalized to global_s0 if present)
    #     t_ds_raw = st.t[ds_idx]
    #     csv_t0   = (global_s0 if have_global_stim else 0.0)
    #     t_ds_csv = (st.t - csv_t0)[ds_idx]

    #     # ---------- containers ----------
    #     all_keys     = self._metric_keys_in_order()
    #     Y_proc_ds    = np.full((M, S), np.nan, dtype=float)
    #     y2_ds_by_key = {k: np.full((M, S), np.nan, dtype=float) for k in all_keys}

    #     peaks_by_sweep, on_by_sweep, off_by_sweep, exm_by_sweep, exo_by_sweep = [], [], [], [], []

    #     # ---------- fill per kept sweep ----------
    #     for col, s in enumerate(kept_sweeps):
    #         y_proc = self._get_processed_for(st.analyze_chan, s)
    #         Y_proc_ds[:, col] = y_proc[ds_idx]

    #         pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #         br  = st.breath_by_sweep.get(s, None)
    #         if br is None and pks.size:
    #             br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #             st.breath_by_sweep[s] = br
    #         if br is None:
    #             br = {
    #                 "onsets":  np.array([], dtype=int),
    #                 "offsets": np.array([], dtype=int),
    #                 "expmins": np.array([], dtype=int),
    #                 "expoffs": np.array([], dtype=int),
    #             }

    #         peaks_by_sweep.append(pks)
    #         on_by_sweep.append(np.asarray(br.get("onsets",  []), dtype=int))
    #         off_by_sweep.append(np.asarray(br.get("offsets", []), dtype=int))
    #         exm_by_sweep.append(np.asarray(br.get("expmins", []), dtype=int))
    #         exo_by_sweep.append(np.asarray(br.get("expoffs", []), dtype=int))

    #         for k in all_keys:
    #             y2 = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #             if y2 is not None and len(y2) == N:
    #                 y2_ds_by_key[k][:, col] = y2[ds_idx]

    #     # ---------- (1) NPZ bundle (downsampled) ----------
    #     base     = self._save_dir / self._save_base
    #     npz_path = base.with_name(base.name + "_bundle.npz")

    #     stim_obj = np.empty(S, dtype=object)
    #     for col, s in enumerate(kept_sweeps):
    #         spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
    #         stim_obj[col] = np.array(spans, dtype=float).reshape(-1, 2) if spans else np.empty((0, 2), dtype=float)

    #     peaks_obj = np.array(peaks_by_sweep, dtype=object)
    #     on_obj    = np.array(on_by_sweep,  dtype=object)
    #     off_obj   = np.array(off_by_sweep, dtype=object)
    #     exm_obj   = np.array(exm_by_sweep, dtype=object)
    #     exo_obj   = np.array(exo_by_sweep, dtype=object)

    #     y2_kwargs_ds = {f"y2_{k}_ds": y2_ds_by_key[k] for k in all_keys}

    #     meta = {
    #         "analyze_channel": st.analyze_chan,
    #         "sr_hz": float(st.sr_hz),
    #         "n_sweeps_total": int(n_sweeps),
    #         "n_sweeps_kept": int(S),
    #         "kept_sweeps": [int(s) for s in kept_sweeps],
    #         "omitted_sweeps": sorted(int(x) for x in getattr(st, "omitted_sweeps", set())),
    #         "abf_path": str(getattr(st, "in_path", "")),
    #         "ui_meta": getattr(self, "_save_meta", {}),
    #         "excluded_for_csv": sorted(list(self._EXCLUDE_FOR_CSV)),
    #         "ds_target_hz": float(DS_TARGET_HZ),
    #         "ds_step": int(ds_step),
    #         "csv_time_zero": float(csv_t0),
    #         "csv_includes_traces": bool(INCLUDE_TRACES),
    #         "norm_window_s": float(NORM_BASELINE_WINDOW_S),
    #     }

    #     np.savez_compressed(
    #         npz_path,
    #         t_ds=t_ds_raw,
    #         Y_proc_ds=Y_proc_ds,
    #         peaks_by_sweep=peaks_obj,
    #         onsets_by_sweep=on_obj,
    #         offsets_by_sweep=off_obj,
    #         expmins_by_sweep=exm_obj,
    #         expoffs_by_sweep=exo_obj,
    #         stim_spans_by_sweep=stim_obj,
    #         meta_json=json.dumps(meta),
    #         **y2_kwargs_ds,
    #     )

    #     # ---------- helpers for normalization ----------
    #     def _per_sweep_baseline_for_time(A_ds: np.ndarray) -> np.ndarray:
    #         b = np.full((A_ds.shape[1],), np.nan, dtype=float)
    #         mask_pre  = (t_ds_csv >= -NORM_BASELINE_WINDOW_S) & (t_ds_csv < 0.0)
    #         mask_post = (t_ds_csv >=  0.0) & (t_ds_csv <= NORM_BASELINE_WINDOW_S)
    #         for sidx in range(A_ds.shape[1]):
    #             col = A_ds[:, sidx]
    #             vals = col[mask_pre]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = col[mask_post]
    #                 vals = vals[np.isfinite(vals)]
    #             if vals.size:
    #                 b[sidx] = float(np.mean(vals))
    #         return b

    #     def _normalize_matrix_by_baseline(A_ds: np.ndarray, b: np.ndarray) -> np.ndarray:
    #         out = np.full_like(A_ds, np.nan)
    #         for sidx in range(A_ds.shape[1]):
    #             bs = b[sidx]
    #             if np.isfinite(bs) and abs(bs) > EPS_BASE:
    #                 out[:, sidx] = A_ds[:, sidx] / bs
    #         return out

    #     # ---------- (2) Per-time CSV (raw + normalized appended) ----------
    #     csv_time_path = base.with_name(base.name + "_means_by_time.csv")
    #     keys_for_csv  = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]

    #     y2_ds_by_key_norm = {}
    #     baseline_by_key   = {}
    #     for k in keys_for_csv:
    #         b = _per_sweep_baseline_for_time(y2_ds_by_key[k])
    #         baseline_by_key[k] = b
    #         y2_ds_by_key_norm[k] = _normalize_matrix_by_baseline(y2_ds_by_key[k], b)

    #     header = ["t"]
    #     for k in keys_for_csv:
    #         if INCLUDE_TRACES:
    #             header += [f"{k}_s{j+1}" for j in range(S)]
    #         header += [f"{k}_mean", f"{k}_sem"]
    #     for k in keys_for_csv:
    #         if INCLUDE_TRACES:
    #             header += [f"{k}_norm_s{j+1}" for j in range(S)]
    #         header += [f"{k}_norm_mean", f"{k}_norm_sem"]

    #     self.setCursor(Qt.CursorShape.WaitCursor)
    #     try:
    #         with open(csv_time_path, "w", newline="") as f:
    #             w = csv.writer(f)
    #             w.writerow(header)
    #             for i in range(M):
    #                 row = [f"{t_ds_csv[i]:.9f}"]
    #                 for k in keys_for_csv:
    #                     col = y2_ds_by_key[k][i, :]
    #                     if INCLUDE_TRACES:
    #                         row += [f"{v:.9g}" if np.isfinite(v) else "" for v in col]
    #                     m, sem = self._mean_sem_1d(col)
    #                     row += [f"{m:.9g}", f"{sem:.9g}"]
    #                 for k in keys_for_csv:
    #                     colN = y2_ds_by_key_norm[k][i, :]
    #                     if INCLUDE_TRACES:
    #                         row += [f"{v:.9g}" if np.isfinite(v) else "" for v in colN]
    #                     mN, semN = self._mean_sem_1d(colN)
    #                     row += [f"{mN:.9g}", f"{semN:.9g}"]
    #                 w.writerow(row)
    #                 if (i % CSV_FLUSH_EVERY) == 0:
    #                     QApplication.processEvents()
    #     finally:
    #         self.unsetCursor()

    #     # ---------- (3) Per-breath CSV (WIDE) ----------
    #     breaths_path = base.with_name(base.name + "_breaths.csv")

    #     # NOTE: We add 'is_sigh' as the last column in each block.
    #     BREATH_COLS = [
    #         "sweep", "breath", "t", "region",
    #         "if", "amp_insp", "amp_exp", "area_insp", "area_exp",
    #         "ti", "te", "vent_proxy",
    #         "is_sigh",                       # NEW
    #     ]
    #     def _headers_for_block(suffix: str | None) -> list[str]:
    #         if not suffix: return BREATH_COLS[:]
    #         return [f"{c}_{suffix}" for c in BREATH_COLS]

    #     def _headers_for_block_norm(suffix: str | None) -> list[str]:
    #         base_cols = _headers_for_block(suffix)
    #         return [h + "_norm" for h in base_cols]

    #     rows_all, rows_bl, rows_st, rows_po = [], [], [], []
    #     rows_all_N, rows_bl_N, rows_st_N, rows_po_N = [], [], [], []

    #     need_keys = ["if", "amp_insp", "amp_exp", "area_insp", "area_exp", "ti", "te", "vent_proxy"]

    #     for s in kept_sweeps:
    #         y_proc = self._get_processed_for(st.analyze_chan, s)
    #         pks    = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #         br     = st.breath_by_sweep.get(s, None)
    #         if br is None and pks.size:
    #             br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #             st.breath_by_sweep[s] = br
    #         if br is None:
    #             br = {"onsets": np.array([], dtype=int)}

    #         on = np.asarray(br.get("onsets", []), dtype=int)
    #         if on.size < 2:
    #             continue

    #         mids = (on[:-1] + on[1:]) // 2
    #         # Sighs for this sweep (peak indices)
    #         sigh_idx = np.asarray(self.state.sigh_by_sweep.get(s, []), dtype=int)

    #         # For each breath interval, mark 1 if any sigh peak falls inside [on[j], on[j+1])
    #         is_sigh_per_breath = np.zeros(len(on) - 1, dtype=int)
    #         if sigh_idx.size:
    #             for j in range(len(on) - 1):
    #                 a, b = int(on[j]), int(on[j+1])
    #                 if np.any((sigh_idx >= a) & (sigh_idx < b)):
    #                     is_sigh_per_breath[j] = 1




    #         # Precompute metric traces for this sweep
    #         traces = {}
    #         for k in need_keys:
    #             if k in metrics.METRICS:
    #                 traces[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #             else:
    #                 traces[k] = None

    #         # Per-breath sigh flags: 1 if any sigh-marked peak lies in [on[i], on[i+1])
    #         # sigh_set = set(int(x) for x in getattr(st, "sighs_by_sweep", {}).get(s, set()))
    #         sigh_set = self._sigh_sample_indices(s, pks)

    #         is_sigh_flags = []
    #         for i in range(len(on) - 1):
    #             a = int(on[i]); b = int(on[i + 1])
    #             sigh_here = any((a <= p < b) for p in sigh_set)
    #             is_sigh_flags.append(1 if sigh_here else 0)

    #         # Baselines for normalization (use breath midpoints)
    #         t_rel_all = (st.t[mids] - (global_s0 if have_global_stim else 0.0)).astype(float)
    #         mask_pre_b  = (t_rel_all >= -NORM_BASELINE_WINDOW_S) & (t_rel_all < 0.0)
    #         mask_post_b = (t_rel_all >=  0.0) & (t_rel_all <= NORM_BASELINE_WINDOW_S)

    #         b_by_k = {}
    #         for k in need_keys:
    #             arr = traces.get(k, None)
    #             if arr is None or len(arr) != N:
    #                 b_by_k[k] = np.nan
    #                 continue
    #             vals = arr[mids[mask_pre_b]]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = arr[mids[mask_post_b]]
    #                 vals = vals[np.isfinite(vals)]
    #             b_by_k[k] = float(np.mean(vals)) if vals.size else np.nan

    #         for i, idx in enumerate(mids, start=1):
    #             t_rel = float(st.t[int(idx)] - (global_s0 if have_global_stim else 0.0))
    #             sigh_flag = is_sigh_flags[i - 1] if (i - 1) < len(is_sigh_flags) else 0

    #             # ----- RAW: ALL
    #             row_all = [str(s + 1), str(i), f"{t_rel:.9g}", "all"]
    #             for k in need_keys:
    #                 v = np.nan
    #                 arr = traces.get(k, None)
    #                 if arr is not None and len(arr) == N:
    #                     v = arr[int(idx)]
    #                 row_all.append(f"{v:.9g}" if np.isfinite(v) else "")
    #             row_all.append(str(int(sigh_flag)))  # is_sigh
    #             rows_all.append(row_all)

    #             # ----- NORM: ALL (same is_sigh value)
    #             row_allN = [str(s + 1), str(i), f"{t_rel:.9g}", "all"]
    #             for k in need_keys:
    #                 v = np.nan
    #                 arr = traces.get(k, None)
    #                 if arr is not None and len(arr) == N:
    #                     v = arr[int(idx)]
    #                 b = b_by_k.get(k, np.nan)
    #                 vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
    #                 row_allN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
    #             row_allN.append(str(int(sigh_flag)))  # is_sigh_norm (kept identical)
    #             rows_all_N.append(row_allN)

    #             if have_global_stim:
    #                 if t_rel < 0:
    #                     tgt_list = rows_bl; tgt_listN = rows_bl_N; region = "Baseline"
    #                 elif 0.0 <= t_rel <= global_dur:
    #                     tgt_list = rows_st; tgt_listN = rows_st_N; region = "Stim"
    #                 else:
    #                     tgt_list = rows_po; tgt_listN = rows_po_N; region = "Post"

    #                 # RAW regional row
    #                 row_reg = [str(s + 1), str(i), f"{t_rel:.9g}", region]
    #                 for k in need_keys:
    #                     v = np.nan
    #                     arr = traces.get(k, None)
    #                     if arr is not None and len(arr) == N:
    #                         v = arr[int(idx)]
    #                     row_reg.append(f"{v:.9g}" if np.isfinite(v) else "")
    #                 row_reg.append(str(int(sigh_flag)))
    #                 tgt_list.append(row_reg)

    #                 # NORM regional row
    #                 row_regN = [str(s + 1), str(i), f"{t_rel:.9g}", region]
    #                 for k in need_keys:
    #                     v = np.nan
    #                     arr = traces.get(k, None)
    #                     if arr is not None and len(arr) == N:
    #                         v = arr[int(idx)]
    #                     b = b_by_k.get(k, np.nan)
    #                     vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
    #                     row_regN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
    #                 row_regN.append(str(int(sigh_flag)))
    #                 tgt_listN.append(row_regN)

    #     def _pad_row(row, want_len):
    #         if row is None: return [""] * want_len
    #         if len(row) < want_len: return row + [""] * (want_len - len(row))
    #         return row

    #     headers_all = _headers_for_block(None)
    #     headers_bl  = _headers_for_block("baseline")
    #     headers_st  = _headers_for_block("stim")
    #     headers_po  = _headers_for_block("post")

    #     headers_allN = _headers_for_block_norm(None)
    #     headers_blN  = _headers_for_block_norm("baseline")
    #     headers_stN  = _headers_for_block_norm("stim")
    #     headers_poN  = _headers_for_block_norm("post")

    #     have_stim_blocks = have_global_stim and (len(rows_bl) + len(rows_st) + len(rows_po) > 0)

    #     with open(breaths_path, "w", newline="") as f:
    #         w = csv.writer(f)
    #         if not have_stim_blocks:
    #             full_header = headers_all + [""] + headers_allN
    #             w.writerow(full_header)
    #             L = max(len(rows_all), len(rows_all_N))
    #             LA = len(headers_all); LAN = len(headers_allN)
    #             for i in range(L):
    #                 ra  = rows_all[i]   if i < len(rows_all)   else None
    #                 raN = rows_all_N[i] if i < len(rows_all_N) else None
    #                 row = _pad_row(ra, LA) + [""] + _pad_row(raN, LAN)
    #                 w.writerow(row)
    #         else:
    #             full_header = (
    #                 headers_all + [""] + headers_bl + [""] + headers_st + [""] + headers_po + [""] +
    #                 headers_allN + [""] + headers_blN + [""] + headers_stN + [""] + headers_poN
    #             )
    #             w.writerow(full_header)

    #             L = max(
    #                 len(rows_all), len(rows_bl), len(rows_st), len(rows_po),
    #                 len(rows_all_N), len(rows_bl_N), len(rows_st_N), len(rows_po_N),
    #             )
    #             LA = len(headers_all); LB = len(headers_bl); LS = len(headers_st); LP = len(headers_po)
    #             LAN = len(headers_allN); LBN = len(headers_blN); LSN = len(headers_stN); LPN = len(headers_poN)

    #             for i in range(L):
    #                 ra  = rows_all[i]   if i < len(rows_all)   else None
    #                 rb  = rows_bl[i]    if i < len(rows_bl)    else None
    #                 rs  = rows_st[i]    if i < len(rows_st)    else None
    #                 rp  = rows_po[i]    if i < len(rows_po)    else None
    #                 raN = rows_all_N[i] if i < len(rows_all_N) else None
    #                 rbN = rows_bl_N[i]  if i < len(rows_bl_N)  else None
    #                 rsN = rows_st_N[i]  if i < len(rows_st_N)  else None
    #                 rpN = rows_po_N[i]  if i < len(rows_po_N)  else None

    #                 row = (
    #                     _pad_row(ra, LA) + [""] +
    #                     _pad_row(rb, LB) + [""] +
    #                     _pad_row(rs, LS) + [""] +
    #                     _pad_row(rp, LP) + [""] +
    #                     _pad_row(raN, LAN) + [""] +
    #                     _pad_row(rbN, LBN) + [""] +
    #                     _pad_row(rsN, LSN) + [""] +
    #                     _pad_row(rpN, LPN)
    #                 )
    #                 w.writerow(row)

    #     # ---------- (4) Summary PDF ----------
    #     keys_for_timeplots = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]
    #     label_by_key = {key: label for (label, key) in metrics.METRIC_SPECS if key in keys_for_timeplots}
    #     pdf_path = base.with_name(base.name + "_summary.pdf")
    #     try:
    #         self._save_metrics_summary_pdf(
    #             out_path=pdf_path,
    #             t_ds_csv=t_ds_csv,
    #             y2_ds_by_key=y2_ds_by_key,
    #             keys_for_csv=keys_for_timeplots,
    #             label_by_key=label_by_key,
    #             stim_zero=(global_s0 if have_global_stim else None),
    #             stim_dur=(global_dur if have_global_stim else None),
    #         )
    #     except Exception as e:
    #         print(f"[save][summary-pdf] skipped: {e}")

    #     msg = f"Saved:\n- {npz_path.name}\n- {csv_time_path.name}\n- {breaths_path.name}\n- {pdf_path.name}"
    #     print("[save]", msg)
    #     try:
    #         self.statusbar.showMessage(msg, 6000)
    #     except Exception:
    #         pass

    def _export_all_analyzed_data(self, preview_only=False, progress_dialog=None):
        """
        Exports (or previews) analyzed data.

        If preview_only=True: Shows interactive PDF preview dialog without saving files.
        If preview_only=False: Prompts for location/name and exports files.

        Exports:
        1) <base>_bundle.npz
            - Downsampled processed trace (kept sweeps only)
            - Downsampled y2 metric traces (all keys)
            - Peaks/breaths/sighs per kept sweep
            - Stim spans per kept sweep
            - Meta

        2) <base>_means_by_time.csv
            - t (relative to global stim start if present)
            - For each metric: optional per-sweep traces, mean, sem
            - Then the same block normalized by per-sweep baseline window (_norm)
            - Then the same block normalized by pooled eupneic baseline (_norm_eupnea)

        3) <base>_breaths.csv
            - Wide layout:
                RAW blocks:  ALL | BASELINE | STIM | POST
                NORM blocks: ALL | BASELINE | STIM | POST (per-sweep time-based)
                NORM_EUPNEA blocks: ALL | BASELINE | STIM | POST (pooled eupneic)
            - Includes `is_sigh` column (1 if any sigh peak in that breath interval)

        4) <base>_events.csv
            - Event intervals: stimulus on/off, apnea episodes, eupnea regions
            - Columns: sweep, event_type, start_time, end_time, duration
            - Times are relative to global stim start if present

        5) <base>_summary.pdf (or preview dialog if preview_only=True)
        """
        import numpy as np, csv, json, time
        from PyQt6.QtWidgets import QApplication

        st = self.state
        if not getattr(st, "in_path", None):
            QMessageBox.information(self, "View Summary" if preview_only else "Save analyzed data", "Open an ABF first.")
            return
        if st.t is None or not st.analyze_chan or st.analyze_chan not in st.sweeps:
            QMessageBox.warning(self, "View Summary" if preview_only else "Save analyzed data", "No analyzed data available.")
            return

        # -------------------- Prompt for save location (skip if preview_only) --------------------
        if not preview_only:
            # --- Build an auto stim string from current sweep metrics, if available ---
            def _auto_stim_from_metrics() -> str:
                s = max(0, min(getattr(st, "sweep_idx", 0), self._sweep_count()-1))
                m = st.stim_metrics_by_sweep.get(s, {}) if getattr(st, "stim_metrics_by_sweep", None) else {}
                if not m:
                    return ""
                def _ri(x):
                    try: return int(round(float(x)))
                    except Exception: return None

                f = _ri(m.get("freq_hz"))
                d = _ri(m.get("duration_s"))
                pw_s = m.get("pulse_width_s")
                n_pulses = m.get("n_pulses", None)

                if f is not None and d is not None and pw_s is not None:
                    pw_str = f"{_ri(pw_s)}s" if pw_s >= 1.0 else f"{_ri(pw_s * 1000)}ms"
                    return f"{f}Hz{d}s{pw_str}"

                if pw_s is not None and (n_pulses == 1 or f is None):
                    return f"{_ri(pw_s)}sPulse" if pw_s >= 1.0 else f"{_ri(pw_s * 1000)}msPulse"
                return ""

            abf_stem = st.in_path.stem
            chan = st.analyze_chan or ""
            auto_stim = _auto_stim_from_metrics()

            # --- Load autocomplete history ---
            history = self._load_save_dialog_history()

            # --- Name builder dialog (with auto stim suggestion) ---
            dlg = self.SaveMetaDialog(abf_name=abf_stem, channel=chan, parent=self, auto_stim=auto_stim, history=history)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return

            vals = dlg.values()

            # --- Update history with new values ---
            self._update_save_dialog_history(vals)
            suggested = self._sanitize_token(vals["preview"]) or "analysis"
            want_picker = bool(vals.get("choose_dir", False))

            target_exact = "Pleth_App_analysis"          # <- required folder name (lowercase 'a')
            target_lower = target_exact.lower()

            def _nearest_analysis_ancestor(p: Path) -> Path | None:
                # Return the closest ancestor (including self) named Pleth_App_analysis (case-insensitive)
                for cand in [p] + list(p.parents):
                    if cand.name.lower() == target_lower:
                        return cand
                return None

            if want_picker:
                # User chooses a folder; we keep your previous smart logic here.
                default_root = Path(self.settings.value("save_root", str(st.in_path.parent)))
                chosen = QFileDialog.getExistingDirectory(
                    self,
                    "Choose a folder (files may go into an existing 'Pleth_App_analysis' here)",
                    str(default_root)
                )
                if not chosen:
                    return
                chosen_path = Path(chosen)

                # 1) If chosen folder is inside an ancestor named Pleth_App_analysis → save THERE (the ancestor)
                anc = _nearest_analysis_ancestor(chosen_path)
                if anc is not None:
                    final_dir = anc
                else:
                    # 2) If the chosen folder already contains 'Pleth_App_analysis' or 'Pleth_App_Analysis' subfolder → use it
                    sub_exact   = chosen_path / "Pleth_App_analysis"
                    sub_variant = chosen_path / "Pleth_App_Analysis"
                    if sub_exact.is_dir():
                        final_dir = sub_exact
                    elif sub_variant.is_dir():
                        final_dir = sub_variant
                    else:
                        # 3) Otherwise, create Pleth_App_analysis directly under the chosen folder
                        final_dir = chosen_path / target_exact
                        try:
                            final_dir.mkdir(parents=True, exist_ok=True)
                        except Exception as e:
                            QMessageBox.critical(self, "Save analyzed data", f"Could not create folder:\n{final_dir}\n\n{e}")
                            return

                # Remember the last *picker* root only when the picker is used
                self.settings.setValue("save_root", str(chosen_path))

            else:
                # UNCHECKED: Always use the CURRENT ABF DIRECTORY (not a remembered one)
                base_root = st.in_path.parent if getattr(st, "in_path", None) else Path.cwd()
                final_dir = base_root / target_exact
                try:
                    final_dir.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    QMessageBox.critical(self, "Save analyzed data", f"Could not create folder:\n{final_dir}\n\n{e}")
                    return
                # IMPORTANT: Do NOT overwrite 'save_root' here — we don't want to "remember" anything for the unchecked case.

            # Set base name + meta, then export
            self._save_dir = final_dir
            self._save_base = suggested
            self._save_meta = vals

            base_path = self._save_dir / self._save_base
            print(f"[save] base path set: {base_path}")
            try:
                self.statusbar.showMessage(f"Saving to: {base_path}", 4000)
            except Exception:
                pass

        # -------------------- knobs --------------------
        DS_TARGET_HZ    = 50.0
        CSV_FLUSH_EVERY = 2000
        INCLUDE_TRACES  = bool(getattr(self, "_csv_include_traces", True))
        NORM_BASELINE_WINDOW_S = float(getattr(self, "_norm_window_s", 10.0))
        EPS_BASE = 1e-12

        # -------------------- basics --------------------
        if progress_dialog:
            progress_dialog.setLabelText("Preparing data...")
            progress_dialog.setValue(5)
            QApplication.processEvents()

        any_ch    = next(iter(st.sweeps.values()))
        n_sweeps  = int(any_ch.shape[1])
        N         = int(len(st.t))

        kept_sweeps = [s for s in range(n_sweeps) if s not in getattr(st, "omitted_sweeps", set())]
        S = len(kept_sweeps)
        if S == 0:
            QMessageBox.warning(self, "Save analyzed data", "All sweeps are omitted. Nothing to save.")
            return

        # Downsample index used for NPZ + time CSV
        ds_step = max(1, int(round(float(st.sr_hz) / DS_TARGET_HZ))) if st.sr_hz else 1
        ds_idx  = np.arange(0, N, ds_step, dtype=int)
        M       = int(len(ds_idx))

        # Global stim zero and duration (union across KEPT sweeps)
        global_s0, global_s1 = None, None
        if st.stim_chan:
            for s in kept_sweeps:
                spans = st.stim_spans_by_sweep.get(s, [])
                if spans:
                    starts = [a for (a, _) in spans]
                    ends   = [b for (_, b) in spans]
                    m0 = float(min(starts)); m1 = float(max(ends))
                    global_s0 = m0 if global_s0 is None else min(global_s0, m0)
                    global_s1 = m1 if global_s1 is None else max(global_s1, m1)
        have_global_stim = (global_s0 is not None and global_s1 is not None)
        global_dur = (global_s1 - global_s0) if have_global_stim else None

        # Time for NPZ (raw) and for CSV (normalized to global_s0 if present)
        t_ds_raw = st.t[ds_idx]
        csv_t0   = (global_s0 if have_global_stim else 0.0)
        t_ds_csv = (st.t - csv_t0)[ds_idx]

        # -------------------- containers --------------------
        all_keys     = self._metric_keys_in_order()
        Y_proc_ds    = np.full((M, S), np.nan, dtype=float)
        y2_ds_by_key = {k: np.full((M, S), np.nan, dtype=float) for k in all_keys}

        peaks_by_sweep, on_by_sweep, off_by_sweep, exm_by_sweep, exo_by_sweep = [], [], [], [], []
        sigh_by_sweep = []

        # -------------------- fill per kept sweep --------------------
        for col, s in enumerate(kept_sweeps):
            y_proc = self._get_processed_for(st.analyze_chan, s)
            Y_proc_ds[:, col] = y_proc[ds_idx]

            pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
            br  = st.breath_by_sweep.get(s, None)
            if br is None and pks.size:
                # backfill breaths for this sweep so exports are consistent
                br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
                st.breath_by_sweep[s] = br
            if br is None:
                br = {
                    "onsets":  np.array([], dtype=int),
                    "offsets": np.array([], dtype=int),
                    "expmins": np.array([], dtype=int),
                    "expoffs": np.array([], dtype=int),
                }

            peaks_by_sweep.append(pks)
            on_by_sweep.append(np.asarray(br.get("onsets",  []), dtype=int))
            off_by_sweep.append(np.asarray(br.get("offsets", []), dtype=int))
            exm_by_sweep.append(np.asarray(br.get("expmins", []), dtype=int))
            exo_by_sweep.append(np.asarray(br.get("expoffs", []), dtype=int))

            # sighs for this sweep (peak indices)
            sighs = np.asarray(st.sigh_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
            # Keep only valid in-range indices
            sighs = sighs[(sighs >= 0) & (sighs < len(y_proc))]
            sigh_by_sweep.append(sighs)

            for k in all_keys:
                y2 = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
                if y2 is not None and len(y2) == N:
                    y2_ds_by_key[k][:, col] = y2[ds_idx]

        if progress_dialog:
            progress_dialog.setLabelText("Computing metrics...")
            progress_dialog.setValue(15)
            QApplication.processEvents()

        # -------------------- Save files (skip if preview_only) --------------------
        if not preview_only:
            # -------------------- (1) NPZ bundle (downsampled) --------------------
            base     = self._save_dir / self._save_base
            npz_path = base.with_name(base.name + "_bundle.npz")
    
            # Pack stim spans in KEPT order (align with columns)
            stim_obj = np.empty(S, dtype=object)
            for col, s in enumerate(kept_sweeps):
                spans = st.stim_spans_by_sweep.get(s, []) if st.stim_chan else []
                stim_obj[col] = np.array(spans, dtype=float).reshape(-1, 2) if spans else np.empty((0, 2), dtype=float)
    
            peaks_obj = np.array(peaks_by_sweep, dtype=object)
            on_obj    = np.array(on_by_sweep,  dtype=object)
            off_obj   = np.array(off_by_sweep, dtype=object)
            exm_obj   = np.array(exm_by_sweep, dtype=object)
            exo_obj   = np.array(exo_by_sweep, dtype=object)
            sigh_obj  = np.array(sigh_by_sweep, dtype=object)

            # Pack sniffing regions (aligned with kept_sweeps)
            sniff_obj = np.empty(S, dtype=object)
            for col, s in enumerate(kept_sweeps):
                regions = st.sniff_regions_by_sweep.get(s, [])
                # Convert to numpy array of shape (N_regions, 2) for start/end times
                sniff_obj[col] = np.array(regions, dtype=float).reshape(-1, 2) if regions else np.empty((0, 2), dtype=float)

            y2_kwargs_ds = {f"y2_{k}_ds": y2_ds_by_key[k] for k in all_keys}
    
            meta = {
                "analyze_channel": st.analyze_chan,
                "sr_hz": float(st.sr_hz),
                "n_sweeps_total": int(n_sweeps),
                "n_sweeps_kept": int(S),
                "kept_sweeps": [int(s) for s in kept_sweeps],                  # original indices
                "omitted_sweeps": sorted(int(x) for x in getattr(st, "omitted_sweeps", set())),
                "abf_path": str(getattr(st, "in_path", "")),
                "ui_meta": getattr(self, "_save_meta", {}),
                "excluded_for_csv": sorted(list(self._EXCLUDE_FOR_CSV)),
                "ds_target_hz": float(DS_TARGET_HZ),
                "ds_step": int(ds_step),
                "csv_time_zero": float(csv_t0),
                "csv_includes_traces": bool(INCLUDE_TRACES),
                "norm_window_s": float(NORM_BASELINE_WINDOW_S),
                # Filter settings (for future NPZ reopening)
                "use_low": bool(st.use_low),
                "use_high": bool(st.use_high),
                "use_mean_sub": bool(st.use_mean_sub),
                "use_invert": bool(st.use_invert),
                "low_hz": float(st.low_hz) if st.low_hz else None,
                "high_hz": float(st.high_hz) if st.high_hz else None,
                "mean_val": float(st.mean_val),
                "filter_order": int(self.filter_order),
                # Notch filter (if active)
                "notch_filter_lower": float(self.notch_filter_lower) if self.notch_filter_lower else None,
                "notch_filter_upper": float(self.notch_filter_upper) if self.notch_filter_upper else None,
                # Channel info (for reopening)
                "channel_names": list(st.channel_names) if st.channel_names else [],
                "stim_chan": str(st.stim_chan) if st.stim_chan else None,
                # Navigation state
                "window_dur_s": float(st.window_dur_s),
            }
    
            # Save enhanced NPZ bundle with timeseries data (for fast consolidation)
            # This will be updated after timeseries normalization is computed
            _npz_timeseries_data = {
                't_ds': t_ds_raw,
                'Y_proc_ds': Y_proc_ds,
                'peaks_by_sweep': peaks_obj,
                'onsets_by_sweep': on_obj,
                'offsets_by_sweep': off_obj,
                'expmins_by_sweep': exm_obj,
                'expoffs_by_sweep': exo_obj,
                'sigh_idx_by_sweep': sigh_obj,
                'sniff_regions_by_sweep': sniff_obj,
                'stim_spans_by_sweep': stim_obj,
                'meta_json': json.dumps(meta),
                **y2_kwargs_ds,
            }
    
            # -------------------- helpers for normalization (time CSV) --------------------
            def _per_sweep_baseline_for_time(A_ds: np.ndarray) -> np.ndarray:
                """
                A_ds: (M,S) downsampled metric matrix.
                Returns b[S]: mean over last NORM_BASELINE_WINDOW_S before 0; fallback to first W after 0.
                """
                b = np.full((A_ds.shape[1],), np.nan, dtype=float)
                mask_pre  = (t_ds_csv >= -NORM_BASELINE_WINDOW_S) & (t_ds_csv < 0.0)
                mask_post = (t_ds_csv >=  0.0) & (t_ds_csv <= NORM_BASELINE_WINDOW_S)
                for sidx in range(A_ds.shape[1]):
                    col = A_ds[:, sidx]
                    vals = col[mask_pre]
                    vals = vals[np.isfinite(vals)]
                    if vals.size == 0:
                        vals = col[mask_post]
                        vals = vals[np.isfinite(vals)]
                    if vals.size:
                        b[sidx] = float(np.mean(vals))
                return b
    
            def _normalize_matrix_by_baseline(A_ds: np.ndarray, b: np.ndarray) -> np.ndarray:
                out = np.full_like(A_ds, np.nan)
                for sidx in range(A_ds.shape[1]):
                    bs = b[sidx]
                    if np.isfinite(bs) and abs(bs) > EPS_BASE:
                        out[:, sidx] = A_ds[:, sidx] / bs
                return out
            
        
    
    
            # -------------------- (2) Per-time CSV (raw + normalized appended) --------------------
            csv_time_path = base.with_name(base.name + "_timeseries.csv")
            keys_for_csv  = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]
    
            if progress_dialog:
                progress_dialog.setLabelText("Computing time-based normalization...")
                progress_dialog.setValue(25)
                QApplication.processEvents()

            # Build normalized stacks per metric (TIME-BASED, per-sweep)
            y2_ds_by_key_norm = {}
            baseline_by_key   = {}
            for k in keys_for_csv:
                b = _per_sweep_baseline_for_time(y2_ds_by_key[k])
                baseline_by_key[k] = b
                y2_ds_by_key_norm[k] = _normalize_matrix_by_baseline(y2_ds_by_key[k], b)

            if progress_dialog:
                progress_dialog.setLabelText("Computing eupnea-based normalization...")
                progress_dialog.setValue(35)
                QApplication.processEvents()

            # Build EUPNEA-BASED normalization (pooled across all sweeps)
            # Use efficient approach: extract from pre-computed matrices
            y2_ds_by_key_norm_eupnea = {}
            eupnea_baseline_by_key = {}

            # Get thresholds from UI
            eupnea_thresh = self._parse_float(self.EupneaThresh) or 5.0  # Hz
            apnea_thresh = self._parse_float(self.ApneaThresh) or 0.5    # seconds

            # Pre-compute eupnea masks once per sweep
            print(f"[CSV-time] Pre-computing eupnea masks for {len(kept_sweeps)} sweeps...")
            eupnea_masks_csv = {}
            for s in kept_sweeps:
                y_proc = self._get_processed_for(st.analyze_chan, s)
                pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                br = st.breath_by_sweep.get(s, None)
                if br is None and pks.size:
                    br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
                    st.breath_by_sweep[s] = br
                if br is None:
                    continue

                on = np.asarray(br.get("onsets", []), dtype=int)
                off = np.asarray(br.get("offsets", []), dtype=int)
                expmins = np.asarray(br.get("expmins", []), dtype=int)
                expoffs = np.asarray(br.get("expoffs", []), dtype=int)

                if on.size >= 2:
                    # Get sniff regions for this sweep
                    sniff_regions = self.state.sniff_regions_by_sweep.get(s, [])
                    eupnea_mask = metrics.detect_eupnic_regions(
                        st.t, y_proc, st.sr_hz, pks, on, off, expmins, expoffs,
                        freq_threshold_hz=eupnea_thresh,
                        min_duration_sec=self.eupnea_min_duration,
                        sniff_regions=sniff_regions
                    )
                    eupnea_masks_csv[s] = eupnea_mask

            # Compute baselines by extracting from y2_ds_by_key matrices
            # OPTIMIZATION: Build index mapping once instead of calling argmin repeatedly
            print(f"[CSV-time] Computing eupnea baselines for {len(keys_for_csv)} metrics...")

            # Pre-compute mapping from downsampled indices to original indices
            t0 = float(global_s0) if have_global_stim else 0.0
            ds_to_orig_idx = np.zeros(len(t_ds_csv), dtype=int)
            for ds_idx in range(len(t_ds_csv)):
                t_target = t_ds_csv[ds_idx] + t0
                ds_to_orig_idx[ds_idx] = np.argmin(np.abs(st.t - t_target))

            # Identify baseline indices (t < 0)
            baseline_ds_mask = t_ds_csv < 0
            poststim_ds_mask = (t_ds_csv >= 0) & (t_ds_csv <= NORM_BASELINE_WINDOW_S)

            for k in keys_for_csv:
                Y_raw = y2_ds_by_key.get(k, None)
                if Y_raw is None or Y_raw.size == 0:
                    eupnea_baseline_by_key[k] = np.nan
                    y2_ds_by_key_norm_eupnea[k] = np.full_like(Y_raw, np.nan) if Y_raw is not None else None
                    continue

                pooled_vals = []

                # Collect eupneic baseline values from downsampled data
                for col_idx, s in enumerate(kept_sweeps):
                    eupnea_mask = eupnea_masks_csv.get(s, None)
                    if eupnea_mask is None:
                        continue

                    # Check baseline time points
                    for ds_idx in np.where(baseline_ds_mask)[0]:
                        i_orig = ds_to_orig_idx[ds_idx]

                        # Check if this point is eupneic
                        if i_orig < len(eupnea_mask) and eupnea_mask[i_orig] > 0:
                            val = Y_raw[ds_idx, col_idx]
                            if np.isfinite(val):
                                pooled_vals.append(val)

                # Fallback: include post-stim eupneic periods if insufficient
                if len(pooled_vals) < 10:
                    for col_idx, s in enumerate(kept_sweeps):
                        eupnea_mask = eupnea_masks_csv.get(s, None)
                        if eupnea_mask is None:
                            continue

                        for ds_idx in np.where(poststim_ds_mask)[0]:
                            i_orig = ds_to_orig_idx[ds_idx]

                            if i_orig < len(eupnea_mask) and eupnea_mask[i_orig] > 0:
                                val = Y_raw[ds_idx, col_idx]
                                if np.isfinite(val):
                                    pooled_vals.append(val)

                # Compute baseline and normalize
                eupnea_b = float(np.mean(pooled_vals)) if len(pooled_vals) > 0 else np.nan
                eupnea_baseline_by_key[k] = eupnea_b

                if np.isfinite(eupnea_b) and abs(eupnea_b) > EPS_BASE:
                    y2_ds_by_key_norm_eupnea[k] = Y_raw / eupnea_b
                else:
                    y2_ds_by_key_norm_eupnea[k] = np.full_like(Y_raw, np.nan)

            # headers: raw first, then time-based norm, then eupnea-based norm
            header = ["t"]
            for k in keys_for_csv:
                if INCLUDE_TRACES:
                    header += [f"{k}_s{j+1}" for j in range(S)]
                header += [f"{k}_mean", f"{k}_sem"]

            for k in keys_for_csv:
                if INCLUDE_TRACES:
                    header += [f"{k}_norm_s{j+1}" for j in range(S)]
                header += [f"{k}_norm_mean", f"{k}_norm_sem"]

            for k in keys_for_csv:
                if INCLUDE_TRACES:
                    header += [f"{k}_norm_eupnea_s{j+1}" for j in range(S)]
                header += [f"{k}_norm_eupnea_mean", f"{k}_norm_eupnea_sem"]
    
            if progress_dialog:
                progress_dialog.setLabelText("Writing time-series CSV...")
                progress_dialog.setValue(50)
                QApplication.processEvents()

            t_start = time.time()
            self.setCursor(Qt.CursorShape.WaitCursor)
            try:
                # Build DataFrame from existing numpy arrays (2-3× faster than row-by-row)
                import pandas as pd

                # Build columns in exact order to match header
                data = {}
                data['t'] = t_ds_csv

                # RAW block
                for k in keys_for_csv:
                    if INCLUDE_TRACES:
                        for j in range(S):
                            data[f'{k}_s{j+1}'] = y2_ds_by_key[k][:, j]
                    # Compute mean and SEM across sweeps
                    means, sems = self._nanmean_sem(y2_ds_by_key[k], axis=1)
                    data[f'{k}_mean'] = means
                    data[f'{k}_sem'] = sems

                # NORMALIZED block (time-based, per-sweep)
                for k in keys_for_csv:
                    if INCLUDE_TRACES:
                        for j in range(S):
                            data[f'{k}_norm_s{j+1}'] = y2_ds_by_key_norm[k][:, j]
                    means_n, sems_n = self._nanmean_sem(y2_ds_by_key_norm[k], axis=1)
                    data[f'{k}_norm_mean'] = means_n
                    data[f'{k}_norm_sem'] = sems_n

                # NORMALIZED block (eupnea-based, pooled)
                for k in keys_for_csv:
                    if INCLUDE_TRACES:
                        for j in range(S):
                            data[f'{k}_norm_eupnea_s{j+1}'] = y2_ds_by_key_norm_eupnea[k][:, j]
                    means_e, sems_e = self._nanmean_sem(y2_ds_by_key_norm_eupnea[k], axis=1)
                    data[f'{k}_norm_eupnea_mean'] = means_e
                    data[f'{k}_norm_eupnea_sem'] = sems_e

                # Create DataFrame and write to CSV
                df = pd.DataFrame(data, columns=header)
                df.to_csv(csv_time_path, index=False, float_format='%.9g', na_rep='')

            finally:
                self.unsetCursor()

            t_elapsed = time.time() - t_start
            print(f"[CSV] ✓ Time-series data written in {t_elapsed:.2f}s")

            # Save enhanced NPZ version 2 with timeseries data (for fast consolidation)
            _npz_timeseries_data['npz_version'] = 2
            _npz_timeseries_data['timeseries_t'] = t_ds_csv
            _npz_timeseries_data['timeseries_keys'] = list(keys_for_csv)
            # Save raw, norm, and eupnea-norm metric matrices
            for k in keys_for_csv:
                _npz_timeseries_data[f'ts_raw_{k}'] = y2_ds_by_key[k]
                _npz_timeseries_data[f'ts_norm_{k}'] = y2_ds_by_key_norm[k]
                _npz_timeseries_data[f'ts_eupnea_{k}'] = y2_ds_by_key_norm_eupnea[k]

            np.savez_compressed(npz_path, **_npz_timeseries_data)

            # Count total sniffing regions across all sweeps
            total_sniff_regions = sum(len(st.sniff_regions_by_sweep.get(s, [])) for s in kept_sweeps)
            print(f"[NPZ] ✓ Enhanced bundle saved (v2) with timeseries data")
            print(f"      - {total_sniff_regions} sniffing region(s) saved across {S} sweep(s)")

            if progress_dialog:
                progress_dialog.setLabelText("Writing breath-by-breath CSV...")
                progress_dialog.setValue(60)
                QApplication.processEvents()

            # -------------------- (3) Per-breath CSV (WIDE; with breath classifications) --------------------
            breaths_path = base.with_name(base.name + "_breaths.csv")

            BREATH_COLS = [
                "sweep", "breath", "t", "region", "is_sigh", "is_sniffing", "is_eupnea", "is_apnea",
                "if", "amp_insp", "amp_exp", "area_insp", "area_exp",
                "ti", "te", "vent_proxy",
            ]
            def _headers_for_block(suffix: str | None) -> list[str]:
                if not suffix: return BREATH_COLS[:]
                return [f"{c}_{suffix}" for c in BREATH_COLS]
    
            def _headers_for_block_norm(suffix: str | None) -> list[str]:
                base_cols = _headers_for_block(suffix)
                return [h + "_norm" for h in base_cols]

            def _headers_for_block_norm_eupnea(suffix: str | None) -> list[str]:
                base_cols = _headers_for_block(suffix)
                return [h + "_norm_eupnea" for h in base_cols]

            rows_all, rows_bl, rows_st, rows_po = [], [], [], []
            rows_all_N, rows_bl_N, rows_st_N, rows_po_N = [], [], [], []
            rows_all_E, rows_bl_E, rows_st_E, rows_po_E = [], [], [], []

            need_keys = ["if", "amp_insp", "amp_exp", "area_insp", "area_exp", "ti", "te", "vent_proxy"]

            # Compute EUPNEA-BASED baselines (pooled across all sweeps)
            # Pre-compute eupnea masks once per sweep to avoid redundant calculations
            eupnea_b_by_k = {}
            eupnea_masks_by_sweep = {}

            # Get thresholds from UI (for breath classification)
            eupnea_thresh = self._parse_float(self.EupneaThresh) or 5.0  # Hz
            apnea_thresh = self._parse_float(self.ApneaThresh) or 0.5    # seconds

            t_start = time.time()
            print(f"[CSV] Pre-computing eupnea masks for {len(kept_sweeps)} sweeps...")
            for s in kept_sweeps:
                y_proc = self._get_processed_for(st.analyze_chan, s)
                pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                br = st.breath_by_sweep.get(s, None)
                if br is None and pks.size:
                    br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
                    st.breath_by_sweep[s] = br
                if br is None:
                    continue

                on = np.asarray(br.get("onsets", []), dtype=int)
                off = np.asarray(br.get("offsets", []), dtype=int)
                expmins = np.asarray(br.get("expmins", []), dtype=int)
                expoffs = np.asarray(br.get("expoffs", []), dtype=int)

                if on.size >= 2:
                    # Get sniff regions for this sweep
                    sniff_regions = st.sniff_regions_by_sweep.get(s, [])

                    # Compute eupnea mask (excluding sniffing regions)
                    eupnea_mask = metrics.detect_eupnic_regions(
                        st.t, y_proc, st.sr_hz, pks, on, off, expmins, expoffs,
                        freq_threshold_hz=eupnea_thresh,
                        min_duration_sec=self.eupnea_min_duration,
                        sniff_regions=sniff_regions
                    )
                    eupnea_masks_by_sweep[s] = eupnea_mask

            t_elapsed = time.time() - t_start
            print(f"[CSV] ✓ Eupnea masks computed in {t_elapsed:.2f}s")

            # Now compute baselines by collecting breath values from eupneic periods
            # IMPORTANT: Cache traces to reuse in main export loop AND PDF generation
            t_start = time.time()
            print(f"[CSV] Computing eupnea baselines and caching traces for {len(need_keys)} metrics...")
            eupnea_baseline_breaths = {k: [] for k in need_keys}

            # Global cache for reuse in PDF generation (stored on self)
            if not hasattr(self, '_global_trace_cache'):
                self._global_trace_cache = {}
            cached_traces_by_sweep = {}  # Cache traces to avoid recomputing

            for s in kept_sweeps:
                # Keep UI responsive during long computation
                if progress_dialog:
                    QApplication.processEvents()

                y_proc = self._get_processed_for(st.analyze_chan, s)
                pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                br = st.breath_by_sweep.get(s, None)
                if br is None or pks.size < 2:
                    continue

                on = np.asarray(br.get("onsets", []), dtype=int)
                if on.size < 2:
                    continue

                eupnea_mask = eupnea_masks_by_sweep.get(s, None)
                if eupnea_mask is None:
                    continue

                mids = (on[:-1] + on[1:]) // 2
                t_rel_all = (st.t[mids] - (global_s0 if have_global_stim else 0.0)).astype(float)

                # Compute traces once for this sweep and CACHE them (both locally and globally)
                traces_for_sweep = {}
                for k in need_keys:
                    if k in metrics.METRICS:
                        traces_for_sweep[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)

                # Store in cache for reuse in main loop
                cached_traces_by_sweep[s] = traces_for_sweep
                # Also store globally for PDF reuse
                self._global_trace_cache[s] = traces_for_sweep

                # Collect baseline eupneic breath values
                for i, mid_idx in enumerate(mids):
                    if t_rel_all[i] >= 0:
                        continue  # Only baseline (t < 0)

                    # Check if this breath is eupneic
                    if mid_idx < len(eupnea_mask) and eupnea_mask[mid_idx] > 0:
                        for k in need_keys:
                            trace = traces_for_sweep.get(k, None)
                            if trace is not None and len(trace) == len(st.t):
                                val = trace[mid_idx]
                                if np.isfinite(val):
                                    eupnea_baseline_breaths[k].append(val)

            # Compute baseline means with fallback if insufficient baseline data
            for k in need_keys:
                if len(eupnea_baseline_breaths[k]) >= 10:
                    eupnea_b_by_k[k] = float(np.mean(eupnea_baseline_breaths[k]))
                elif len(eupnea_baseline_breaths[k]) > 0:
                    eupnea_b_by_k[k] = float(np.mean(eupnea_baseline_breaths[k]))
                else:
                    # Fallback: include post-stim eupneic breaths using CACHED traces
                    fallback_vals = []
                    for s in kept_sweeps:
                        traces_cached = cached_traces_by_sweep.get(s, None)
                        if traces_cached is None:
                            continue

                        br = st.breath_by_sweep.get(s, None)
                        if br is None:
                            continue

                        on = np.asarray(br.get("onsets", []), dtype=int)
                        if on.size < 2:
                            continue

                        eupnea_mask = eupnea_masks_by_sweep.get(s, None)
                        if eupnea_mask is None:
                            continue

                        mids = (on[:-1] + on[1:]) // 2
                        t_rel_all = (st.t[mids] - (global_s0 if have_global_stim else 0.0)).astype(float)

                        trace = traces_cached.get(k, None)
                        if trace is None or len(trace) != len(st.t):
                            continue

                        for i, mid_idx in enumerate(mids):
                            if 0 <= t_rel_all[i] <= NORM_BASELINE_WINDOW_S:
                                if mid_idx < len(eupnea_mask) and eupnea_mask[mid_idx] > 0:
                                    val = trace[mid_idx]
                                    if np.isfinite(val):
                                        fallback_vals.append(val)

                    eupnea_b_by_k[k] = float(np.mean(fallback_vals)) if len(fallback_vals) > 0 else np.nan

            t_elapsed = time.time() - t_start
            print(f"[CSV] ✓ Baselines and traces computed in {t_elapsed:.2f}s")

            t_start = time.time()
            print(f"[CSV] Writing breath-by-breath data...")
            for s in kept_sweeps:
                # Use cached traces if available, otherwise compute
                traces = cached_traces_by_sweep.get(s, None)
                if traces is None:
                    # Compute if not cached (shouldn't happen, but safety fallback)
                    y_proc = self._get_processed_for(st.analyze_chan, s)
                    pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                    br = st.breath_by_sweep.get(s, None)
                    if br is None and pks.size:
                        br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
                        st.breath_by_sweep[s] = br
                    if br is None:
                        br = {"onsets": np.array([], dtype=int)}

                    traces = {}
                    for k in need_keys:
                        if k in metrics.METRICS:
                            traces[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
                        else:
                            traces[k] = None
                else:
                    # Use cached data
                    br = st.breath_by_sweep.get(s, None)
                    if br is None:
                        br = {"onsets": np.array([], dtype=int)}

                on = np.asarray(br.get("onsets", []), dtype=int)
                if on.size < 2:
                    continue

                mids = (on[:-1] + on[1:]) // 2
    
                # Per-sweep breath-based baselines (use breath midpoints)
                t_rel_all = (st.t[mids] - (global_s0 if have_global_stim else 0.0)).astype(float)
                mask_pre_b  = (t_rel_all >= -NORM_BASELINE_WINDOW_S) & (t_rel_all < 0.0)
                mask_post_b = (t_rel_all >=  0.0) & (t_rel_all <= NORM_BASELINE_WINDOW_S)
    
                b_by_k = {}
                for k in need_keys:
                    arr = traces.get(k, None)
                    if arr is None or len(arr) != N:
                        b_by_k[k] = np.nan
                        continue
                    vals = arr[mids[mask_pre_b]]
                    vals = vals[np.isfinite(vals)]
                    if vals.size == 0:
                        vals = arr[mids[mask_post_b]]
                        vals = vals[np.isfinite(vals)]
                    b_by_k[k] = float(np.mean(vals)) if vals.size else np.nan
    
                # Breath classification flags: sigh, sniffing, eupnea, apnea
                # - is_sigh: Any sigh peak within breath interval [on[j], on[j+1])
                # - is_sniffing: Breath midpoint falls within marked sniffing region
                # - is_eupnea: Breath midpoint is part of eupneic (normal) breathing pattern
                # - is_apnea: Breath is preceded by long gap (recovery breath after apnea)

                # Sigh: Any sigh peak within breath interval [on[j], on[j+1])
                sigh_idx = np.asarray(st.sigh_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                sigh_idx = sigh_idx[(sigh_idx >= 0) & (sigh_idx < len(y_proc))]
                is_sigh_per_breath = np.zeros(on.size - 1, dtype=int)
                if sigh_idx.size:
                    for j in range(on.size - 1):
                        a = int(on[j]); b = int(on[j+1])
                        if np.any((sigh_idx >= a) & (sigh_idx < b)):
                            is_sigh_per_breath[j] = 1

                # Sniffing: Breath midpoint falls within any sniffing region
                sniff_regions = st.sniff_regions_by_sweep.get(s, [])
                is_sniffing_per_breath = np.zeros(on.size - 1, dtype=int)
                if sniff_regions:
                    for j, mid_idx in enumerate(mids):
                        t_mid = st.t[int(mid_idx)]
                        for sn_start, sn_end in sniff_regions:
                            if sn_start <= t_mid <= sn_end:
                                is_sniffing_per_breath[j] = 1
                                break

                # Eupnea: Breath midpoint marked as eupneic
                eupnea_mask = eupnea_masks_by_sweep.get(s, None)
                is_eupnea_per_breath = np.zeros(on.size - 1, dtype=int)
                if eupnea_mask is not None:
                    for j, mid_idx in enumerate(mids):
                        if int(mid_idx) < len(eupnea_mask) and eupnea_mask[int(mid_idx)] > 0:
                            is_eupnea_per_breath[j] = 1

                # Apnea: Breath preceded by long inter-breath interval (recovery breath after apnea)
                is_apnea_per_breath = np.zeros(on.size - 1, dtype=int)
                for j in range(len(mids)):
                    if j > 0:  # Need a previous onset to check inter-breath interval
                        ibi = st.t[int(on[j])] - st.t[int(on[j-1])]  # Time since last breath
                        if ibi > apnea_thresh:
                            # This breath (starting at on[j]) comes after a long gap
                            is_apnea_per_breath[j] = 1

                for i, idx in enumerate(mids, start=1):
                    t_rel = float(st.t[int(idx)] - (global_s0 if have_global_stim else 0.0))
                    breath_idx = i - 1  # 0-indexed for accessing arrays
                    sigh_flag = str(int(is_sigh_per_breath[breath_idx]))
                    sniff_flag = str(int(is_sniffing_per_breath[breath_idx]))
                    eupnea_flag = str(int(is_eupnea_per_breath[breath_idx]))
                    apnea_flag = str(int(is_apnea_per_breath[breath_idx]))
    
                    # ----- RAW: ALL
                    row_all = [str(s + 1), str(i), f"{t_rel:.9g}", "all", sigh_flag, sniff_flag, eupnea_flag, apnea_flag]
                    for k in need_keys:
                        v = np.nan
                        arr = traces.get(k, None)
                        if arr is not None and len(arr) == N:
                            v = arr[int(idx)]
                        row_all.append(f"{v:.9g}" if np.isfinite(v) else "")
                    rows_all.append(row_all)

                    # ----- NORM: ALL (time-based, per-sweep)
                    row_allN = [str(s + 1), str(i), f"{t_rel:.9g}", "all", sigh_flag, sniff_flag, eupnea_flag, apnea_flag]
                    for k in need_keys:
                        v = np.nan
                        arr = traces.get(k, None)
                        if arr is not None and len(arr) == N:
                            v = arr[int(idx)]
                        b = b_by_k.get(k, np.nan)
                        vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
                        row_allN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
                    rows_all_N.append(row_allN)

                    # ----- NORM_EUPNEA: ALL (eupnea-based, pooled)
                    row_allE = [str(s + 1), str(i), f"{t_rel:.9g}", "all", sigh_flag, sniff_flag, eupnea_flag, apnea_flag]
                    for k in need_keys:
                        v = np.nan
                        arr = traces.get(k, None)
                        if arr is not None and len(arr) == N:
                            v = arr[int(idx)]
                        eb = eupnea_b_by_k.get(k, np.nan)
                        ve = (v / eb) if (np.isfinite(v) and np.isfinite(eb) and abs(eb) > EPS_BASE) else np.nan
                        row_allE.append(f"{ve:.9g}" if np.isfinite(ve) else "")
                    rows_all_E.append(row_allE)

                    if have_global_stim:
                        region = "Baseline" if t_rel < 0 else ("Stim" if t_rel <= global_dur else "Post")

                        # RAW regional row
                        row_reg = [str(s + 1), str(i), f"{t_rel:.9g}", region, sigh_flag, sniff_flag, eupnea_flag, apnea_flag]
                        for k in need_keys:
                            v = np.nan
                            arr = traces.get(k, None)
                            if arr is not None and len(arr) == N:
                                v = arr[int(idx)]
                            row_reg.append(f"{v:.9g}" if np.isfinite(v) else "")
                        if region == "Baseline": rows_bl.append(row_reg)
                        elif region == "Stim":  rows_st.append(row_reg)
                        else:                   rows_po.append(row_reg)

                        # NORM regional row (time-based, per-sweep)
                        row_regN = [str(s + 1), str(i), f"{t_rel:.9g}", region, sigh_flag, sniff_flag, eupnea_flag, apnea_flag]
                        for k in need_keys:
                            v = np.nan
                            arr = traces.get(k, None)
                            if arr is not None and len(arr) == N:
                                v = arr[int(idx)]
                            b = b_by_k.get(k, np.nan)
                            vn = (v / b) if (np.isfinite(v) and np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
                            row_regN.append(f"{vn:.9g}" if np.isfinite(vn) else "")
                        if region == "Baseline": rows_bl_N.append(row_regN)
                        elif region == "Stim":  rows_st_N.append(row_regN)
                        else:                   rows_po_N.append(row_regN)

                        # NORM_EUPNEA regional row (eupnea-based, pooled)
                        row_regE = [str(s + 1), str(i), f"{t_rel:.9g}", region, sigh_flag, sniff_flag, eupnea_flag, apnea_flag]
                        for k in need_keys:
                            v = np.nan
                            arr = traces.get(k, None)
                            if arr is not None and len(arr) == N:
                                v = arr[int(idx)]
                            eb = eupnea_b_by_k.get(k, np.nan)
                            ve = (v / eb) if (np.isfinite(v) and np.isfinite(eb) and abs(eb) > EPS_BASE) else np.nan
                            row_regE.append(f"{ve:.9g}" if np.isfinite(ve) else "")
                        if region == "Baseline": rows_bl_E.append(row_regE)
                        elif region == "Stim":  rows_st_E.append(row_regE)
                        else:                   rows_po_E.append(row_regE)
    
            def _pad_row(row, want_len):
                if row is None: return [""] * want_len
                if len(row) < want_len: return row + [""] * (want_len - len(row))
                return row
    
            headers_all = _headers_for_block(None)
            headers_bl  = _headers_for_block("baseline")
            headers_st  = _headers_for_block("stim")
            headers_po  = _headers_for_block("post")

            headers_allN = _headers_for_block_norm(None)
            headers_blN  = _headers_for_block_norm("baseline")
            headers_stN  = _headers_for_block_norm("stim")
            headers_poN  = _headers_for_block_norm("post")

            headers_allE = _headers_for_block_norm_eupnea(None)
            headers_blE  = _headers_for_block_norm_eupnea("baseline")
            headers_stE  = _headers_for_block_norm_eupnea("stim")
            headers_poE  = _headers_for_block_norm_eupnea("post")

            have_stim_blocks = have_global_stim and (len(rows_bl) + len(rows_st) + len(rows_po) > 0)

            # Build DataFrames from row lists (1.5-2× faster than row-by-row)
            import pandas as pd

            if not have_stim_blocks:
                # RAW + NORM + NORM_EUPNEA (ALL only)
                df_all = pd.DataFrame(rows_all, columns=headers_all) if rows_all else pd.DataFrame(columns=headers_all)
                df_all_N = pd.DataFrame(rows_all_N, columns=headers_allN) if rows_all_N else pd.DataFrame(columns=headers_allN)
                df_all_E = pd.DataFrame(rows_all_E, columns=headers_allE) if rows_all_E else pd.DataFrame(columns=headers_allE)

                # Add separator columns and concatenate
                sep1 = pd.DataFrame({'': [''] * len(df_all)}) if len(df_all) > 0 else pd.DataFrame({'': []})
                sep2 = pd.DataFrame({'': [''] * len(df_all)}) if len(df_all) > 0 else pd.DataFrame({'': []})

                df_combined = pd.concat([df_all, sep1, df_all_N, sep2, df_all_E], axis=1)
            else:
                # RAW blocks, then NORM blocks, then NORM_EUPNEA blocks
                df_all = pd.DataFrame(rows_all, columns=headers_all) if rows_all else pd.DataFrame(columns=headers_all)
                df_bl = pd.DataFrame(rows_bl, columns=headers_bl) if rows_bl else pd.DataFrame(columns=headers_bl)
                df_st = pd.DataFrame(rows_st, columns=headers_st) if rows_st else pd.DataFrame(columns=headers_st)
                df_po = pd.DataFrame(rows_po, columns=headers_po) if rows_po else pd.DataFrame(columns=headers_po)

                df_all_N = pd.DataFrame(rows_all_N, columns=headers_allN) if rows_all_N else pd.DataFrame(columns=headers_allN)
                df_bl_N = pd.DataFrame(rows_bl_N, columns=headers_blN) if rows_bl_N else pd.DataFrame(columns=headers_blN)
                df_st_N = pd.DataFrame(rows_st_N, columns=headers_stN) if rows_st_N else pd.DataFrame(columns=headers_stN)
                df_po_N = pd.DataFrame(rows_po_N, columns=headers_poN) if rows_po_N else pd.DataFrame(columns=headers_poN)

                df_all_E = pd.DataFrame(rows_all_E, columns=headers_allE) if rows_all_E else pd.DataFrame(columns=headers_allE)
                df_bl_E = pd.DataFrame(rows_bl_E, columns=headers_blE) if rows_bl_E else pd.DataFrame(columns=headers_blE)
                df_st_E = pd.DataFrame(rows_st_E, columns=headers_stE) if rows_st_E else pd.DataFrame(columns=headers_stE)
                df_po_E = pd.DataFrame(rows_po_E, columns=headers_poE) if rows_po_E else pd.DataFrame(columns=headers_poE)

                # Create separator columns (use max length for consistency)
                max_len = max(
                    len(df_all), len(df_bl), len(df_st), len(df_po),
                    len(df_all_N), len(df_bl_N), len(df_st_N), len(df_po_N),
                    len(df_all_E), len(df_bl_E), len(df_st_E), len(df_po_E)
                ) if any([len(df_all), len(df_bl), len(df_st), len(df_po)]) else 0

                sep = pd.DataFrame({'': [''] * max_len}) if max_len > 0 else pd.DataFrame({'': []})

                # Concatenate with separators
                df_combined = pd.concat([
                    df_all, sep.copy(), df_bl, sep.copy(), df_st, sep.copy(), df_po, sep.copy(),
                    df_all_N, sep.copy(), df_bl_N, sep.copy(), df_st_N, sep.copy(), df_po_N, sep.copy(),
                    df_all_E, sep.copy(), df_bl_E, sep.copy(), df_st_E, sep.copy(), df_po_E
                ], axis=1)

            # Write to CSV
            df_combined.to_csv(breaths_path, index=False, na_rep='')

            t_elapsed = time.time() - t_start
            print(f"[CSV] ✓ Breath data written in {t_elapsed:.2f}s")

            if progress_dialog:
                progress_dialog.setLabelText("Writing events CSV...")
                progress_dialog.setValue(70)
                QApplication.processEvents()

            # -------------------- (4) Events CSV (stimulus, apnea, eupnea intervals) --------------------
            t_start = time.time()
            events_path = base.with_name(base.name + "_events.csv")

            events_rows = []

            # Get thresholds from UI
            eupnea_thresh = self._parse_float(self.EupneaThresh) or 5.0  # Hz
            apnea_thresh = self._parse_float(self.ApneaThresh) or 0.5    # seconds

            for s in kept_sweeps:
                y_proc = self._get_processed_for(st.analyze_chan, s)
                pks    = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                br     = st.breath_by_sweep.get(s, None)
                if br is None and pks.size:
                    br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
                if br is None:
                    br = {"onsets": np.array([], dtype=int)}

                on = np.asarray(br.get("onsets", []), dtype=int)
                off = np.asarray(br.get("offsets", []), dtype=int)
                expmins = np.asarray(br.get("expmins", []), dtype=int)
                expoffs = np.asarray(br.get("expoffs", []), dtype=int)

                # Stimulus events
                stim_on_idx = st.stim_onsets_by_sweep.get(s, None)
                stim_off_idx = st.stim_offsets_by_sweep.get(s, None)

                if stim_on_idx is not None and len(stim_on_idx) > 0:
                    for i, on_idx in enumerate(stim_on_idx):
                        start_time = float(st.t[int(on_idx)])
                        if stim_off_idx is not None and i < len(stim_off_idx):
                            end_time = float(st.t[int(stim_off_idx[i])])
                        else:
                            end_time = np.nan

                        # Convert to relative time if global stim available
                        if have_global_stim:
                            start_time -= global_s0
                            if np.isfinite(end_time):
                                end_time -= global_s0

                        duration = end_time - start_time if np.isfinite(end_time) else np.nan

                        events_rows.append([
                            str(s + 1),
                            "stimulus",
                            f"{start_time:.9g}",
                            f"{end_time:.9g}" if np.isfinite(end_time) else "",
                            f"{duration:.9g}" if np.isfinite(duration) else ""
                        ])

                # Apnea and eupnea events (only if we have breath data)
                if on.size >= 2:
                    # Detect apnea regions
                    apnea_mask = metrics.detect_apneas(
                        st.t, y_proc, st.sr_hz, pks, on, off, expmins, expoffs,
                        min_apnea_duration_sec=apnea_thresh
                    )

                    # Detect eupnea regions
                    eupnea_mask = metrics.detect_eupnic_regions(
                        st.t, y_proc, st.sr_hz, pks, on, off, expmins, expoffs,
                        freq_threshold_hz=eupnea_thresh
                    )

                    # Convert masks to interval lists
                    def mask_to_intervals(mask):
                        """Convert boolean mask to list of (start_idx, end_idx) intervals."""
                        intervals = []
                        in_region = False
                        start_idx = 0

                        for i in range(len(mask)):
                            if mask[i] and not in_region:
                                # Start of new region
                                start_idx = i
                                in_region = True
                            elif not mask[i] and in_region:
                                # End of region
                                intervals.append((start_idx, i - 1))
                                in_region = False

                        # Handle case where region extends to end
                        if in_region:
                            intervals.append((start_idx, len(mask) - 1))

                        return intervals

                    # Add apnea intervals
                    apnea_intervals = mask_to_intervals(apnea_mask > 0)
                    for start_idx, end_idx in apnea_intervals:
                        start_time = float(st.t[int(start_idx)])
                        end_time = float(st.t[int(end_idx)])

                        # Convert to relative time if global stim available
                        if have_global_stim:
                            start_time -= global_s0
                            end_time -= global_s0

                        duration = end_time - start_time

                        events_rows.append([
                            str(s + 1),
                            "apnea",
                            f"{start_time:.9g}",
                            f"{end_time:.9g}",
                            f"{duration:.9g}"
                        ])

                    # Add eupnea intervals
                    eupnea_intervals = mask_to_intervals(eupnea_mask > 0)
                    for start_idx, end_idx in eupnea_intervals:
                        start_time = float(st.t[int(start_idx)])
                        end_time = float(st.t[int(end_idx)])

                        # Convert to relative time if global stim available
                        if have_global_stim:
                            start_time -= global_s0
                            end_time -= global_s0

                        duration = end_time - start_time

                        events_rows.append([
                            str(s + 1),
                            "eupnea",
                            f"{start_time:.9g}",
                            f"{end_time:.9g}",
                            f"{duration:.9g}"
                        ])

                # Add sniffing bout intervals
                sniff_regions = st.sniff_regions_by_sweep.get(s, [])
                for start_time, end_time in sniff_regions:
                    # start_time and end_time are already in seconds

                    # Convert to relative time if global stim available
                    if have_global_stim:
                        start_time_rel = start_time - global_s0
                        end_time_rel = end_time - global_s0
                    else:
                        start_time_rel = start_time
                        end_time_rel = end_time

                    duration = end_time - start_time

                    events_rows.append([
                        str(s + 1),
                        "sniffing",
                        f"{start_time_rel:.9g}",
                        f"{end_time_rel:.9g}",
                        f"{duration:.9g}"
                    ])

            # Write events CSV using pandas
            import pandas as pd
            df_events = pd.DataFrame(
                events_rows,
                columns=["sweep", "event_type", "start_time", "end_time", "duration"]
            ) if events_rows else pd.DataFrame(columns=["sweep", "event_type", "start_time", "end_time", "duration"])
            df_events.to_csv(events_path, index=False, na_rep='')

            # Count event types
            if events_rows:
                event_counts = {}
                for row in events_rows:
                    event_type = row[1]
                    event_counts[event_type] = event_counts.get(event_type, 0) + 1
                event_summary = ", ".join([f"{count} {etype}" for etype, count in sorted(event_counts.items())])
            else:
                event_summary = "no events"

            t_elapsed = time.time() - t_start
            print(f"[CSV] ✓ Events data written in {t_elapsed:.2f}s ({event_summary})")

        # -------------------- (4) Summary PDF or Preview --------------------
        t_start = time.time()
        if progress_dialog:
            progress_dialog.setLabelText("Generating summary figures..." if preview_only else "Generating PDF...")
            progress_dialog.setValue(80)
            QApplication.processEvents()

        keys_for_timeplots = [k for k in all_keys if k not in self._EXCLUDE_FOR_CSV]
        label_by_key = {key: label for (label, key) in metrics.METRIC_SPECS if key in keys_for_timeplots}

        if preview_only:
            # Show interactive preview dialog instead of saving
            try:
                self._show_summary_preview_dialog(
                    t_ds_csv=t_ds_csv,
                    y2_ds_by_key=y2_ds_by_key,
                    keys_for_csv=keys_for_timeplots,
                    label_by_key=label_by_key,
                    stim_zero=(global_s0 if have_global_stim else None),
                    stim_dur=(global_dur if have_global_stim else None),
                )
            except Exception as e:
                QMessageBox.critical(self, "View Summary", f"Error generating preview:\n{e}")
                import traceback
                traceback.print_exc()
        else:
            # Save PDF to disk
            pdf_path = base.with_name(base.name + "_summary.pdf")
            try:
                self._save_metrics_summary_pdf(
                    out_path=pdf_path,
                    t_ds_csv=t_ds_csv,
                    y2_ds_by_key=y2_ds_by_key,
                    keys_for_csv=keys_for_timeplots,
                    label_by_key=label_by_key,
                    stim_zero=(global_s0 if have_global_stim else None),
                    stim_dur=(global_dur if have_global_stim else None),
                )
            except Exception as e:
                print(f"[save][summary-pdf] skipped: {e}")

            t_elapsed = time.time() - t_start
            if preview_only:
                print(f"[PDF] ✓ Preview generated in {t_elapsed:.2f}s")
            else:
                print(f"[PDF] ✓ PDF saved in {t_elapsed:.2f}s")

            # -------------------- done --------------------
            if progress_dialog:
                progress_dialog.setValue(100)
                QApplication.processEvents()

            msg = f"Saved:\n- {npz_path.name}\n- {csv_time_path.name}\n- {breaths_path.name}\n- {events_path.name}\n- {pdf_path.name}"
            print("[save]", msg)
            try:
                self.statusbar.showMessage(msg, 6000)
            except Exception:
                pass

            # Show success dialog
            QMessageBox.information(
                self,
                "Save Successful",
                f"Files saved successfully to:\n{self._save_dir}\n\n{msg}"
            )


    def _mean_sem_1d(self, arr: np.ndarray):
        """Finite-only mean and SEM (ddof=1) for a 1D array. Returns (mean, sem).
        If no finite values -> (nan, nan). If only 1 finite value -> (mean, nan)."""
        arr = np.asarray(arr, dtype=float)
        finite = np.isfinite(arr)
        n = int(finite.sum())
        if n == 0:
            return (np.nan, np.nan)
        vals = arr[finite]
        m = float(np.mean(vals))
        if n >= 2:
            s = float(np.std(vals, ddof=1))
            sem = s / np.sqrt(n)
        else:
            sem = np.nan
        return (m, sem)

    # def _save_metrics_summary_pdf(
    #     self,
    #     out_path,
    #     t_ds_csv: np.ndarray,
    #     y2_ds_by_key: dict,
    #     keys_for_csv: list[str],
    #     label_by_key: dict[str, str],
    #     stim_zero: float | None,
    #     stim_dur: float | None,
    #     ):
    #     """
    #     Build a two-page PDF:
    #     • Page 1: rows = metrics, cols = [all sweeps | mean±SEM | histograms] using RAW data
    #     • Page 2: same layout, using NORMALIZED data (per sweep, per metric baseline)

    #     Normalization baseline per sweep:
    #     - mean over the last W seconds before t=0 (W = self._norm_window_s, default 10.0)
    #     - fallback to first W seconds after 0 if no pre-stim samples exist
    #     - value_norm = value / baseline; unstable divisions → NaN
    #     """
    #     import numpy as np
    #     import matplotlib.pyplot as plt
    #     from matplotlib.backends.backend_pdf import PdfPages

    #     st = self.state
    #     n_sweeps = next(iter(y2_ds_by_key.values())).shape[1] if y2_ds_by_key else 0
    #     M = len(t_ds_csv)
    #     have_stim = (stim_zero is not None and stim_dur is not None)

    #     # --- normalization knobs (match exporter) ---
    #     NORM_BASELINE_WINDOW_S = float(getattr(self, "_norm_window_s", 10.0))
    #     EPS_BASE = 1e-12

    #     # ---------- Helpers ----------
    #     def _per_sweep_baseline_time(A_ds: np.ndarray) -> np.ndarray:
    #         """Baseline per sweep (time-series): mean over last W sec before 0; fallback to first W sec after 0."""
    #         b = np.full((A_ds.shape[1],), np.nan, dtype=float)
    #         mask_pre  = (t_ds_csv >= -NORM_BASELINE_WINDOW_S) & (t_ds_csv < 0.0)
    #         mask_post = (t_ds_csv >=  0.0) & (t_ds_csv <= NORM_BASELINE_WINDOW_S)
    #         for s in range(A_ds.shape[1]):
    #             col = A_ds[:, s]
    #             vals = col[mask_pre]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = col[mask_post]
    #                 vals = vals[np.isfinite(vals)]
    #             if vals.size:
    #                 b[s] = float(np.mean(vals))
    #         return b

    #     def _normalize_matrix(A_ds: np.ndarray, b: np.ndarray) -> np.ndarray:
    #         out = np.full_like(A_ds, np.nan)
    #         for s in range(A_ds.shape[1]):
    #             bs = b[s]
    #             if np.isfinite(bs) and abs(bs) > EPS_BASE:
    #                 out[:, s] = A_ds[:, s] / bs
    #         return out

    #     def _build_hist_vals_raw_and_norm():
    #         """Collect per-breath RAW and NORM values for histograms by metric/region."""
    #         hist_raw = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
    #         hist_nrm = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
    #         need_keys = set(keys_for_csv)

    #         for s in range(n_sweeps):
    #             y_proc = self._get_processed_for(st.analyze_chan, s)
    #             pks    = np.asarray(self.state.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #             br     = self.state.breath_by_sweep.get(s, None)
    #             if br is None and pks.size:
    #                 br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #                 self.state.breath_by_sweep[s] = br
    #             if br is None:
    #                 br = {"onsets": np.array([], dtype=int)}

    #             on = np.asarray(br.get("onsets", []), dtype=int)
    #             if on.size < 2:
    #                 continue
    #             mids = (on[:-1] + on[1:]) // 2

    #             # Precompute metric traces for this sweep
    #             traces = {}
    #             for k in need_keys:
    #                 try:
    #                     traces[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #                 except TypeError:
    #                     traces[k] = None

    #             # Per-sweep breath-based baselines (use breath midpoints; match exporter)
    #             t_rel_all = (st.t[mids] - (stim_zero or 0.0)).astype(float)
    #             mask_pre_b  = (t_rel_all >= -NORM_BASELINE_WINDOW_S) & (t_rel_all < 0.0)
    #             mask_post_b = (t_rel_all >=  0.0) & (t_rel_all <= NORM_BASELINE_WINDOW_S)

    #             b_by_k = {}
    #             for k in need_keys:
    #                 arr = traces.get(k, None)
    #                 if arr is None or len(arr) != len(st.t):
    #                     b_by_k[k] = np.nan
    #                     continue
    #                 vals = arr[mids[mask_pre_b]]
    #                 vals = vals[np.isfinite(vals)]
    #                 if vals.size == 0:
    #                     vals = arr[mids[mask_post_b]]
    #                     vals = vals[np.isfinite(vals)]
    #                 b_by_k[k] = float(np.mean(vals)) if vals.size else np.nan

    #             # Fill raw + normalized buckets
    #             for idx, t_rel in zip(mids, t_rel_all):
    #                 for k in need_keys:
    #                     arr = traces.get(k, None)
    #                     if arr is None or len(arr) != len(st.t):
    #                         continue
    #                     v = float(arr[int(idx)])
    #                     if not np.isfinite(v):
    #                         continue
    #                     # raw
    #                     hist_raw[k]["all"].append(v)
    #                     # norm
    #                     b = b_by_k.get(k, np.nan)
    #                     vn = (v / b) if (np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
    #                     if np.isfinite(vn):
    #                         hist_nrm[k]["all"].append(vn)

    #                     if have_stim:
    #                         if t_rel < 0:
    #                             hist_raw[k]["baseline"].append(v)
    #                             if np.isfinite(vn): hist_nrm[k]["baseline"].append(vn)
    #                         elif 0.0 <= t_rel <= stim_dur:
    #                             hist_raw[k]["stim"].append(v)
    #                             if np.isfinite(vn): hist_nrm[k]["stim"].append(vn)
    #                         else:
    #                             hist_raw[k]["post"].append(v)
    #                             if np.isfinite(vn): hist_nrm[k]["post"].append(vn)

    #         return hist_raw, hist_nrm

    #     def _plot_grid(fig, axes, Y_by_key, hist_vals, title_suffix):
    #         """Render one page (grid) given series & histogram data dicts."""
    #         nrows = max(1, len(keys_for_csv))
    #         for r, k in enumerate(keys_for_csv):
    #             label = label_by_key.get(k, k)

    #             # --- col 1: all sweeps overlaid ---
    #             ax1 = axes[r, 0]
    #             Y = Y_by_key.get(k, None)
    #             if Y is not None and Y.shape[0] == M:
    #                 for s in range(n_sweeps):
    #                     ax1.plot(t_ds_csv, Y[:, s], lw=0.8, alpha=0.45)
    #             if have_stim:
    #                 ax1.axvspan(0.0, stim_dur, color="#4c78a8", alpha=0.12)
    #                 ax1.axvline(0.0, color="#4c78a8", lw=1.0, alpha=0.6)
    #                 ax1.axvline(stim_dur, color="#4c78a8", lw=1.0, alpha=0.6, ls="--")
    #             ax1.set_title(f"{label} — all sweeps{title_suffix}")
    #             if r == nrows - 1:
    #                 ax1.set_xlabel("Time (s, rel. stim onset)")

    #             # --- col 2: mean ± SEM ---
    #             ax2 = axes[r, 1]
    #             if Y is not None and Y.shape[0] == M:
    #                 with np.errstate(invalid="ignore"):
    #                     mean = np.nanmean(Y, axis=1)
    #                     n    = np.sum(np.isfinite(Y), axis=1)
    #                     std  = np.nanstd(Y, axis=1, ddof=1)
    #                     sem  = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #                 ax2.plot(t_ds_csv, mean, lw=1.8)
    #                 ax2.fill_between(t_ds_csv, mean - sem, mean + sem, alpha=0.25, linewidth=0)
    #             if have_stim:
    #                 ax2.axvspan(0.0, stim_dur, color="#4c78a8", alpha=0.12)
    #                 ax2.axvline(0.0, color="#4c78a8", lw=1.0, alpha=0.6)
    #                 ax2.axvline(stim_dur, color="#4c78a8", lw=1.0, alpha=0.6, ls="--")
    #             ax2.set_title(f"{label} — mean ± SEM{title_suffix}")
    #             if r == nrows - 1:
    #                 ax2.set_xlabel("Time (s, rel. stim onset)")

    #             # --- col 3: line histograms (density) ---
    #             ax3 = axes[r, 2]
    #             # Build common bin edges across groups for this metric
    #             groups = []
    #             for nm in ("all", "baseline", "stim", "post"):
    #                 vals = np.asarray(hist_vals[k][nm], dtype=float)
    #                 if vals.size:
    #                     groups.append(vals)
    #             if len(groups):
    #                 combined = np.concatenate(groups)
    #                 edges = np.histogram_bin_edges(combined, bins="auto")
    #                 centers = 0.5 * (edges[:-1] + edges[1:])

    #                 def _plot_line(vals, lbl, style_kw):
    #                     vals = np.asarray(vals, dtype=float)
    #                     if vals.size == 0:
    #                         return
    #                     dens, _ = np.histogram(vals, bins=edges, density=True)
    #                     ax3.plot(centers, dens, **style_kw, label=lbl)

    #                 _plot_line(hist_vals[k]["all"], "All", dict(lw=1.8))
    #                 if have_stim:
    #                     _plot_line(hist_vals[k]["baseline"], "Baseline", dict(lw=1.6))
    #                     _plot_line(hist_vals[k]["stim"],     "Stim",     dict(lw=1.6, ls="--"))
    #                     _plot_line(hist_vals[k]["post"],     "Post",     dict(lw=1.6, ls=":"))

    #             ax3.set_title(f"{label} — distribution (density){title_suffix}")
    #             ax3.set_ylabel("Density")
    #             if len(ax3.lines):
    #                 ax3.legend(loc="best", fontsize=8)

    #         fig.tight_layout()

    #     # ---------- Prepare both RAW and NORMALIZED datasets ----------
    #     # RAW time-series already provided: y2_ds_by_key

    #     # NORMALIZED time-series per metric
    #     y2_ds_by_key_norm = {}
    #     for k in keys_for_csv:
    #         Y = y2_ds_by_key.get(k, None)
    #         if Y is None or not Y.size:
    #             y2_ds_by_key_norm[k] = None
    #             continue
    #         b = _per_sweep_baseline_time(Y)
    #         y2_ds_by_key_norm[k] = _normalize_matrix(Y, b)

    #     # RAW + NORMALIZED histogram values
    #     hist_vals_raw, hist_vals_norm = _build_hist_vals_raw_and_norm()

    #     # ---------- Create two-page PDF ----------
    #     nrows = max(1, len(keys_for_csv))
    #     fig_w = 13
    #     fig_h = max(4.0, 2.8 * nrows)

    #     with PdfPages(out_path) as pdf:
    #         # Page 1 — RAW
    #         fig1, axes1 = plt.subplots(nrows, 3, figsize=(fig_w, fig_h), squeeze=False)
    #         plt.subplots_adjust(hspace=0.6, wspace=0.25)
    #         _plot_grid(fig1, axes1, y2_ds_by_key, hist_vals_raw, title_suffix="")
    #         fig1.suptitle("PlethApp summary — raw", y=0.995, fontsize=12)
    #         pdf.savefig(fig1, dpi=150)
    #         plt.close(fig1)

    #         # Page 2 — NORMALIZED
    #         fig2, axes2 = plt.subplots(nrows, 3, figsize=(fig_w, fig_h), squeeze=False)
    #         plt.subplots_adjust(hspace=0.6, wspace=0.25)
    #         _plot_grid(fig2, axes2, y2_ds_by_key_norm, hist_vals_norm, title_suffix=" (norm)")
    #         fig2.suptitle("PlethApp summary — normalized", y=0.995, fontsize=12)
    #         pdf.savefig(fig2, dpi=150)
    #         plt.close(fig2)

    # def _save_metrics_summary_pdf(
    #     self,
    #     out_path,
    #     t_ds_csv: np.ndarray,
    #     y2_ds_by_key: dict,
    #     keys_for_csv: list[str],
    #     label_by_key: dict[str, str],
    #     stim_zero: float | None,
    #     stim_dur: float | None,
    #     ):
    #     """
    #     Build a two-page PDF:
    #     • Page 1: rows = metrics, cols = [all sweeps | mean±SEM | histograms] using RAW data
    #     • Page 2: same layout, using NORMALIZED data (per sweep, per metric baseline)
    #     • NEW: overlay orange star markers at times where sighs occurred (first two columns)
    #     """
    #     import numpy as np
    #     import matplotlib.pyplot as plt
    #     from matplotlib.backends.backend_pdf import PdfPages

    #     st = self.state
    #     n_sweeps = next(iter(y2_ds_by_key.values())).shape[1] if y2_ds_by_key else 0
    #     M = len(t_ds_csv)
    #     have_stim = (stim_zero is not None and stim_dur is not None)

    #     # --- normalization knobs (match exporter) ---
    #     NORM_BASELINE_WINDOW_S = float(getattr(self, "_norm_window_s", 10.0))
    #     EPS_BASE = 1e-12

    #     # ---------- Helpers ----------
    #     def _per_sweep_baseline_time(A_ds: np.ndarray) -> np.ndarray:
    #         b = np.full((A_ds.shape[1],), np.nan, dtype=float)
    #         mask_pre  = (t_ds_csv >= -NORM_BASELINE_WINDOW_S) & (t_ds_csv < 0.0)
    #         mask_post = (t_ds_csv >=  0.0) & (t_ds_csv <= NORM_BASELINE_WINDOW_S)
    #         for s in range(A_ds.shape[1]):
    #             col = A_ds[:, s]
    #             vals = col[mask_pre]
    #             vals = vals[np.isfinite(vals)]
    #             if vals.size == 0:
    #                 vals = col[mask_post]
    #                 vals = vals[np.isfinite(vals)]
    #             if vals.size:
    #                 b[s] = float(np.mean(vals))
    #         return b

    #     def _normalize_matrix(A_ds: np.ndarray, b: np.ndarray) -> np.ndarray:
    #         out = np.full_like(A_ds, np.nan)
    #         for s in range(A_ds.shape[1]):
    #             bs = b[s]
    #             if np.isfinite(bs) and abs(bs) > EPS_BASE:
    #                 out[:, s] = A_ds[:, s] / bs
    #         return out

    #     # Collect sigh times (relative to stim_zero if present) across kept sweeps
    #     # def _collect_sigh_times_rel() -> list[float]:
    #     #     kept = [s for s in range(next(iter(st.sweeps.values())).shape[1])
    #     #             if s not in getattr(st, "omitted_sweeps", set())]
    #     #     t0 = float(stim_zero) if stim_zero is not None else 0.0
    #     #     ts = []
    #     #     for s in kept:
    #     #         for idx in getattr(st, "sighs_by_sweep", {}).get(s, set()):
    #     #             try:
    #     #                 ts.append(float(st.t[int(idx)] - t0))
    #     #             except Exception:
    #     #                 pass
    #     #     if not ts:
    #     #         return []
    #     #     ts = sorted(ts)
    #     #     # de-duplicate very near events (within ~1 sample)
    #     #     dt = float(np.median(np.diff(st.t))) if len(st.t) > 1 else (1.0 / float(st.sr_hz) if st.sr_hz else 0.01)
    #     #     eps = max(1e-9, 2.0 * dt)
    #     #     out = []
    #     #     for t in ts:
    #     #         if not out or abs(t - out[-1]) > eps:
    #     #             out.append(t)
    #     #     return out

    #     def _collect_sigh_times_rel() -> list[float]:
    #         kept = [s for s in range(next(iter(st.sweeps.values())).shape[1])
    #                 if s not in getattr(st, "omitted_sweeps", set())]
    #         t0 = float(stim_zero) if stim_zero is not None else 0.0
    #         ts = []
    #         for s in kept:
    #             # Get peaks for this sweep for mapping if needed
    #             pks = np.asarray(self.state.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #             sigh_idxs = self._sigh_sample_indices(s, pks)
    #             for i in sigh_idxs:
    #                 ts.append(float(st.t[int(i)] - t0))
    #         if not ts:
    #             return []
    #         ts = sorted(ts)
    #         # Deduplicate very-close events
    #         dt = float(np.median(np.diff(st.t))) if len(st.t) > 1 else (1.0 / float(st.sr_hz) if st.sr_hz else 0.01)
    #         eps = max(1e-9, 2.0 * dt)
    #         out = []
    #         for t in ts:
    #             if not out or abs(t - out[-1]) > eps:
    #                 out.append(t)
    #         return out


    #     sigh_times_rel = _collect_sigh_times_rel()

    #     def _build_hist_vals_raw_and_norm():
    #         hist_raw = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
    #         hist_nrm = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
    #         need_keys = set(keys_for_csv)

    #         for s in range(next(iter(st.sweeps.values())).shape[1]):
    #             if s in getattr(st, "omitted_sweeps", set()):
    #                 continue
    #             y_proc = self._get_processed_for(st.analyze_chan, s)
    #             pks    = np.asarray(self.state.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
    #             br     = self.state.breath_by_sweep.get(s, None)
    #             if br is None and pks.size:
    #                 br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
    #                 self.state.breath_by_sweep[s] = br
    #             if br is None:
    #                 br = {"onsets": np.array([], dtype=int)}

    #             on = np.asarray(br.get("onsets", []), dtype=int)
    #             if on.size < 2:
    #                 continue
    #             mids = (on[:-1] + on[1:]) // 2

    #             traces = {}
    #             for k in need_keys:
    #                 try:
    #                     traces[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
    #                 except TypeError:
    #                     traces[k] = None

    #             t_rel_all = (st.t[mids] - (stim_zero or 0.0)).astype(float)
    #             mask_pre_b  = (t_rel_all >= -NORM_BASELINE_WINDOW_S) & (t_rel_all < 0.0)
    #             mask_post_b = (t_rel_all >=  0.0) & (t_rel_all <= NORM_BASELINE_WINDOW_S)

    #             b_by_k = {}
    #             for k in need_keys:
    #                 arr = traces.get(k, None)
    #                 if arr is None or len(arr) != len(st.t):
    #                     b_by_k[k] = np.nan
    #                     continue
    #                 vals = arr[mids[mask_pre_b]]
    #                 vals = vals[np.isfinite(vals)]
    #                 if vals.size == 0:
    #                     vals = arr[mids[mask_post_b]]
    #                     vals = vals[np.isfinite(vals)]
    #                 b_by_k[k] = float(np.mean(vals)) if vals.size else np.nan

    #             for idx, t_rel in zip(mids, t_rel_all):
    #                 for k in need_keys:
    #                     arr = traces.get(k, None)
    #                     if arr is None or len(arr) != len(st.t):
    #                         continue
    #                     v = float(arr[int(idx)])
    #                     if not np.isfinite(v):
    #                         continue
    #                     hist_raw[k]["all"].append(v)
    #                     b = b_by_k.get(k, np.nan)
    #                     vn = (v / b) if (np.isfinite(b) and abs(b) > 0) else np.nan
    #                     if np.isfinite(vn):
    #                         hist_nrm[k]["all"].append(vn)

    #                     if have_stim:
    #                         if t_rel < 0:
    #                             hist_raw[k]["baseline"].append(v)
    #                             if np.isfinite(vn): hist_nrm[k]["baseline"].append(vn)
    #                         elif 0.0 <= t_rel <= stim_dur:
    #                             hist_raw[k]["stim"].append(v)
    #                             if np.isfinite(vn): hist_nrm[k]["stim"].append(vn)
    #                         else:
    #                             hist_raw[k]["post"].append(v)
    #                             if np.isfinite(vn): hist_nrm[k]["post"].append(vn)

    #         return hist_raw, hist_nrm

    #     def _plot_sigh_stars(ax):
    #         """Overlay stars at sigh times near the top of axis without changing limits."""
    #         if not sigh_times_rel:
    #             return
    #         ylim = ax.get_ylim()
    #         y_star = ylim[1] - 0.06 * (ylim[1] - ylim[0])
    #         ax.plot(
    #             sigh_times_rel,
    #             [y_star] * len(sigh_times_rel),
    #             linestyle="none",
    #             marker="*",
    #             markersize=9,
    #             color="#ff9d00",
    #             alpha=0.95,
    #             zorder=6,
    #         )
    #         ax.set_ylim(*ylim)

    #     def _plot_grid(fig, axes, Y_by_key, hist_vals, title_suffix):
    #         """Render one page (grid) given series & histogram data dicts."""
    #         nrows = max(1, len(keys_for_csv))
    #         for r, k in enumerate(keys_for_csv):
    #             label = label_by_key.get(k, k)

    #             # --- col 1: all sweeps overlaid ---
    #             ax1 = axes[r, 0]
    #             Y = Y_by_key.get(k, None)
    #             if Y is not None and Y.shape[0] == M:
    #                 for s in range(Y.shape[1]):
    #                     ax1.plot(t_ds_csv, Y[:, s], lw=0.8, alpha=0.45)
    #             if have_stim:
    #                 ax1.axvspan(0.0, stim_dur, color="#4c78a8", alpha=0.12)
    #                 ax1.axvline(0.0, color="#4c78a8", lw=1.0, alpha=0.6)
    #                 ax1.axvline(stim_dur, color="#4c78a8", lw=1.0, alpha=0.6, ls="--")
    #             _plot_sigh_stars(ax1)  # NEW
    #             ax1.set_title(f"{label} — all sweeps{title_suffix}")
    #             if r == nrows - 1:
    #                 ax1.set_xlabel("Time (s, rel. stim onset)")

    #             # --- col 2: mean ± SEM ---
    #             ax2 = axes[r, 1]
    #             if Y is not None and Y.shape[0] == M:
    #                 with np.errstate(invalid="ignore"):
    #                     mean = np.nanmean(Y, axis=1)
    #                     n    = np.sum(np.isfinite(Y), axis=1)
    #                     std  = np.nanstd(Y, axis=1, ddof=1)
    #                     sem  = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #                 ax2.plot(t_ds_csv, mean, lw=1.8)
    #                 ax2.fill_between(t_ds_csv, mean - sem, mean + sem, alpha=0.25, linewidth=0)
    #             if have_stim:
    #                 ax2.axvspan(0.0, stim_dur, color="#4c78a8", alpha=0.12)
    #                 ax2.axvline(0.0, color="#4c78a8", lw=1.0, alpha=0.6)
    #                 ax2.axvline(stim_dur, color="#4c78a8", lw=1.0, alpha=0.6, ls="--")
    #             _plot_sigh_stars(ax2)  # NEW
    #             ax2.set_title(f"{label} — mean ± SEM{title_suffix}")
    #             if r == nrows - 1:
    #                 ax2.set_xlabel("Time (s, rel. stim onset)")

    #             # --- col 3: line histograms (density) ---
    #             ax3 = axes[r, 2]
    #             groups = []
    #             for nm in ("all", "baseline", "stim", "post"):
    #                 vals = np.asarray(hist_vals[k][nm], dtype=float)
    #                 if vals.size:
    #                     groups.append(vals)
    #             if len(groups):
    #                 combined = np.concatenate(groups)
    #                 edges = np.histogram_bin_edges(combined, bins="auto")
    #                 centers = 0.5 * (edges[:-1] + edges[1:])

    #                 def _plot_line(vals, lbl, style_kw):
    #                     vals = np.asarray(vals, dtype=float)
    #                     if vals.size == 0:
    #                         return
    #                     dens, _ = np.histogram(vals, bins=edges, density=True)
    #                     ax3.plot(centers, dens, **style_kw, label=lbl)

    #                 _plot_line(hist_vals[k]["all"], "All", dict(lw=1.8))
    #                 if have_stim:
    #                     _plot_line(hist_vals[k]["baseline"], "Baseline", dict(lw=1.6))
    #                     _plot_line(hist_vals[k]["stim"],     "Stim",     dict(lw=1.6, ls="--"))
    #                     _plot_line(hist_vals[k]["post"],     "Post",     dict(lw=1.6, ls=":"))

    #             ax3.set_title(f"{label} — distribution (density){title_suffix}")
    #             ax3.set_ylabel("Density")
    #             if len(ax3.lines):
    #                 ax3.legend(loc="best", fontsize=8)

    #         fig.tight_layout()

    #     # ---------- Prepare both RAW and NORMALIZED datasets ----------
    #     y2_ds_by_key_norm = {}
    #     for k in keys_for_csv:
    #         Y = y2_ds_by_key.get(k, None)
    #         if Y is None or not Y.size:
    #             y2_ds_by_key_norm[k] = None
    #             continue
    #         b = _per_sweep_baseline_time(Y)
    #         y2_ds_by_key_norm[k] = _normalize_matrix(Y, b)

    #     hist_vals_raw, hist_vals_norm = _build_hist_vals_raw_and_norm()

    #     # ---------- Create two-page PDF ----------
    #     nrows = max(1, len(keys_for_csv))
    #     fig_w = 13
    #     fig_h = max(4.0, 2.8 * nrows)

    #     with PdfPages(out_path) as pdf:
    #         # Page 1 — RAW
    #         fig1, axes1 = plt.subplots(nrows, 3, figsize=(fig_w, fig_h), squeeze=False)
    #         plt.subplots_adjust(hspace=0.6, wspace=0.25)
    #         _plot_grid(fig1, axes1, y2_ds_by_key, hist_vals_raw, title_suffix="")
    #         fig1.suptitle("PlethApp summary — raw", y=0.995, fontsize=12)
    #         pdf.savefig(fig1, dpi=150)
    #         plt.close(fig1)

    #         # Page 2 — NORMALIZED
    #         fig2, axes2 = plt.subplots(nrows, 3, figsize=(fig_w, fig_h), squeeze=False)
    #         plt.subplots_adjust(hspace=0.6, wspace=0.25)
    #         _plot_grid(fig2, axes2, y2_ds_by_key_norm, hist_vals_norm, title_suffix=" (norm)")
    #         fig2.suptitle("PlethApp summary — normalized", y=0.995, fontsize=12)
    #         pdf.savefig(fig2, dpi=150)
    #         plt.close(fig2)

    def _save_metrics_summary_pdf(
        self,
        out_path,
        t_ds_csv: np.ndarray,
        y2_ds_by_key: dict,
        keys_for_csv: list[str],
        label_by_key: dict[str, str],
        stim_zero: float | None,
        stim_dur: float | None,
        return_figures: bool = False,
         ):
        """
        Build a three-page PDF (or return figures for preview).

        If return_figures=True: Returns (fig1, fig2, fig3) tuple instead of saving.
        If return_figures=False: Saves PDF to out_path.

        • Page 1: rows = metrics, cols = [all sweeps | mean±SEM | histograms] using RAW data
        • Page 2: same layout, using NORMALIZED data (time-based, per-sweep baseline)
        • Page 3: same layout, using NORMALIZED data (eupnea-based, pooled baseline)
        • NEW: overlay orange star markers at times where sighs occurred (first two columns)
            and at x = metric value of sigh breaths (histogram column).
        """
        import numpy as np
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages

        st = self.state
        n_sweeps_total = next(iter(y2_ds_by_key.values())).shape[1] if y2_ds_by_key else 0
        M = len(t_ds_csv)
        have_stim = (stim_zero is not None and stim_dur is not None)

        # --- normalization knobs (match exporter) ---
        NORM_BASELINE_WINDOW_S = float(getattr(self, "_norm_window_s", 10.0))
        EPS_BASE = 1e-12

        # ---------- Helpers ----------
        def _per_sweep_baseline_time(A_ds: np.ndarray) -> np.ndarray:
            b = np.full((A_ds.shape[1],), np.nan, dtype=float)
            mask_pre  = (t_ds_csv >= -NORM_BASELINE_WINDOW_S) & (t_ds_csv < 0.0)
            mask_post = (t_ds_csv >=  0.0) & (t_ds_csv <= NORM_BASELINE_WINDOW_S)
            for s in range(A_ds.shape[1]):
                col = A_ds[:, s]
                vals = col[mask_pre]
                vals = vals[np.isfinite(vals)]
                if vals.size == 0:
                    vals = col[mask_post]
                    vals = vals[np.isfinite(vals)]
                if vals.size:
                    b[s] = float(np.mean(vals))
            return b

        def _normalize_matrix(A_ds: np.ndarray, b: np.ndarray) -> np.ndarray:
            out = np.full_like(A_ds, np.nan)
            for s in range(A_ds.shape[1]):
                bs = b[s]
                if np.isfinite(bs) and abs(bs) > EPS_BASE:
                    out[:, s] = A_ds[:, s] / bs
            return out

        def _plot_sigh_time_stars(ax, sigh_times_rel):
            """Overlay stars at sigh times near the top of a time-series axis without changing limits."""
            if not sigh_times_rel:
                return
            ylim = ax.get_ylim()
            y_star = ylim[1] - 0.06 * (ylim[1] - ylim[0])
            ax.plot(
                sigh_times_rel,
                [y_star] * len(sigh_times_rel),
                linestyle="none",
                marker="*",
                markersize=9,
                color="#ff9f1a",
                markeredgecolor="#a35400",
                alpha=0.95,
                zorder=6,
            )
            ax.set_ylim(*ylim)

        def _plot_hist_stars(ax, star_x_vals):
            """Overlay stars along the top of histogram axes at given x positions (metric values on sigh breaths)."""
            if not star_x_vals:
                return
            ylim = ax.get_ylim()
            y_star = ylim[1] - 0.06 * (ylim[1] - ylim[0])
            ax.scatter(
                np.asarray(star_x_vals, dtype=float),
                np.full(len(star_x_vals), y_star, dtype=float),
                marker="*",
                s=70,
                linewidths=0.9,
                edgecolors="#a35400",
                facecolors="#ff9f1a",
                zorder=6,
                clip_on=False,
            )
            ax.set_ylim(*ylim)

        # Build histogram pools AND collect (a) sigh times (for time-series stars)
        # and (b) metric values at sigh breaths (for histogram stars).
        def _build_hist_vals_raw_and_norm():
            hist_raw = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
            hist_nrm = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
            sigh_vals_raw_by_key  = {k: [] for k in keys_for_csv}   # x-positions for stars on histogram (raw)
            sigh_vals_norm_by_key = {k: [] for k in keys_for_csv}   # x-positions for stars on histogram (norm)
            sigh_times_rel = []  # for time-series stars (seconds, relative to stim_zero if provided)

            kept = [s for s in range(next(iter(st.sweeps.values())).shape[1])
                    if s not in getattr(st, "omitted_sweeps", set())]
            t0 = float(stim_zero) if stim_zero is not None else 0.0

            for s in kept:
                y_proc = self._get_processed_for(st.analyze_chan, s)
                pks    = np.asarray(self.state.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                br     = self.state.breath_by_sweep.get(s, None)
                if br is None and pks.size:
                    try:
                        br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
                    except TypeError:
                        br = peakdet.compute_breath_events(y_proc, pks)
                    self.state.breath_by_sweep[s] = br
                if br is None:
                    br = {"onsets": np.array([], dtype=int)}

                on = np.asarray(br.get("onsets", []), dtype=int)
                if on.size < 2:
                    continue

                mids = (on[:-1] + on[1:]) // 2
                t_rel_all = (st.t[mids] - t0).astype(float)

                # Whether each breath interval [on[j], on[j+1]) contains a sigh peak
                sigh_idx = np.asarray(self.state.sigh_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                is_sigh_breath = np.zeros(len(mids), dtype=bool)
                if sigh_idx.size:
                    for j in range(len(mids)):
                        a = int(on[j]); b = int(on[j + 1])
                        if np.any((sigh_idx >= a) & (sigh_idx < b)):
                            is_sigh_breath[j] = True
                            # time for time-series star (use breath midpoint)
                            sigh_times_rel.append(float(st.t[int(mids[j])] - t0))

                # Build metric traces - use global cache if available
                traces = None
                if hasattr(self, '_global_trace_cache'):
                    traces = self._global_trace_cache.get(s, None)

                if traces is None:
                    # Compute if not cached
                    traces = {}
                    for k in keys_for_csv:
                        try:
                            traces[k] = self._compute_metric_trace(k, st.t, y_proc, st.sr_hz, pks, br)
                        except TypeError:
                            traces[k] = None

                # Per-sweep baselines for normalization (breath-based)
                mask_pre_b  = (t_rel_all >= -NORM_BASELINE_WINDOW_S) & (t_rel_all < 0.0)
                mask_post_b = (t_rel_all >=  0.0) & (t_rel_all <= NORM_BASELINE_WINDOW_S)
                b_by_k = {}
                for k in keys_for_csv:
                    arr = traces.get(k, None)
                    if arr is None or len(arr) != len(st.t):
                        b_by_k[k] = np.nan
                        continue
                    vals = arr[mids[mask_pre_b]]
                    vals = vals[np.isfinite(vals)]
                    if vals.size == 0:
                        vals = arr[mids[mask_post_b]]
                        vals = vals[np.isfinite(vals)]
                    b_by_k[k] = float(np.mean(vals)) if vals.size else np.nan

                # Fill histogram pools & record sigh metric values
                for j, (idx_mid, t_rel) in enumerate(zip(mids, t_rel_all)):
                    for k in keys_for_csv:
                        arr = traces.get(k, None)
                        if arr is None or len(arr) != len(st.t):
                            continue
                        v = float(arr[int(idx_mid)])
                        if not np.isfinite(v):
                            continue

                        # RAW pools
                        hist_raw[k]["all"].append(v)
                        if have_stim:
                            if t_rel < 0:
                                hist_raw[k]["baseline"].append(v)
                            elif 0.0 <= t_rel <= stim_dur:
                                hist_raw[k]["stim"].append(v)
                            else:
                                hist_raw[k]["post"].append(v)

                        # NORM pools
                        b = b_by_k.get(k, np.nan)
                        vn = (v / b) if (np.isfinite(b) and abs(b) > EPS_BASE) else np.nan
                        if np.isfinite(vn):
                            hist_nrm[k]["all"].append(vn)
                            if have_stim:
                                if t_rel < 0:
                                    hist_nrm[k]["baseline"].append(vn)
                                elif 0.0 <= t_rel <= stim_dur:
                                    hist_nrm[k]["stim"].append(vn)
                                else:
                                    hist_nrm[k]["post"].append(vn)

                        # SIGH stars (hist column): collect x = metric values on sigh breaths
                        if is_sigh_breath[j]:
                            sigh_vals_raw_by_key[k].append(v)
                            if np.isfinite(vn):
                                sigh_vals_norm_by_key[k].append(vn)

            # Dedup very-close times for cleaner star rows in time plots
            if len(st.t) > 1:
                dt = float(np.median(np.diff(st.t)))
            else:
                dt = (1.0 / float(st.sr_hz)) if getattr(st, "sr_hz", None) else 0.01
            eps = max(1e-9, 2.0 * dt)
            sigh_times_rel = sorted(sigh_times_rel)
            sigh_times_clean = []
            for tt in sigh_times_rel:
                if not sigh_times_clean or abs(tt - sigh_times_clean[-1]) > eps:
                    sigh_times_clean.append(tt)

            return hist_raw, hist_nrm, sigh_vals_raw_by_key, sigh_vals_norm_by_key, sigh_times_clean

        def _rowwise_mean_sem(Y2D: np.ndarray):
                    import numpy as np
                    # Y2D shape: (M, S) = (time, sweeps)
                    if Y2D is None or Y2D.size == 0:
                        M = 0
                        return np.array([]), np.array([])
                    M = Y2D.shape[0]
                    n = np.sum(np.isfinite(Y2D), axis=1)          # (M,)
                    mean = np.full(M, np.nan, float)
                    sem  = np.full(M, np.nan, float)

                    msk_mean = n > 0
                    if np.any(msk_mean):
                        mean[msk_mean] = np.nanmean(Y2D[msk_mean, :], axis=1)

                    msk_sem = n >= 2
                    if np.any(msk_sem):
                        std = np.nanstd(Y2D[msk_sem, :], axis=1, ddof=1)
                        sem[msk_sem] = std / np.sqrt(n[msk_sem])
                    return mean, sem

        def _plot_grid(fig, axes, Y_by_key, hist_vals, sigh_hist_vals_by_key, sigh_times_rel, title_suffix):
            """Render one page (grid) given series & histogram data dicts + star overlays."""
            nrows = max(1, len(keys_for_csv))
            for r, k in enumerate(keys_for_csv):
                label = label_by_key.get(k, k)
                is_bottom_row = (r == nrows - 1)
                is_top_row = (r == 0)

                # --- col 1: all sweeps overlaid ---
                ax1 = axes[r, 0]
                Y = Y_by_key.get(k, None)
                if Y is not None and Y.shape[0] == M:
                    for s in range(Y.shape[1]):
                        ax1.plot(t_ds_csv, Y[:, s], lw=0.8, alpha=0.45)
                if have_stim:
                    ax1.axvspan(0.0, stim_dur, color="#2E5090", alpha=0.25)
                    ax1.axvline(0.0, color="#2E5090", lw=1.0, alpha=0.8)
                    ax1.axvline(stim_dur, color="#2E5090", lw=1.0, alpha=0.8, ls="--")
                _plot_sigh_time_stars(ax1, sigh_times_rel)
                # Use less padding on top row to make room for figure title
                title_pad = 3 if is_top_row else 8
                ax1.set_title(f"{label} — all sweeps{title_suffix}", fontsize=9, pad=title_pad)
                ax1.set_ylabel(label, fontsize=8)  # Add y-label with metric name
                ax1.set_xlabel("Time (s)", fontsize=8)

                # --- col 2: mean ± SEM ---
                ax2 = axes[r, 1]
                if Y is not None and Y.shape[0] == M:
                    mean, sem = _rowwise_mean_sem(Y)
                    ax2.plot(t_ds_csv, mean, lw=1.8)
                    ax2.fill_between(t_ds_csv, mean - sem, mean + sem, alpha=0.25, linewidth=0)

                if have_stim:
                    ax2.axvspan(0.0, stim_dur, color="#2E5090", alpha=0.25)
                    ax2.axvline(0.0, color="#2E5090", lw=1.0, alpha=0.8)
                    ax2.axvline(stim_dur, color="#2E5090", lw=1.0, alpha=0.8, ls="--")
                _plot_sigh_time_stars(ax2, sigh_times_rel)
                ax2.set_title(f"{label} — mean ± SEM{title_suffix}", fontsize=9, pad=title_pad)
                ax2.set_ylabel(label, fontsize=8)  # Add y-label with metric name
                ax2.set_xlabel("Time (s)", fontsize=8)

                # --- col 3: line histograms (density) + stars at sigh metric values ---
                ax3 = axes[r, 2]
                groups = []
                for nm in ("all", "baseline", "stim", "post"):
                    vals = np.asarray(hist_vals[k][nm], dtype=float)
                    if vals.size:
                        groups.append(vals)

                if len(groups):
                    combined = np.concatenate(groups)
                    edges = np.histogram_bin_edges(combined, bins="auto")
                    centers = 0.5 * (edges[:-1] + edges[1:])

                    def _plot_line(vals, lbl, style_kw):
                        vals = np.asarray(vals, dtype=float)
                        if vals.size == 0:
                            return
                        dens, _ = np.histogram(vals, bins=edges, density=True)
                        ax3.plot(centers, dens, **style_kw, label=lbl)

                    _plot_line(hist_vals[k]["all"], "All", dict(lw=1.8))
                    if have_stim:
                        _plot_line(hist_vals[k]["baseline"], "Baseline", dict(lw=1.6))
                        _plot_line(hist_vals[k]["stim"],     "Stim",     dict(lw=1.6, ls="--"))
                        _plot_line(hist_vals[k]["post"],     "Post",     dict(lw=1.6, ls=":"))

                ax3.set_title(f"{label} — distribution (density){title_suffix}", fontsize=9, pad=title_pad)
                ax3.set_ylabel("Density", fontsize=8)
                ax3.set_xlabel(label, fontsize=8)  # Add metric name as x-label
                # Only show legend on top row to save space
                if len(ax3.lines) and is_top_row:
                    ax3.legend(loc="best", fontsize=7)

                # NEW: stars for histogram at sigh metric values (use "all" sigh values)
                _plot_hist_stars(ax3, sigh_hist_vals_by_key.get(k, []))

                # Reduce tick label font size for all axes
                ax1.tick_params(labelsize=7)
                ax2.tick_params(labelsize=7)
                ax3.tick_params(labelsize=7)

            fig.tight_layout()

        # ---------- Prepare both RAW and NORMALIZED datasets ----------
        y2_ds_by_key_norm = {}
        for k in keys_for_csv:
            Y = y2_ds_by_key.get(k, None)
            if Y is None or not Y.size:
                y2_ds_by_key_norm[k] = None
                continue
            b = _per_sweep_baseline_time(Y)
            y2_ds_by_key_norm[k] = _normalize_matrix(Y, b)

        # Prepare EUPNEA-BASED normalized datasets
        y2_ds_by_key_norm_eupnea = {}
        eupnea_baselines_by_key = {}  # Store computed baselines for histogram building
        eupnea_thresh = self._parse_float(self.EupneaThresh) or 5.0  # Hz

        # OPTIMIZATION: Compute eupnea masks once per sweep, reuse for all metrics
        eupnea_masks_by_sweep = {}
        kept = [s for s in range(next(iter(st.sweeps.values())).shape[1])
                if s not in getattr(st, "omitted_sweeps", set())]

        print(f"[PDF] Computing eupnea masks for {len(kept)} sweeps...")
        for s in kept:
            y_proc = self._get_processed_for(st.analyze_chan, s)
            pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
            br = st.breath_by_sweep.get(s, None)
            if br is None and pks.size:
                try:
                    br = peakdet.compute_breath_events(y_proc, pks, st.sr_hz)
                except TypeError:
                    br = peakdet.compute_breath_events(y_proc, pks)

            if br is not None:
                on = np.asarray(br.get("onsets", []), dtype=int)
                off = np.asarray(br.get("offsets", []), dtype=int)
                expmins = np.asarray(br.get("expmins", []), dtype=int)
                expoffs = np.asarray(br.get("expoffs", []), dtype=int)

                if on.size >= 2:
                    eupnea_mask = metrics.detect_eupnic_regions(
                        st.t, y_proc, st.sr_hz, pks, on, off, expmins, expoffs,
                        freq_threshold_hz=eupnea_thresh,
                        min_duration_sec=self.eupnea_min_duration
                    )
                    eupnea_masks_by_sweep[s] = eupnea_mask

        print(f"[PDF] Computed {len(eupnea_masks_by_sweep)} eupnea masks")

        print(f"[PDF] Building histograms (raw and time-normalized)...")
        # Pools + sigh overlays
        (hist_vals_raw,
        hist_vals_norm,
        sigh_vals_raw_by_key,
        sigh_vals_norm_by_key,
        sigh_times_rel) = _build_hist_vals_raw_and_norm()
        print(f"[PDF] Built raw and time-normalized histograms")

        # Build EUPNEA-normalized histograms using a simple breath-based approach
        print(f"[PDF] Building eupnea-normalized histograms...")

        def _build_eupnea_normalized_hists():
            """
            Simple approach:
            1. Collect all breath metric values from eupneic baseline periods
            2. Compute mean baseline per metric
            3. Normalize all breath values by that baseline
            """
            # Step 1: Collect eupneic baseline breath values
            eupnea_baseline_breaths = {k: [] for k in keys_for_csv}
            all_breath_values_raw = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
            sigh_vals_raw = {k: [] for k in keys_for_csv}

            kept = [s for s in range(next(iter(st.sweeps.values())).shape[1])
                    if s not in getattr(st, "omitted_sweeps", set())]
            t0 = float(stim_zero) if stim_zero is not None else 0.0

            for s in kept:
                y_proc = self._get_processed_for(st.analyze_chan, s)
                pks = np.asarray(st.peaks_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                br = st.breath_by_sweep.get(s, None)
                if br is None or pks.size < 2:
                    continue

                on = np.asarray(br.get("onsets", []), dtype=int)
                if on.size < 2:
                    continue

                # Get eupnea mask for this sweep
                eupnea_mask = eupnea_masks_by_sweep.get(s, None)
                if eupnea_mask is None:
                    continue

                mids = (on[:-1] + on[1:]) // 2
                t_rel_all = (st.t[mids] - t0).astype(float)

                # Check which breaths are sighs
                sigh_idx = np.asarray(st.sigh_by_sweep.get(s, np.array([], dtype=int)), dtype=int)
                is_sigh_breath = np.zeros(len(mids), dtype=bool)
                if sigh_idx.size:
                    for j in range(len(mids)):
                        a = int(on[j]); b = int(on[j + 1])
                        if np.any((sigh_idx >= a) & (sigh_idx < b)):
                            is_sigh_breath[j] = True

                # For each breath, check if midpoint is eupneic and collect metric values
                for i, mid_idx in enumerate(mids):
                    t_rel = t_rel_all[i]
                    is_eupneic = eupnea_mask[mid_idx] > 0 if mid_idx < len(eupnea_mask) else False
                    is_baseline = t_rel < 0

                    # Determine region
                    if have_stim:
                        if t_rel < 0:
                            region = "baseline"
                        elif t_rel <= stim_dur:
                            region = "stim"
                        else:
                            region = "post"
                    else:
                        region = "all"

                    # Extract metric values from raw histograms (already computed by _build_hist_vals_raw_and_norm)
                    # We'll use the same raw values that were already extracted
                    # But we need to recompute - let's just use the data from hist_vals_raw
                    # Actually, we can't easily map back. Let's sample from y2_ds_by_key at this time point

                    for k in keys_for_csv:
                        Y_raw = y2_ds_by_key.get(k, None)
                        if Y_raw is None or Y_raw.size == 0:
                            continue

                        # Find closest downsampled time point
                        ds_idx = np.argmin(np.abs(t_ds_csv - t_rel))
                        if ds_idx >= Y_raw.shape[0]:
                            continue

                        col_idx = kept.index(s)
                        val = Y_raw[ds_idx, col_idx]

                        if not np.isfinite(val):
                            continue

                        # Collect for baseline calculation if eupneic and baseline
                        if is_eupneic and is_baseline:
                            eupnea_baseline_breaths[k].append(val)

                        # Collect all breath values (will normalize later)
                        all_breath_values_raw[k]["all"].append(val)
                        if have_stim:
                            all_breath_values_raw[k][region].append(val)

                        # Collect sigh values
                        if is_sigh_breath[i]:
                            sigh_vals_raw[k].append(val)

            # Step 2: Compute eupneic baseline means
            eupnea_baselines = {}
            for k in keys_for_csv:
                if len(eupnea_baseline_breaths[k]) >= 10:
                    eupnea_baselines[k] = np.mean(eupnea_baseline_breaths[k])
                elif len(eupnea_baseline_breaths[k]) > 0:
                    # Use what we have if < 10
                    eupnea_baselines[k] = np.mean(eupnea_baseline_breaths[k])
                else:
                    eupnea_baselines[k] = np.nan

            # Step 3: Normalize all values
            hist_norm_eupnea = {k: {"all": [], "baseline": [], "stim": [], "post": []} for k in keys_for_csv}
            sigh_vals_norm_eupnea = {k: [] for k in keys_for_csv}

            for k in keys_for_csv:
                baseline = eupnea_baselines.get(k, np.nan)
                if not np.isfinite(baseline) or abs(baseline) < EPS_BASE:
                    continue

                for region in ["all", "baseline", "stim", "post"]:
                    hist_norm_eupnea[k][region] = [v / baseline for v in all_breath_values_raw[k][region]]

                sigh_vals_norm_eupnea[k] = [v / baseline for v in sigh_vals_raw[k]]

            # Also store baselines and normalized time series for plotting
            y2_norm_eupnea = {}
            for k in keys_for_csv:
                Y_raw = y2_ds_by_key.get(k, None)
                baseline = eupnea_baselines.get(k, np.nan)
                if Y_raw is not None and np.isfinite(baseline) and abs(baseline) > EPS_BASE:
                    y2_norm_eupnea[k] = Y_raw / baseline
                else:
                    y2_norm_eupnea[k] = None

            return hist_norm_eupnea, sigh_vals_norm_eupnea, eupnea_baselines, y2_norm_eupnea

        hist_vals_norm_eupnea, sigh_vals_norm_eupnea_by_key, eupnea_baselines_by_key, y2_ds_by_key_norm_eupnea = _build_eupnea_normalized_hists()
        print(f"[PDF] Built eupnea-normalized histograms")

        # ---------- Create three-page PDF ----------
        nrows = max(1, len(keys_for_csv))
        fig_w = 13
        fig_h = max(4.0, 2.6 * nrows)  # Reduced from 5.25 to 2.6 per row (half of 5.25)

        # Page 1 — RAW
        fig1, axes1 = plt.subplots(nrows, 3, figsize=(fig_w, fig_h), squeeze=False)
        _plot_grid(fig1, axes1, y2_ds_by_key, hist_vals_raw, sigh_vals_raw_by_key, sigh_times_rel, title_suffix="")
        fig1.suptitle("Summary — raw", fontsize=11)
        fig1.tight_layout(rect=[0, 0, 1, 0.99])  # Leave 1% at top for suptitle

        # Page 2 — NORMALIZED (time-based)
        fig2, axes2 = plt.subplots(nrows, 3, figsize=(fig_w, fig_h), squeeze=False)
        _plot_grid(fig2, axes2, y2_ds_by_key_norm, hist_vals_norm, sigh_vals_norm_by_key, sigh_times_rel, title_suffix=" (norm)")
        fig2.suptitle("Summary — normalized (time-based)", fontsize=11)
        fig2.tight_layout(rect=[0, 0, 1, 0.99])  # Leave 1% at top for suptitle

        # Page 3 — NORMALIZED (eupnea-based)
        fig3, axes3 = plt.subplots(nrows, 3, figsize=(fig_w, fig_h), squeeze=False)
        _plot_grid(fig3, axes3, y2_ds_by_key_norm_eupnea, hist_vals_norm_eupnea, sigh_vals_norm_eupnea_by_key, sigh_times_rel, title_suffix=" (norm eupnea)")
        fig3.suptitle("Summary — normalized (eupnea-based)", fontsize=11)
        fig3.tight_layout(rect=[0, 0, 1, 0.99])  # Leave 1% at top for suptitle

        # Either return figures or save to PDF
        if return_figures:
            return fig1, fig2, fig3
        else:
            with PdfPages(out_path) as pdf:
                pdf.savefig(fig1, dpi=150)
                pdf.savefig(fig2, dpi=150)
                pdf.savefig(fig3, dpi=150)
            plt.close(fig1)
            plt.close(fig2)
            plt.close(fig3)


    def _show_summary_preview_dialog(self, t_ds_csv, y2_ds_by_key, keys_for_csv, label_by_key, stim_zero, stim_dur):
        """Display interactive preview dialog with the three summary figures."""
        import matplotlib.pyplot as plt
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QScrollArea
        from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas

        # Generate figures using the same method that saves PDFs, but get figures back instead
        fig1, fig2, fig3 = self._save_metrics_summary_pdf(
            out_path=None,  # Not used when return_figures=True
            t_ds_csv=t_ds_csv,
            y2_ds_by_key=y2_ds_by_key,
            keys_for_csv=keys_for_csv,
            label_by_key=label_by_key,
            stim_zero=stim_zero,
            stim_dur=stim_dur,
            return_figures=True,
        )

        # -------------------- Display in dialog --------------------
        dialog = QDialog(self)
        dialog.setWindowTitle("Summary Preview")
        dialog.resize(1300, 900)  # Slightly larger window

        main_layout = QVBoxLayout(dialog)

        # Page selector controls at top
        control_layout = QHBoxLayout()
        page_label = QLabel("Page 1 of 3 (Raw)")
        prev_btn = QPushButton("← Previous")
        next_btn = QPushButton("Next →")
        close_btn = QPushButton("Close")

        control_layout.addWidget(prev_btn)
        control_layout.addWidget(page_label)
        control_layout.addWidget(next_btn)
        control_layout.addStretch()
        control_layout.addWidget(close_btn)

        main_layout.addLayout(control_layout)

        # Create scroll area for the canvas with mouse wheel support
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)  # Allow canvas to resize to fit width
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        # Enable mouse wheel scrolling by setting focus policy
        scroll_area.setFocusPolicy(Qt.FocusPolicy.WheelFocus)

        # Canvas for displaying matplotlib figures
        canvas1 = FigureCanvas(fig1)
        canvas2 = FigureCanvas(fig2)
        canvas3 = FigureCanvas(fig3)

        # Set size policies to allow width scaling but keep height fixed
        from PyQt6.QtWidgets import QSizePolicy
        from PyQt6.QtCore import QEvent
        for canvas in [canvas1, canvas2, canvas3]:
            canvas.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            canvas.setMinimumWidth(800)  # Minimum width

        # Create a container widget to hold all three canvases (so they don't get deleted)
        from PyQt6.QtWidgets import QWidget, QStackedWidget
        canvas_stack = QStackedWidget()
        canvas_stack.addWidget(canvas1)  # index 0
        canvas_stack.addWidget(canvas2)  # index 1
        canvas_stack.addWidget(canvas3)  # index 2
        canvas_stack.setCurrentIndex(0)  # Start with page 1

        # Install event filter on canvases to forward wheel events to scroll area
        class WheelEventFilter:
            def __init__(self, scroll_area):
                self.scroll_area = scroll_area

            def eventFilter(self, obj, event):
                if event.type() == QEvent.Type.Wheel:
                    # Forward wheel event to scroll area's viewport
                    scrollbar = self.scroll_area.verticalScrollBar()
                    scrollbar.setValue(scrollbar.value() - event.angleDelta().y())
                    return True  # Event handled
                return False  # Let other events pass through

        wheel_filter = WheelEventFilter(scroll_area)
        from PyQt6.QtCore import QObject

        class EventFilterObject(QObject):
            def __init__(self, filter_func):
                super().__init__()
                self.filter_func = filter_func

            def eventFilter(self, obj, event):
                return self.filter_func(obj, event)

        filter_obj = EventFilterObject(wheel_filter.eventFilter)
        canvas1.installEventFilter(filter_obj)
        canvas2.installEventFilter(filter_obj)
        canvas3.installEventFilter(filter_obj)

        # Put the stacked widget in the scroll area
        scroll_area.setWidget(canvas_stack)
        main_layout.addWidget(scroll_area)

        # Page navigation
        current_page = [1]  # Use list to allow modification in nested function

        def update_page():
            if current_page[0] == 1:
                canvas_stack.setCurrentIndex(0)  # Show canvas1
                page_label.setText("Page 1 of 3 (Raw)")
                prev_btn.setEnabled(False)
                next_btn.setEnabled(True)
            elif current_page[0] == 2:
                canvas_stack.setCurrentIndex(1)  # Show canvas2
                page_label.setText("Page 2 of 3 (Normalized - Time-based)")
                prev_btn.setEnabled(True)
                next_btn.setEnabled(True)
            else:
                canvas_stack.setCurrentIndex(2)  # Show canvas3
                page_label.setText("Page 3 of 3 (Normalized - Eupnea-based)")
                prev_btn.setEnabled(True)
                next_btn.setEnabled(False)
            # Reset scroll position to top when changing pages
            scroll_area.verticalScrollBar().setValue(0)

        def go_prev():
            current_page[0] = max(1, current_page[0] - 1)
            update_page()

        def go_next():
            current_page[0] = min(3, current_page[0] + 1)
            update_page()

        prev_btn.clicked.connect(go_prev)
        next_btn.clicked.connect(go_next)
        close_btn.clicked.connect(dialog.close)

        update_page()
        dialog.exec()

        # Cleanup
        plt.close(fig1)
        plt.close(fig2)
        plt.close(fig3)


    def _sigh_sample_indices(self, s: int, pks: np.ndarray | None) -> set[int]:
        """
        Return a set of SAMPLE indices (into st.t / y) for sigh-marked peaks on sweep s,
        regardless of how they were originally stored.

        Accepts any of these storage patterns per sweep:
        • sample indices (ints 0..N-1)
        • indices INTO the peaks list (ints 0..len(pks)-1), which we map via pks[idx]
        • times in seconds (floats), which we map to nearest sample via searchsorted
        • numpy array / list / set in any of the above forms
        """
        st = self.state
        N = len(st.t)

        # Prefer the name you've been using; try some alternates if needed.
        candidates = None
        for name in ("sighs_by_sweep", "sigh_indices_by_sweep", "sigh_peaks_by_sweep", "sigh_mask_by_sweep"):
            if hasattr(st, name):
                candidates = getattr(st, name).get(s, None)
                if candidates is not None:
                    break

        if candidates is None:
            return set()

        arr = np.asarray(list(candidates))
        out: set[int] = set()

        # If integer-like
        if arr.dtype.kind in "iu":
            arr = arr.astype(int)
            if pks is not None and arr.size and arr.max(initial=-1) < len(pks):
                # Looks like indexes INTO peaks -> map to sample indexes
                for idx in arr:
                    if 0 <= idx < len(pks):
                        i = int(pks[idx])
                        if 0 <= i < N:
                            out.add(i)
            else:
                # Already sample indexes
                for i in arr:
                    if 0 <= i < N:
                        out.add(int(i))
            return out

        # If float-like -> assume times (seconds)
        if arr.dtype.kind in "f":
            t = st.t
            for val in arr:
                try:
                    i = int(np.clip(np.searchsorted(t, float(val)), 0, N - 1))
                    out.add(i)
                except Exception:
                    pass
            return out

        # Fallback: try to coerce each element
        for v in arr:
            try:
                i = int(v)
                if 0 <= i < N:
                    out.add(i)
            except Exception:
                try:
                    i = int(np.clip(np.searchsorted(st.t, float(v)), 0, N - 1))
                    out.add(i)
                except Exception:
                    pass
        return out



    ##################################################
    ##Curration                                     ##
    ##################################################
    # def on_curation_choose_dir_clicked(self):
    #     base = QFileDialog.getExistingDirectory(self, "Choose a folder to scan", self._last_scanned_dir if hasattr(self, "_last_scanned_dir") else str(Path.home()))
    #     if not base:
    #         return
    #     self._last_scanned_dir = base
    #     groups = self._scan_csv_groups(Path(base))
    #     self._populate_file_list_from_groups(groups)

    def on_curation_choose_dir_clicked(self):
        # Get last used directory from settings, default to home
        last_dir = self.settings.value("curation_last_dir", str(Path.home()))
        
        # Ensure the path exists, otherwise fall back to home
        if not Path(last_dir).exists():
            last_dir = str(Path.home())
        
        base = QFileDialog.getExistingDirectory(
            self, "Choose a folder to scan", last_dir
        )
        
        if not base:
            return
        
        # Save the selected directory for next time
        self.settings.setValue("curation_last_dir", base)
        
        groups = self._scan_csv_groups(Path(base))
        self._populate_file_list_from_groups(groups)

    def _scan_csv_groups(self, base_dir: Path):
        """
        Walk base_dir recursively and group CSVs by common root:
        root + '_breaths.csv'
        root + '_timeseries.csv'
        root + '_events.csv'
        Returns a list of dicts: {key, root, dir, breaths, means, events}
        """
        groups = {}
        for dirpath, _, filenames in os.walk(str(base_dir)):
            for fn in filenames:
                lower = fn.lower()
                if not lower.endswith(".csv"):
                    continue

                kind = None
                if lower.endswith("_breaths.csv"):
                    root = fn[:-len("_breaths.csv")]
                    kind = "breaths"
                elif lower.endswith("_timeseries.csv"):
                    root = fn[:-len("_timeseries.csv")]
                    kind = "means"
                elif lower.endswith("_means_by_time.csv"):  # Legacy support
                    root = fn[:-len("_means_by_time.csv")]
                    kind = "means"
                elif lower.endswith("_events.csv"):
                    root = fn[:-len("_events.csv")]
                    kind = "events"

                if kind is None:
                    continue

                dir_p = Path(dirpath)
                key = str((dir_p / root).resolve()).lower()  # unique per dir+root (case-insensitive on Win)
                entry = groups.get(key)
                if entry is None:
                    entry = {"key": key, "root": root, "dir": dir_p, "breaths": None, "means": None, "events": None}
                    groups[key] = entry
                entry[kind] = str(dir_p / fn)

        # Return as a stable, sorted list
        return sorted(groups.values(), key=lambda e: (str(e["dir"]).lower(), e["root"].lower()))


    def _populate_file_list_from_groups(self, groups: list[dict]):
        """
        Fill left list (FileList) with one item per root. Display only name,
        store both full paths in UserRole for later consolidation.
        """
        from PyQt6.QtCore import Qt
        from PyQt6.QtWidgets import QListWidgetItem

        self.FileList.clear()
        # Do not clear the right list automatically so users don't lose selections:
        # self.FilestoConsolidateList.clear()

        for g in groups:
            root = g["root"]
            has_b = bool(g["breaths"])
            has_m = bool(g["means"])
            has_e = bool(g["events"])

            # Build suffix showing what files are present
            parts = []
            if has_b:
                parts.append("breaths")
            if has_m:
                parts.append("timeseries")
            if has_e:
                parts.append("events")

            if parts:
                suffix = f"[{' + '.join(parts)}]"
            else:
                # Shouldn't happen; skip if nothing is present
                continue

            item = QListWidgetItem(f"{root}  {suffix}")
            tt_lines = [f"Root: {root}", f"Dir:  {g['dir']}"]
            if g["breaths"]:
                tt_lines.append(f"breaths:    {g['breaths']}")
            if g["means"]:
                tt_lines.append(f"timeseries: {g['means']}")
            if g["events"]:
                tt_lines.append(f"events:     {g['events']}")
            item.setToolTip("\n".join(tt_lines))

            # Store full metadata for later use
            item.setData(Qt.ItemDataRole.UserRole, g)  # {'key', 'root', 'dir', 'breaths', 'means', 'events'}

            self.FileList.addItem(item)

        # Optional: sort visually
        self.FileList.sortItems()

    def _filter_file_list(self, text: str):
        """Show/hide items in FileList based on search text.

        Supports multiple search modes:
        - Single keyword: 'gfp' - shows files containing 'gfp'
        - Multiple keywords (AND): 'gfp 2.5mW' - shows files containing BOTH 'gfp' AND '2.5mW'
        - Multiple keywords (OR): 'gfp, chr2' - shows files containing EITHER 'gfp' OR 'chr2'
        """
        search_text = text.strip().lower()

        # Determine search mode
        if ',' in search_text:
            # OR mode: split by comma
            keywords = [k.strip() for k in search_text.split(',') if k.strip()]
            search_mode = 'OR'
        else:
            # AND mode: split by whitespace
            keywords = [k.strip() for k in search_text.split() if k.strip()]
            search_mode = 'AND'

        for i in range(self.FileList.count()):
            item = self.FileList.item(i)
            if not item:
                continue

            # Get the display text
            item_text = item.text().lower()

            # Also search in tooltip (which contains full path)
            tooltip = (item.toolTip() or "").lower()
            combined_text = f"{item_text} {tooltip}"

            # Show item if search text is empty
            if not keywords:
                item.setHidden(False)
                continue

            # Apply search logic
            if search_mode == 'AND':
                # ALL keywords must be present
                matches = all(kw in combined_text for kw in keywords)
            else:  # OR mode
                # ANY keyword must be present
                matches = any(kw in combined_text for kw in keywords)

            item.setHidden(not matches)

    def _curation_scan_and_fill(self, root: Path):
        """Scan for matching CSVs and fill FileList with filenames (store full paths in item data)."""
        from PyQt6.QtWidgets import QListWidgetItem
        from PyQt6.QtCore import Qt

        # Clear existing items
        self.FileList.clear()

        # Patterns to include (recursive)
        patterns = ["*_breaths.csv", "*_timeseries.csv", "*_means_by_time.csv", "*_events.csv"]

        files = []
        try:
            for pat in patterns:
                files.extend(root.rglob(pat))
        except Exception as e:
            QMessageBox.critical(self, "Scan error", f"Failed to scan folder:\n{root}\n\n{e}")
            return

        # Deduplicate & sort (by name, then path for stability)
        uniq = {}
        for p in files:
            try:
                # Only include files (ignore dirs, weird links)
                if p.is_file():
                    # keep all—even if names clash—because display is name-only,
                    # but we keep full path in item data and tooltip
                    uniq[str(p)] = p
            except Exception:
                pass

        files_sorted = sorted(uniq.values(), key=lambda x: (x.name.lower(), str(x).lower()))

        if not files_sorted:
            try:
                self.statusbar.showMessage("No matching CSV files found in the selected folder.", 4000)
            except Exception:
                pass
            return

        for p in files_sorted:
            item = QListWidgetItem(p.name)
            item.setToolTip(str(p))  # show full path on hover
            item.setData(Qt.ItemDataRole.UserRole, str(p))  # keep full path for later use
            self.FileList.addItem(item)

        # Optional: sort in the widget (already sorted, but harmless)
        self.FileList.sortItems()

    def _list_has_path(self, lw, full_path: str) -> bool:
        """Return True if any item in lw has UserRole == full_path."""
        for i in range(lw.count()):
            it = lw.item(i)
            if it and it.data(Qt.ItemDataRole.UserRole) == full_path:
                return True
        return False

    def _list_has_key(self, lw, key: str) -> bool:
        """True if any item in lw has the same group key."""
        from PyQt6.QtCore import Qt
        for i in range(lw.count()):
            it = lw.item(i)
            if not it:
                continue
            meta = it.data(Qt.ItemDataRole.UserRole) or {}
            if isinstance(meta, dict) and meta.get("key", "").lower() == key.lower():
                return True
        return False


    def _move_items(self, src_lw, dst_lw, rows_to_move: list[int]):
        """
        Move grouped items by root from src_lw to dst_lw.
        Duplicate check is by 'key' (dir+root), not by file path.
        """
        from PyQt6.QtCore import Qt
        if not rows_to_move:
            return 0, 0

        plan = []
        for r in rows_to_move:
            it = src_lw.item(r)
            if it is None:
                continue
            meta = it.data(Qt.ItemDataRole.UserRole) or {}
            key = (meta.get("key") or "").lower()
            is_dup = self._list_has_key(dst_lw, key)
            plan.append((r, is_dup))

        taken = []
        skipped_dups = 0
        for r, is_dup in sorted(plan, key=lambda x: x[0], reverse=True):
            if is_dup:
                skipped_dups += 1
                continue
            it = src_lw.takeItem(r)
            if it is not None:
                taken.append((r, it))

        moved = 0
        for _, it in sorted(taken, key=lambda x: x[0]):
            dst_lw.addItem(it)
            moved += 1

        src_lw.sortItems()
        dst_lw.sortItems()
        return moved, skipped_dups


    # def _move_items(self, src_lw, dst_lw, rows_to_move: list[int]):
    #     """
    #     Move given rows from src_lw to dst_lw, skipping duplicates by full path.
    #     Preserves per-item tooltip and UserRole. Returns (moved_count, skipped_dups).
    #     """
    #     if not rows_to_move:
    #         return 0, 0

    #     # Build a plan: check duplicates before removing anything
    #     plan = []
    #     for r in rows_to_move:
    #         it = src_lw.item(r)
    #         if it is None:
    #             continue
    #         full_path = it.data(Qt.ItemDataRole.UserRole)
    #         is_dup = self._list_has_path(dst_lw, full_path)
    #         plan.append((r, is_dup))

    #     # Take items in descending row order to avoid index shifting
    #     taken = []
    #     skipped_dups = 0
    #     for r, is_dup in sorted(plan, key=lambda x: x[0], reverse=True):
    #         if is_dup:
    #             skipped_dups += 1
    #             continue
    #         it = src_lw.takeItem(r)  # this detaches the item from src
    #         if it is not None:
    #             taken.append((r, it))

    #     # Add to destination in ascending original order
    #     moved = 0
    #     for _, it in sorted(taken, key=lambda x: x[0]):
    #         dst_lw.addItem(it)
    #         moved += 1

    #     # Optional: keep lists sorted alphabetically by name
    #     src_lw.sortItems()
    #     dst_lw.sortItems()

    #     return moved, skipped_dups

    def on_move_selected_right(self):
        """Move selected from left (FileList) to right (FilestoConsolidateList)."""
        src = self.FileList
        dst = self.FilestoConsolidateList
        rows = [src.row(it) for it in src.selectedItems()]
        moved, skipped = self._move_items(src, dst, rows)
        try:
            if moved or skipped:
                self.statusbar.showMessage(f"Moved {moved} item(s) to right. Skipped {skipped} duplicate(s).", 3000)
        except Exception:
            pass

    def on_move_all_right(self):
        """Move ALL VISIBLE from left to right."""
        src = self.FileList
        dst = self.FilestoConsolidateList
        # Only move visible (non-hidden) items
        rows = [i for i in range(src.count()) if not src.item(i).isHidden()]
        moved, skipped = self._move_items(src, dst, rows)
        try:
            if moved or skipped:
                self.statusbar.showMessage(f"Moved {moved} visible item(s) to right. Skipped {skipped} duplicate(s).", 3000)
        except Exception:
            pass

    def on_move_selected_left(self):
        """Move selected from right back to left."""
        src = self.FilestoConsolidateList
        dst = self.FileList
        rows = [src.row(it) for it in src.selectedItems()]
        moved, skipped = self._move_items(src, dst, rows)
        try:
            if moved or skipped:
                self.statusbar.showMessage(f"Moved {moved} item(s) to left. Skipped {skipped} duplicate(s).", 3000)
        except Exception:
            pass

    def on_move_all_left(self):
        """Move ALL VISIBLE from right back to left."""
        src = self.FilestoConsolidateList
        dst = self.FileList
        # Only move visible (non-hidden) items
        rows = [i for i in range(src.count()) if not src.item(i).isHidden()]
        moved, skipped = self._move_items(src, dst, rows)
        try:
            if moved or skipped:
                self.statusbar.showMessage(f"Moved {moved} visible item(s) to left. Skipped {skipped} duplicate(s).", 3000)
        except Exception:
            pass


    # def on_consolidate_save_data_clicked(self):
    #     """Consolidate data from selected files."""
    #     from PyQt6.QtCore import Qt
    #     import pandas as pd
    #     from pathlib import Path
        
    #     # Get all selected files from right list
    #     items = []
    #     for i in range(self.FilestoConsolidateList.count()):
    #         item = self.FilestoConsolidateList.item(i)
    #         if item:
    #             items.append(item)
        
    #     if not items:
    #         QMessageBox.warning(self, "Consolidate", "No files selected to consolidate.")
    #         return
        
    #     # Separate by file type
    #     means_files = []
    #     breaths_files = []
        
    #     for item in items:
    #         meta = item.data(Qt.ItemDataRole.UserRole) or {}
    #         if meta.get("means"):
    #             means_files.append((meta["root"], Path(meta["means"])))
    #         if meta.get("breaths"):
    #             breaths_files.append((meta["root"], Path(meta["breaths"])))
        
    #     if not means_files:
    #         QMessageBox.warning(self, "Consolidate", "No *_means_by_time.csv files selected.")
    #         return
        
    #     # Process means files
    #     try:
    #         consolidated_data = self._consolidate_means_files(means_files)
            
    #         # Save consolidated data
    #         save_dir = QFileDialog.getExistingDirectory(
    #             self, "Choose folder to save consolidated data",
    #             str(means_files[0][1].parent)
    #         )
            
    #         if save_dir:
    #             self._save_consolidated_data(consolidated_data, Path(save_dir))
    #             QMessageBox.information(
    #                 self, "Success", 
    #                 f"Consolidated {len(means_files)} files.\nSaved to: {save_dir}"
    #             )
        
    #     except Exception as e:
    #         QMessageBox.critical(self, "Consolidation Error", f"Failed to consolidate:\n{e}")
    #         import traceback
    #         traceback.print_exc()


    # def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Consolidate means_by_time CSV files.
    #     Returns dict: {metric_name: DataFrame with 't' and columns for each file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract (mean columns only)
    #     metrics = [
    #         'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 
    #         'ti', 'te', 'vent_proxy'
    #     ]
        
    #     consolidated = {}
        
    #     # Read first file to get time column
    #     first_root, first_path = files[0]
    #     df_first = pd.read_csv(first_path)
    #     t_values = df_first['t'].values
        
    #     for metric in metrics:
    #         metric_mean_col = f"{metric}_mean"
            
    #         # Check if this metric exists in first file
    #         if metric_mean_col not in df_first.columns:
    #             continue
            
    #         # Initialize dataframe with time column
    #         result_df = pd.DataFrame({'t': t_values})
            
    #         # Add column from each file
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             # Verify time alignment
    #             if not np.allclose(df['t'].values, t_values, rtol=1e-5, atol=1e-8):
    #                 print(f"Warning: Time mismatch in {root}")
                
    #             if metric_mean_col in df.columns:
    #                 # Use root name as column identifier
    #                 result_df[f"{metric}_mean_{root}"] = df[metric_mean_col].values
            
    #         consolidated[metric] = result_df
        
    #     # Also do normalized versions
    #     for metric in metrics:
    #         metric_norm_col = f"{metric}_norm_mean"
            
    #         if metric_norm_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_values})
            
    #         for root, path in files:
    #             df = pd.read_csv(path)
    #             if metric_norm_col in df.columns:
    #                 result_df[f"{metric}_norm_mean_{root}"] = df[metric_norm_col].values
            
    #         consolidated[f"{metric}_norm"] = result_df
        
    #     return consolidated


    # def _save_consolidated_data(self, consolidated: dict, save_dir: Path):
    #     """Save consolidated dataframes to CSV files."""
    #     for metric_name, df in consolidated.items():
    #         out_path = save_dir / f"consolidated_{metric_name}_means.csv"
    #         df.to_csv(out_path, index=False)
    #         print(f"Saved: {out_path}")

    # def on_consolidate_save_data_clicked(self):
    #     """Consolidate data from selected files into a single Excel file."""
    #     from PyQt6.QtCore import Qt
    #     import pandas as pd
    #     from pathlib import Path
        
    #     # Get all selected files from right list
    #     items = []
    #     for i in range(self.FilestoConsolidateList.count()):
    #         item = self.FilestoConsolidateList.item(i)
    #         if item:
    #             items.append(item)
        
    #     if not items:
    #         QMessageBox.warning(self, "Consolidate", "No files selected to consolidate.")
    #         return
        
    #     # Separate by file type
    #     means_files = []
    #     breaths_files = []
        
    #     for item in items:
    #         meta = item.data(Qt.ItemDataRole.UserRole) or {}
    #         if meta.get("means"):
    #             means_files.append((meta["root"], Path(meta["means"])))
    #         if meta.get("breaths"):
    #             breaths_files.append((meta["root"], Path(meta["breaths"])))
        
    #     if not means_files:
    #         QMessageBox.warning(self, "Consolidate", "No *_means_by_time.csv files selected.")
    #         return
        
    #     # Process means files
    #     try:
    #         consolidated_data = self._consolidate_means_files(means_files)
            
    #         # Choose save location
    #         save_path, _ = QFileDialog.getSaveFileName(
    #             self, "Save consolidated data as...",
    #             str(means_files[0][1].parent / "consolidated_data.xlsx"),
    #             "Excel Files (*.xlsx)"
    #         )
            
    #         if save_path:
    #             self._save_consolidated_to_excel(consolidated_data, Path(save_path))
    #             QMessageBox.information(
    #                 self, "Success", 
    #                 f"Consolidated {len(means_files)} files.\nSaved to: {save_path}"
    #             )
        
    #     except Exception as e:
    #         QMessageBox.critical(self, "Consolidation Error", f"Failed to consolidate:\n{e}")
    #         import traceback
    #         traceback.print_exc()


    # def on_consolidate_save_data_clicked(self):
    #     """Consolidate data from selected files into a single Excel file."""
    #     from PyQt6.QtCore import Qt
    #     import pandas as pd
    #     from pathlib import Path
        
    #     # Get all selected files from right list
    #     items = []
    #     for i in range(self.FilestoConsolidateList.count()):
    #         item = self.FilestoConsolidateList.item(i)
    #         if item:
    #             items.append(item)
        
    #     if not items:
    #         QMessageBox.warning(self, "Consolidate", "No files selected to consolidate.")
    #         return
        
    #     # Separate by file type
    #     means_files = []
    #     breaths_files = []
        
    #     for item in items:
    #         meta = item.data(Qt.ItemDataRole.UserRole) or {}
    #         if meta.get("means"):
    #             means_files.append((meta["root"], Path(meta["means"])))
    #         if meta.get("breaths"):
    #             breaths_files.append((meta["root"], Path(meta["breaths"])))
        
    #     if not means_files and not breaths_files:
    #         QMessageBox.warning(self, "Consolidate", "No CSV files selected.")
    #         return
        
    #     # Process files
    #     try:
    #         consolidated_data = {}
            
    #         if means_files:
    #             consolidated_data.update(self._consolidate_means_files(means_files))
            
    #         if breaths_files:
    #             histogram_data = self._consolidate_breaths_histograms(breaths_files)
    #             consolidated_data.update(histogram_data)
            
    #         # Choose save location
    #         default_name = "consolidated_data.xlsx"
    #         if means_files:
    #             default_name = str(means_files[0][1].parent / "consolidated_data.xlsx")
    #         elif breaths_files:
    #             default_name = str(breaths_files[0][1].parent / "consolidated_data.xlsx")
                
    #         save_path, _ = QFileDialog.getSaveFileName(
    #             self, "Save consolidated data as...",
    #             default_name,
    #             "Excel Files (*.xlsx)"
    #         )
            
    #         if save_path:
    #             self._save_consolidated_to_excel(consolidated_data, Path(save_path))
    #             n_files = len(means_files) + len(breaths_files)
    #             QMessageBox.information(
    #                 self, "Success", 
    #                 f"Consolidated {n_files} files.\nSaved to: {save_path}"
    #             )
        
    #     except Exception as e:
    #         QMessageBox.critical(self, "Consolidation Error", f"Failed to consolidate:\n{e}")
    #         import traceback
    #         traceback.print_exc()

    def _propose_consolidated_filename(self, files: list) -> tuple[str, list[str]]:
        """Generate a descriptive filename based on the files being consolidated.

        Returns:
            tuple: (proposed_filename, list of warnings)
        """
        warnings = []

        if not files:
            return "consolidated_data.xlsx", warnings

        # Parse filenames to extract common metadata
        # Expected format: Strain_Virus_Location_Sex_Animal_Stim_Power_ABF_Channel_*.csv
        # Also support older format without Location
        parsed_files = []
        for _, filepath in files:
            stem = filepath.stem  # filename without extension
            parts = stem.split('_')

            # Try to intelligently parse - check if we have location field
            # Location field is typically anatomical abbreviations like preBotC, RTN, etc.
            # Sex is typically M, F, or Unknown (short)
            # We'll use heuristics: if part[2] is 1-2 chars, it's probably sex (old format)

            if len(parts) >= 7 and len(parts[3]) <= 2:
                # New format with location: Strain_Virus_Location_Sex_Animal_Stim_Power_...
                parsed_files.append({
                    'strain': parts[0] if len(parts) > 0 else '',
                    'virus': parts[1] if len(parts) > 1 else '',
                    'location': parts[2] if len(parts) > 2 else '',
                    'sex': parts[3] if len(parts) > 3 else '',
                    'animal': parts[4] if len(parts) > 4 else '',
                    'stim': parts[5] if len(parts) > 5 else '',
                    'power': parts[6] if len(parts) > 6 else '',
                })
            elif len(parts) >= 6 and len(parts[2]) <= 2:
                # Old format without location: Strain_Virus_Sex_Animal_Stim_Power_...
                parsed_files.append({
                    'strain': parts[0] if len(parts) > 0 else '',
                    'virus': parts[1] if len(parts) > 1 else '',
                    'location': '',
                    'sex': parts[2] if len(parts) > 2 else '',
                    'animal': parts[3] if len(parts) > 3 else '',
                    'stim': parts[4] if len(parts) > 4 else '',
                    'power': parts[5] if len(parts) > 5 else '',
                })

        if not parsed_files:
            return "consolidated_data.xlsx", warnings

        # Find common values across all files and check for variations
        common = {}
        for key in ['strain', 'virus', 'location', 'sex', 'stim', 'power']:
            values = set(f[key] for f in parsed_files if f.get(key))
            if len(values) == 1:
                common[key] = values.pop()
            elif len(values) > 1:
                # Warn about different stimulation parameters
                if key == 'stim':
                    warnings.append(f"Multiple stimulation types detected: {', '.join(sorted(values))}")
                elif key == 'power':
                    warnings.append(f"Multiple laser powers detected: {', '.join(sorted(values))}")

        # Build descriptive filename from common fields
        parts = []
        if common.get('strain'):
            parts.append(common['strain'])
        if common.get('virus'):
            parts.append(common['virus'])
        if common.get('location'):
            parts.append(common['location'])
        if common.get('sex'):
            parts.append(common['sex'])

        # If multiple animals, indicate that
        animals = set(f['animal'] for f in parsed_files if f.get('animal'))
        if len(animals) == 1:
            parts.append(animals.pop())
        elif len(animals) > 1:
            parts.append(f"N{len(animals)}")  # Capital N

        if common.get('stim'):
            parts.append(common['stim'])
        if common.get('power'):
            parts.append(common['power'])

        # Add "consolidated" suffix
        parts.append("consolidated")

        if parts:
            return "_".join(parts) + ".xlsx", warnings
        else:
            return "consolidated_data.xlsx", warnings

    def on_consolidate_save_data_clicked(self):
        """Consolidate data from selected files into a single Excel file."""
        from PyQt6.QtCore import Qt
        from PyQt6.QtWidgets import QProgressDialog
        import pandas as pd
        from pathlib import Path

        # Get all selected files from right list
        items = []
        for i in range(self.FilestoConsolidateList.count()):
            item = self.FilestoConsolidateList.item(i)
            if item:
                items.append(item)

        if not items:
            QMessageBox.warning(self, "Consolidate", "No files selected to consolidate.")
            return

        # Separate by file type
        means_files = []
        breaths_files = []
        events_files = []

        for item in items:
            meta = item.data(Qt.ItemDataRole.UserRole) or {}
            if meta.get("means"):
                means_files.append((meta["root"], Path(meta["means"])))
            if meta.get("breaths"):
                breaths_files.append((meta["root"], Path(meta["breaths"])))
            if meta.get("events"):
                events_files.append((meta["root"], Path(meta["events"])))

        if not means_files and not breaths_files and not events_files:
            QMessageBox.warning(self, "Consolidate", "No CSV files selected.")
            return

        # Choose save location first with intelligent default name
        files_for_naming = means_files or breaths_files
        proposed_filename, warnings = self._propose_consolidated_filename(files_for_naming)

        # Show warnings if any
        if warnings:
            warning_msg = "Warning about files being consolidated:\n\n" + "\n".join(f"• {w}" for w in warnings)
            warning_msg += "\n\nDo you want to continue?"
            reply = QMessageBox.question(
                self, "Consolidation Warning",
                warning_msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        if means_files:
            default_name = str(means_files[0][1].parent / proposed_filename)
        elif breaths_files:
            default_name = str(breaths_files[0][1].parent / proposed_filename)
        else:
            default_name = proposed_filename

        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save consolidated data as...",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if not save_path:
            return

        # Create progress dialog
        n_total_files = len(means_files) + len(breaths_files)
        progress = QProgressDialog("Consolidating data...", "Cancel", 0, 100, self)
        progress.setWindowTitle("PlethAnalysis")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        # Process files
        try:
            consolidated_data = {}

            if means_files:
                progress.setLabelText(f"Processing time series data ({len(means_files)} files)...")
                progress.setValue(10)
                QApplication.processEvents()
                if progress.wasCanceled():
                    return

                consolidated_data.update(self._consolidate_means_files(means_files))
                progress.setValue(40)
                QApplication.processEvents()

            if breaths_files:
                if progress.wasCanceled():
                    return

                progress.setLabelText(f"Processing breath histograms ({len(breaths_files)} files)...")
                progress.setValue(50)
                QApplication.processEvents()

                histogram_data = self._consolidate_breaths_histograms(breaths_files)
                consolidated_data.update(histogram_data)

                progress.setValue(70)
                QApplication.processEvents()
                if progress.wasCanceled():
                    return

                # Extract sigh data
                progress.setLabelText("Extracting sigh data...")
                sighs_df = self._consolidate_breaths_sighs(breaths_files)
                consolidated_data['sighs'] = {
                    'time_series': sighs_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'windows': []
                }
                progress.setValue(80)
                QApplication.processEvents()

            if events_files:
                if progress.wasCanceled():
                    return

                progress.setLabelText(f"Processing events data ({len(events_files)} files)...")
                progress.setValue(82)
                QApplication.processEvents()

                events_df = self._consolidate_events(events_files)
                print(f"Events DataFrame shape: {events_df.shape}")
                print(f"Events columns: {events_df.columns.tolist()}")
                if len(events_df) > 0:
                    print(f"First few event types: {events_df['event_type'].unique()[:5]}")
                consolidated_data['events'] = {
                    'time_series': events_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

                # Process stimulus events separately
                progress.setLabelText(f"Processing stimulus data ({len(events_files)} files)...")
                progress.setValue(83)
                QApplication.processEvents()

                stimulus_df, stim_warnings = self._consolidate_stimulus(events_files)
                print(f"Stimulus DataFrame shape: {stimulus_df.shape}")
                if len(stimulus_df) > 0:
                    print(f"Stimulus columns: {stimulus_df.columns.tolist()}")
                    print(f"Stimulus rows: {len(stimulus_df)}")
                consolidated_data['stimulus'] = {
                    'time_series': stimulus_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

                # Add stimulus warnings to consolidated warnings
                if stim_warnings:
                    if '_warnings' not in consolidated_data:
                        consolidated_data['_warnings'] = []
                    consolidated_data['_warnings'].extend(stim_warnings)

            if progress.wasCanceled():
                return

            progress.setLabelText("Saving Excel file and generating charts...")
            progress.setValue(85)
            QApplication.processEvents()

            self._save_consolidated_to_excel(consolidated_data, Path(save_path))
            progress.setValue(100)

            n_files = len(means_files) + len(breaths_files)

            # Check for warnings from consolidation
            if '_warnings' in consolidated_data:
                warnings_text = consolidated_data['_warnings']
                msg_box = QMessageBox(self)
                msg_box.setIcon(QMessageBox.Icon.Warning)
                msg_box.setWindowTitle("Consolidation Completed with Warnings")
                msg_box.setText(f"Consolidated {n_files} files successfully.\nSaved to: {save_path}\n\nHowever, some files required special handling:")
                msg_box.setDetailedText(warnings_text)
                msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
                msg_box.exec()
            else:
                QMessageBox.information(
                    self, "Success",
                    f"Consolidated {n_files} files.\nSaved to: {save_path}"
                )

        except Exception as e:
            progress.close()
            QMessageBox.critical(self, "Consolidation Error", f"Failed to consolidate:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            progress.close()

    # def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Process breaths CSV files and create histograms for each metric.
    #     Returns dict: {metric_name_region: DataFrame with histogram bins and counts per file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract from breaths data
    #     metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
    #     # Regions in the breaths CSV
    #     regions = {
    #         'all': '',  # no suffix for "all" block
    #         'baseline': '_baseline',
    #         'stim': '_stim', 
    #         'post': '_post'
    #     }
        
    #     consolidated = {}
        
    #     for metric in metrics:
    #         for region_name, suffix in regions.items():
    #             # Column name in the CSV
    #             col_name = f"{metric}{suffix}" if suffix else metric
                
    #             # Collect all data for this metric/region across files
    #             all_data = []
    #             file_roots = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data.append(values)
    #                             file_roots.append(root)
    #                         else:
    #                             print(f"No data for {col_name} in {root}")
    #                     else:
    #                         print(f"Column {col_name} not found in {root}")
                            
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data:
    #                 continue
                
    #             # Determine common bin edges across all files
    #             all_combined = np.concatenate(all_data)
                
    #             # Use automatic binning or fixed number of bins
    #             n_bins = 30  # You can adjust this
    #             bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
                
    #             # Calculate histogram for each file
    #             hist_data = {'bin_center': (bin_edges[:-1] + bin_edges[1:]) / 2}
                
    #             for root, values in zip(file_roots, all_data):
    #                 counts, _ = np.histogram(values, bins=bin_edges)
    #                 hist_data[root] = counts
                
    #             # Create DataFrame
    #             hist_df = pd.DataFrame(hist_data)
                
    #             # Store with descriptive key
    #             key = f"{metric}_histogram_{region_name}"
    #             consolidated[key] = {'time_series': hist_df, 'raw_summary': {}, 'norm_summary': {}, 'windows': []}
        
    #     return consolidated

    # def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Process breaths CSV files and create density histograms for each metric.
    #     Returns dict: {metric_name: DataFrame with histogram bins and densities per file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract from breaths data
    #     metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
    #     # Regions in the breaths CSV
    #     regions = {
    #         'all': '',
    #         'baseline': '_baseline',
    #         'stim': '_stim', 
    #         'post': '_post'
    #     }
        
    #     consolidated = {}
        
    #     # Helper to calculate mean and SEM
    #     def calc_mean_sem(data_array):
    #         mean = np.nanmean(data_array, axis=1)
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #         return mean, sem
        
    #     for metric in metrics:
    #         # For IF, combine all regions into one sheet
    #         if metric == 'if':
    #             combined_df = None
                
    #             for region_name, suffix in regions.items():
    #                 col_name = f"{metric}{suffix}" if suffix else metric
                    
    #                 # Collect data for this region
    #                 all_data = []
    #                 file_roots = []
                    
    #                 for root, path in files:
    #                     try:
    #                         df = pd.read_csv(path)
                            
    #                         if col_name in df.columns:
    #                             values = df[col_name].dropna().values
    #                             if len(values) > 0:
    #                                 all_data.append(values)
    #                                 file_roots.append(root)
    #                     except Exception as e:
    #                         print(f"Error reading {path}: {e}")
                    
    #                 if not all_data:
    #                     continue
                    
    #                 # Common bin edges for this region
    #                 all_combined = np.concatenate(all_data)
    #                 n_bins = 30
    #                 bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
    #                 bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                    
    #                 # Calculate density histogram for each file
    #                 region_data = {f'bin_center_{region_name}': bin_centers}
                    
    #                 density_arrays = []
    #                 for root, values in zip(file_roots, all_data):
    #                     counts, _ = np.histogram(values, bins=bin_edges)
    #                     # Convert to density (normalize so integral = 1)
    #                     bin_widths = np.diff(bin_edges)
    #                     density = counts / (counts.sum() * bin_widths)
    #                     region_data[f'{root}_{region_name}'] = density
    #                     density_arrays.append(density)
                    
    #                 # Calculate mean and SEM for this region
    #                 if density_arrays:
    #                     density_matrix = np.column_stack(density_arrays)
    #                     mean, sem = calc_mean_sem(density_matrix)
    #                     region_data[f'mean_{region_name}'] = mean
    #                     region_data[f'sem_{region_name}'] = sem
                    
    #                 # Create DataFrame for this region
    #                 region_df = pd.DataFrame(region_data)
                    
    #                 # Combine with other regions
    #                 if combined_df is None:
    #                     combined_df = region_df
    #                 else:
    #                     # Add blank column separator
    #                     combined_df[''] = ''
    #                     # Merge horizontally
    #                     combined_df = pd.concat([combined_df, region_df], axis=1)
                
    #             if combined_df is not None:
    #                 consolidated[f'{metric}_histogram'] = {
    #                     'time_series': combined_df, 
    #                     'raw_summary': {}, 
    #                     'norm_summary': {}, 
    #                     'windows': []
    #                 }
            
    #         else:
    #             # For other metrics, keep separate sheets per region
    #             for region_name, suffix in regions.items():
    #                 col_name = f"{metric}{suffix}" if suffix else metric
                    
    #                 all_data = []
    #                 file_roots = []
                    
    #                 for root, path in files:
    #                     try:
    #                         df = pd.read_csv(path)
                            
    #                         if col_name in df.columns:
    #                             values = df[col_name].dropna().values
    #                             if len(values) > 0:
    #                                 all_data.append(values)
    #                                 file_roots.append(root)
    #                     except Exception as e:
    #                         print(f"Error reading {path}: {e}")
                    
    #                 if not all_data:
    #                     continue
                    
    #                 all_combined = np.concatenate(all_data)
    #                 n_bins = 30
    #                 bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
                    
    #                 hist_data = {'bin_center': (bin_edges[:-1] + bin_edges[1:]) / 2}
                    
    #                 density_arrays = []
    #                 for root, values in zip(file_roots, all_data):
    #                     counts, _ = np.histogram(values, bins=bin_edges)
    #                     bin_widths = np.diff(bin_edges)
    #                     density = counts / (counts.sum() * bin_widths)
    #                     hist_data[root] = density
    #                     density_arrays.append(density)
                    
    #                 # Calculate mean and SEM
    #                 if density_arrays:
    #                     density_matrix = np.column_stack(density_arrays)
    #                     mean, sem = calc_mean_sem(density_matrix)
    #                     hist_data['mean'] = mean
    #                     hist_data['sem'] = sem
                    
    #                 hist_df = pd.DataFrame(hist_data)
                    
    #                 key = f"{metric}_histogram_{region_name}"
    #                 consolidated[key] = {
    #                     'time_series': hist_df, 
    #                     'raw_summary': {}, 
    #                     'norm_summary': {}, 
    #                     'windows': []
    #                 }
        
    #     return consolidated


    # def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Process breaths CSV files and create density histograms for each metric.
    #     Returns dict: {metric_name: DataFrame with histogram bins and densities per file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract from breaths data
    #     metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
    #     # Regions in the breaths CSV
    #     regions = {
    #         'all': '',
    #         'baseline': '_baseline',
    #         'stim': '_stim', 
    #         'post': '_post'
    #     }
        
    #     consolidated = {}
        
    #     # Helper to calculate mean and SEM
    #     def calc_mean_sem(data_array):
    #         mean = np.nanmean(data_array, axis=1)
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #         return mean, sem
        
    #     # Process each metric (combine all regions into one sheet)
    #     for metric in metrics:
    #         combined_df = None
            
    #         for region_name, suffix in regions.items():
    #             col_name = f"{metric}{suffix}" if suffix else metric
                
    #             # Collect data for this region
    #             all_data = []
    #             file_roots = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data.append(values)
    #                             file_roots.append(root)
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data:
    #                 continue
                
    #             # Common bin edges for this region
    #             all_combined = np.concatenate(all_data)
    #             n_bins = 30
    #             bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
    #             bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                
    #             # Calculate density histogram for each file
    #             region_data = {f'bin_center_{region_name}': bin_centers}
                
    #             density_arrays = []
    #             for root, values in zip(file_roots, all_data):
    #                 counts, _ = np.histogram(values, bins=bin_edges)
    #                 # Convert to density (normalize so integral = 1)
    #                 bin_widths = np.diff(bin_edges)
    #                 density = counts / (counts.sum() * bin_widths)
    #                 region_data[f'{root}_{region_name}'] = density
    #                 density_arrays.append(density)
                
    #             # Calculate mean and SEM for this region
    #             if density_arrays:
    #                 density_matrix = np.column_stack(density_arrays)
    #                 mean, sem = calc_mean_sem(density_matrix)
    #                 region_data[f'mean_{region_name}'] = mean
    #                 region_data[f'sem_{region_name}'] = sem
                
    #             # Create DataFrame for this region
    #             region_df = pd.DataFrame(region_data)
                
    #             # Combine with other regions
    #             if combined_df is None:
    #                 combined_df = region_df
    #             else:
    #                 # Add blank column separator between regions
    #                 combined_df[''] = ''
    #                 # Merge horizontally
    #                 combined_df = pd.concat([combined_df, region_df], axis=1)
            
    #         if combined_df is not None:
    #             consolidated[f'{metric}_histogram'] = {
    #                 'time_series': combined_df, 
    #                 'raw_summary': {}, 
    #                 'norm_summary': {}, 
    #                 'windows': []
    #             }
        
    #     return consolidated

    # def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Process breaths CSV files and create density histograms for each metric.
    #     Returns dict: {metric_name: DataFrame with histogram bins and densities per file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract from breaths data
    #     metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
    #     # Regions in the breaths CSV
    #     regions = {
    #         'all': '',
    #         'baseline': '_baseline',
    #         'stim': '_stim', 
    #         'post': '_post'
    #     }
        
    #     consolidated = {}
        
    #     # Helper to calculate mean and SEM
    #     def calc_mean_sem(data_array):
    #         mean = np.nanmean(data_array, axis=1)
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #         return mean, sem
        
    #     # Process each metric (combine all regions into one sheet)
    #     for metric in metrics:
    #         combined_df = None
            
    #         # Process RAW data
    #         for region_name, suffix in regions.items():
    #             col_name = f"{metric}{suffix}" if suffix else metric
                
    #             # Collect data for this region
    #             all_data = []
    #             file_roots = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data.append(values)
    #                             file_roots.append(root)
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data:
    #                 continue
                
    #             # Common bin edges for this region
    #             all_combined = np.concatenate(all_data)
    #             n_bins = 30
    #             bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
    #             bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                
    #             # Calculate density histogram for each file
    #             region_data = {f'bin_center_{region_name}': bin_centers}
                
    #             density_arrays = []
    #             for root, values in zip(file_roots, all_data):
    #                 counts, _ = np.histogram(values, bins=bin_edges)
    #                 # Convert to density (normalize so integral = 1)
    #                 bin_widths = np.diff(bin_edges)
    #                 density = counts / (counts.sum() * bin_widths)
    #                 region_data[f'{root}_{region_name}'] = density
    #                 density_arrays.append(density)
                
    #             # Calculate mean and SEM for this region
    #             if density_arrays:
    #                 density_matrix = np.column_stack(density_arrays)
    #                 mean, sem = calc_mean_sem(density_matrix)
    #                 region_data[f'mean_{region_name}'] = mean
    #                 region_data[f'sem_{region_name}'] = sem
                
    #             # Create DataFrame for this region
    #             region_df = pd.DataFrame(region_data)
                
    #             # Combine with other regions
    #             if combined_df is None:
    #                 combined_df = region_df
    #             else:
    #                 # Add blank column separator between regions
    #                 combined_df[''] = ''
    #                 # Merge horizontally
    #                 combined_df = pd.concat([combined_df, region_df], axis=1)
            
    #         # Add extra blank column before normalized data
    #         if combined_df is not None:
    #             combined_df['  '] = ''
            
    #         # Process NORMALIZED data
    #         for region_name, suffix in regions.items():
    #             col_name = f"{metric}{suffix}_norm" if suffix else f"{metric}_norm"
                
    #             # Collect normalized data for this region
    #             all_data_norm = []
    #             file_roots_norm = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data_norm.append(values)
    #                             file_roots_norm.append(root)
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data_norm:
    #                 continue
                
    #             # Common bin edges for normalized region
    #             all_combined_norm = np.concatenate(all_data_norm)
    #             n_bins = 30
    #             bin_edges_norm = np.histogram_bin_edges(all_combined_norm, bins=n_bins)
    #             bin_centers_norm = (bin_edges_norm[:-1] + bin_edges_norm[1:]) / 2
                
    #             # Calculate density histogram for each file
    #             region_data_norm = {f'bin_center_{region_name}_norm': bin_centers_norm}
                
    #             density_arrays_norm = []
    #             for root, values in zip(file_roots_norm, all_data_norm):
    #                 counts, _ = np.histogram(values, bins=bin_edges_norm)
    #                 bin_widths = np.diff(bin_edges_norm)
    #                 density = counts / (counts.sum() * bin_widths)
    #                 region_data_norm[f'{root}_{region_name}_norm'] = density
    #                 density_arrays_norm.append(density)
                
    #             # Calculate mean and SEM for normalized region
    #             if density_arrays_norm:
    #                 density_matrix_norm = np.column_stack(density_arrays_norm)
    #                 mean_norm, sem_norm = calc_mean_sem(density_matrix_norm)
    #                 region_data_norm[f'mean_{region_name}_norm'] = mean_norm
    #                 region_data_norm[f'sem_{region_name}_norm'] = sem_norm
                
    #             # Create DataFrame for this normalized region
    #             region_df_norm = pd.DataFrame(region_data_norm)
                
    #             # Combine with existing data
    #             if combined_df is not None:
    #                 # Add blank column separator between regions
    #                 combined_df[''] = ''
    #                 # Merge horizontally
    #                 combined_df = pd.concat([combined_df, region_df_norm], axis=1)
            
    #         if combined_df is not None:
    #             consolidated[f'{metric}_histogram'] = {
    #                 'time_series': combined_df, 
    #                 'raw_summary': {}, 
    #                 'norm_summary': {}, 
    #                 'windows': []
    #             }
        
    #     return consolidated

    # def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Process breaths CSV files and create density histograms for each metric.
    #     Returns dict: {metric_name: DataFrame with histogram bins and densities per file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract from breaths data
    #     metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
    #     # Regions in the breaths CSV
    #     regions = {
    #         'all': '',
    #         'baseline': '_baseline',
    #         'stim': '_stim', 
    #         'post': '_post'
    #     }
        
    #     consolidated = {}
        
    #     # Helper to calculate mean and SEM
    #     def calc_mean_sem(data_array):
    #         mean = np.nanmean(data_array, axis=1)
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #         return mean, sem
        
    #     # Process each metric (combine all regions into one sheet)
    #     for metric in metrics:
    #         combined_df = None
            
    #         # Process RAW data
    #         for region_name, suffix in regions.items():
    #             col_name = f"{metric}{suffix}" if suffix else metric
                
    #             # Collect data for this region
    #             all_data = []
    #             file_roots = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data.append(values)
    #                             file_roots.append(root)
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data:
    #                 continue
                
    #             # Common bin edges for this region
    #             all_combined = np.concatenate(all_data)
    #             n_bins = 30
    #             bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
    #             bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                
    #             # Calculate density histogram for each file
    #             region_data = {f'bin_center_{region_name}': bin_centers}
                
    #             density_arrays = []
    #             for root, values in zip(file_roots, all_data):
    #                 counts, _ = np.histogram(values, bins=bin_edges)
    #                 # Convert to density (normalize so integral = 1)
    #                 bin_widths = np.diff(bin_edges)
    #                 density = counts / (counts.sum() * bin_widths)
    #                 region_data[f'{root}_{region_name}'] = density
    #                 density_arrays.append(density)
                
    #             # Calculate mean and SEM for this region
    #             if density_arrays:
    #                 density_matrix = np.column_stack(density_arrays)
    #                 mean, sem = calc_mean_sem(density_matrix)
    #                 region_data[f'mean_{region_name}'] = mean
    #                 region_data[f'sem_{region_name}'] = sem
                
    #             # Create DataFrame for this region
    #             region_df = pd.DataFrame(region_data)
                
    #             # Combine with other regions
    #             if combined_df is None:
    #                 combined_df = region_df
    #             else:
    #                 # Add blank column separator between regions
    #                 combined_df[''] = ''
    #                 # Merge horizontally
    #                 combined_df = pd.concat([combined_df, region_df], axis=1)
            
    #         # Add extra blank column before normalized data
    #         if combined_df is not None:
    #             combined_df['  '] = ''
            
    #         # Process NORMALIZED data
    #         for region_name, suffix in regions.items():
    #             col_name = f"{metric}{suffix}_norm" if suffix else f"{metric}_norm"
                
    #             # Collect normalized data for this region
    #             all_data_norm = []
    #             file_roots_norm = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data_norm.append(values)
    #                             file_roots_norm.append(root)
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data_norm:
    #                 continue
                
    #             # Common bin edges for normalized region
    #             all_combined_norm = np.concatenate(all_data_norm)
    #             n_bins = 30
    #             bin_edges_norm = np.histogram_bin_edges(all_combined_norm, bins=n_bins)
    #             bin_centers_norm = (bin_edges_norm[:-1] + bin_edges_norm[1:]) / 2
                
    #             # Calculate density histogram for each file
    #             region_data_norm = {f'bin_center_{region_name}_norm': bin_centers_norm}
                
    #             density_arrays_norm = []
    #             for root, values in zip(file_roots_norm, all_data_norm):
    #                 counts, _ = np.histogram(values, bins=bin_edges_norm)
    #                 bin_widths = np.diff(bin_edges_norm)
    #                 density = counts / (counts.sum() * bin_widths)
    #                 region_data_norm[f'{root}_{region_name}_norm'] = density
    #                 density_arrays_norm.append(density)
                
    #             # Calculate mean and SEM for normalized region
    #             if density_arrays_norm:
    #                 density_matrix_norm = np.column_stack(density_arrays_norm)
    #                 mean_norm, sem_norm = calc_mean_sem(density_matrix_norm)
    #                 region_data_norm[f'mean_{region_name}_norm'] = mean_norm
    #                 region_data_norm[f'sem_{region_name}_norm'] = sem_norm
                
    #             # Create DataFrame for this normalized region
    #             region_df_norm = pd.DataFrame(region_data_norm)
                
    #             # Combine with existing data
    #             if combined_df is not None:
    #                 # Add blank column separator between regions
    #                 combined_df[''] = ''
    #                 # Merge horizontally
    #                 combined_df = pd.concat([combined_df, region_df_norm], axis=1)
            
    #         if combined_df is not None:
    #             consolidated[f'{metric}_histogram'] = {
    #                 'time_series': combined_df, 
    #                 'raw_summary': {}, 
    #                 'norm_summary': {}, 
    #                 'windows': []
    #             }
        
    #     return consolidated

    # def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Process breaths CSV files and create density histograms for each metric.
    #     Returns dict: {metric_name: DataFrame with histogram bins and densities per file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract from breaths data
    #     metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
    #     # Regions in the breaths CSV
    #     regions = {
    #         'all': '',
    #         'baseline': '_baseline',
    #         'stim': '_stim', 
    #         'post': '_post'
    #     }
        
    #     consolidated = {}
        
    #     # Helper to calculate mean and SEM
    #     def calc_mean_sem(data_array):
    #         mean = np.nanmean(data_array, axis=1)
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #         return mean, sem
        
    #     # Process each metric (combine all regions into one sheet)
    #     for metric in metrics:
    #         combined_df = None
            
    #         # Process RAW data
    #         for region_name, suffix in regions.items():
    #             col_name = f"{metric}{suffix}" if suffix else metric
                
    #             # Collect data for this region
    #             all_data = []
    #             file_roots = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data.append(values)
    #                             file_roots.append(root)
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data:
    #                 continue
                
    #             # Common bin edges for this region
    #             all_combined = np.concatenate(all_data)
    #             n_bins = 30
    #             bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
    #             bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                
    #             # Calculate density histogram for each file
    #             region_data = {f'bin_center_{region_name}': bin_centers}
                
    #             density_arrays = []
    #             for root, values in zip(file_roots, all_data):
    #                 counts, _ = np.histogram(values, bins=bin_edges)
    #                 # Convert to density (normalize so integral = 1)
    #                 bin_widths = np.diff(bin_edges)
    #                 density = counts / (counts.sum() * bin_widths)
    #                 region_data[f'{root}_{region_name}'] = density
    #                 density_arrays.append(density)
                
    #             # Calculate mean and SEM for this region
    #             if density_arrays:
    #                 density_matrix = np.column_stack(density_arrays)
    #                 mean, sem = calc_mean_sem(density_matrix)
    #                 region_data[f'mean_{region_name}'] = mean
    #                 region_data[f'sem_{region_name}'] = sem
                
    #             # Create DataFrame for this region
    #             region_df = pd.DataFrame(region_data)
                
    #             # Combine with other regions
    #             if combined_df is None:
    #                 combined_df = region_df
    #             else:
    #                 # Add blank column separator between regions
    #                 combined_df[''] = ''
    #                 # Merge horizontally
    #                 combined_df = pd.concat([combined_df, region_df], axis=1)
            
    #         # Add extra blank column before normalized data
    #         if combined_df is not None:
    #             combined_df['  '] = ''
            
    #         # Process NORMALIZED data
    #         for region_name, suffix in regions.items():
    #             col_name = f"{metric}{suffix}_norm" if suffix else f"{metric}_norm"
                
    #             # Collect normalized data for this region
    #             all_data_norm = []
    #             file_roots_norm = []
                
    #             for root, path in files:
    #                 try:
    #                     df = pd.read_csv(path)
                        
    #                     if col_name in df.columns:
    #                         values = df[col_name].dropna().values
    #                         if len(values) > 0:
    #                             all_data_norm.append(values)
    #                             file_roots_norm.append(root)
    #                 except Exception as e:
    #                     print(f"Error reading {path}: {e}")
                
    #             if not all_data_norm:
    #                 continue
                
    #             # Common bin edges for normalized region
    #             all_combined_norm = np.concatenate(all_data_norm)
    #             n_bins = 30
    #             bin_edges_norm = np.histogram_bin_edges(all_combined_norm, bins=n_bins)
    #             bin_centers_norm = (bin_edges_norm[:-1] + bin_edges_norm[1:]) / 2
                
    #             # Calculate density histogram for each file
    #             region_data_norm = {f'bin_center_{region_name}_norm': bin_centers_norm}
                
    #             density_arrays_norm = []
    #             for root, values in zip(file_roots_norm, all_data_norm):
    #                 counts, _ = np.histogram(values, bins=bin_edges_norm)
    #                 bin_widths = np.diff(bin_edges_norm)
    #                 density = counts / (counts.sum() * bin_widths)
    #                 region_data_norm[f'{root}_{region_name}_norm'] = density
    #                 density_arrays_norm.append(density)
                
    #             # Calculate mean and SEM for normalized region
    #             if density_arrays_norm:
    #                 density_matrix_norm = np.column_stack(density_arrays_norm)
    #                 mean_norm, sem_norm = calc_mean_sem(density_matrix_norm)
    #                 region_data_norm[f'mean_{region_name}_norm'] = mean_norm
    #                 region_data_norm[f'sem_{region_name}_norm'] = sem_norm
                
    #             # Create DataFrame for this normalized region
    #             region_df_norm = pd.DataFrame(region_data_norm)
                
    #             # Combine with existing data
    #             if combined_df is not None:
    #                 # Add blank column separator between regions
    #                 combined_df[''] = ''
    #                 # Merge horizontally
    #                 combined_df = pd.concat([combined_df, region_df_norm], axis=1)
            
    #         if combined_df is not None:
    #             consolidated[f'{metric}_histogram'] = {
    #                 'time_series': combined_df, 
    #                 'raw_summary': {}, 
    #                 'norm_summary': {}, 
    #                 'windows': []
    #             }
        
    #     return consolidated

    def _consolidate_breaths_histograms(self, files: list[tuple[str, Path]]) -> dict:
        """
        Process breaths CSV files and create density histograms for each metric.
        Returns dict: {metric_name: DataFrame with histogram bins and densities per file}
        """
        import pandas as pd
        import numpy as np
        
        # Metrics to extract from breaths data
        metrics = ['if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 'ti', 'te', 'vent_proxy']
        
        # Regions in the breaths CSV
        regions = {
            'all': '',
            'baseline': '_baseline',
            'stim': '_stim', 
            'post': '_post'
        }
        
        consolidated = {}
        
        # Helper to calculate mean and SEM
        def calc_mean_sem(data_array):
            mean = np.nanmean(data_array, axis=1)
            n = np.sum(np.isfinite(data_array), axis=1)
            std = np.nanstd(data_array, axis=1, ddof=1)
            sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
            return mean, sem
        
        # Process each metric (combine all regions into one sheet)
        for metric in metrics:
            combined_df = None
            
            # Process RAW data
            for region_name, suffix in regions.items():
                col_name = f"{metric}{suffix}" if suffix else metric
                
                # Collect data for this region
                all_data = []
                file_roots = []
                
                for root, path in files:
                    try:
                        df = pd.read_csv(path, low_memory=False)

                        if col_name in df.columns:
                            values = df[col_name].dropna().values
                            if len(values) > 0:
                                all_data.append(values)
                                file_roots.append(root)
                    except Exception as e:
                        print(f"Error reading {path}: {e}")
                
                if not all_data:
                    continue
                
                # Common bin edges for this region
                all_combined = np.concatenate(all_data)
                n_bins = 30
                bin_edges = np.histogram_bin_edges(all_combined, bins=n_bins)
                bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                
                # Calculate density histogram for each file
                region_data = {f'bin_center_{region_name}': bin_centers}
                
                density_arrays = []
                for root, values in zip(file_roots, all_data):
                    counts, _ = np.histogram(values, bins=bin_edges)
                    # Convert to density (normalize so integral = 1)
                    bin_widths = np.diff(bin_edges)
                    density = counts / (counts.sum() * bin_widths)
                    region_data[f'{root}_{region_name}'] = density
                    density_arrays.append(density)
                
                # Calculate mean and SEM for this region
                if density_arrays:
                    density_matrix = np.column_stack(density_arrays)
                    mean, sem = calc_mean_sem(density_matrix)
                    region_data[f'mean_{region_name}'] = mean
                    region_data[f'sem_{region_name}'] = sem
                
                # Create DataFrame for this region
                region_df = pd.DataFrame(region_data)
                
                # Combine with other regions
                if combined_df is None:
                    combined_df = region_df
                else:
                    # Merge horizontally
                    combined_df = pd.concat([combined_df, region_df], axis=1)
                
                # Add blank column separator AFTER each region
                combined_df[''] = ''
            
            # Process NORMALIZED data
            for region_name, suffix in regions.items():
                col_name = f"{metric}{suffix}_norm" if suffix else f"{metric}_norm"
                
                # Collect normalized data for this region
                all_data_norm = []
                file_roots_norm = []
                
                for root, path in files:
                    try:
                        df = pd.read_csv(path, low_memory=False)

                        if col_name in df.columns:
                            values = df[col_name].dropna().values
                            if len(values) > 0:
                                all_data_norm.append(values)
                                file_roots_norm.append(root)
                    except Exception as e:
                        print(f"Error reading {path}: {e}")
                
                if not all_data_norm:
                    continue
                
                # Common bin edges for normalized region
                all_combined_norm = np.concatenate(all_data_norm)
                n_bins = 30
                bin_edges_norm = np.histogram_bin_edges(all_combined_norm, bins=n_bins)
                bin_centers_norm = (bin_edges_norm[:-1] + bin_edges_norm[1:]) / 2
                
                # Calculate density histogram for each file
                region_data_norm = {f'bin_center_{region_name}_norm': bin_centers_norm}
                
                density_arrays_norm = []
                for root, values in zip(file_roots_norm, all_data_norm):
                    counts, _ = np.histogram(values, bins=bin_edges_norm)
                    bin_widths = np.diff(bin_edges_norm)
                    density = counts / (counts.sum() * bin_widths)
                    region_data_norm[f'{root}_{region_name}_norm'] = density
                    density_arrays_norm.append(density)
                
                # Calculate mean and SEM for normalized region
                if density_arrays_norm:
                    density_matrix_norm = np.column_stack(density_arrays_norm)
                    mean_norm, sem_norm = calc_mean_sem(density_matrix_norm)
                    region_data_norm[f'mean_{region_name}_norm'] = mean_norm
                    region_data_norm[f'sem_{region_name}_norm'] = sem_norm
                
                # Create DataFrame for this normalized region
                region_df_norm = pd.DataFrame(region_data_norm)
                
                # Merge horizontally
                combined_df = pd.concat([combined_df, region_df_norm], axis=1)
                
                # Add blank column separator AFTER each normalized region
                combined_df[''] = ''

            # Process EUPNEA-NORMALIZED data
            for region_name, suffix in regions.items():
                col_name = f"{metric}{suffix}_norm_eupnea" if suffix else f"{metric}_norm_eupnea"

                # Collect eupnea-normalized data for this region
                all_data_eupnea = []
                file_roots_eupnea = []

                for root, path in files:
                    try:
                        df = pd.read_csv(path, low_memory=False)

                        if col_name in df.columns:
                            values = df[col_name].dropna().values
                            if len(values) > 0:
                                all_data_eupnea.append(values)
                                file_roots_eupnea.append(root)
                    except Exception as e:
                        print(f"Error reading {path}: {e}")

                if not all_data_eupnea:
                    continue

                # Common bin edges for eupnea-normalized region
                all_combined_eupnea = np.concatenate(all_data_eupnea)
                n_bins = 30
                bin_edges_eupnea = np.histogram_bin_edges(all_combined_eupnea, bins=n_bins)
                bin_centers_eupnea = (bin_edges_eupnea[:-1] + bin_edges_eupnea[1:]) / 2

                # Calculate density histogram for each file
                region_data_eupnea = {f'bin_center_{region_name}_eupnea': bin_centers_eupnea}

                density_arrays_eupnea = []
                for root, values in zip(file_roots_eupnea, all_data_eupnea):
                    counts, _ = np.histogram(values, bins=bin_edges_eupnea)
                    bin_widths = np.diff(bin_edges_eupnea)
                    density = counts / (counts.sum() * bin_widths)
                    region_data_eupnea[f'{root}_{region_name}_eupnea'] = density
                    density_arrays_eupnea.append(density)

                # Calculate mean and SEM for eupnea-normalized region
                if density_arrays_eupnea:
                    density_matrix_eupnea = np.column_stack(density_arrays_eupnea)
                    mean_eupnea, sem_eupnea = calc_mean_sem(density_matrix_eupnea)
                    region_data_eupnea[f'mean_{region_name}_eupnea'] = mean_eupnea
                    region_data_eupnea[f'sem_{region_name}_eupnea'] = sem_eupnea

                # Create DataFrame for this eupnea-normalized region
                region_df_eupnea = pd.DataFrame(region_data_eupnea)

                # Merge horizontally
                combined_df = pd.concat([combined_df, region_df_eupnea], axis=1)

                # Add blank column separator AFTER each eupnea-normalized region
                combined_df[''] = ''

            if combined_df is not None:
                consolidated[f'{metric}_histogram'] = {
                    'time_series': combined_df,
                    'raw_summary': {},
                    'norm_summary': {},
                    'eupnea_summary': {},
                    'windows': []
                }

        return consolidated


    def _consolidate_events(self, files: list[tuple[str, Path]]) -> pd.DataFrame:
        """
        Consolidate events CSV files from multiple experiments (excluding stimulus events).
        Adds experiment, experiment_number, and global_sweep_number columns.
        """
        import pandas as pd
        import numpy as np

        all_events = []

        for exp_num, (root, path) in enumerate(files, start=1):
            try:
                df = pd.read_csv(path, low_memory=False)

                if len(df) > 0:
                    # Filter out stimulus events
                    if 'event_type' in df.columns:
                        df = df[~df['event_type'].str.contains('stimulus', case=False, na=False)]

                    if len(df) > 0:  # Check if we still have data after filtering
                        # Add experiment identifier columns
                        df.insert(0, 'experiment', root)
                        df.insert(1, 'experiment_number', exp_num)

                        # Calculate global sweep number
                        # Each experiment has sweeps numbered 1, 2, 3, etc.
                        # We need to add an offset based on previous experiments
                        if 'sweep' in df.columns:
                            # Find the maximum sweep number in this experiment
                            max_sweep = df['sweep'].max()
                            # Calculate offset (total sweeps from previous experiments)
                            sweep_offset = sum([pd.read_csv(files[i][1], low_memory=False)['sweep'].max()
                                              for i in range(exp_num - 1)
                                              if 'sweep' in pd.read_csv(files[i][1], low_memory=False).columns])
                            df.insert(2, 'global_sweep_number', df['sweep'] + sweep_offset)

                        all_events.append(df)

            except Exception as e:
                print(f"Error reading events file {path}: {e}")

        if all_events:
            return pd.concat(all_events, ignore_index=True)
        else:
            return pd.DataFrame()

    def _consolidate_stimulus(self, files: list[tuple[str, Path]]) -> tuple[pd.DataFrame, list[str]]:
        """
        Extract stimulus events and validate consistency across all experiments/sweeps.
        Returns: (stimulus_df with one instance, list of warnings)
        """
        import pandas as pd
        import numpy as np

        all_stimulus = []
        warnings = []
        reference_stim = None

        for exp_num, (root, path) in enumerate(files, start=1):
            try:
                df = pd.read_csv(path, low_memory=False)

                if 'event_type' in df.columns:
                    # Extract only stimulus events
                    stim_events = df[df['event_type'].str.contains('stimulus', case=False, na=False)].copy()

                    if len(stim_events) > 0:
                        # Determine which column names are used (could be start_time/end_time or t_start/t_end)
                        start_col = 'start_time' if 'start_time' in stim_events.columns else 't_start'
                        end_col = 'end_time' if 'end_time' in stim_events.columns else 't_end'

                        # Get unique stimulus events (should be same across all sweeps)
                        unique_stim = stim_events.drop_duplicates(subset=['event_type', start_col, end_col])

                        if reference_stim is None:
                            # First experiment sets the reference
                            reference_stim = unique_stim[[start_col, end_col]].values
                            # Store with experiment identifier
                            result_df = unique_stim.copy()
                            result_df.insert(0, 'experiment', root)
                            all_stimulus.append(result_df)
                        else:
                            # Validate against reference
                            current_stim = unique_stim[[start_col, end_col]].values

                            # Check if stimulus events match
                            if not (len(reference_stim) == len(current_stim) and
                                    np.allclose(reference_stim.astype(float),
                                               current_stim.astype(float),
                                               rtol=1e-5)):
                                warnings.append(f"WARNING: Stimulus timing differs in experiment '{root}'")
                                print(f"  Stimulus mismatch in {root}")

            except Exception as e:
                print(f"Error reading stimulus from {path}: {e}")
                import traceback
                traceback.print_exc()

        if all_stimulus:
            # Return only the first instance (they should all be identical)
            return all_stimulus[0], warnings
        else:
            return pd.DataFrame(), warnings

    # def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Consolidate means_by_time CSV files.
    #     Returns dict: {metric_name: DataFrame with 't' and columns for each file}
    #     """
    #     import pandas as pd
    #     import numpy as np
        
    #     # Metrics to extract (mean columns only)
    #     metrics = [
    #         'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 
    #         'ti', 'te', 'vent_proxy'
    #     ]
        
    #     consolidated = {}
        
    #     # Read first file to get time column
    #     first_root, first_path = files[0]
    #     df_first = pd.read_csv(first_path)
    #     t_values = df_first['t'].values
        
    #     # Process raw means
    #     for metric in metrics:
    #         metric_mean_col = f"{metric}_mean"
            
    #         if metric_mean_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_values})
            
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             if not np.allclose(df['t'].values, t_values, rtol=1e-5, atol=1e-8):
    #                 print(f"Warning: Time mismatch in {root}")
                
    #             if metric_mean_col in df.columns:
    #                 result_df[root] = df[metric_mean_col].values
            
    #         consolidated[metric] = result_df
        
    #     # Process normalized means
    #     for metric in metrics:
    #         metric_norm_col = f"{metric}_norm_mean"
            
    #         if metric_norm_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_values})
            
    #         for root, path in files:
    #             df = pd.read_csv(path)
    #             if metric_norm_col in df.columns:
    #                 result_df[root] = df[metric_norm_col].values
            
    #         consolidated[f"{metric}_norm"] = result_df
        
    #     return consolidated

    # def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Consolidate means_by_time CSV files.
    #     Interpolates all data to a common time base.
    #     Returns dict: {metric_name: DataFrame with 't' and columns for each file}
    #     """
    #     import pandas as pd
    #     import numpy as np
    #     from scipy.interpolate import interp1d
        
    #     metrics = [
    #         'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 
    #         'ti', 'te', 'vent_proxy'
    #     ]
        
    #     # Determine common time base (use first file's time)
    #     first_root, first_path = files[0]
    #     df_first = pd.read_csv(first_path)
    #     t_common = df_first['t'].values
        
    #     consolidated = {}
        
    #     # Process raw means
    #     for metric in metrics:
    #         metric_mean_col = f"{metric}_mean"
            
    #         if metric_mean_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_common})
            
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             if metric_mean_col not in df.columns:
    #                 print(f"Warning: {metric_mean_col} not found in {root}")
    #                 continue
                
    #             t_file = df['t'].values
    #             y_file = df[metric_mean_col].values
                
    #             # Check if time arrays match
    #             if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                 # Direct assignment if times match
    #                 result_df[root] = y_file
    #             else:
    #                 # Interpolate to common time base
    #                 print(f"Interpolating {root} to common time base for {metric}")
                    
    #                 # Remove NaN values before interpolation
    #                 mask = np.isfinite(y_file)
    #                 if mask.sum() < 2:
    #                     print(f"Warning: Not enough valid data in {root} for {metric}")
    #                     result_df[root] = np.nan
    #                     continue
                    
    #                 try:
    #                     f_interp = interp1d(
    #                         t_file[mask], y_file[mask], 
    #                         kind='linear', 
    #                         bounds_error=False, 
    #                         fill_value=np.nan
    #                     )
    #                     result_df[root] = f_interp(t_common)
    #                 except Exception as e:
    #                     print(f"Error interpolating {root} for {metric}: {e}")
    #                     result_df[root] = np.nan
            
    #         consolidated[metric] = result_df
        
    #     # Process normalized means (same logic)
    #     for metric in metrics:
    #         metric_norm_col = f"{metric}_norm_mean"
            
    #         if metric_norm_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_common})
            
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             if metric_norm_col not in df.columns:
    #                 continue
                
    #             t_file = df['t'].values
    #             y_file = df[metric_norm_col].values
                
    #             if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                 result_df[root] = y_file
    #             else:
    #                 mask = np.isfinite(y_file)
    #                 if mask.sum() >= 2:
    #                     try:
    #                         f_interp = interp1d(
    #                             t_file[mask], y_file[mask], 
    #                             kind='linear', 
    #                             bounds_error=False, 
    #                             fill_value=np.nan
    #                         )
    #                         result_df[root] = f_interp(t_common)
    #                     except:
    #                         result_df[root] = np.nan
    #                 else:
    #                     result_df[root] = np.nan
            
    #         consolidated[f"{metric}_norm"] = result_df
        
    #     return consolidated

    # def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Consolidate means_by_time CSV files.
    #     Interpolates all data to a common time base.
    #     Returns dict: {metric_name: DataFrame with 't', file columns, 'mean', and 'sem'}
    #     """
    #     import pandas as pd
    #     import numpy as np
    #     from scipy.interpolate import interp1d
        
    #     metrics = [
    #         'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 
    #         'ti', 'te', 'vent_proxy'
    #     ]
        
    #     # Determine common time base (use first file's time)
    #     first_root, first_path = files[0]
    #     df_first = pd.read_csv(first_path)
    #     t_common = df_first['t'].values
        
    #     consolidated = {}
        
    #     # Helper function to calculate mean and SEM
    #     def add_mean_sem(df, t_col='t'):
    #         """Add mean and SEM columns to dataframe, excluding the time column."""
    #         data_cols = [col for col in df.columns if col != t_col]
    #         data_array = df[data_cols].values
            
    #         # Mean across files (axis=1 means across columns)
    #         df['mean'] = np.nanmean(data_array, axis=1)
            
    #         # SEM across files
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         df['sem'] = np.where(n >= 2, std / np.sqrt(n), np.nan)
            
    #         return df
        
    #     # Process raw means
    #     for metric in metrics:
    #         metric_mean_col = f"{metric}_mean"
            
    #         if metric_mean_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_common})
            
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             if metric_mean_col not in df.columns:
    #                 print(f"Warning: {metric_mean_col} not found in {root}")
    #                 continue
                
    #             t_file = df['t'].values
    #             y_file = df[metric_mean_col].values
                
    #             # Check if time arrays match
    #             if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                 result_df[root] = y_file
    #             else:
    #                 # Interpolate to common time base
    #                 print(f"Interpolating {root} to common time base for {metric}")
                    
    #                 mask = np.isfinite(y_file)
    #                 if mask.sum() < 2:
    #                     print(f"Warning: Not enough valid data in {root} for {metric}")
    #                     result_df[root] = np.nan
    #                     continue
                    
    #                 try:
    #                     f_interp = interp1d(
    #                         t_file[mask], y_file[mask], 
    #                         kind='linear', 
    #                         bounds_error=False, 
    #                         fill_value=np.nan
    #                     )
    #                     result_df[root] = f_interp(t_common)
    #                 except Exception as e:
    #                     print(f"Error interpolating {root} for {metric}: {e}")
    #                     result_df[root] = np.nan
            
    #         # Add mean and SEM columns
    #         result_df = add_mean_sem(result_df)
    #         consolidated[metric] = result_df
        
    #     # Process normalized means
    #     for metric in metrics:
    #         metric_norm_col = f"{metric}_norm_mean"
            
    #         if metric_norm_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_common})
            
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             if metric_norm_col not in df.columns:
    #                 continue
                
    #             t_file = df['t'].values
    #             y_file = df[metric_norm_col].values
                
    #             if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                 result_df[root] = y_file
    #             else:
    #                 mask = np.isfinite(y_file)
    #                 if mask.sum() >= 2:
    #                     try:
    #                         f_interp = interp1d(
    #                             t_file[mask], y_file[mask], 
    #                             kind='linear', 
    #                             bounds_error=False, 
    #                             fill_value=np.nan
    #                         )
    #                         result_df[root] = f_interp(t_common)
    #                     except:
    #                         result_df[root] = np.nan
    #                 else:
    #                     result_df[root] = np.nan
            
    #         # Add mean and SEM columns
    #         result_df = add_mean_sem(result_df)
    #         consolidated[f"{metric}_norm"] = result_df
        
    #     return consolidated

    # def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Consolidate means_by_time CSV files.
    #     Interpolates all data to a common time base.
    #     Returns dict: {metric_name: DataFrame with 't', raw data, raw mean/sem, norm data, norm mean/sem}
    #     """
    #     import pandas as pd
    #     import numpy as np
    #     from scipy.interpolate import interp1d
        
    #     metrics = [
    #         'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 
    #         'ti', 'te', 'vent_proxy'
    #     ]
        
    #     # Determine common time base (use first file's time)
    #     first_root, first_path = files[0]
    #     df_first = pd.read_csv(first_path)
    #     t_common = df_first['t'].values
        
    #     consolidated = {}
        
    #     # Helper function to calculate mean and SEM
    #     def calc_mean_sem(data_array):
    #         """Calculate mean and SEM from array of values."""
    #         mean = np.nanmean(data_array, axis=1)
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #         return mean, sem
        
    #     # Process each metric (combining raw and normalized in same sheet)
    #     for metric in metrics:
    #         metric_mean_col = f"{metric}_mean"
    #         metric_norm_col = f"{metric}_norm_mean"
            
    #         # Check if metric exists
    #         if metric_mean_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_common})
            
    #         # Collect raw data from all files
    #         raw_data_cols = []
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             if metric_mean_col not in df.columns:
    #                 print(f"Warning: {metric_mean_col} not found in {root}")
    #                 continue
                
    #             t_file = df['t'].values
    #             y_file = df[metric_mean_col].values
                
    #             if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                 result_df[root] = y_file
    #             else:
    #                 print(f"Interpolating {root} to common time base for {metric}")
    #                 mask = np.isfinite(y_file)
    #                 if mask.sum() >= 2:
    #                     try:
    #                         f_interp = interp1d(
    #                             t_file[mask], y_file[mask], 
    #                             kind='linear', 
    #                             bounds_error=False, 
    #                             fill_value=np.nan
    #                         )
    #                         result_df[root] = f_interp(t_common)
    #                     except Exception as e:
    #                         print(f"Error interpolating {root} for {metric}: {e}")
    #                         result_df[root] = np.nan
    #                 else:
    #                     result_df[root] = np.nan
                
    #             raw_data_cols.append(root)
            
    #         # Calculate raw mean and SEM
    #         if raw_data_cols:
    #             raw_data = result_df[raw_data_cols].values
    #             result_df['mean'], result_df['sem'] = calc_mean_sem(raw_data)
            
    #         # Collect normalized data from all files
    #         if metric_norm_col in df_first.columns:
    #             norm_data_cols = []
    #             for root, path in files:
    #                 df = pd.read_csv(path)
                    
    #                 if metric_norm_col not in df.columns:
    #                     continue
                    
    #                 t_file = df['t'].values
    #                 y_file = df[metric_norm_col].values
                    
    #                 norm_col_name = f"{root}_norm"
                    
    #                 if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                     result_df[norm_col_name] = y_file
    #                 else:
    #                     mask = np.isfinite(y_file)
    #                     if mask.sum() >= 2:
    #                         try:
    #                             f_interp = interp1d(
    #                                 t_file[mask], y_file[mask], 
    #                                 kind='linear', 
    #                                 bounds_error=False, 
    #                                 fill_value=np.nan
    #                             )
    #                             result_df[norm_col_name] = f_interp(t_common)
    #                         except:
    #                             result_df[norm_col_name] = np.nan
    #                     else:
    #                         result_df[norm_col_name] = np.nan
                    
    #                 norm_data_cols.append(norm_col_name)
                
    #             # Calculate normalized mean and SEM
    #             if norm_data_cols:
    #                 norm_data = result_df[norm_data_cols].values
    #                 result_df['mean_norm'], result_df['sem_norm'] = calc_mean_sem(norm_data)
            
    #         consolidated[metric] = result_df
        
    #     return consolidated

    # def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
    #     """
    #     Consolidate means_by_time CSV files.
    #     Interpolates all data to a common time base and adds summary statistics.
    #     Returns dict: {metric_name: DataFrame with time series + summary stats}
    #     """
    #     import pandas as pd
    #     import numpy as np
    #     from scipy.interpolate import interp1d
        
    #     metrics = [
    #         'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp', 
    #         'ti', 'te', 'vent_proxy'
    #     ]
        
    #     # Determine common time base (use first file's time)
    #     first_root, first_path = files[0]
    #     df_first = pd.read_csv(first_path)
    #     t_common = df_first['t'].values
        
    #     consolidated = {}
        
    #     # Helper function to calculate mean and SEM
    #     def calc_mean_sem(data_array):
    #         """Calculate mean and SEM from array of values."""
    #         mean = np.nanmean(data_array, axis=1)
    #         n = np.sum(np.isfinite(data_array), axis=1)
    #         std = np.nanstd(data_array, axis=1, ddof=1)
    #         sem = np.where(n >= 2, std / np.sqrt(n), np.nan)
    #         return mean, sem
        
    #     # Helper function to calculate window mean
    #     def window_mean(t, y, t_start, t_end):
    #         """Calculate mean of y values where t is between t_start and t_end."""
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     # Define time windows for summary stats
    #     windows = [
    #         ('Baseline (-10 to 0s)', -10.0, 0.0),
    #         ('Baseline (-5 to 0s)', -5.0, 0.0),
    #         ('Stim (0-15s)', 0.0, 15.0),
    #         ('Stim (0-5s)', 0.0, 5.0),
    #         ('Stim (5-10s)', 5.0, 10.0),
    #         ('Stim (10-15s)', 10.0, 15.0),
    #         ('Post (15-25s)', 15.0, 25.0),
    #         ('Post (15-20s)', 15.0, 20.0),
    #         ('Post (20-25s)', 20.0, 25.0),
    #         ('Post (25-30s)', 25.0, 30.0),
    #     ]
        
    #     # Process each metric (combining raw and normalized in same sheet)
    #     for metric in metrics:
    #         metric_mean_col = f"{metric}_mean"
    #         metric_norm_col = f"{metric}_norm_mean"
            
    #         # Check if metric exists
    #         if metric_mean_col not in df_first.columns:
    #             continue
            
    #         result_df = pd.DataFrame({'t': t_common})
            
    #         # Store data for summary calculations
    #         raw_data_dict = {}
    #         norm_data_dict = {}
            
    #         # Collect raw data from all files
    #         raw_data_cols = []
    #         for root, path in files:
    #             df = pd.read_csv(path)
                
    #             if metric_mean_col not in df.columns:
    #                 print(f"Warning: {metric_mean_col} not found in {root}")
    #                 continue
                
    #             t_file = df['t'].values
    #             y_file = df[metric_mean_col].values
                
    #             if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                 result_df[root] = y_file
    #                 raw_data_dict[root] = (t_common, y_file)
    #             else:
    #                 print(f"Interpolating {root} to common time base for {metric}")
    #                 mask = np.isfinite(y_file)
    #                 if mask.sum() >= 2:
    #                     try:
    #                         f_interp = interp1d(
    #                             t_file[mask], y_file[mask], 
    #                             kind='linear', 
    #                             bounds_error=False, 
    #                             fill_value=np.nan
    #                         )
    #                         y_interp = f_interp(t_common)
    #                         result_df[root] = y_interp
    #                         raw_data_dict[root] = (t_common, y_interp)
    #                     except Exception as e:
    #                         print(f"Error interpolating {root} for {metric}: {e}")
    #                         result_df[root] = np.nan
    #                 else:
    #                     result_df[root] = np.nan
                
    #             raw_data_cols.append(root)
            
    #         # Calculate raw mean and SEM
    #         if raw_data_cols:
    #             raw_data = result_df[raw_data_cols].values
    #             result_df['mean'], result_df['sem'] = calc_mean_sem(raw_data)
            
    #         # Collect normalized data from all files
    #         if metric_norm_col in df_first.columns:
    #             norm_data_cols = []
    #             for root, path in files:
    #                 df = pd.read_csv(path)
                    
    #                 if metric_norm_col not in df.columns:
    #                     continue
                    
    #                 t_file = df['t'].values
    #                 y_file = df[metric_norm_col].values
                    
    #                 norm_col_name = f"{root}_norm"
                    
    #                 if np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
    #                     result_df[norm_col_name] = y_file
    #                     norm_data_dict[root] = (t_common, y_file)
    #                 else:
    #                     mask = np.isfinite(y_file)
    #                     if mask.sum() >= 2:
    #                         try:
    #                             f_interp = interp1d(
    #                                 t_file[mask], y_file[mask], 
    #                                 kind='linear', 
    #                                 bounds_error=False, 
    #                                 fill_value=np.nan
    #                             )
    #                             y_interp = f_interp(t_common)
    #                             result_df[norm_col_name] = y_interp
    #                             norm_data_dict[root] = (t_common, y_interp)
    #                         except:
    #                             result_df[norm_col_name] = np.nan
    #                     else:
    #                         result_df[norm_col_name] = np.nan
                    
    #                 norm_data_cols.append(norm_col_name)
                
    #             # Calculate normalized mean and SEM
    #             if norm_data_cols:
    #                 norm_data = result_df[norm_data_cols].values
    #                 result_df['mean_norm'], result_df['sem_norm'] = calc_mean_sem(norm_data)
            
    #         # Insert blank column before normalized data
    #         norm_start_idx = None
    #         for i, col in enumerate(result_df.columns):
    #             if '_norm' in str(col):
    #                 norm_start_idx = i
    #                 break
    #         if norm_start_idx is not None:
    #             result_df.insert(norm_start_idx, '', '')
            
    #         # Add another blank column before summary stats
    #         result_df['  '] = ''
            
    #         # Build summary statistics table
    #         # RAW summary stats
    #         summary_rows = []
    #         for root in raw_data_dict.keys():
    #             t, y = raw_data_dict[root]
    #             row = {'File': root}
    #             for window_name, t_start, t_end in windows:
    #                 row[window_name] = window_mean(t, y, t_start, t_end)
    #             summary_rows.append(row)
            
    #         if summary_rows:
    #             summary_df = pd.DataFrame(summary_rows)
    #             # Add to result_df (will create columns to the right)
    #             for col in summary_df.columns:
    #                 if col != 'File':
    #                     result_df[col] = summary_df[col].values
    #             # Add file names as a column
    #             result_df.insert(len(result_df.columns) - len(windows), 'File', summary_df['File'].values)
            
    #         # NORMALIZED summary stats (if they exist)
    #         if norm_data_dict:
    #             result_df['   '] = ''  # Another blank separator
                
    #             norm_summary_rows = []
    #             for root in norm_data_dict.keys():
    #                 t, y = norm_data_dict[root]
    #                 row = {'File_norm': root}
    #                 for window_name, t_start, t_end in windows:
    #                     row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                 norm_summary_rows.append(row)
                
    #             if norm_summary_rows:
    #                 norm_summary_df = pd.DataFrame(norm_summary_rows)
    #                 for col in norm_summary_df.columns:
    #                     if col != 'File_norm':
    #                         result_df[col] = norm_summary_df[col].values
    #                 result_df.insert(len(result_df.columns) - len(windows), 'File_norm', norm_summary_df['File_norm'].values)
            
    #         consolidated[metric] = result_df
        
    #     return consolidated

    # -------------------- NPZ-based fast consolidation helpers --------------------

    def _try_load_npz_v2(self, npz_path: Path) -> dict | None:
        """
        Try to load NPZ version 2 bundle with full timeseries data.
        Returns dict with data if successful, None otherwise.
        """
        # Silently skip if file doesn't exist (expected for old format)
        if not npz_path.exists():
            return None

        try:
            data = np.load(npz_path, allow_pickle=True)

            # Check if this is version 2
            version = data.get('npz_version', None)
            if version is None or int(version) < 2:
                # Silently skip v1 NPZ files (will use CSV fallback)
                return None

            # Verify required keys exist
            required_keys = ['timeseries_t', 'timeseries_keys']
            if not all(k in data for k in required_keys):
                print(f"[NPZ] Warning: {npz_path.name} missing required keys, using CSV fallback")
                return None

            return dict(data)
        except Exception as e:
            # Only warn on actual load errors (not missing files)
            print(f"[NPZ] Warning: Failed to load {npz_path.name}: {e}")
            return None

    def _extract_timeseries_from_npz(self, npz_data: dict, metric: str, variant: str = 'raw') -> tuple[np.ndarray, np.ndarray]:
        """
        Extract timeseries data for a specific metric from NPZ bundle.

        Args:
            npz_data: Loaded NPZ data dictionary
            metric: Metric name (e.g., 'if', 'amp_insp')
            variant: One of 'raw', 'norm', 'eupnea'

        Returns:
            (t, Y) where t is time vector and Y is (M, S) metric matrix
        """
        t = npz_data['timeseries_t']

        if variant == 'raw':
            key = f'ts_raw_{metric}'
        elif variant == 'norm':
            key = f'ts_norm_{metric}'
        elif variant == 'eupnea':
            key = f'ts_eupnea_{metric}'
        else:
            raise ValueError(f"Unknown variant: {variant}")

        Y = npz_data[key]
        return t, Y

    def _consolidate_from_npz_v2(self, npz_data_by_root: dict, files: list[tuple[str, Path]], metrics: list[str]) -> dict:
        """
        Fast consolidation using NPZ v2 bundles with pre-computed timeseries data.
        Much faster than CSV because:
        1. No CSV parsing overhead
        2. No interpolation needed (all files share same time base from stimulus alignment)
        3. Direct numpy array operations
        """
        import pandas as pd
        import numpy as np

        # Helper function to calculate mean and SEM
        def calc_mean_sem(data_array):
            """Calculate mean and SEM from array of values."""
            n = np.sum(np.isfinite(data_array), axis=1)
            mean = np.full(data_array.shape[0], np.nan)
            sem = np.full(data_array.shape[0], np.nan)
            valid_rows = n > 0
            if valid_rows.any():
                mean[valid_rows] = np.nanmean(data_array[valid_rows, :], axis=1)
                sem_rows = n >= 2
                if sem_rows.any():
                    std = np.nanstd(data_array[sem_rows, :], axis=1, ddof=1)
                    sem[sem_rows] = std / np.sqrt(n[sem_rows])
            return mean, sem

        # Helper function to calculate window mean
        def window_mean(t, y, t_start, t_end):
            """Calculate mean of y values where t is between t_start and t_end."""
            mask = (t >= t_start) & (t < t_end)
            if mask.sum() == 0:
                return np.nan
            return np.nanmean(y[mask])

        # Define time windows for summary stats
        windows = [
            ('Baseline (-10 to 0s)', -10.0, 0.0),
            ('Baseline (-5 to 0s)', -5.0, 0.0),
            ('Stim (0-15s)', 0.0, 15.0),
            ('Stim (0-5s)', 0.0, 5.0),
            ('Stim (5-10s)', 5.0, 10.0),
            ('Stim (10-15s)', 10.0, 15.0),
            ('Post (15-25s)', 15.0, 25.0),
            ('Post (15-20s)', 15.0, 20.0),
            ('Post (20-25s)', 20.0, 25.0),
            ('Post (25-30s)', 25.0, 30.0),
        ]

        # Determine common time base by scanning all files
        # This allows experiments with different durations to be consolidated
        all_t_mins = []
        all_t_maxs = []
        all_steps = []

        for root, _ in files:
            t_file = npz_data_by_root[root]['timeseries_t']
            all_t_mins.append(t_file.min())
            all_t_maxs.append(t_file.max())
            if len(t_file) > 1:
                all_steps.append(np.median(np.diff(t_file)))

        # Create common time grid spanning the union of all time ranges
        t_common_min = min(all_t_mins)
        t_common_max = max(all_t_maxs)
        t_common_step = np.median(all_steps) if all_steps else 0.1

        # Generate uniform time grid
        t_common = np.arange(t_common_min, t_common_max + t_common_step/2, t_common_step)

        print(f"[NPZ] Common time grid: {t_common_min:.2f}s to {t_common_max:.2f}s, step={t_common_step:.4f}s ({len(t_common)} points)")

        consolidated = {}

        # Get first file root for checking metric existence
        first_root = files[0][0]

        # Process each metric
        for metric in metrics:
            # Check if metric exists in first NPZ
            test_key = f'ts_raw_{metric}'
            if test_key not in npz_data_by_root[first_root]:
                continue

            # Build all columns as dict first to avoid fragmentation
            data_dict = {'t': t_common}

            # Process each variant: raw, norm, eupnea
            # Note: eupnea suffix is '_eupnea' not '_norm_eupnea' to match CSV consolidation
            for variant, suffix in [('raw', ''), ('norm', '_norm'), ('eupnea', '_eupnea')]:
                # Collect data from all files
                all_means = []
                file_means = {}  # Store for window calculations

                for root, _ in files:
                    t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, variant)

                    # Compute mean across sweeps (axis 1)
                    y_mean_file = np.nanmean(Y, axis=1)

                    # Interpolate to common time grid if needed
                    if len(t_file) != len(t_common) or not np.allclose(t_file, t_common, rtol=0.01):
                        from scipy.interpolate import interp1d
                        # Use linear interpolation, extrapolate with NaN outside range
                        interp_func = interp1d(t_file, y_mean_file, kind='linear',
                                              bounds_error=False, fill_value=np.nan)
                        y_mean = interp_func(t_common)
                    else:
                        y_mean = y_mean_file

                    all_means.append(y_mean)
                    file_means[root] = (t_file, y_mean_file)  # Store original for window calculations

                    # Add individual file column
                    data_dict[f'{root}{suffix}'] = y_mean

                # Stack all means into matrix (M, num_files)
                all_means_matrix = np.column_stack(all_means)

                # Compute mean and SEM across files
                mean, sem = calc_mean_sem(all_means_matrix)
                data_dict[f'mean{suffix}'] = mean
                data_dict[f'sem{suffix}'] = sem

                # Add separator column
                data_dict[f' {suffix}'] = ''

            # Create DataFrame once with all columns
            result_df = pd.DataFrame(data_dict)

            # Build summary data dicts (for Excel summary section)
            # These use the original (non-interpolated) data for window calculations
            # We need to collect these from each variant's file_means
            raw_summary = {}
            norm_summary = {}
            eupnea_summary = {}

            # Re-extract data for each variant to populate summary dicts
            for root, _ in files:
                # Raw data
                t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, 'raw')
                y_mean_file = np.nanmean(Y, axis=1)
                raw_summary[root] = (t_file, y_mean_file)

                # Normalized data
                t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, 'norm')
                y_mean_file = np.nanmean(Y, axis=1)
                norm_summary[root] = (t_file, y_mean_file)

                # Eupnea-normalized data
                t_file, Y = self._extract_timeseries_from_npz(npz_data_by_root[root], metric, 'eupnea')
                y_mean_file = np.nanmean(Y, axis=1)
                eupnea_summary[root] = (t_file, y_mean_file)

            # Package result in same format as CSV consolidation
            consolidated[metric] = {
                'time_series': result_df,
                'raw_summary': raw_summary,
                'norm_summary': norm_summary,
                'eupnea_summary': eupnea_summary,
                'windows': windows
            }

        print(f"[NPZ] ✓ Fast consolidation complete")
        return consolidated

    def _consolidate_means_files(self, files: list[tuple[str, Path]]) -> dict:
        """
        Consolidate means_by_time CSV files.
        Interpolates all data to a common time base and adds summary statistics.
        Returns dict: {metric_name: DataFrame with time series + summary stats}
        """
        import pandas as pd
        import numpy as np
        from scipy.interpolate import interp1d

        metrics = [
            'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp',
            'ti', 'te', 'vent_proxy'
        ]

        # Try fast NPZ-based consolidation first (5-10× faster)
        npz_data_by_root = {}
        npz_success_count = 0
        for root, csv_path in files:
            # Look for NPZ bundle next to the CSV (handle both naming patterns)
            if csv_path.name.endswith('_timeseries.csv'):
                npz_name = csv_path.name.replace('_timeseries.csv', '_bundle.npz')
            elif csv_path.name.endswith('_means_by_time.csv'):
                npz_name = csv_path.name.replace('_means_by_time.csv', '_bundle.npz')
            else:
                continue  # Unknown CSV pattern

            npz_path = csv_path.parent / npz_name
            npz_data = self._try_load_npz_v2(npz_path)
            if npz_data is not None:
                npz_data_by_root[root] = npz_data
                npz_success_count += 1

        # If all files have NPZ v2, use fast path (with interpolation support)
        if npz_success_count == len(files) and npz_success_count > 0:
            print(f"[NPZ Fast Path] Loading {npz_success_count} files from NPZ bundles (v2)...")
            return self._consolidate_from_npz_v2(npz_data_by_root, files, metrics)
        elif npz_success_count > 0:
            print(f"[NPZ] Only {npz_success_count}/{len(files)} files have NPZ v2, falling back to CSV...")
        else:
            print(f"[CSV] No NPZ v2 bundles found, using CSV files...")

        # Determine common time base by scanning all files
        # This allows experiments with different durations to be consolidated
        print("Scanning all files to determine common time range...")
        all_t_mins = []
        all_t_maxs = []
        all_steps = []

        for root, path in files:
            df_temp = pd.read_csv(path, low_memory=False)
            t_temp = df_temp['t'].values
            all_t_mins.append(t_temp.min())
            all_t_maxs.append(t_temp.max())
            if len(t_temp) > 1:
                all_steps.append(np.median(np.diff(t_temp)))

        # Create common time grid spanning the union of all time ranges
        t_common_min = min(all_t_mins)
        t_common_max = max(all_t_maxs)
        t_common_step = np.median(all_steps) if all_steps else 0.1

        # Generate uniform time grid
        t_common = np.arange(t_common_min, t_common_max + t_common_step/2, t_common_step)

        print(f"Common time grid: {t_common_min:.2f}s to {t_common_max:.2f}s, step={t_common_step:.4f}s ({len(t_common)} points)")

        # Load first file for column checking
        first_root, first_path = files[0]
        df_first = pd.read_csv(first_path, low_memory=False)

        # Track files with potential issues
        warning_messages = []
        files_needing_interpolation = []
        files_with_poor_overlap = []
        files_with_different_sampling = []

        consolidated = {}
        
        # Helper function to calculate mean and SEM
        def calc_mean_sem(data_array):
            """Calculate mean and SEM from array of values."""
            # Count valid values per row
            n = np.sum(np.isfinite(data_array), axis=1)

            # Initialize outputs with NaN
            mean = np.full(data_array.shape[0], np.nan)
            sem = np.full(data_array.shape[0], np.nan)

            # Only calculate for rows with at least one valid value
            valid_rows = n > 0
            if valid_rows.any():
                mean[valid_rows] = np.nanmean(data_array[valid_rows, :], axis=1)

                # Only calculate SEM for rows with at least 2 values
                sem_rows = n >= 2
                if sem_rows.any():
                    std = np.nanstd(data_array[sem_rows, :], axis=1, ddof=1)
                    sem[sem_rows] = std / np.sqrt(n[sem_rows])

            return mean, sem
        
        # Helper function to calculate window mean
        def window_mean(t, y, t_start, t_end):
            """Calculate mean of y values where t is between t_start and t_end."""
            mask = (t >= t_start) & (t < t_end)
            if mask.sum() == 0:
                return np.nan
            return np.nanmean(y[mask])
        
        # Define time windows for summary stats
        windows = [
            ('Baseline (-10 to 0s)', -10.0, 0.0),
            ('Baseline (-5 to 0s)', -5.0, 0.0),
            ('Stim (0-15s)', 0.0, 15.0),
            ('Stim (0-5s)', 0.0, 5.0),
            ('Stim (5-10s)', 5.0, 10.0),
            ('Stim (10-15s)', 10.0, 15.0),
            ('Post (15-25s)', 15.0, 25.0),
            ('Post (15-20s)', 15.0, 20.0),
            ('Post (20-25s)', 20.0, 25.0),
            ('Post (25-30s)', 25.0, 30.0),
        ]
        
        # Process each metric (combining raw, time-normalized, and eupnea-normalized in same sheet)
        for metric in metrics:
            metric_mean_col = f"{metric}_mean"
            metric_norm_col = f"{metric}_norm_mean"
            metric_eupnea_col = f"{metric}_norm_eupnea_mean"

            # Check if metric exists
            if metric_mean_col not in df_first.columns:
                continue

            result_df = pd.DataFrame({'t': t_common})

            # Store data for summary calculations
            raw_data_dict = {}
            norm_data_dict = {}
            eupnea_data_dict = {}
            
            # Collect raw data from all files
            raw_data_cols = []
            for root, path in files:
                df = pd.read_csv(path, low_memory=False)

                if metric_mean_col not in df.columns:
                    print(f"Warning: {metric_mean_col} not found in {root}")
                    continue

                t_file = df['t'].values
                y_file = df[metric_mean_col].values

                # Check time range and sampling
                t_file_min, t_file_max = t_file.min(), t_file.max()
                t_file_step = np.median(np.diff(t_file)) if len(t_file) > 1 else np.nan

                # Calculate overlap percentage
                overlap_start = max(t_common_min, t_file_min)
                overlap_end = min(t_common_max, t_file_max)
                overlap_range = overlap_end - overlap_start
                common_range = t_common_max - t_common_min
                overlap_pct = 100 * overlap_range / common_range if common_range > 0 else 0

                # Check for different sampling rates
                if not np.isnan(t_file_step) and not np.isnan(t_common_step):
                    sampling_diff_pct = 100 * abs(t_file_step - t_common_step) / t_common_step
                    if sampling_diff_pct > 10 and root != first_root:
                        files_with_different_sampling.append(
                            f"{root}: {t_file_step:.4f}s vs reference {t_common_step:.4f}s ({sampling_diff_pct:.1f}% difference)"
                        )

                # Check for poor overlap
                if overlap_pct < 80 and root != first_root:
                    files_with_poor_overlap.append(
                        f"{root}: {overlap_pct:.1f}% overlap (range: {t_file_min:.1f} to {t_file_max:.1f}s)"
                    )

                # Always interpolate to common time grid (different files may have different durations)
                mask = np.isfinite(y_file)
                if mask.sum() >= 2:
                    try:
                        # Check if exact match first (optimization)
                        if len(t_file) == len(t_common) and np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
                            result_df[root] = y_file
                            raw_data_dict[root] = (t_common, y_file)
                        else:
                            # Interpolate to common grid
                            if root not in files_needing_interpolation:
                                files_needing_interpolation.append(root)
                            print(f"Interpolating {root} to common time base for {metric}")
                            f_interp = interp1d(
                                t_file[mask], y_file[mask],
                                kind='linear',
                                bounds_error=False,
                                fill_value=np.nan
                            )
                            y_interp = f_interp(t_common)
                            result_df[root] = y_interp
                            raw_data_dict[root] = (t_common, y_interp)

                            # Count how many points were extrapolated (NaN after interpolation)
                            n_extrapolated = np.sum(np.isnan(y_interp) & ~np.isnan(t_common))
                            if n_extrapolated > 0:
                                extrap_pct = 100 * n_extrapolated / len(t_common)
                                if extrap_pct > 5:
                                    print(f"  Warning: {extrap_pct:.1f}% of points extrapolated (outside data range)")
                    except Exception as e:
                        print(f"Error interpolating {root} for {metric}: {e}")
                        result_df[root] = np.nan
                else:
                    print(f"  Warning: Insufficient data points for interpolation in {root}")
                    result_df[root] = np.nan

                raw_data_cols.append(root)
            
            # Calculate raw mean and SEM
            if raw_data_cols:
                raw_data = result_df[raw_data_cols].values
                result_df['mean'], result_df['sem'] = calc_mean_sem(raw_data)
            
            # Collect normalized data from all files
            if metric_norm_col in df_first.columns:
                norm_data_cols = []
                for root, path in files:
                    df = pd.read_csv(path, low_memory=False)
                    
                    if metric_norm_col not in df.columns:
                        continue
                    
                    t_file = df['t'].values
                    y_file = df[metric_norm_col].values
                    
                    norm_col_name = f"{root}_norm"

                    # Always interpolate to common time grid
                    mask = np.isfinite(y_file)
                    if mask.sum() >= 2:
                        try:
                            # Check if exact match first (optimization)
                            if len(t_file) == len(t_common) and np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
                                result_df[norm_col_name] = y_file
                                norm_data_dict[root] = (t_common, y_file)
                            else:
                                f_interp = interp1d(
                                    t_file[mask], y_file[mask],
                                    kind='linear',
                                    bounds_error=False,
                                    fill_value=np.nan
                                )
                                y_interp = f_interp(t_common)
                                result_df[norm_col_name] = y_interp
                                norm_data_dict[root] = (t_common, y_interp)
                        except:
                            result_df[norm_col_name] = np.nan
                    else:
                        result_df[norm_col_name] = np.nan
                    
                    norm_data_cols.append(norm_col_name)
                
                # Calculate normalized mean and SEM
                if norm_data_cols:
                    norm_data = result_df[norm_data_cols].values
                    result_df['mean_norm'], result_df['sem_norm'] = calc_mean_sem(norm_data)

            # Collect eupnea-normalized data from all files
            if metric_eupnea_col in df_first.columns:
                eupnea_data_cols = []
                for root, path in files:
                    df = pd.read_csv(path, low_memory=False)

                    if metric_eupnea_col not in df.columns:
                        continue

                    t_file = df['t'].values
                    y_file = df[metric_eupnea_col].values

                    eupnea_col_name = f"{root}_eupnea"

                    # Always interpolate to common time grid
                    mask = np.isfinite(y_file)
                    if mask.sum() >= 2:
                        try:
                            # Check if exact match first (optimization)
                            if len(t_file) == len(t_common) and np.allclose(t_file, t_common, rtol=1e-5, atol=1e-8):
                                result_df[eupnea_col_name] = y_file
                                eupnea_data_dict[root] = (t_common, y_file)
                            else:
                                f_interp = interp1d(
                                    t_file[mask], y_file[mask],
                                    kind='linear',
                                    bounds_error=False,
                                    fill_value=np.nan
                                )
                                y_interp = f_interp(t_common)
                                result_df[eupnea_col_name] = y_interp
                                eupnea_data_dict[root] = (t_common, y_interp)
                        except:
                            result_df[eupnea_col_name] = np.nan
                    else:
                        result_df[eupnea_col_name] = np.nan

                    eupnea_data_cols.append(eupnea_col_name)

                # Calculate eupnea-normalized mean and SEM
                if eupnea_data_cols:
                    eupnea_data = result_df[eupnea_data_cols].values
                    result_df['mean_eupnea'], result_df['sem_eupnea'] = calc_mean_sem(eupnea_data)

            # Insert blank columns between data blocks
            # First blank: before time-normalized data
            norm_start_idx = None
            for i, col in enumerate(result_df.columns):
                if '_norm' in str(col) and '_eupnea' not in str(col):
                    norm_start_idx = i
                    break
            if norm_start_idx is not None:
                result_df.insert(norm_start_idx, '', '')

            # Second blank: before eupnea-normalized data
            eupnea_start_idx = None
            for i, col in enumerate(result_df.columns):
                if '_eupnea' in str(col):
                    eupnea_start_idx = i
                    break
            if eupnea_start_idx is not None:
                result_df.insert(eupnea_start_idx, ' ', '')

            # Build summary statistics (as rows below the time series)
            # This will be saved as a separate section in Excel
            consolidated[metric] = {
                'time_series': result_df,
                'raw_summary': raw_data_dict,
                'norm_summary': norm_data_dict,
                'eupnea_summary': eupnea_data_dict,
                'windows': windows
            }

        # Build warning summary
        if files_needing_interpolation or files_with_poor_overlap or files_with_different_sampling:
            warning_parts = []

            if files_needing_interpolation:
                warning_parts.append("FILES REQUIRING INTERPOLATION:")
                warning_parts.append(f"Reference file (no interpolation): {first_root}")
                warning_parts.append(f"Time range: {t_common_min:.2f} to {t_common_max:.2f}s")
                warning_parts.append(f"Sample interval: {t_common_step:.4f}s\n")
                for f in files_needing_interpolation:
                    warning_parts.append(f"  • {f}")
                warning_parts.append("")

            if files_with_different_sampling:
                warning_parts.append("FILES WITH DIFFERENT SAMPLING RATES:")
                for msg in files_with_different_sampling:
                    warning_parts.append(f"  • {msg}")
                warning_parts.append("")

            if files_with_poor_overlap:
                warning_parts.append("FILES WITH POOR TIME OVERLAP (<80%):")
                for msg in files_with_poor_overlap:
                    warning_parts.append(f"  • {msg}")
                warning_parts.append("")

            # Store warning message for display after processing
            consolidated['_warnings'] = '\n'.join(warning_parts)

        return consolidated


    def _consolidate_breaths_sighs(self, files: list[tuple[str, Path]]) -> pd.DataFrame:
        """
        Extract all breaths marked as sighs (is_sigh == 1) from breaths CSV files.
        Adds experiment_number and global_sweep_number columns.
        Returns DataFrame with sigh breaths from all files.
        """
        import pandas as pd
        import numpy as np

        # Columns to extract for raw data
        raw_cols = ['sweep', 'breath', 't', 'region', 'is_sigh',
                    'if', 'amp_insp', 'amp_exp', 'area_insp', 'area_exp',
                    'ti', 'te', 'vent_proxy']

        # Columns to extract for normalized data
        norm_cols = ['sweep', 'breath', 't', 'region', 'is_sigh',
                    'if_norm', 'amp_insp_norm', 'amp_exp_norm', 'area_insp_norm', 'area_exp_norm',
                    'ti_norm', 'te_norm', 'vent_proxy_norm']

        all_sighs_raw = []
        all_sighs_norm = []

        for exp_num, (root, path) in enumerate(files, start=1):
            try:
                df = pd.read_csv(path, low_memory=False)

                # Filter for sighs (is_sigh == 1)
                if 'is_sigh' in df.columns:
                    sigh_mask = df['is_sigh'] == 1

                    # Extract raw sigh data
                    available_raw_cols = [col for col in raw_cols if col in df.columns]
                    if available_raw_cols and sigh_mask.sum() > 0:
                        sigh_df_raw = df.loc[sigh_mask, available_raw_cols].copy()
                        sigh_df_raw.insert(0, 'file', root)
                        sigh_df_raw.insert(1, 'experiment_number', exp_num)

                        # Calculate global sweep number
                        if 'sweep' in sigh_df_raw.columns:
                            # Calculate offset from previous experiments
                            sweep_offset = sum([pd.read_csv(files[i][1], low_memory=False)['sweep'].max()
                                              for i in range(exp_num - 1)
                                              if 'sweep' in pd.read_csv(files[i][1], low_memory=False).columns])
                            sigh_df_raw.insert(2, 'global_sweep_number', sigh_df_raw['sweep'] + sweep_offset)

                        all_sighs_raw.append(sigh_df_raw)

                    # Extract normalized sigh data
                    available_norm_cols = [col for col in norm_cols if col in df.columns]
                    if available_norm_cols and sigh_mask.sum() > 0:
                        sigh_df_norm = df.loc[sigh_mask, available_norm_cols].copy()
                        sigh_df_norm.insert(0, 'file', root)
                        all_sighs_norm.append(sigh_df_norm)

            except Exception as e:
                print(f"Error reading sighs from {path}: {e}")

        # Combine all sigh data
        if all_sighs_raw:
            combined_raw = pd.concat(all_sighs_raw, ignore_index=True)
        else:
            combined_raw = pd.DataFrame(columns=['file', 'experiment_number', 'global_sweep_number'] + raw_cols)

        if all_sighs_norm:
            combined_norm = pd.concat(all_sighs_norm, ignore_index=True)
        else:
            combined_norm = pd.DataFrame(columns=['file'] + norm_cols)

        # Combine raw and normalized with blank column separator
        combined_raw[''] = ''
        combined_sighs = pd.concat([combined_raw, combined_norm], axis=1)

        return combined_sighs

    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
        
    #     # First, write all dataframes to Excel
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, df in consolidated.items():
    #             # Insert blank column between raw and normalized data
    #             # Find where normalized columns start (first column with '_norm' in name)
    #             norm_start_idx = None
    #             for i, col in enumerate(df.columns):
    #                 if '_norm' in str(col):
    #                     norm_start_idx = i
    #                     break
                
    #             if norm_start_idx is not None:
    #                 # Insert blank column before normalized data
    #                 df.insert(norm_start_idx, '', '')
                
    #             sheet_name = metric_name[:31]
    #             df.to_excel(writer, sheet_name=sheet_name, index=False)
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Now apply bold formatting
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Get header row
    #         header_row = ws[1]
            
    #         # Bold specific columns: t, mean, sem, mean_norm, sem_norm
    #         bold_columns = {'t', 'mean', 'sem', 'mean_norm', 'sem_norm'}
            
    #         for cell in header_row:
    #             if cell.value in bold_columns:
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column
    #                 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting. Consolidated Excel file saved: {save_path}")


    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, df in consolidated.items():
    #             # Excel sheet names have a 31 character limit
    #             sheet_name = metric_name[:31]
    #             df.to_excel(writer, sheet_name=sheet_name, index=False)
    #             print(f"Saved sheet: {sheet_name}")
        
    #     print(f"Consolidated Excel file saved: {save_path}")


    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
        
    #     # Helper to calculate window means
    #     def window_mean(t, y, t_start, t_end):
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, data_dict in consolidated.items():
    #             time_series_df = data_dict['time_series']
    #             raw_summary = data_dict['raw_summary']
    #             norm_summary = data_dict['norm_summary']
    #             windows = data_dict['windows']
                
    #             sheet_name = metric_name[:31]
                
    #             # Write time series data
    #             time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                
    #             # Calculate starting row for summary (after time series + 2 blank rows)
    #             summary_start_row = len(time_series_df) + 3
                
    #             # Build raw summary DataFrame
    #             raw_summary_rows = []
    #             for root in raw_summary.keys():
    #                 t, y = raw_summary[root]
    #                 row = {'File': root}
    #                 for window_name, t_start, t_end in windows:
    #                     row[window_name] = window_mean(t, y, t_start, t_end)
    #                 raw_summary_rows.append(row)
                
    #             if raw_summary_rows:
    #                 raw_summary_df = pd.DataFrame(raw_summary_rows)
    #                 raw_summary_df.to_excel(writer, sheet_name=sheet_name, 
    #                                     index=False, startrow=summary_start_row)
                
    #             # Build normalized summary DataFrame  
    #             if norm_summary:
    #                 norm_start_row = summary_start_row + len(raw_summary_rows) + 3
                    
    #                 norm_summary_rows = []
    #                 for root in norm_summary.keys():
    #                     t, y = norm_summary[root]
    #                     row = {'File': root}
    #                     for window_name, t_start, t_end in windows:
    #                         row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                     norm_summary_rows.append(row)
                    
    #                 if norm_summary_rows:
    #                     norm_summary_df = pd.DataFrame(norm_summary_rows)
    #                     norm_summary_df.to_excel(writer, sheet_name=sheet_name, 
    #                                         index=False, startrow=norm_start_row)
                
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Apply bold formatting
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Bold columns in time series: t, mean, sem, mean_norm, sem_norm
    #         header_row = ws[1]
    #         bold_columns = {'t', 'mean', 'sem', 'mean_norm', 'sem_norm'}
            
    #         for cell in header_row:
    #             if cell.value in bold_columns:
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column (only in time series section)
    #                 for row in ws.iter_rows(min_row=1, max_row=len(time_series_df)+1, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting. Consolidated Excel file saved: {save_path}")

    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     import numpy as np
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
    #     from openpyxl.utils.dataframe import dataframe_to_rows
        
    #     # Helper to calculate window means
    #     def window_mean(t, y, t_start, t_end):
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, data_dict in consolidated.items():
    #             time_series_df = data_dict['time_series']
    #             raw_summary = data_dict['raw_summary']
    #             norm_summary = data_dict['norm_summary']
    #             windows = data_dict['windows']
                
    #             sheet_name = metric_name[:31]
                
    #             # Write time series data starting at A1
    #             time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
    #             # Calculate starting column for summary (after time series + 2 blank columns)
    #             summary_start_col = len(time_series_df.columns) + 2
                
    #             # Get the worksheet to write summary data
    #             worksheet = writer.sheets[sheet_name]
                
    #             # Build raw summary DataFrame
    #             raw_summary_rows = []
    #             for root in raw_summary.keys():
    #                 t, y = raw_summary[root]
    #                 row = {'File': root}
    #                 for window_name, t_start, t_end in windows:
    #                     row[window_name] = window_mean(t, y, t_start, t_end)
    #                 raw_summary_rows.append(row)
                
    #             if raw_summary_rows:
    #                 raw_summary_df = pd.DataFrame(raw_summary_rows)
                    
    #                 # Write raw summary starting at top right
    #                 for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
    #                     for c_idx, value in enumerate(row):
    #                         worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                
    #             # Build normalized summary DataFrame  
    #             if norm_summary:
    #                 # Start normalized summary after raw summary + 2 blank columns
    #                 norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                    
    #                 norm_summary_rows = []
    #                 for root in norm_summary.keys():
    #                     t, y = norm_summary[root]
    #                     row = {'File': root}
    #                     for window_name, t_start, t_end in windows:
    #                         row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                     norm_summary_rows.append(row)
                    
    #                 if norm_summary_rows:
    #                     norm_summary_df = pd.DataFrame(norm_summary_rows)
                        
    #                     # Write normalized summary to the right of raw summary
    #                     for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
    #                         for c_idx, value in enumerate(row):
    #                             worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)
                
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Apply bold formatting
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Bold columns in time series: t, mean, sem, mean_norm, sem_norm
    #         header_row = ws[1]
    #         bold_columns = {'t', 'mean', 'sem', 'mean_norm', 'sem_norm'}
            
    #         for cell in header_row:
    #             if cell.value in bold_columns:
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column (only for time series height)
    #                 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting. Consolidated Excel file saved: {save_path}")

    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     import numpy as np
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
    #     from openpyxl.utils.dataframe import dataframe_to_rows
        
    #     # Helper to calculate window means
    #     def window_mean(t, y, t_start, t_end):
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, data_dict in consolidated.items():
    #             time_series_df = data_dict['time_series']
    #             raw_summary = data_dict.get('raw_summary', {})
    #             norm_summary = data_dict.get('norm_summary', {})
    #             windows = data_dict.get('windows', [])
                
    #             sheet_name = metric_name[:31]
                
    #             # Write time series data starting at A1
    #             time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
    #             # Only add summary if it exists (means files have it, histogram files don't)
    #             if raw_summary:
    #                 # Calculate starting column for summary (after time series + 2 blank columns)
    #                 summary_start_col = len(time_series_df.columns) + 2
                    
    #                 # Get the worksheet to write summary data
    #                 worksheet = writer.sheets[sheet_name]
                    
    #                 # Build raw summary DataFrame
    #                 raw_summary_rows = []
    #                 for root in raw_summary.keys():
    #                     t, y = raw_summary[root]
    #                     row = {'File': root}
    #                     for window_name, t_start, t_end in windows:
    #                         row[window_name] = window_mean(t, y, t_start, t_end)
    #                     raw_summary_rows.append(row)
                    
    #                 if raw_summary_rows:
    #                     raw_summary_df = pd.DataFrame(raw_summary_rows)
                        
    #                     # Write raw summary starting at top right
    #                     for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
    #                         for c_idx, value in enumerate(row):
    #                             worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                    
    #                 # Build normalized summary DataFrame  
    #                 if norm_summary:
    #                     # Start normalized summary after raw summary + 2 blank columns
    #                     norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                        
    #                     norm_summary_rows = []
    #                     for root in norm_summary.keys():
    #                         t, y = norm_summary[root]
    #                         row = {'File': root}
    #                         for window_name, t_start, t_end in windows:
    #                             row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                         norm_summary_rows.append(row)
                        
    #                     if norm_summary_rows:
    #                         norm_summary_df = pd.DataFrame(norm_summary_rows)
                            
    #                         # Write normalized summary to the right of raw summary
    #                         for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
    #                             for c_idx, value in enumerate(row):
    #                                 worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)
                
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Apply bold formatting
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Bold columns in time series: t, mean, sem, mean_norm, sem_norm, bin_center
    #         header_row = ws[1]
    #         bold_columns = {'t', 'mean', 'sem', 'mean_norm', 'sem_norm', 'bin_center'}
            
    #         for cell in header_row:
    #             if cell.value in bold_columns:
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column
    #                 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting. Consolidated Excel file saved: {save_path}")


    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     import numpy as np
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
    #     from openpyxl.utils.dataframe import dataframe_to_rows
        
    #     # Helper to calculate window means
    #     def window_mean(t, y, t_start, t_end):
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, data_dict in consolidated.items():
    #             time_series_df = data_dict['time_series']
    #             raw_summary = data_dict.get('raw_summary', {})
    #             norm_summary = data_dict.get('norm_summary', {})
    #             windows = data_dict.get('windows', [])
                
    #             sheet_name = metric_name[:31]
                
    #             # Write time series data starting at A1
    #             time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
    #             # Only add summary if it exists (means files have it, histogram files don't)
    #             if raw_summary:
    #                 # Calculate starting column for summary (after time series + 2 blank columns)
    #                 summary_start_col = len(time_series_df.columns) + 2
                    
    #                 # Get the worksheet to write summary data
    #                 worksheet = writer.sheets[sheet_name]
                    
    #                 # Build raw summary DataFrame
    #                 raw_summary_rows = []
    #                 for root in raw_summary.keys():
    #                     t, y = raw_summary[root]
    #                     row = {'File': root}
    #                     for window_name, t_start, t_end in windows:
    #                         row[window_name] = window_mean(t, y, t_start, t_end)
    #                     raw_summary_rows.append(row)
                    
    #                 if raw_summary_rows:
    #                     raw_summary_df = pd.DataFrame(raw_summary_rows)
                        
    #                     # Write raw summary starting at top right
    #                     for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
    #                         for c_idx, value in enumerate(row):
    #                             worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                    
    #                 # Build normalized summary DataFrame  
    #                 if norm_summary:
    #                     # Start normalized summary after raw summary + 2 blank columns
    #                     norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                        
    #                     norm_summary_rows = []
    #                     for root in norm_summary.keys():
    #                         t, y = norm_summary[root]
    #                         row = {'File': root}
    #                         for window_name, t_start, t_end in windows:
    #                             row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                         norm_summary_rows.append(row)
                        
    #                     if norm_summary_rows:
    #                         norm_summary_df = pd.DataFrame(norm_summary_rows)
                            
    #                         # Write normalized summary to the right of raw summary
    #                         for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
    #                             for c_idx, value in enumerate(row):
    #                                 worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)
                
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Apply bold formatting
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Bold columns: t, mean, sem, mean_norm, sem_norm, bin_center, and histogram mean/sem
    #         header_row = ws[1]
            
    #         for cell in header_row:
    #             cell_val = str(cell.value) if cell.value else ''
                
    #             # Bold if column name contains 'mean', 'sem', starts with 't' or 'bin_center'
    #             if (cell_val in {'t', 'mean', 'sem', 'mean_norm', 'sem_norm'} or 
    #                 cell_val.startswith('bin_center') or 
    #                 cell_val.startswith('mean_') or 
    #                 cell_val.startswith('sem_')):
                    
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column
    #                 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting. Consolidated Excel file saved: {save_path}")

    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     import numpy as np
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
    #     from openpyxl.utils.dataframe import dataframe_to_rows
    #     from openpyxl.chart import ScatterChart, Reference, Series
        
    #     # Helper to calculate window means
    #     def window_mean(t, y, t_start, t_end):
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, data_dict in consolidated.items():
    #             time_series_df = data_dict['time_series']
    #             raw_summary = data_dict.get('raw_summary', {})
    #             norm_summary = data_dict.get('norm_summary', {})
    #             windows = data_dict.get('windows', [])
                
    #             sheet_name = metric_name[:31]
                
    #             # Write time series data starting at A1
    #             time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
    #             # Only add summary if it exists (means files have it, histogram files don't)
    #             if raw_summary:
    #                 # Calculate starting column for summary (after time series + 2 blank columns)
    #                 summary_start_col = len(time_series_df.columns) + 2
                    
    #                 # Get the worksheet to write summary data
    #                 worksheet = writer.sheets[sheet_name]
                    
    #                 # Build raw summary DataFrame
    #                 raw_summary_rows = []
    #                 for root in raw_summary.keys():
    #                     t, y = raw_summary[root]
    #                     row = {'File': root}
    #                     for window_name, t_start, t_end in windows:
    #                         row[window_name] = window_mean(t, y, t_start, t_end)
    #                     raw_summary_rows.append(row)
                    
    #                 if raw_summary_rows:
    #                     raw_summary_df = pd.DataFrame(raw_summary_rows)
                        
    #                     # Write raw summary starting at top right
    #                     for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
    #                         for c_idx, value in enumerate(row):
    #                             worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                    
    #                 # Build normalized summary DataFrame  
    #                 if norm_summary:
    #                     # Start normalized summary after raw summary + 2 blank columns
    #                     norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                        
    #                     norm_summary_rows = []
    #                     for root in norm_summary.keys():
    #                         t, y = norm_summary[root]
    #                         row = {'File': root}
    #                         for window_name, t_start, t_end in windows:
    #                             row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                         norm_summary_rows.append(row)
                        
    #                     if norm_summary_rows:
    #                         norm_summary_df = pd.DataFrame(norm_summary_rows)
                            
    #                         # Write normalized summary to the right of raw summary
    #                         for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
    #                             for c_idx, value in enumerate(row):
    #                                 worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)
                
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Apply bold formatting and add charts to histogram sheets
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Bold columns: t, mean, sem, mean_norm, sem_norm, bin_center, and histogram mean/sem
    #         header_row = ws[1]
            
    #         for cell in header_row:
    #             cell_val = str(cell.value) if cell.value else ''
                
    #             # Bold if column name contains 'mean', 'sem', starts with 't' or 'bin_center'
    #             if (cell_val in {'t', 'mean', 'sem', 'mean_norm', 'sem_norm'} or 
    #                 cell_val.startswith('bin_center') or 
    #                 cell_val.startswith('mean_') or 
    #                 cell_val.startswith('sem_')):
                    
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column
    #                 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
            
    #         # Add charts for histogram sheets
    #         if '_histogram' in sheet_name:
    #             # Find mean columns for raw and normalized data
    #             regions = ['all', 'baseline', 'stim', 'post']
                
    #             # Chart 1: Raw means overlay
    #             chart1 = ScatterChart()
    #             chart1.title = f"{sheet_name} - Raw Mean Histograms"
    #             chart1.x_axis.title = "Bin Center"
    #             chart1.y_axis.title = "Density"
    #             chart1.style = 2
                
    #             for region in regions:
    #                 bin_col = None
    #                 mean_col = None
                    
    #                 # Find the bin_center and mean columns for this region
    #                 for idx, cell in enumerate(header_row, start=1):
    #                     if cell.value == f'bin_center_{region}':
    #                         bin_col = idx
    #                     elif cell.value == f'mean_{region}':
    #                         mean_col = idx
                    
    #                 if bin_col and mean_col:
    #                     # Create x values (bin centers)
    #                     xvalues = Reference(ws, min_col=bin_col, min_row=2, max_row=ws.max_row)
    #                     # Create y values (means)
    #                     yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                        
    #                     series = Series(yvalues, xvalues, title=region)
    #                     chart1.series.append(series)
                
    #             # Position chart below data
    #             chart1.width = 20
    #             chart1.height = 12
    #             ws.add_chart(chart1, f"A{ws.max_row + 3}")
                
    #             # Chart 2: Normalized means overlay
    #             chart2 = ScatterChart()
    #             chart2.title = f"{sheet_name} - Normalized Mean Histograms"
    #             chart2.x_axis.title = "Bin Center (normalized)"
    #             chart2.y_axis.title = "Density"
    #             chart2.style = 2
                
    #             for region in regions:
    #                 bin_col_norm = None
    #                 mean_col_norm = None
                    
    #                 # Find the bin_center and mean columns for normalized region
    #                 for idx, cell in enumerate(header_row, start=1):
    #                     if cell.value == f'bin_center_{region}_norm':
    #                         bin_col_norm = idx
    #                     elif cell.value == f'mean_{region}_norm':
    #                         mean_col_norm = idx
                    
    #                 if bin_col_norm and mean_col_norm:
    #                     xvalues = Reference(ws, min_col=bin_col_norm, min_row=2, max_row=ws.max_row)
    #                     yvalues = Reference(ws, min_col=mean_col_norm, min_row=2, max_row=ws.max_row)
                        
    #                     series = Series(yvalues, xvalues, title=f"{region}_norm")
    #                     chart2.series.append(series)
                
    #             # Position chart to the right of first chart
    #             chart2.width = 20
    #             chart2.height = 12
    #             ws.add_chart(chart2, f"M{ws.max_row + 3}")
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting and charts. Consolidated Excel file saved: {save_path}")


    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     import numpy as np
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
    #     from openpyxl.utils.dataframe import dataframe_to_rows
    #     from openpyxl.chart import ScatterChart, Reference, Series
    #     from openpyxl.chart.axis import NumericAxis
        
    #     # Helper to calculate window means
    #     def window_mean(t, y, t_start, t_end):
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, data_dict in consolidated.items():
    #             time_series_df = data_dict['time_series']
    #             raw_summary = data_dict.get('raw_summary', {})
    #             norm_summary = data_dict.get('norm_summary', {})
    #             windows = data_dict.get('windows', [])
                
    #             sheet_name = metric_name[:31]
                
    #             # Write time series data starting at A1
    #             time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
    #             # Only add summary if it exists (means files have it, histogram files don't)
    #             if raw_summary:
    #                 # Calculate starting column for summary (after time series + 2 blank columns)
    #                 summary_start_col = len(time_series_df.columns) + 2
                    
    #                 # Get the worksheet to write summary data
    #                 worksheet = writer.sheets[sheet_name]
                    
    #                 # Build raw summary DataFrame
    #                 raw_summary_rows = []
    #                 for root in raw_summary.keys():
    #                     t, y = raw_summary[root]
    #                     row = {'File': root}
    #                     for window_name, t_start, t_end in windows:
    #                         row[window_name] = window_mean(t, y, t_start, t_end)
    #                     raw_summary_rows.append(row)
                    
    #                 if raw_summary_rows:
    #                     raw_summary_df = pd.DataFrame(raw_summary_rows)
                        
    #                     # Write raw summary starting at top right
    #                     for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
    #                         for c_idx, value in enumerate(row):
    #                             worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                    
    #                 # Build normalized summary DataFrame  
    #                 if norm_summary:
    #                     # Start normalized summary after raw summary + 2 blank columns
    #                     norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                        
    #                     norm_summary_rows = []
    #                     for root in norm_summary.keys():
    #                         t, y = norm_summary[root]
    #                         row = {'File': root}
    #                         for window_name, t_start, t_end in windows:
    #                             row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                         norm_summary_rows.append(row)
                        
    #                     if norm_summary_rows:
    #                         norm_summary_df = pd.DataFrame(norm_summary_rows)
                            
    #                         # Write normalized summary to the right of raw summary
    #                         for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
    #                             for c_idx, value in enumerate(row):
    #                                 worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)
                
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Apply bold formatting and add charts
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Bold columns: t, mean, sem, mean_norm, sem_norm, bin_center, and histogram mean/sem
    #         header_row = ws[1]
            
    #         for cell in header_row:
    #             cell_val = str(cell.value) if cell.value else ''
                
    #             # Bold if column name contains 'mean', 'sem', starts with 't' or 'bin_center'
    #             if (cell_val in {'t', 'mean', 'sem', 'mean_norm', 'sem_norm'} or 
    #                 cell_val.startswith('bin_center') or 
    #                 cell_val.startswith('mean_') or 
    #                 cell_val.startswith('sem_')):
                    
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column
    #                 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
            
    #         # Add charts for histogram sheets
    #         if '_histogram' in sheet_name:
    #             regions = ['all', 'baseline', 'stim', 'post']
                
    #             # Chart 1: Raw means overlay
    #             chart1 = ScatterChart()
    #             chart1.title = f"{sheet_name} - Raw Mean Histograms"
    #             chart1.style = 2
                
    #             # Configure axes with tick marks
    #             chart1.x_axis.title = "Bin Center"
    #             chart1.x_axis.majorGridlines = None
    #             chart1.x_axis.tickLblPos = "low"
                
    #             chart1.y_axis.title = "Density"
    #             chart1.y_axis.majorGridlines = None
    #             chart1.y_axis.tickLblPos = "low"
                
    #             for region in regions:
    #                 bin_col = None
    #                 mean_col = None
                    
    #                 for idx, cell in enumerate(header_row, start=1):
    #                     if cell.value == f'bin_center_{region}':
    #                         bin_col = idx
    #                     elif cell.value == f'mean_{region}':
    #                         mean_col = idx
                    
    #                 if bin_col and mean_col:
    #                     xvalues = Reference(ws, min_col=bin_col, min_row=2, max_row=ws.max_row)
    #                     yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                        
    #                     series = Series(yvalues, xvalues, title=region)
    #                     chart1.series.append(series)
                
    #             chart1.width = 20
    #             chart1.height = 12
    #             ws.add_chart(chart1, f"A{ws.max_row + 3}")
                
    #             # Chart 2: Normalized means overlay
    #             chart2 = ScatterChart()
    #             chart2.title = f"{sheet_name} - Normalized Mean Histograms"
    #             chart2.style = 2
                
    #             chart2.x_axis.title = "Bin Center (normalized)"
    #             chart2.x_axis.majorGridlines = None
    #             chart2.x_axis.tickLblPos = "low"
                
    #             chart2.y_axis.title = "Density"
    #             chart2.y_axis.majorGridlines = None
    #             chart2.y_axis.tickLblPos = "low"
                
    #             for region in regions:
    #                 bin_col_norm = None
    #                 mean_col_norm = None
                    
    #                 for idx, cell in enumerate(header_row, start=1):
    #                     if cell.value == f'bin_center_{region}_norm':
    #                         bin_col_norm = idx
    #                     elif cell.value == f'mean_{region}_norm':
    #                         mean_col_norm = idx
                    
    #                 if bin_col_norm and mean_col_norm:
    #                     xvalues = Reference(ws, min_col=bin_col_norm, min_row=2, max_row=ws.max_row)
    #                     yvalues = Reference(ws, min_col=mean_col_norm, min_row=2, max_row=ws.max_row)
                        
    #                     series = Series(yvalues, xvalues, title=f"{region}_norm")
    #                     chart2.series.append(series)
                
    #             chart2.width = 20
    #             chart2.height = 12
    #             ws.add_chart(chart2, f"M{ws.max_row + 3}")
            
    #         # Add charts for time series sheets (not histograms)
    #         elif '_histogram' not in sheet_name:
    #             t_col = None
    #             mean_col = None
    #             mean_norm_col = None
                
    #             # Find t, mean, and mean_norm columns
    #             for idx, cell in enumerate(header_row, start=1):
    #                 if cell.value == 't':
    #                     t_col = idx
    #                 elif cell.value == 'mean':
    #                     mean_col = idx
    #                 elif cell.value == 'mean_norm':
    #                     mean_norm_col = idx
                
    #             # Chart 1: Raw mean vs time
    #             if t_col and mean_col:
    #                 chart1 = ScatterChart()
    #                 chart1.title = f"{sheet_name} - Mean vs Time (Raw)"
    #                 chart1.style = 2
                    
    #                 chart1.x_axis.title = "Time (s)"
    #                 chart1.x_axis.majorGridlines = None
    #                 chart1.x_axis.tickLblPos = "low"
                    
    #                 chart1.y_axis.title = sheet_name
    #                 chart1.y_axis.majorGridlines = None
    #                 chart1.y_axis.tickLblPos = "low"
                    
    #                 xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)
    #                 yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                    
    #                 series = Series(yvalues, xvalues, title="Mean")
    #                 chart1.series.append(series)
                    
    #                 chart1.width = 20
    #                 chart1.height = 12
    #                 ws.add_chart(chart1, f"A{ws.max_row + 3}")
                
    #             # Chart 2: Normalized mean vs time
    #             if t_col and mean_norm_col:
    #                 chart2 = ScatterChart()
    #                 chart2.title = f"{sheet_name} - Mean vs Time (Normalized)"
    #                 chart2.style = 2
                    
    #                 chart2.x_axis.title = "Time (s)"
    #                 chart2.x_axis.majorGridlines = None
    #                 chart2.x_axis.tickLblPos = "low"
                    
    #                 chart2.y_axis.title = f"{sheet_name} (normalized)"
    #                 chart2.y_axis.majorGridlines = None
    #                 chart2.y_axis.tickLblPos = "low"
                    
    #                 xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)
    #                 yvalues = Reference(ws, min_col=mean_norm_col, min_row=2, max_row=ws.max_row)
                    
    #                 series = Series(yvalues, xvalues, title="Mean (normalized)")
    #                 chart2.series.append(series)
                    
    #                 chart2.width = 20
    #                 chart2.height = 12
    #                 ws.add_chart(chart2, f"M{ws.max_row + 3}")
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting and charts. Consolidated Excel file saved: {save_path}")

    # def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
    #     """Save consolidated dataframes to a single Excel file with multiple sheets."""
    #     import pandas as pd
    #     import numpy as np
    #     from openpyxl import load_workbook
    #     from openpyxl.styles import Font
    #     from openpyxl.utils.dataframe import dataframe_to_rows
    #     from openpyxl.chart import ScatterChart, Reference, Series
    #     from openpyxl.chart.marker import Marker
        
    #     # Helper to calculate window means
    #     def window_mean(t, y, t_start, t_end):
    #         mask = (t >= t_start) & (t < t_end)
    #         if mask.sum() == 0:
    #             return np.nan
    #         return np.nanmean(y[mask])
        
    #     with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    #         for metric_name, data_dict in consolidated.items():
    #             time_series_df = data_dict['time_series']
    #             raw_summary = data_dict.get('raw_summary', {})
    #             norm_summary = data_dict.get('norm_summary', {})
    #             windows = data_dict.get('windows', [])
                
    #             sheet_name = metric_name[:31]
                
    #             # Write time series data starting at A1
    #             time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
    #             # Only add summary if it exists (means files have it, histogram files don't)
    #             if raw_summary:
    #                 # Calculate starting column for summary (after time series + 2 blank columns)
    #                 summary_start_col = len(time_series_df.columns) + 2
                    
    #                 # Get the worksheet to write summary data
    #                 worksheet = writer.sheets[sheet_name]
                    
    #                 # Build raw summary DataFrame
    #                 raw_summary_rows = []
    #                 for root in raw_summary.keys():
    #                     t, y = raw_summary[root]
    #                     row = {'File': root}
    #                     for window_name, t_start, t_end in windows:
    #                         row[window_name] = window_mean(t, y, t_start, t_end)
    #                     raw_summary_rows.append(row)
                    
    #                 if raw_summary_rows:
    #                     raw_summary_df = pd.DataFrame(raw_summary_rows)
                        
    #                     # Write raw summary starting at top right
    #                     for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
    #                         for c_idx, value in enumerate(row):
    #                             worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                    
    #                 # Build normalized summary DataFrame  
    #                 if norm_summary:
    #                     # Start normalized summary after raw summary + 2 blank columns
    #                     norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                        
    #                     norm_summary_rows = []
    #                     for root in norm_summary.keys():
    #                         t, y = norm_summary[root]
    #                         row = {'File': root}
    #                         for window_name, t_start, t_end in windows:
    #                             row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
    #                         norm_summary_rows.append(row)
                        
    #                     if norm_summary_rows:
    #                         norm_summary_df = pd.DataFrame(norm_summary_rows)
                            
    #                         # Write normalized summary to the right of raw summary
    #                         for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
    #                             for c_idx, value in enumerate(row):
    #                                 worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)
                
    #             print(f"Saved sheet: {sheet_name}")
        
    #     # Apply bold formatting and add charts
    #     wb = load_workbook(save_path)
    #     bold_font = Font(bold=True)
        
    #     for sheet_name in wb.sheetnames:
    #         ws = wb[sheet_name]
            
    #         # Bold columns: t, mean, sem, mean_norm, sem_norm, bin_center, and histogram mean/sem
    #         header_row = ws[1]
            
    #         for cell in header_row:
    #             cell_val = str(cell.value) if cell.value else ''
                
    #             # Bold if column name contains 'mean', 'sem', starts with 't' or 'bin_center'
    #             if (cell_val in {'t', 'mean', 'sem', 'mean_norm', 'sem_norm'} or 
    #                 cell_val.startswith('bin_center') or 
    #                 cell_val.startswith('mean_') or 
    #                 cell_val.startswith('sem_')):
                    
    #                 col_letter = cell.column_letter
    #                 # Bold the entire column
    #                 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
    #                                     min_col=cell.column, max_col=cell.column):
    #                     for c in row:
    #                         c.font = bold_font
            
    #         # Add charts for histogram sheets
    #         if '_histogram' in sheet_name:
    #             regions = ['all', 'baseline', 'stim', 'post']
                
    #             # Chart 1: Raw means overlay
    #             chart1 = ScatterChart()
    #             chart1.title = f"{sheet_name} - Raw Mean Histograms"
    #             chart1.style = 13
                
    #             # Configure axes
    #             chart1.x_axis.title = "Bin Center"
    #             chart1.x_axis.tickLblPos = "low"
    #             chart1.x_axis.majorTickMark = "out"
                
    #             chart1.y_axis.title = "Density"
    #             chart1.y_axis.tickLblPos = "low"
    #             chart1.y_axis.majorTickMark = "out"
                
    #             # Remove legend if not needed
    #             chart1.legend.position = 'r'
                
    #             for region in regions:
    #                 bin_col = None
    #                 mean_col = None
                    
    #                 for idx, cell in enumerate(header_row, start=1):
    #                     if cell.value == f'bin_center_{region}':
    #                         bin_col = idx
    #                     elif cell.value == f'mean_{region}':
    #                         mean_col = idx
                    
    #                 if bin_col and mean_col:
    #                     xvalues = Reference(ws, min_col=bin_col, min_row=2, max_row=ws.max_row)
    #                     yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                        
    #                     series = Series(yvalues, xvalues, title=region)
    #                     series.marker = Marker('none')  # No markers, just lines
    #                     series.smooth = True
    #                     chart1.series.append(series)
                
    #             chart1.width = 20
    #             chart1.height = 12
    #             ws.add_chart(chart1, f"A{ws.max_row + 3}")
                
    #             # Chart 2: Normalized means overlay
    #             chart2 = ScatterChart()
    #             chart2.title = f"{sheet_name} - Normalized Mean Histograms"
    #             chart2.style = 13
                
    #             chart2.x_axis.title = "Bin Center (normalized)"
    #             chart2.x_axis.tickLblPos = "low"
    #             chart2.x_axis.majorTickMark = "out"
                
    #             chart2.y_axis.title = "Density"
    #             chart2.y_axis.tickLblPos = "low"
    #             chart2.y_axis.majorTickMark = "out"
                
    #             chart2.legend.position = 'r'
                
    #             for region in regions:
    #                 bin_col_norm = None
    #                 mean_col_norm = None
                    
    #                 for idx, cell in enumerate(header_row, start=1):
    #                     if cell.value == f'bin_center_{region}_norm':
    #                         bin_col_norm = idx
    #                     elif cell.value == f'mean_{region}_norm':
    #                         mean_col_norm = idx
                    
    #                 if bin_col_norm and mean_col_norm:
    #                     xvalues = Reference(ws, min_col=bin_col_norm, min_row=2, max_row=ws.max_row)
    #                     yvalues = Reference(ws, min_col=mean_col_norm, min_row=2, max_row=ws.max_row)
                        
    #                     series = Series(yvalues, xvalues, title=f"{region}_norm")
    #                     series.marker = Marker('none')
    #                     series.smooth = True
    #                     chart2.series.append(series)
                
    #             chart2.width = 20
    #             chart2.height = 12
    #             ws.add_chart(chart2, f"M{ws.max_row + 3}")
            
    #         # Add charts for time series sheets (not histograms)
    #         elif '_histogram' not in sheet_name:
    #             t_col = None
    #             mean_col = None
    #             mean_norm_col = None
                
    #             # Find t, mean, and mean_norm columns
    #             for idx, cell in enumerate(header_row, start=1):
    #                 if cell.value == 't':
    #                     t_col = idx
    #                 elif cell.value == 'mean':
    #                     mean_col = idx
    #                 elif cell.value == 'mean_norm':
    #                     mean_norm_col = idx
                
    #             # Chart 1: Raw mean vs time
    #             if t_col and mean_col:
    #                 chart1 = ScatterChart()
    #                 chart1.title = f"{sheet_name} - Mean vs Time (Raw)"
    #                 chart1.style = 13
                    
    #                 chart1.x_axis.title = "Time (s)"
    #                 chart1.x_axis.tickLblPos = "low"
    #                 chart1.x_axis.majorTickMark = "out"
                    
    #                 chart1.y_axis.title = sheet_name
    #                 chart1.y_axis.tickLblPos = "low"
    #                 chart1.y_axis.majorTickMark = "out"
                    
    #                 # Hide legend
    #                 chart1.legend = None
                    
    #                 xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)
    #                 yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                    
    #                 series = Series(yvalues, xvalues, title="Mean")
    #                 series.marker = Marker('none')  # No markers, just line
    #                 series.smooth = True
    #                 series.graphicalProperties.line.solidFill = "4472C4"  # Solid blue line
    #                 chart1.series.append(series)
                    
    #                 chart1.width = 20
    #                 chart1.height = 12
    #                 ws.add_chart(chart1, f"A{ws.max_row + 3}")
                
    #             # Chart 2: Normalized mean vs time
    #             if t_col and mean_norm_col:
    #                 chart2 = ScatterChart()
    #                 chart2.title = f"{sheet_name} - Mean vs Time (Normalized)"
    #                 chart2.style = 13
                    
    #                 chart2.x_axis.title = "Time (s)"
    #                 chart2.x_axis.tickLblPos = "low"
    #                 chart2.x_axis.majorTickMark = "out"
                    
    #                 chart2.y_axis.title = f"{sheet_name} (normalized)"
    #                 chart2.y_axis.tickLblPos = "low"
    #                 chart2.y_axis.majorTickMark = "out"
                    
    #                 # Hide legend
    #                 chart2.legend = None
                    
    #                 xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)
    #                 yvalues = Reference(ws, min_col=mean_norm_col, min_row=2, max_row=ws.max_row)
                    
    #                 series = Series(yvalues, xvalues, title="Mean (normalized)")
    #                 series.marker = Marker('none')
    #                 series.smooth = True
    #                 series.graphicalProperties.line.solidFill = "ED7D31"  # Solid orange line
    #                 chart2.series.append(series)
                    
    #                 chart2.width = 20
    #                 chart2.height = 12
    #                 ws.add_chart(chart2, f"M{ws.max_row + 3}")
        
    #     wb.save(save_path)
    #     print(f"Applied bold formatting and charts. Consolidated Excel file saved: {save_path}")

    def _save_consolidated_to_excel(self, consolidated: dict, save_path: Path):
        """Save consolidated dataframes to a single Excel file with multiple sheets."""
        import pandas as pd
        import numpy as np
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.chart.marker import Marker  # Fixed typo here
        
        # Helper to calculate window means
        def window_mean(t, y, t_start, t_end):
            mask = (t >= t_start) & (t < t_end)
            if mask.sum() == 0:
                return np.nan
            return np.nanmean(y[mask])
        
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            # Define sheet order: time series metrics first, then sighs, then events/stimulus, then histograms
            time_series_metrics = []
            sighs_sheets = []
            event_sheets = []
            histogram_sheets = []

            for metric_name in consolidated.keys():
                if metric_name == '_warnings':
                    continue
                elif metric_name == 'sighs':
                    sighs_sheets.append(metric_name)
                elif metric_name in ['events', 'stimulus']:
                    event_sheets.append(metric_name)
                elif '_histogram' in metric_name:
                    histogram_sheets.append(metric_name)
                else:
                    time_series_metrics.append(metric_name)

            # Process sheets in desired order: time series -> sighs -> events/stimulus -> histograms
            ordered_sheets = time_series_metrics + sighs_sheets + event_sheets + histogram_sheets

            for metric_name in ordered_sheets:
                data_dict = consolidated[metric_name]
                time_series_df = data_dict['time_series']
                raw_summary = data_dict.get('raw_summary', {})
                norm_summary = data_dict.get('norm_summary', {})
                eupnea_summary = data_dict.get('eupnea_summary', {})
                windows = data_dict.get('windows', [])
                
                sheet_name = metric_name[:31]
                
                # Write time series data starting at A1
                time_series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                
                # Only add summary if it exists (means files have it, histogram files don't)
                if raw_summary:
                    # Calculate starting column for summary (after time series + 2 blank columns)
                    summary_start_col = len(time_series_df.columns) + 2
                    
                    # Get the worksheet to write summary data
                    worksheet = writer.sheets[sheet_name]
                    
                    # Build raw summary DataFrame
                    raw_summary_rows = []
                    for root in raw_summary.keys():
                        t, y = raw_summary[root]
                        row = {'File': root}
                        for window_name, t_start, t_end in windows:
                            row[window_name] = window_mean(t, y, t_start, t_end)
                        raw_summary_rows.append(row)
                    
                    if raw_summary_rows:
                        raw_summary_df = pd.DataFrame(raw_summary_rows)
                        
                        # Write raw summary starting at top right
                        for r_idx, row in enumerate(dataframe_to_rows(raw_summary_df, index=False, header=True)):
                            for c_idx, value in enumerate(row):
                                worksheet.cell(row=r_idx + 1, column=summary_start_col + c_idx, value=value)
                    
                    # Build normalized summary DataFrame  
                    if norm_summary:
                        # Start normalized summary after raw summary + 2 blank columns
                        norm_start_col = summary_start_col + len(raw_summary_df.columns) + 2
                        
                        norm_summary_rows = []
                        for root in norm_summary.keys():
                            t, y = norm_summary[root]
                            row = {'File': root}
                            for window_name, t_start, t_end in windows:
                                row[f"{window_name}_norm"] = window_mean(t, y, t_start, t_end)
                            norm_summary_rows.append(row)
                        
                        if norm_summary_rows:
                            norm_summary_df = pd.DataFrame(norm_summary_rows)

                            # Write normalized summary to the right of raw summary
                            for r_idx, row in enumerate(dataframe_to_rows(norm_summary_df, index=False, header=True)):
                                for c_idx, value in enumerate(row):
                                    worksheet.cell(row=r_idx + 1, column=norm_start_col + c_idx, value=value)

                    # Build eupnea-normalized summary DataFrame
                    if eupnea_summary:
                        # Start eupnea summary after normalized summary + 2 blank columns
                        eupnea_start_col = norm_start_col + len(norm_summary_df.columns) + 2 if norm_summary else summary_start_col

                        eupnea_summary_rows = []
                        for root in eupnea_summary.keys():
                            t, y = eupnea_summary[root]
                            row = {'File': root}
                            for window_name, t_start, t_end in windows:
                                row[f"{window_name}_eupnea"] = window_mean(t, y, t_start, t_end)
                            eupnea_summary_rows.append(row)

                        if eupnea_summary_rows:
                            eupnea_summary_df = pd.DataFrame(eupnea_summary_rows)

                            # Write eupnea summary to the right of normalized summary
                            for r_idx, row in enumerate(dataframe_to_rows(eupnea_summary_df, index=False, header=True)):
                                for c_idx, value in enumerate(row):
                                    worksheet.cell(row=r_idx + 1, column=eupnea_start_col + c_idx, value=value)

                print(f"Saved sheet: {sheet_name}")
        
        # Apply bold formatting and add charts
        wb = load_workbook(save_path)
        bold_font = Font(bold=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Processing sheet: '{sheet_name}'")

            # Bold header row only (much faster than bolding entire columns)
            header_row = ws[1]

            for cell in header_row:
                cell.font = bold_font
            
            # Add charts based on sheet type
            if sheet_name == 'sighs':
                print(f"Matched sighs sheet condition for '{sheet_name}'")
                # Create sigh timeline chart
                self._add_sighs_chart(ws, header_row)

            elif sheet_name == 'events':
                print(f"Matched events sheet condition for '{sheet_name}'")
                # Create eupnea and apnea timeline charts
                self._add_events_charts(ws, header_row)

            # Stimulus sheet - no charts needed (just tabular data)
            elif sheet_name == 'stimulus':
                pass  # No charts for stimulus

            # Add charts for histogram sheets
            elif '_histogram' in sheet_name:
                regions = ['all', 'baseline', 'stim', 'post']
                
                # Chart 1: Raw means overlay (reverted to original style)
                chart1 = ScatterChart()
                chart1.title = f"{sheet_name} - Raw Mean Histograms"
                chart1.style = 2
                chart1.x_axis.title = "Bin Center"
                chart1.y_axis.title = "Density"

                # Enable axes display
                chart1.x_axis.delete = False
                chart1.y_axis.delete = False

                # Enable axis tick marks and labels (major only, no minor)
                chart1.x_axis.tickLblPos = "nextTo"  # Changed from "low" to fix label positioning
                chart1.y_axis.tickLblPos = "nextTo"
                chart1.x_axis.majorTickMark = "out"
                chart1.y_axis.majorTickMark = "out"
                chart1.x_axis.minorTickMark = "none"
                chart1.y_axis.minorTickMark = "none"

                # Disable gridlines
                chart1.x_axis.majorGridlines = None
                chart1.y_axis.majorGridlines = None

                # Set y-axis to start at 0
                chart1.y_axis.scaling.min = 0
                
                for region in regions:
                    bin_col = None
                    mean_col = None
                    
                    for idx, cell in enumerate(header_row, start=1):
                        if cell.value == f'bin_center_{region}':
                            bin_col = idx
                        elif cell.value == f'mean_{region}':
                            mean_col = idx
                    
                    if bin_col and mean_col:
                        xvalues = Reference(ws, min_col=bin_col, min_row=2, max_row=ws.max_row)
                        yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                        
                        series = Series(yvalues, xvalues, title=region)
                        chart1.series.append(series)
                
                chart1.width = 10
                chart1.height = 6
                ws.add_chart(chart1, f"A{ws.max_row + 3}")

                # Chart 2: Time-normalized means overlay
                chart2 = ScatterChart()
                chart2.title = f"{sheet_name} - Time-Normalized Mean Histograms"
                chart2.style = 2
                chart2.x_axis.title = "Bin Center (time-normalized)"
                chart2.y_axis.title = "Density"

                # Enable axes display
                chart2.x_axis.delete = False
                chart2.y_axis.delete = False

                # Enable axis tick marks and labels (major only, no minor)
                chart2.x_axis.tickLblPos = "nextTo"
                chart2.y_axis.tickLblPos = "nextTo"
                chart2.x_axis.majorTickMark = "out"
                chart2.y_axis.majorTickMark = "out"
                chart2.x_axis.minorTickMark = "none"
                chart2.y_axis.minorTickMark = "none"

                # Disable gridlines
                chart2.x_axis.majorGridlines = None
                chart2.y_axis.majorGridlines = None

                # Set y-axis to start at 0
                chart2.y_axis.scaling.min = 0

                for region in regions:
                    bin_col_norm = None
                    mean_col_norm = None

                    for idx, cell in enumerate(header_row, start=1):
                        if cell.value == f'bin_center_{region}_norm':
                            bin_col_norm = idx
                        elif cell.value == f'mean_{region}_norm':
                            mean_col_norm = idx

                    if bin_col_norm and mean_col_norm:
                        xvalues = Reference(ws, min_col=bin_col_norm, min_row=2, max_row=ws.max_row)
                        yvalues = Reference(ws, min_col=mean_col_norm, min_row=2, max_row=ws.max_row)

                        series = Series(yvalues, xvalues, title=f"{region}_norm")
                        chart2.series.append(series)

                chart2.width = 10
                chart2.height = 6
                ws.add_chart(chart2, f"K{ws.max_row + 3}")

                # Chart 3: Eupnea-normalized means overlay
                chart3 = ScatterChart()
                chart3.title = f"{sheet_name} - Eupnea-Normalized Mean Histograms"
                chart3.style = 2
                chart3.x_axis.title = "Bin Center (eupnea-normalized)"
                chart3.y_axis.title = "Density"

                # Enable axes display
                chart3.x_axis.delete = False
                chart3.y_axis.delete = False

                # Enable axis tick marks and labels (major only, no minor)
                chart3.x_axis.tickLblPos = "nextTo"
                chart3.y_axis.tickLblPos = "nextTo"
                chart3.x_axis.majorTickMark = "out"
                chart3.y_axis.majorTickMark = "out"
                chart3.x_axis.minorTickMark = "none"
                chart3.y_axis.minorTickMark = "none"

                # Disable gridlines
                chart3.x_axis.majorGridlines = None
                chart3.y_axis.majorGridlines = None

                # Set y-axis to start at 0
                chart3.y_axis.scaling.min = 0

                for region in regions:
                    bin_col_eupnea = None
                    mean_col_eupnea = None

                    for idx, cell in enumerate(header_row, start=1):
                        if cell.value == f'bin_center_{region}_eupnea':
                            bin_col_eupnea = idx
                        elif cell.value == f'mean_{region}_eupnea':
                            mean_col_eupnea = idx

                    if bin_col_eupnea and mean_col_eupnea:
                        xvalues = Reference(ws, min_col=bin_col_eupnea, min_row=2, max_row=ws.max_row)
                        yvalues = Reference(ws, min_col=mean_col_eupnea, min_row=2, max_row=ws.max_row)

                        series = Series(yvalues, xvalues, title=f"{region}_eupnea")
                        chart3.series.append(series)

                chart3.width = 10
                chart3.height = 6
                ws.add_chart(chart3, f"U{ws.max_row + 3}")  # Position to the right of chart2

            # Add charts for time series sheets (not histograms)
            elif '_histogram' not in sheet_name:
                t_col = None
                mean_col = None
                sem_col = None
                mean_norm_col = None
                sem_norm_col = None
                mean_eupnea_col = None
                sem_eupnea_col = None

                # Find t, mean, sem, mean_norm, sem_norm, mean_eupnea, and sem_eupnea columns
                for idx, cell in enumerate(header_row, start=1):
                    if cell.value == 't':
                        t_col = idx
                    elif cell.value == 'mean':
                        mean_col = idx
                    elif cell.value == 'sem':
                        sem_col = idx
                    elif cell.value == 'mean_norm':
                        mean_norm_col = idx
                    elif cell.value == 'sem_norm':
                        sem_norm_col = idx
                    elif cell.value == 'mean_eupnea':
                        mean_eupnea_col = idx
                    elif cell.value == 'sem_eupnea':
                        sem_eupnea_col = idx
                
                # Position charts near top of sheet (row 5)
                chart_row = 5

                # Find individual file columns for plotting
                # Raw data columns: between 'mean' and first '_norm' column
                raw_file_cols = []
                norm_file_cols = []
                eupnea_file_cols = []

                for idx, cell in enumerate(header_row, start=1):
                    col_name = str(cell.value or '')
                    # Skip time, mean, sem, and blank columns
                    if col_name in ['t', 'mean', 'sem', 'mean_norm', 'sem_norm', 'mean_eupnea', 'sem_eupnea', '', 'None']:
                        continue
                    # Skip columns with numeric or empty names
                    if not col_name or col_name.isspace():
                        continue
                    if '_eupnea' in col_name:
                        eupnea_file_cols.append(idx)
                    elif '_norm' in col_name:
                        norm_file_cols.append(idx)
                    else:
                        # Must be a raw data file column
                        if t_col and idx > t_col:  # Make sure it's after the time column
                            raw_file_cols.append(idx)

                # Chart 1: Raw mean vs time
                if t_col and mean_col:
                    chart1 = ScatterChart()
                    chart1.title = f"{sheet_name} - Mean vs Time (Raw)"
                    chart1.style = 13

                    chart1.x_axis.title = "Time (s)"
                    chart1.y_axis.title = sheet_name

                    # Enable axes display
                    chart1.x_axis.delete = False
                    chart1.y_axis.delete = False

                    # Enable axis tick marks and labels (major only, no minor)
                    chart1.x_axis.tickLblPos = "nextTo"
                    chart1.y_axis.tickLblPos = "nextTo"
                    chart1.x_axis.majorTickMark = "out"
                    chart1.y_axis.majorTickMark = "out"
                    chart1.x_axis.minorTickMark = "none"
                    chart1.y_axis.minorTickMark = "none"

                    # Position y-axis on the left side
                    chart1.y_axis.crosses = "min"

                    # Disable gridlines
                    chart1.x_axis.majorGridlines = None
                    chart1.y_axis.majorGridlines = None

                    # Hide legend
                    chart1.legend = None

                    xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)

                    # Add individual file traces in light gray
                    for file_col in raw_file_cols:
                        yvalues_file = Reference(ws, min_col=file_col, min_row=2, max_row=ws.max_row)
                        series_file = Series(yvalues_file, xvalues, title="")
                        series_file.marker = Marker('none')
                        series_file.smooth = True
                        series_file.graphicalProperties.line.solidFill = "D3D3D3"  # Light gray
                        series_file.graphicalProperties.line.width = 8000  # Thinner than mean
                        chart1.series.append(series_file)

                    # Add mean line on top
                    yvalues = Reference(ws, min_col=mean_col, min_row=2, max_row=ws.max_row)
                    series = Series(yvalues, xvalues, title="Mean")
                    series.marker = Marker('none')
                    series.smooth = True
                    series.graphicalProperties.line.solidFill = "4472C4"  # Solid blue line
                    series.graphicalProperties.line.width = 12700  # 1pt line width
                    chart1.series.append(series)

                    chart1.width = 10
                    chart1.height = 6
                    ws.add_chart(chart1, f"A{chart_row}")
                
                # Chart 2: Time-normalized mean vs time
                if t_col and mean_norm_col:
                    chart2 = ScatterChart()
                    chart2.title = f"{sheet_name} - Mean vs Time (Time-Normalized)"
                    chart2.style = 13

                    chart2.x_axis.title = "Time (s)"
                    chart2.y_axis.title = f"{sheet_name} (time-normalized)"

                    # Enable axes display
                    chart2.x_axis.delete = False
                    chart2.y_axis.delete = False

                    # Enable axis tick marks and labels (major only, no minor)
                    chart2.x_axis.tickLblPos = "nextTo"
                    chart2.y_axis.tickLblPos = "nextTo"
                    chart2.x_axis.majorTickMark = "out"
                    chart2.y_axis.majorTickMark = "out"
                    chart2.x_axis.minorTickMark = "none"
                    chart2.y_axis.minorTickMark = "none"

                    # Position y-axis on the left side
                    chart2.y_axis.crosses = "min"

                    # Disable gridlines
                    chart2.x_axis.majorGridlines = None
                    chart2.y_axis.majorGridlines = None

                    # Hide legend
                    chart2.legend = None

                    xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)

                    # Add individual file traces in light gray
                    for file_col in norm_file_cols:
                        yvalues_file = Reference(ws, min_col=file_col, min_row=2, max_row=ws.max_row)
                        series_file = Series(yvalues_file, xvalues, title="")
                        series_file.marker = Marker('none')
                        series_file.smooth = True
                        series_file.graphicalProperties.line.solidFill = "D3D3D3"  # Light gray
                        series_file.graphicalProperties.line.width = 8000  # Thinner than mean
                        chart2.series.append(series_file)

                    # Add mean line on top
                    yvalues = Reference(ws, min_col=mean_norm_col, min_row=2, max_row=ws.max_row)
                    series = Series(yvalues, xvalues, title="Mean")
                    series.marker = Marker('none')
                    series.smooth = True
                    series.graphicalProperties.line.solidFill = "ED7D31"  # Solid orange line
                    series.graphicalProperties.line.width = 12700  # 1pt line width
                    chart2.series.append(series)

                    chart2.width = 10
                    chart2.height = 6
                    ws.add_chart(chart2, f"K{chart_row}")

                # Chart 3: Eupnea-normalized mean vs time
                if t_col and mean_eupnea_col:
                    chart3 = ScatterChart()
                    chart3.title = f"{sheet_name} - Mean vs Time (Eupnea-Normalized)"
                    chart3.style = 13

                    chart3.x_axis.title = "Time (s)"
                    chart3.y_axis.title = f"{sheet_name} (eupnea-normalized)"

                    # Enable axes display
                    chart3.x_axis.delete = False
                    chart3.y_axis.delete = False

                    # Enable axis tick marks and labels (major only, no minor)
                    chart3.x_axis.tickLblPos = "nextTo"
                    chart3.y_axis.tickLblPos = "nextTo"
                    chart3.x_axis.majorTickMark = "out"
                    chart3.y_axis.majorTickMark = "out"
                    chart3.x_axis.minorTickMark = "none"
                    chart3.y_axis.minorTickMark = "none"

                    # Position y-axis on the left side
                    chart3.y_axis.crosses = "min"

                    # Disable gridlines
                    chart3.x_axis.majorGridlines = None
                    chart3.y_axis.majorGridlines = None

                    # Hide legend
                    chart3.legend = None

                    xvalues = Reference(ws, min_col=t_col, min_row=2, max_row=ws.max_row)

                    # Add individual file traces in light gray
                    for file_col in eupnea_file_cols:
                        yvalues_file = Reference(ws, min_col=file_col, min_row=2, max_row=ws.max_row)
                        series_file = Series(yvalues_file, xvalues, title="")
                        series_file.marker = Marker('none')
                        series_file.smooth = True
                        series_file.graphicalProperties.line.solidFill = "D3D3D3"  # Light gray
                        series_file.graphicalProperties.line.width = 8000  # Thinner than mean
                        chart3.series.append(series_file)

                    # Add mean line on top
                    yvalues = Reference(ws, min_col=mean_eupnea_col, min_row=2, max_row=ws.max_row)
                    series = Series(yvalues, xvalues, title="Mean")
                    series.marker = Marker('none')
                    series.smooth = True
                    series.graphicalProperties.line.solidFill = "70AD47"  # Solid green line
                    series.graphicalProperties.line.width = 12700  # 1pt line width
                    chart3.series.append(series)

                    chart3.width = 10
                    chart3.height = 6
                    ws.add_chart(chart3, f"U{chart_row}")  # Position to the right of chart2

        wb.save(save_path)
        print(f"Applied bold formatting and charts. Consolidated Excel file saved: {save_path}")


    def _add_events_charts(self, ws, header_row):
        """Add eupnea, apnea, and sniffing timeline charts to the events sheet."""
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.chart.marker import Marker
        import matplotlib.pyplot as plt
        import matplotlib.patches as patches
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        import io
        import numpy as np

        print(f"_add_events_charts called for sheet with {ws.max_row} rows")

        # Find required columns (handle both naming conventions)
        exp_num_col = None
        sweep_col = None
        global_sweep_col = None
        event_type_col = None
        t_start_col = None
        t_end_col = None

        for idx, cell in enumerate(header_row, start=1):
            if cell.value == 'experiment_number':
                exp_num_col = idx
            elif cell.value == 'sweep':
                sweep_col = idx
            elif cell.value == 'global_sweep_number':
                global_sweep_col = idx
            elif cell.value == 'event_type':
                event_type_col = idx
            elif cell.value in ['t_start', 'start_time']:
                t_start_col = idx
            elif cell.value in ['t_end', 'end_time']:
                t_end_col = idx

        if not all([exp_num_col, global_sweep_col, event_type_col, t_start_col, t_end_col]):
            print(f"Events sheet missing required columns for charts. Found columns: {[cell.value for cell in header_row]}")
            return

        # Read data from sheet
        events = []
        for row_idx in range(2, ws.max_row + 1):
            exp_num = ws.cell(row=row_idx, column=exp_num_col).value
            global_sweep = ws.cell(row=row_idx, column=global_sweep_col).value
            event_type = ws.cell(row=row_idx, column=event_type_col).value
            t_start = ws.cell(row=row_idx, column=t_start_col).value
            t_end = ws.cell(row=row_idx, column=t_end_col).value

            if all([exp_num is not None, global_sweep is not None,
                    event_type is not None, t_start is not None, t_end is not None]):
                events.append({
                    'exp_num': int(exp_num),
                    'global_sweep': int(global_sweep),
                    'event_type': str(event_type).lower(),
                    't_start': float(t_start),
                    't_end': float(t_end)
                })

        if not events:
            print("No events found for chart generation")
            return

        # Separate eupnea, apnea, and sniffing events
        eupnea_events = [e for e in events if 'eupnea' in e['event_type']]
        apnea_events = [e for e in events if 'apnea' in e['event_type']]
        sniffing_events = [e for e in events if 'sniff' in e['event_type']]

        if not eupnea_events and not apnea_events and not sniffing_events:
            print("No eupnea, apnea, or sniffing events found for charts")
            return

        # Create colormap for experiments
        n_experiments = max([e['exp_num'] for e in events])
        colors = plt.cm.tab10(np.linspace(0, 1, n_experiments))

        # Position charts at the top (starting at row 1)
        # We'll place them to the right of the data columns
        chart_start_row = 1

        # Create eupnea chart
        if eupnea_events:
            fig1, ax1 = plt.subplots(figsize=(10, 6))
            for event in eupnea_events:
                color = colors[event['exp_num'] - 1]
                ax1.plot([event['t_start'], event['t_end']],
                        [event['global_sweep'], event['global_sweep']],
                        color=color, linewidth=2, solid_capstyle='butt')

            ax1.set_xlabel('Time (s)')
            ax1.set_ylabel('Global Sweep Number')
            ax1.set_title('Eupnea Periods Across Experiments')
            ax1.grid(True, alpha=0.3)

            # Save figure to bytes
            buf1 = io.BytesIO()
            fig1.savefig(buf1, format='png', dpi=100, bbox_inches='tight')
            buf1.seek(0)
            plt.close(fig1)

            # Insert image into Excel
            img1 = XLImage(buf1)
            img1.width = 600
            img1.height = 360
            ws.add_image(img1, f"A{chart_start_row}")
            print(f"Added eupnea chart at A{chart_start_row}")

        # Create apnea chart
        if apnea_events:
            fig2, ax2 = plt.subplots(figsize=(10, 6))
            for event in apnea_events:
                color = colors[event['exp_num'] - 1]
                ax2.plot([event['t_start'], event['t_end']],
                        [event['global_sweep'], event['global_sweep']],
                        color=color, linewidth=2, solid_capstyle='butt')

            ax2.set_xlabel('Time (s)')
            ax2.set_ylabel('Global Sweep Number')
            ax2.set_title('Apnea Periods Across Experiments')
            ax2.grid(True, alpha=0.3)

            # Save figure to bytes
            buf2 = io.BytesIO()
            fig2.savefig(buf2, format='png', dpi=100, bbox_inches='tight')
            buf2.seek(0)
            plt.close(fig2)

            # Insert image into Excel
            img2 = XLImage(buf2)
            img2.width = 600
            img2.height = 360
            # Position to the right of first chart (column K)
            ws.add_image(img2, f"K{chart_start_row}")
            print(f"Added apnea chart at K{chart_start_row}")

        # Create sniffing chart
        if sniffing_events:
            fig3, ax3 = plt.subplots(figsize=(10, 6))
            for event in sniffing_events:
                color = colors[event['exp_num'] - 1]
                ax3.plot([event['t_start'], event['t_end']],
                        [event['global_sweep'], event['global_sweep']],
                        color=color, linewidth=2, solid_capstyle='butt')

            ax3.set_xlabel('Time (s)')
            ax3.set_ylabel('Global Sweep Number')
            ax3.set_title('Sniffing Bouts Across Experiments')
            ax3.grid(True, alpha=0.3)

            # Save figure to bytes
            buf3 = io.BytesIO()
            fig3.savefig(buf3, format='png', dpi=100, bbox_inches='tight')
            buf3.seek(0)
            plt.close(fig3)

            # Insert image into Excel
            img3 = XLImage(buf3)
            img3.width = 600
            img3.height = 360
            # Position to the right of apnea chart (column U)
            ws.add_image(img3, f"U{chart_start_row}")
            print(f"Added sniffing chart at U{chart_start_row}")

    def _add_sighs_chart(self, ws, header_row):
        """Add sigh timeline scatter plot to the sighs sheet."""
        import matplotlib.pyplot as plt
        from openpyxl.drawing.image import Image as XLImage
        import io
        import numpy as np

        print(f"_add_sighs_chart called for sheet with {ws.max_row} rows")

        # Find required columns
        exp_num_col = None
        global_sweep_col = None
        t_col = None

        for idx, cell in enumerate(header_row, start=1):
            if cell.value == 'experiment_number':
                exp_num_col = idx
            elif cell.value == 'global_sweep_number':
                global_sweep_col = idx
            elif cell.value == 't':
                t_col = idx

        if not all([exp_num_col, global_sweep_col, t_col]):
            print(f"Sighs sheet missing required columns for chart. Found columns: {[cell.value for cell in header_row]}")
            return

        # Read data from sheet
        sighs = []
        for row_idx in range(2, ws.max_row + 1):
            exp_num = ws.cell(row=row_idx, column=exp_num_col).value
            global_sweep = ws.cell(row=row_idx, column=global_sweep_col).value
            t = ws.cell(row=row_idx, column=t_col).value

            if all([exp_num is not None, global_sweep is not None, t is not None]):
                sighs.append({
                    'exp_num': int(exp_num),
                    'global_sweep': int(global_sweep),
                    't': float(t)
                })

        if not sighs:
            print("No sigh data found for chart generation")
            return

        # Create colormap for experiments
        n_experiments = max([s['exp_num'] for s in sighs])
        colors = plt.cm.tab10(np.linspace(0, 1, n_experiments))

        # Position chart at the top (starting at row 1)
        chart_start_row = 1

        # Create sigh scatter plot
        fig, ax = plt.subplots(figsize=(10, 6))

        # Plot each sigh as a yellow asterisk, colored by experiment
        for sigh in sighs:
            color = colors[sigh['exp_num'] - 1]
            ax.scatter(sigh['t'], sigh['global_sweep'],
                      marker='*', s=200, color='gold', edgecolors=color, linewidths=1.5)

        ax.set_xlabel('Time (s)')
        ax.set_ylabel('Global Sweep Number')
        ax.set_title('Sigh Events Across Experiments')
        ax.grid(True, alpha=0.3)

        # Save figure to bytes
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        plt.close(fig)

        # Insert image into Excel
        img = XLImage(buf)
        img.width = 600
        img.height = 360
        ws.add_image(img, f"A{chart_start_row}")
        print(f"Added sigh chart at A{chart_start_row}")

if __name__ == "__main__":
    from PyQt6.QtWidgets import QSplashScreen, QProgressBar, QVBoxLayout, QLabel, QWidget
    from PyQt6.QtGui import QPixmap
    from PyQt6.QtCore import Qt, QTimer

    app = QApplication(sys.argv)

    # Create splash screen
    # Try to load icon (with fallback path handling)
    splash_paths = [
        Path(__file__).parent / "images" / "plethapp_splash_dark-01.png",
        Path(__file__).parent / "images" / "plethapp_splash.png",
        Path(__file__).parent / "images" / "plethapp_thumbnail_dark_round.ico",
        Path(__file__).parent / "assets" / "plethapp_thumbnail_dark_round.ico",
    ]

    splash_pix = None
    for splash_path in splash_paths:
        if splash_path.exists():
            splash_pix = QPixmap(str(splash_path))
            break

    if splash_pix is None or splash_pix.isNull():
        # Fallback: create simple splash with text
        splash_pix = QPixmap(200, 150)
        splash_pix.fill(Qt.GlobalColor.darkGray)

    # Scale to smaller size for faster display
    splash_pix = splash_pix.scaled(150, 150, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)

    splash = QSplashScreen(splash_pix, Qt.WindowType.WindowStaysOnTopHint)
    splash.setWindowFlags(Qt.WindowType.WindowStaysOnTopHint | Qt.WindowType.FramelessWindowHint)

    # Add loading message
    splash.showMessage(
        "Loading PlethApp...",
        Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignCenter,
        Qt.GlobalColor.white
    )
    splash.show()
    app.processEvents()

    # Create main window (this is where the loading time happens)
    w = MainWindow()

    # Close splash and show main window
    splash.finish(w)
    w.show()

    sys.exit(app.exec())
